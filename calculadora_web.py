# -*- coding: utf-8 -*-
"""
Calculadora de Lucro Shopee - Versao Web (porta 5050)
Converte a calculadora_v4.py desktop para Flask web app.
Processa XMLs de NF-e + planilha de custos e gera relatorio Excel.
"""

import os
import io
import uuid
import tempfile
import zipfile
from datetime import datetime

import pandas as pd
import xmltodict
from flask import Flask, request, jsonify, send_file, render_template_string
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

# Diretorio temporario para uploads
UPLOAD_DIR = os.path.join(tempfile.gettempdir(), "calc_shopee_uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)


HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Calculadora de Lucro Shopee</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<style>
:root {
    --bg: #0f1117; --bg2: #1a1d27; --bg3: #242836;
    --text: #e4e6eb; --muted: #8b8d93; --border: #2d3140;
    --primary: #3b82f6; --green: #25D366; --red: #e74c3c; --gold: #FFD700;
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Segoe UI', sans-serif; background: var(--bg); color: var(--text); min-height: 100vh; }
.container { max-width: 800px; margin: 0 auto; padding: 24px; }
h1 { font-size: 22px; margin-bottom: 8px; }
h1 i { color: var(--primary); }
.subtitle { color: var(--muted); font-size: 13px; margin-bottom: 24px; }
.card { background: var(--bg2); border: 1px solid var(--border); border-radius: 12px; padding: 20px; margin-bottom: 16px; }
.card h3 { font-size: 15px; margin-bottom: 14px; color: var(--text); }
.card h3 i { color: var(--primary); margin-right: 6px; }
.form-row { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-bottom: 12px; }
label { font-size: 12px; font-weight: 600; color: var(--muted); display: block; margin-bottom: 4px; }
input[type="number"], input[type="text"] {
    width: 100%; padding: 10px 12px; background: var(--bg3); border: 1px solid var(--border);
    border-radius: 8px; color: var(--text); font-size: 14px; outline: none;
}
input:focus { border-color: var(--primary); }
.upload-area {
    border: 2px dashed var(--border); border-radius: 10px; padding: 24px; text-align: center;
    cursor: pointer; transition: all 0.2s; margin-bottom: 12px; position: relative;
}
.upload-area:hover, .upload-area.dragover { border-color: var(--primary); background: rgba(59,130,246,0.05); }
.upload-area input { position: absolute; inset: 0; opacity: 0; cursor: pointer; }
.upload-area i { font-size: 28px; color: var(--muted); display: block; margin-bottom: 8px; }
.upload-area .label { font-size: 13px; color: var(--muted); }
.upload-area .filename { font-size: 12px; color: var(--green); margin-top: 6px; display: none; }
.btn {
    width: 100%; padding: 14px; border: none; border-radius: 10px; font-size: 16px; font-weight: 700;
    cursor: pointer; transition: all 0.2s; display: flex; align-items: center; justify-content: center; gap: 8px;
}
.btn-primary { background: var(--green); color: #fff; }
.btn-primary:hover { background: #1db954; transform: scale(1.01); }
.btn-primary:disabled { background: var(--border); color: var(--muted); cursor: not-allowed; transform: none; }
.log { background: var(--bg); border: 1px solid var(--border); border-radius: 8px; padding: 12px; min-height: 80px; max-height: 200px; overflow-y: auto; font-size: 12px; font-family: monospace; color: var(--muted); margin-top: 12px; }
.log .ok { color: var(--green); } .log .err { color: var(--red); } .log .info { color: var(--primary); }
.result-card { background: var(--bg3); border-radius: 8px; padding: 12px; margin-top: 8px; }
.result-row { display: flex; justify-content: space-between; padding: 4px 0; font-size: 13px; }
.result-row .val { font-weight: 700; }
.result-row .val.pos { color: var(--green); } .result-row .val.neg { color: var(--red); }
.download-btn {
    display: inline-flex; align-items: center; gap: 6px; padding: 8px 16px; background: var(--primary);
    color: #fff; border-radius: 6px; font-size: 13px; text-decoration: none; margin-top: 10px;
}
.download-btn:hover { background: #2563eb; }
.hidden { display: none; }
table { width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 10px; }
th { background: var(--bg3); padding: 8px; text-align: left; font-weight: 600; border-bottom: 2px solid var(--border); }
td { padding: 6px 8px; border-bottom: 1px solid var(--border); }
tr:hover td { background: rgba(59,130,246,0.05); }
.table-scroll { max-height: 300px; overflow-y: auto; }
</style>
</head>
<body>
<div class="container">
    <h1><i class="fas fa-calculator"></i> Calculadora de Lucro</h1>
    <p class="subtitle">Processe XMLs de NF-e + planilha de custos para calcular lucro por SKU</p>

    <form id="calcForm" enctype="multipart/form-data">
        <!-- Uploads -->
        <div class="card">
            <h3><i class="fas fa-file-upload"></i> Arquivos</h3>
            <div class="upload-area" id="xmlArea">
                <input type="file" name="xmls" id="xmlInput" multiple accept=".xml,.zip" onchange="fileSelected(this,'xmlName')">
                <i class="fas fa-file-code"></i>
                <div class="label">Arraste XMLs ou ZIP aqui (ou clique para selecionar)</div>
                <div class="filename" id="xmlName"></div>
            </div>
            <div class="upload-area" id="excelArea">
                <input type="file" name="custos" id="custosInput" accept=".xlsx,.xls,.csv" onchange="fileSelected(this,'custosName')">
                <i class="fas fa-file-excel"></i>
                <div class="label">Planilha de Custos (.xlsx)</div>
                <div class="filename" id="custosName"></div>
            </div>
        </div>

        <!-- Parametros -->
        <div class="card">
            <h3><i class="fas fa-sliders-h"></i> Parametros</h3>
            <div class="form-row">
                <div>
                    <label>% Valor Declarado</label>
                    <input type="number" name="perc_declarado" value="100" step="1" min="1" max="100">
                </div>
                <div>
                    <label>Taxa Shopee (%)</label>
                    <input type="number" name="taxa_shopee" value="18" step="0.1" min="0" max="100">
                </div>
            </div>
            <div class="form-row">
                <div>
                    <label>Imposto Simples (%)</label>
                    <input type="number" name="taxa_imposto" value="4" step="0.1" min="0" max="100">
                </div>
                <div>
                    <label>Custo Fixo (R$)</label>
                    <input type="number" name="custo_fixo" value="3" step="0.5" min="0">
                </div>
            </div>
        </div>

        <button type="submit" class="btn btn-primary" id="btnGerar">
            <i class="fas fa-chart-line"></i> GERAR RELATORIO
        </button>
    </form>

    <!-- Log -->
    <div class="card hidden" id="logCard">
        <h3><i class="fas fa-terminal"></i> Log</h3>
        <div class="log" id="logBox"></div>
    </div>

    <!-- Resultado -->
    <div class="card hidden" id="resultCard">
        <h3><i class="fas fa-chart-pie"></i> Resultado</h3>
        <div id="resultSummary" class="result-card"></div>
        <a id="downloadLink" class="download-btn hidden" href="#"><i class="fas fa-download"></i> Baixar Excel</a>
        <div class="table-scroll" id="tableArea"></div>
    </div>
</div>

<script>
function fileSelected(input, labelId) {
    var el = document.getElementById(labelId);
    if (input.files.length > 0) {
        var names = Array.from(input.files).map(f => f.name);
        el.textContent = names.length > 3 ? names.length + ' arquivos selecionados' : names.join(', ');
        el.style.display = 'block';
        input.closest('.upload-area').style.borderColor = '#25D366';
    }
}

// Drag & drop
['xmlArea','excelArea'].forEach(function(id) {
    var el = document.getElementById(id);
    el.addEventListener('dragover', function(e) { e.preventDefault(); el.classList.add('dragover'); });
    el.addEventListener('dragleave', function() { el.classList.remove('dragover'); });
    el.addEventListener('drop', function(e) {
        e.preventDefault(); el.classList.remove('dragover');
        var input = el.querySelector('input');
        input.files = e.dataTransfer.files;
        var evt = new Event('change'); input.dispatchEvent(evt);
    });
});

function addLog(msg, cls) {
    var box = document.getElementById('logBox');
    box.innerHTML += '<div class="' + (cls||'') + '">' + msg + '</div>';
    box.scrollTop = box.scrollHeight;
    document.getElementById('logCard').classList.remove('hidden');
}

function formatBRL(v) { return 'R$ ' + v.toFixed(2).replace('.', ','); }

document.getElementById('calcForm').addEventListener('submit', async function(e) {
    e.preventDefault();
    var btn = document.getElementById('btnGerar');
    btn.disabled = true; btn.innerHTML = '<i class="fas fa-spinner fa-spin"></i> Processando...';
    document.getElementById('logBox').innerHTML = '';
    document.getElementById('resultCard').classList.add('hidden');
    document.getElementById('downloadLink').classList.add('hidden');
    addLog('Enviando arquivos...', 'info');

    var fd = new FormData(this);
    try {
        var res = await fetch('/api/calcular', { method: 'POST', body: fd });
        var data = await res.json();

        if (data.sucesso) {
            addLog('Processados ' + data.total_itens + ' itens de ' + data.total_xmls + ' XMLs', 'ok');
            addLog('Relatorio gerado com sucesso!', 'ok');

            // Summary
            var s = data.resumo;
            var lucroClass = s.lucro >= 0 ? 'pos' : 'neg';
            document.getElementById('resultSummary').innerHTML =
                '<div class="result-row"><span>Receita Real</span><span class="val pos">' + formatBRL(s.receita) + '</span></div>' +
                '<div class="result-row"><span>Custo Produtos</span><span class="val neg">' + formatBRL(s.custo) + '</span></div>' +
                '<div class="result-row"><span>Taxa Shopee</span><span class="val neg">' + formatBRL(s.shopee) + '</span></div>' +
                '<div class="result-row"><span>Imposto</span><span class="val neg">' + formatBRL(s.imposto) + '</span></div>' +
                '<div class="result-row" style="border-top:2px solid var(--border);padding-top:8px;margin-top:4px;">' +
                '<span style="font-weight:700;">LUCRO</span><span class="val ' + lucroClass + '" style="font-size:16px;">' + formatBRL(s.lucro) + '</span></div>';

            // Download link
            var dl = document.getElementById('downloadLink');
            dl.href = '/api/download/' + data.arquivo_id;
            dl.classList.remove('hidden');

            // Table
            if (data.tabela && data.tabela.length > 0) {
                var cols = Object.keys(data.tabela[0]);
                var html = '<table><thead><tr>' + cols.map(c => '<th>' + c + '</th>').join('') + '</tr></thead><tbody>';
                data.tabela.forEach(function(row) {
                    html += '<tr>' + cols.map(function(c) {
                        var v = row[c];
                        if (typeof v === 'number') v = v.toFixed(2);
                        return '<td>' + v + '</td>';
                    }).join('') + '</tr>';
                });
                html += '</tbody></table>';
                document.getElementById('tableArea').innerHTML = html;
            }

            document.getElementById('resultCard').classList.remove('hidden');
        } else {
            addLog('Erro: ' + (data.erro || 'desconhecido'), 'err');
        }
    } catch(err) {
        addLog('Erro de conexao: ' + err.message, 'err');
    }
    btn.disabled = false; btn.innerHTML = '<i class="fas fa-chart-line"></i> GERAR RELATORIO';
});
</script>
</body>
</html>'''


def formatar_excel(caminho):
    """Formata o Excel gerado com cores e bordas."""
    wb = load_workbook(caminho)
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    lucro_pos = Font(color="006100", bold=True)
    lucro_neg = Font(color="9C0006", bold=True)
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    max_row = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.col_idx >= 3:
                cell.number_format = 'R$ #,##0.00'
            if cell.col_idx == ws.max_column:
                if isinstance(cell.value, (int, float)) and cell.value >= 0:
                    cell.font = lucro_pos
                else:
                    cell.font = lucro_neg

    for cell in ws[max_row]:
        cell.fill = total_fill
        cell.font = Font(bold=True)

    for column in ws.columns:
        col_cells = [cell for cell in column]
        max_length = 0
        for cell in col_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_length + 2

    wb.save(caminho)


@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route('/api/calcular', methods=['POST'])
def calcular():
    try:
        # Parametros
        perc_declarado = float(request.form.get('perc_declarado', 100)) / 100
        taxa_shopee = float(request.form.get('taxa_shopee', 18)) / 100
        taxa_imposto = float(request.form.get('taxa_imposto', 4)) / 100
        custo_fixo = float(request.form.get('custo_fixo', 3))

        # Planilha de custos
        custos_file = request.files.get('custos')
        dict_custos = {}
        if custos_file and custos_file.filename:
            df_custos = pd.read_excel(custos_file)
            dict_custos = {
                str(row.iloc[0]).strip(): float(row.iloc[1])
                for _, row in df_custos.iterrows()
                if pd.notnull(row.iloc[1])
            }

        # XMLs (podem vir como multiplos arquivos ou ZIP)
        xml_files = request.files.getlist('xmls')
        xml_contents = []

        for f in xml_files:
            if f.filename.lower().endswith('.zip'):
                # Extrair XMLs do ZIP
                with zipfile.ZipFile(io.BytesIO(f.read())) as zf:
                    for name in zf.namelist():
                        if name.lower().endswith('.xml'):
                            xml_contents.append(zf.read(name))
            elif f.filename.lower().endswith('.xml'):
                xml_contents.append(f.read())

        if not xml_contents:
            return jsonify({"sucesso": False, "erro": "Nenhum XML encontrado"})

        # Processar XMLs
        lista_dados = []
        for xml_data in xml_contents:
            try:
                doc = xmltodict.parse(xml_data)
            except Exception:
                continue

            if "nfeProc" in doc:
                nfe = doc["nfeProc"]["NFe"]["infNFe"]
            elif "NFe" in doc:
                nfe = doc["NFe"]["infNFe"]
            else:
                continue

            dets = nfe.get("det", [])
            if not isinstance(dets, list):
                dets = [dets]

            for item in dets:
                prod = item["prod"]
                sku = str(prod.get("cProd", "")).strip()
                qtd = float(prod.get("qCom", 1))

                v_declarado = float(prod.get("vProd", 0))
                v_real = v_declarado / perc_declarado if perc_declarado > 0 else v_declarado

                c_imposto = v_declarado * taxa_imposto
                c_shopee = (v_real * taxa_shopee) + (custo_fixo * qtd)
                c_produto = dict_custos.get(sku, 0.0) * qtd

                lucro = v_real - c_imposto - c_shopee - c_produto

                lista_dados.append({
                    "SKU": sku,
                    "Qtd": qtd,
                    "V. Real": round(v_real, 2),
                    "V. Decl.": round(v_declarado, 2),
                    "Custo": round(c_produto, 2),
                    "Shopee": round(c_shopee, 2),
                    "Imposto": round(c_imposto, 2),
                    "LUCRO": round(lucro, 2),
                })

        if not lista_dados:
            return jsonify({"sucesso": False, "erro": "Nenhum item encontrado nos XMLs"})

        # Criar DataFrame
        df = pd.DataFrame(lista_dados)
        totais = df.sum(numeric_only=True)
        totais["SKU"] = "TOTAIS"

        # Gerar Excel
        arquivo_id = str(uuid.uuid4())[:8]
        nome_arq = f"Lucro_Shopee_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho = os.path.join(UPLOAD_DIR, f"{arquivo_id}_{nome_arq}")

        df_total = pd.concat([df, totais.to_frame().T], ignore_index=True)
        df_total.to_excel(caminho, index=False)
        formatar_excel(caminho)

        # Resumo
        resumo = {
            "receita": round(float(totais.get("V. Real", 0)), 2),
            "custo": round(float(totais.get("Custo", 0)), 2),
            "shopee": round(float(totais.get("Shopee", 0)), 2),
            "imposto": round(float(totais.get("Imposto", 0)), 2),
            "lucro": round(float(totais.get("LUCRO", 0)), 2),
        }

        # Tabela (max 100 linhas para preview)
        tabela = df.head(100).to_dict(orient="records")

        return jsonify({
            "sucesso": True,
            "resumo": resumo,
            "tabela": tabela,
            "total_itens": len(lista_dados),
            "total_xmls": len(xml_contents),
            "arquivo_id": arquivo_id,
            "arquivo_nome": nome_arq,
        })

    except Exception as e:
        return jsonify({"sucesso": False, "erro": str(e)})


@app.route('/api/download/<arquivo_id>')
def download(arquivo_id):
    """Download do Excel gerado."""
    for f in os.listdir(UPLOAD_DIR):
        if f.startswith(arquivo_id):
            caminho = os.path.join(UPLOAD_DIR, f)
            nome = f[len(arquivo_id) + 1:]  # remover prefixo id_
            return send_file(caminho, as_attachment=True, download_name=nome)
    return jsonify({"erro": "Arquivo nao encontrado"}), 404


if __name__ == '__main__':
    print("=" * 50)
    print("  Calculadora de Lucro Shopee - Web")
    print("  http://localhost:5050")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5050, debug=False)
