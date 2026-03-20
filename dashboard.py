# -*- coding: utf-8 -*-
"""
Dashboard Interativo - Beka MultiPlace
Backend Flask com API REST + Autenticacao JWT + Multi-usuario
"""

import os
import sys
import json
import time
import threading

# Forcar UTF-8 no stdout/stderr do Windows (evita crash 'charmap' em print com acentos)
# Quando roda via pythonw.exe, stdout/stderr podem ser None — redirecionar para devnull
if sys.platform == "win32":
    import io as _io
    for _stream_name in ("stdout", "stderr"):
        _stream = getattr(sys, _stream_name, None)
        if _stream is None or (hasattr(_stream, 'closed') and _stream.closed):
            # pythonw.exe: sem console, redirecionar para devnull
            setattr(sys, _stream_name, open(os.devnull, "w", encoding="utf-8"))
        elif hasattr(_stream, "buffer"):
            try:
                setattr(sys, _stream_name, _io.TextIOWrapper(_stream.buffer, encoding="utf-8", errors="replace"))
            except Exception:
                setattr(sys, _stream_name, open(os.devnull, "w", encoding="utf-8"))
import subprocess
import shutil
import hmac
import hashlib
import smtplib
import secrets
import re as _re
import unicodedata
from contextlib import contextmanager
from urllib.parse import urlparse, quote_plus, urlencode
from datetime import datetime, timedelta, timezone

# Fuso horario de Brasilia (UTC-3) - usado em TODOS os timestamps visiveis ao usuario
_FUSO_BRASILIA = timezone(timedelta(hours=-3))


def _agora_brasil():
    """Retorna datetime atual no fuso de Brasilia (UTC-3), sem tzinfo (naive)."""
    return datetime.now(_FUSO_BRASILIA).replace(tzinfo=None)
from collections import defaultdict
from flask import Flask, request, jsonify, send_from_directory, send_file, redirect, make_response
from flask_cors import CORS
from flask_jwt_extended import JWTManager, jwt_required, get_jwt_identity, get_jwt, create_access_token
import xmltodict
import pandas as pd
import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from etiquetas_shopee import ProcessadorEtiquetasShopee
from models import (db, bcrypt, User, Session, WhatsAppContact, Schedule,
                    UpSellerConfig, ExecutionLog, Loja, EmailContact,
                    WhatsAppQueueItem, MarketplaceApiConfig, MarketplaceLoja,
                    TimeLote, encrypt_value, decrypt_value,
                    Funcionario, FolhaPagamento, ValeParcela)
from auth import auth_bp
from email_utils import enviar_email_com_anexo, enviar_email_com_anexos, smtp_configurado, get_smtp_config
from payments import payments_bp
from scheduler import beka_scheduler
from whatsapp_service import WhatsAppService
from whatsapp_delivery import montar_entregas_por_resultado, montar_destinos_por_resultado

# PyInstaller frozen path support
if getattr(sys, 'frozen', False):
    _BASE_DIR = sys._MEIPASS
else:
    _BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, static_folder=os.path.join(_BASE_DIR, 'static'))
CORS(app)

# ----------------------------------------------------------------
# CONFIGURACAO DO APP
# ----------------------------------------------------------------
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'chave-secreta-trocar-em-producao')
app.config['JWT_SECRET_KEY'] = os.environ.get('JWT_SECRET_KEY', 'jwt-secret-trocar-em-producao')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(days=30)

# Banco de dados SQLite — usar volume persistente do Railway
# Railway monta o volume no path definido em RAILWAY_VOLUME_MOUNT_PATH
# IMPORTANTE: sempre usar o diretorio principal do projeto para o banco,
# nunca um worktree, para evitar bancos duplicados/dessincronizados
_MAIN_PROJECT_DIR = os.path.normpath(os.path.join(os.path.dirname(os.path.abspath(__file__))))
# Se estamos dentro de um worktree (.claude/worktrees/), usar o DB do projeto principal
if '.claude' + os.sep + 'worktrees' in _MAIN_PROJECT_DIR:
    _MAIN_PROJECT_DIR = _MAIN_PROJECT_DIR.split('.claude' + os.sep + 'worktrees')[0].rstrip(os.sep)
_VOLUME_PATH = os.environ.get('RAILWAY_VOLUME_MOUNT_PATH', os.environ.get('DB_DIR', os.path.join(_MAIN_PROJECT_DIR, 'data')))
os.makedirs(_VOLUME_PATH, exist_ok=True)
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{os.path.join(_VOLUME_PATH, 'app.db')}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
_WEBVIEW_STORAGE_DIR = os.path.join(_VOLUME_PATH, 'webview_profile')
os.makedirs(_WEBVIEW_STORAGE_DIR, exist_ok=True)


def _normalizar_base_url(raw: str) -> str:
    txt = str(raw or "").strip().rstrip("/")
    if not txt:
        return ""
    if "://" not in txt:
        txt = "https://" + txt
    return txt


def _detectar_ngrok_auto() -> str:
    """Detecta tunnel ngrok ativo via API local (localhost:4040)."""
    try:
        import requests as _req
        r = _req.get("http://localhost:4040/api/tunnels", timeout=2)
        if r.status_code == 200:
            tunnels = r.json().get("tunnels", [])
            for t in tunnels:
                url = str(t.get("public_url") or "").strip()
                if url.startswith("https://"):
                    return url
    except Exception:
        pass
    return ""


def _detectar_base_publica() -> str:
    """
    Detecta base publica para callback OAuth/testes:
    - desenvolvimento: NGROK_URL / NGROK_PUBLIC_URL / auto-detect ngrok
    - producao: SHOPEE_REDIRECT_BASE_URL / PUBLIC_BASE_URL / RAILWAY_PUBLIC_DOMAIN
    """
    candidatos = [
        os.environ.get("SHOPEE_REDIRECT_BASE_URL"),
        os.environ.get("NGROK_URL"),
        os.environ.get("NGROK_PUBLIC_URL"),
        os.environ.get("PUBLIC_BASE_URL"),
        os.environ.get("RAILWAY_STATIC_URL"),
        os.environ.get("RAILWAY_PUBLIC_DOMAIN"),
    ]
    for c in candidatos:
        base = _normalizar_base_url(c)
        if base:
            return base

    # Auto-detectar ngrok ativo na maquina local
    ngrok_url = _detectar_ngrok_auto()
    if ngrok_url:
        return _normalizar_base_url(ngrok_url) or ngrok_url

    # Fallback em runtime (quando requisicao chega com host publico).
    try:
        host_url = _normalizar_base_url((request.host_url if request else ""))
        if host_url:
            return host_url
    except Exception:
        pass
    return ""


def _get_shopee_redirect_url() -> str:
    base = _detectar_base_publica()
    return (base + "/api/marketplace/shopee/callback") if base else ""


def _get_shopee_redirect_domain() -> str:
    ru = _get_shopee_redirect_url()
    if not ru:
        return ""
    try:
        return (urlparse(ru).netloc or "").strip()
    except Exception:
        return ""


def _shopee_oauth_state_secret() -> bytes:
    """Segredo para assinar state do OAuth Shopee."""
    secret_txt = str(
        app.config.get("JWT_SECRET_KEY")
        or app.config.get("SECRET_KEY")
        or "beka-shopee-state-secret"
    )
    return secret_txt.encode("utf-8")


def _build_shopee_oauth_state(user_id: int) -> str:
    """Cria state assinado contendo user_id e timestamp (anti-tamper)."""
    uid = int(user_id)
    ts = int(time.time())
    nonce = secrets.token_hex(8)
    payload = f"{uid}:{ts}:{nonce}"
    sig = hmac.new(
        _shopee_oauth_state_secret(),
        payload.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()[:32]
    return f"{payload}:{sig}"


def _parse_shopee_oauth_state(state: str, max_age_sec: int = 1800):
    """Valida state assinado e retorna user_id."""
    txt = str(state or "").strip()
    parts = txt.split(":")
    if len(parts) != 4:
        return None, "state_invalido"
    uid_s, ts_s, nonce, sig = parts
    payload = f"{uid_s}:{ts_s}:{nonce}"
    expected = hmac.new(
        _shopee_oauth_state_secret(),
        payload.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()[:32]
    if not hmac.compare_digest(sig, expected):
        return None, "state_assinatura_invalida"
    try:
        uid = int(uid_s)
        ts = int(ts_s)
    except Exception:
        return None, "state_formato_invalido"
    now = int(time.time())
    if ts > now + 300:
        return None, "state_tempo_invalido"
    if now - ts > max(60, int(max_age_sec or 1800)):
        return None, "state_expirado"
    return uid, ""


# OAuth pendente: persiste no DB (marketplace_api_config.oauth_pending_at).
# Funciona cross-worker no Railway (multiplos processos Gunicorn).
_PENDING_OAUTH_MAX_AGE = 1800  # 30 minutos


def _register_pending_oauth(user_id: int, **_kwargs):
    """Marca no DB que user_id iniciou OAuth (seta oauth_pending_at = agora)."""
    try:
        cfg = MarketplaceApiConfig.query.filter_by(
            user_id=int(user_id), marketplace="shopee"
        ).first()
        if cfg:
            cfg.oauth_pending_at = _agora_brasil()
            db.session.commit()
    except Exception as e:
        print(f"[OAUTH] _register_pending_oauth FALHOU: {e}", flush=True, file=sys.stderr)


def _find_pending_oauth_cfg():
    """Encontra o MarketplaceApiConfig com OAuth pendente mais recente.
    Retorna o objeto cfg direto (ja com partner_key_enc no modelo).
    Funciona cross-worker porque le do DB."""
    try:
        cutoff = _agora_brasil() - timedelta(seconds=_PENDING_OAUTH_MAX_AGE)
        cfg = MarketplaceApiConfig.query.filter(
            MarketplaceApiConfig.marketplace == "shopee",
            MarketplaceApiConfig.oauth_pending_at.isnot(None),
            MarketplaceApiConfig.oauth_pending_at > cutoff,
        ).order_by(MarketplaceApiConfig.oauth_pending_at.desc()).first()
        return cfg
    except Exception as e:
        print(f"[OAUTH] _find_pending_oauth_cfg FALHOU: {e}", flush=True, file=sys.stderr)
        return None


def _consume_pending_oauth(user_id: int):
    """Limpa oauth_pending_at apos uso."""
    try:
        cfg = MarketplaceApiConfig.query.filter_by(
            user_id=int(user_id), marketplace="shopee"
        ).first()
        if cfg and cfg.oauth_pending_at:
            cfg.oauth_pending_at = None
            db.session.commit()
    except Exception:
        pass


# Inicializar extensoes
db.init_app(app)
bcrypt.init_app(app)
jwt = JWTManager(app)


# ----------------------------------------------------------------
# Permitir JWT via query param para endpoints de download (pywebview compat)
# ----------------------------------------------------------------
@app.before_request
def _jwt_query_param_fallback():
    """Injeta JWT do query param no header para endpoints de download.
    pywebview/WebView2 nao suporta blob download via JS, entao usamos
    window.open(url?jwt=...) que precisa do token no query string."""
    if '/download' in (request.path or '') or '/export' in (request.path or ''):
        qt = request.args.get('jwt') or request.args.get('qt')
        if qt and 'authorization' not in {k.lower() for k in request.headers.keys()}:
            request.environ['HTTP_AUTHORIZATION'] = f'Bearer {qt}'


# Registrar blueprints
app.register_blueprint(auth_bp)
app.register_blueprint(payments_bp)

from aios_routes import aios_bp
app.register_blueprint(aios_bp)

from funcionarios_routes import funcionarios_bp
app.register_blueprint(funcionarios_bp)

from shopee_monitor_routes import shopee_monitor_bp
app.register_blueprint(shopee_monitor_bp)


# ----------------------------------------------------------------
# Auto-login local (desktop) — elimina tela de login
# ----------------------------------------------------------------
@app.route('/api/auth/local-token', methods=['POST'])
def _local_auto_token():
    """Retorna token JWT para o primeiro usuario local.
    Usado pelo frontend desktop para pular tela de login."""
    from flask_jwt_extended import create_access_token
    user = User.query.first()
    if not user:
        # Criar usuario local padrao se nao existir
        user = User(email='local@beka.app', nome='Beka Local')
        user.set_password('local')
        user.is_active = True
        user.email_verified = True
        user.plano = 'vitalicio'
        db.session.add(user)
        db.session.commit()
    # Criar/reusar sessao no banco para passar na validacao do blocklist
    token_id = "local-desktop"
    sessao = Session.query.filter_by(user_id=user.id, token_id=token_id).first()
    if not sessao:
        sessao = Session(user_id=user.id, token_id=token_id, ip="127.0.0.1")
        db.session.add(sessao)
        db.session.commit()
    token = create_access_token(
        identity=str(user.id),
        additional_claims={"sid": token_id}
    )
    return jsonify({"token": token, "user": user.to_dict()})


# Migrar banco existente (adicionar colunas novas em tabelas existentes)
def _migrate_db():
    """SQLite nao adiciona colunas automaticamente em tabelas existentes.
    Esta funcao verifica e adiciona colunas faltantes."""
    import sqlalchemy
    inspector = sqlalchemy.inspect(db.engine)

    if 'users' in inspector.get_table_names():
        colunas = [c['name'] for c in inspector.get_columns('users')]
        with db.engine.begin() as conn:
            if 'email_verified' not in colunas:
                conn.execute(sqlalchemy.text('ALTER TABLE users ADD COLUMN email_verified BOOLEAN DEFAULT 0'))
            if 'email_code' not in colunas:
                conn.execute(sqlalchemy.text("ALTER TABLE users ADD COLUMN email_code VARCHAR(6) DEFAULT ''"))
            if 'email_code_expires' not in colunas:
                conn.execute(sqlalchemy.text('ALTER TABLE users ADD COLUMN email_code_expires DATETIME'))
            if 'google_id' not in colunas:
                conn.execute(sqlalchemy.text('ALTER TABLE users ADD COLUMN google_id VARCHAR(255)'))
            if 'reset_code' not in colunas:
                conn.execute(sqlalchemy.text("ALTER TABLE users ADD COLUMN reset_code VARCHAR(6) DEFAULT ''"))
            if 'reset_code_expires' not in colunas:
                conn.execute(sqlalchemy.text('ALTER TABLE users ADD COLUMN reset_code_expires DATETIME'))
            if 'cupom_indicacao' not in colunas:
                conn.execute(sqlalchemy.text('ALTER TABLE users ADD COLUMN cupom_indicacao VARCHAR(20)'))
            if 'indicado_por' not in colunas:
                conn.execute(sqlalchemy.text('ALTER TABLE users ADD COLUMN indicado_por INTEGER'))
            if 'meses_gratis' not in colunas:
                conn.execute(sqlalchemy.text('ALTER TABLE users ADD COLUMN meses_gratis INTEGER DEFAULT 0'))
            if 'plano_expira' not in colunas:
                conn.execute(sqlalchemy.text('ALTER TABLE users ADD COLUMN plano_expira DATETIME'))
            if 'auto_send_whatsapp' not in colunas:
                conn.execute(sqlalchemy.text('ALTER TABLE users ADD COLUMN auto_send_whatsapp BOOLEAN DEFAULT 0'))
            if 'email_remetente' not in colunas:
                conn.execute(sqlalchemy.text("ALTER TABLE users ADD COLUMN email_remetente VARCHAR(200) DEFAULT ''"))
            if 'nome_remetente' not in colunas:
                conn.execute(sqlalchemy.text("ALTER TABLE users ADD COLUMN nome_remetente VARCHAR(200) DEFAULT ''"))
            if 'smtp_host' not in colunas:
                conn.execute(sqlalchemy.text("ALTER TABLE users ADD COLUMN smtp_host VARCHAR(200) DEFAULT ''"))
            if 'smtp_port' not in colunas:
                conn.execute(sqlalchemy.text("ALTER TABLE users ADD COLUMN smtp_port INTEGER DEFAULT 587"))
            if 'smtp_user' not in colunas:
                conn.execute(sqlalchemy.text("ALTER TABLE users ADD COLUMN smtp_user VARCHAR(200) DEFAULT ''"))
            if 'smtp_pass_enc' not in colunas:
                conn.execute(sqlalchemy.text("ALTER TABLE users ADD COLUMN smtp_pass_enc TEXT DEFAULT ''"))
            if 'smtp_from' not in colunas:
                conn.execute(sqlalchemy.text("ALTER TABLE users ADD COLUMN smtp_from VARCHAR(200) DEFAULT ''"))
            if 'pasta_avulsas' not in colunas:
                conn.execute(sqlalchemy.text("ALTER TABLE users ADD COLUMN pasta_avulsas VARCHAR(500) DEFAULT ''"))

    # Migrar tabela lojas
    if 'lojas' in inspector.get_table_names():
        colunas_lojas = [c['name'] for c in inspector.get_columns('lojas')]
        with db.engine.begin() as conn:
            if 'notas_pendentes' not in colunas_lojas:
                conn.execute(sqlalchemy.text('ALTER TABLE lojas ADD COLUMN notas_pendentes INTEGER DEFAULT 0'))
            if 'etiquetas_pendentes' not in colunas_lojas:
                conn.execute(sqlalchemy.text('ALTER TABLE lojas ADD COLUMN etiquetas_pendentes INTEGER DEFAULT 0'))

    # Migrar contatos WhatsApp
    if 'whatsapp_contacts' in inspector.get_table_names():
        cols = [c['name'] for c in inspector.get_columns('whatsapp_contacts')]
        with db.engine.begin() as conn:
            if 'lojas_json' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE whatsapp_contacts ADD COLUMN lojas_json TEXT DEFAULT '[]'"))
            if 'grupos_json' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE whatsapp_contacts ADD COLUMN grupos_json TEXT DEFAULT '[]'"))
            if 'horario' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE whatsapp_contacts ADD COLUMN horario VARCHAR(5) DEFAULT ''"))
            if 'dias_semana' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE whatsapp_contacts ADD COLUMN dias_semana TEXT DEFAULT '[]'"))
            if 'horarios_json' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE whatsapp_contacts ADD COLUMN horarios_json TEXT DEFAULT '[]'"))
                # Migrar dados existentes de horario+dias_semana para horarios_json
                try:
                    rows = conn.execute(sqlalchemy.text(
                        "SELECT id, horario, dias_semana FROM whatsapp_contacts WHERE horario IS NOT NULL AND horario != ''"
                    )).fetchall()
                    for row in rows:
                        _id, _hor, _dias = row[0], row[1], row[2]
                        if _hor and str(_hor).strip():
                            dias_list = []
                            try:
                                dias_list = json.loads(_dias) if _dias else []
                            except Exception:
                                dias_list = [d.strip() for d in str(_dias).split(',') if d.strip()] if _dias else []
                            if not dias_list:
                                dias_list = ["seg","ter","qua","qui","sex","sab","dom"]
                            h_json = json.dumps([{"dias": dias_list, "horas": [str(_hor).strip()]}], ensure_ascii=False)
                            conn.execute(sqlalchemy.text(
                                "UPDATE whatsapp_contacts SET horarios_json = :hj WHERE id = :id"
                            ), {"hj": h_json, "id": _id})
                except Exception as e_mig:
                    print(f"[Migrate] Aviso ao migrar horarios whatsapp: {e_mig}")

    # Migrar contatos de email
    if 'email_contacts' in inspector.get_table_names():
        cols = [c['name'] for c in inspector.get_columns('email_contacts')]
        with db.engine.begin() as conn:
            if 'lojas_json' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE email_contacts ADD COLUMN lojas_json TEXT DEFAULT '[]'"))
            if 'grupos_json' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE email_contacts ADD COLUMN grupos_json TEXT DEFAULT '[]'"))
            if 'horario' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE email_contacts ADD COLUMN horario VARCHAR(5) DEFAULT ''"))
            if 'dias_semana' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE email_contacts ADD COLUMN dias_semana TEXT DEFAULT '[]'"))
            if 'horarios_json' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE email_contacts ADD COLUMN horarios_json TEXT DEFAULT '[]'"))
                # Migrar dados existentes de horario+dias_semana para horarios_json
                try:
                    rows = conn.execute(sqlalchemy.text(
                        "SELECT id, horario, dias_semana FROM email_contacts WHERE horario IS NOT NULL AND horario != ''"
                    )).fetchall()
                    for row in rows:
                        _id, _hor, _dias = row[0], row[1], row[2]
                        if _hor and str(_hor).strip():
                            dias_list = []
                            try:
                                dias_list = json.loads(_dias) if _dias else []
                            except Exception:
                                dias_list = [d.strip() for d in str(_dias).split(',') if d.strip()] if _dias else []
                            if not dias_list:
                                dias_list = ["seg","ter","qua","qui","sex","sab","dom"]
                            h_json = json.dumps([{"dias": dias_list, "horas": [str(_hor).strip()]}], ensure_ascii=False)
                            conn.execute(sqlalchemy.text(
                                "UPDATE email_contacts SET horarios_json = :hj WHERE id = :id"
                            ), {"hj": h_json, "id": _id})
                except Exception as e_mig:
                    print(f"[Migrate] Aviso ao migrar horarios email: {e_mig}")

    # Migrar agendamentos
    if 'schedules' in inspector.get_table_names():
        cols = [c['name'] for c in inspector.get_columns('schedules')]
        with db.engine.begin() as conn:
            if 'lojas_json' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE schedules ADD COLUMN lojas_json TEXT DEFAULT '[]'"))
            if 'grupos_json' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE schedules ADD COLUMN grupos_json TEXT DEFAULT '[]'"))
            if 'modo_pipeline' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE schedules ADD COLUMN modo_pipeline VARCHAR(20) DEFAULT 'completo'"))
            if 'enviar_email' not in cols:
                conn.execute(sqlalchemy.text("ALTER TABLE schedules ADD COLUMN enviar_email BOOLEAN DEFAULT 0"))

    # Migrar contatos WhatsApp — adicionar lote_ids_json e agendamento_ativo
    if 'whatsapp_contacts' in inspector.get_table_names():
        cols_wc = [c['name'] for c in inspector.get_columns('whatsapp_contacts')]
        with db.engine.begin() as conn:
            if 'lote_ids_json' not in cols_wc:
                conn.execute(sqlalchemy.text("ALTER TABLE whatsapp_contacts ADD COLUMN lote_ids_json TEXT DEFAULT '[]'"))
            if 'agendamento_ativo' not in cols_wc:
                conn.execute(sqlalchemy.text("ALTER TABLE whatsapp_contacts ADD COLUMN agendamento_ativo BOOLEAN DEFAULT 1"))

    # Migrar contatos Email — adicionar lote_ids_json e agendamento_ativo
    if 'email_contacts' in inspector.get_table_names():
        cols_ec = [c['name'] for c in inspector.get_columns('email_contacts')]
        with db.engine.begin() as conn:
            if 'lote_ids_json' not in cols_ec:
                conn.execute(sqlalchemy.text("ALTER TABLE email_contacts ADD COLUMN lote_ids_json TEXT DEFAULT '[]'"))
            if 'agendamento_ativo' not in cols_ec:
                conn.execute(sqlalchemy.text("ALTER TABLE email_contacts ADD COLUMN agendamento_ativo BOOLEAN DEFAULT 1"))

    # Migrar marketplace_api_config — adicionar oauth_pending_at
    if 'marketplace_api_config' in inspector.get_table_names():
        cols_mkt = [c['name'] for c in inspector.get_columns('marketplace_api_config')]
        with db.engine.begin() as conn:
            if 'oauth_pending_at' not in cols_mkt:
                conn.execute(sqlalchemy.text("ALTER TABLE marketplace_api_config ADD COLUMN oauth_pending_at DATETIME"))

# Criar tabelas
with app.app_context():
    _migrate_db()
    db.create_all()

# Inicializar scheduler de automacao
beka_scheduler.init_app(app)
app._beka_scheduler = beka_scheduler  # Referencia para uso nos endpoints de lote

# Recuperar execucoes que ficaram travadas (servidor morreu no meio)
with app.app_context():
    _stuck = ExecutionLog.query.filter_by(status="executando").all()
    for _ex in _stuck:
        _ex.status = "erro"
        _ex.fim = _agora_brasil()
        _ex.detalhes = json.dumps(
            {"erro": "Execucao interrompida por reinicio do servidor"},
            ensure_ascii=False
        )
    if _stuck:
        db.session.commit()
        print(f"[Startup] {len(_stuck)} execucao(oes) travada(s) marcada(s) como erro")

# Job scheduler: enviar resumo geral consolidado no fim do dia (21:00)
def _job_resumo_geral_diario():
    """Envia resumo geral consolidado (XLSX + JPEG) para contatos com flag resumo_geral=True."""
    with app.app_context():
        users = User.query.filter_by(is_active=True).all()
        for user in users:
            try:
                pasta = user.get_pasta_saida() if hasattr(user, 'get_pasta_saida') else os.path.join("C:\\tmp\\users", str(user.id), "Etiquetas prontas")
                result = _enviar_resumo_geral_whatsapp(user.id, pasta, consolidado=True)
                if result.get("ok"):
                    print(f"[ResumoGeral] Consolidado enviado para user {user.id}")
            except Exception as e:
                print(f"[ResumoGeral] Erro para user {user.id}: {e}")

with app.app_context():
    from apscheduler.triggers.cron import CronTrigger
    try:
        existing = beka_scheduler.scheduler.get_job('beka_resumo_geral_diario')
        if existing:
            beka_scheduler.scheduler.remove_job('beka_resumo_geral_diario')
        beka_scheduler.scheduler.add_job(
            _job_resumo_geral_diario,
            trigger=CronTrigger(hour=21, minute=0, timezone="America/Sao_Paulo"),
            id='beka_resumo_geral_diario',
            name='Resumo Geral Diario',
            replace_existing=True,
        )
        print("[Startup] Job resumo geral diario registrado (21:00)")
    except Exception as e_sched:
        print(f"[Startup] Erro ao registrar job resumo geral: {e_sched}")

# Emails com acesso vitalicio (plano empresarial permanente)
EMAILS_VITALICIO = [
    "charlescr9@gmail.com",
]

with app.app_context():
    for email_vip in EMAILS_VITALICIO:
        user_vip = User.query.filter_by(email=email_vip).first()
        if user_vip and user_vip.plano != "empresarial":
            user_vip.plano = "empresarial"
            db.session.commit()
            print(f"Acesso vitalicio ativado para {email_vip}")


# Verificar sessao valida em TODA request protegida
@jwt.additional_claims_loader
def add_claims(identity):
    return {}


@jwt.token_in_blocklist_loader
def check_session_valid(jwt_header, jwt_payload):
    """Retorna True se o token NAO e mais valido (sessao removida)."""
    token_id = jwt_payload.get("sid", "")
    if not token_id:
        return False  # tokens antigos sem sid passam
    user_id = jwt_payload.get("sub", "")
    if not user_id:
        return False
    try:
        sessao = Session.query.filter_by(user_id=int(user_id), token_id=token_id).first()
        return sessao is None  # True = bloqueado (sessao nao existe mais)
    except Exception:
        # Se o banco estiver com lock (ex: callback Shopee escrevendo),
        # nao bloquear o token — permitir a request.
        return False

# ----------------------------------------------------------------
# ESTADO POR USUARIO (em memoria)
# ----------------------------------------------------------------
estados = {}  # {user_id: {processando, logs, ultimo_resultado, ...}}


def _get_estado(user_id):
    """Retorna o estado do usuario, criando se nao existir."""
    uid = int(user_id)
    if uid not in estados:
        user = User.query.get(uid)
        if not user:
            return None
        estados[uid] = {
            "processando": False,
            "logs": [],
            "ultimo_resultado": None,
            "historico": [],
            "agrupamentos": [],
            "configuracoes": {
                "pasta_entrada": user.get_pasta_entrada(),
                "pasta_saida": user.get_pasta_saida(),
                "pasta_lucro": user.get_pasta_lucro(),
                "largura_mm": 150,
                "altura_mm": 230,
                "margem_esq": 8,
                "margem_dir": 8,
                "margem_topo": 5,
                "margem_inf": 5,
                "fonte_produto": 7,
                "exibicao_produto": "sku",
                "perc_declarado": 100,
                "taxa_shopee": 18,
                "imposto_simples": 4,
                "custo_fixo": 3.0,
                "planilha_custos": "",
                "lucro_por_loja": {},
            }
        }
        # Tentar carregar config salva
        _carregar_config_usuario(uid)
    return estados[uid]


def _config_path(user_id):
    """Caminho do arquivo de config do usuario."""
    user = User.query.get(int(user_id))
    if not user:
        return None
    pasta = user.get_pasta_entrada()
    return os.path.join(pasta, "_config.json")


def _agrupamentos_path(user_id):
    """Caminho do arquivo de agrupamentos persistidos do usuario."""
    user = User.query.get(int(user_id))
    if not user:
        return None
    pasta = user.get_pasta_entrada()
    return os.path.join(pasta, "_agrupamentos.json")


def _salvar_config_usuario(user_id):
    """Salva config do usuario em JSON."""
    try:
        estado = estados.get(int(user_id))
        if not estado:
            return
        path = _config_path(user_id)
        if path:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(estado["configuracoes"], f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Aviso: nao salvou config user {user_id}: {e}")


def _salvar_agrupamentos_usuario(user_id):
    """Salva agrupamentos do usuario em JSON separado (persistente)."""
    try:
        estado = estados.get(int(user_id))
        if not estado:
            return
        path = _agrupamentos_path(user_id)
        if path:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(estado.get("agrupamentos", []) or [], f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Aviso: nao salvou agrupamentos user {user_id}: {e}")


def _carregar_agrupamentos_usuario(user_id):
    """Carrega agrupamentos persistidos do usuario (se existir)."""
    try:
        path = _agrupamentos_path(user_id)
        if path and os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                grupos = json.load(f)
            estado = estados.get(int(user_id))
            if estado and isinstance(grupos, list):
                estado["agrupamentos"] = grupos
    except Exception as e:
        print(f"Aviso: nao carregou agrupamentos user {user_id}: {e}")


def _carregar_config_usuario(user_id):
    """Carrega config do usuario se existir."""
    try:
        path = _config_path(user_id)
        if path and os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                config_salva = json.load(f)
            estado = estados.get(int(user_id))
            if estado:
                for chave, valor in config_salva.items():
                    if chave in estado["configuracoes"]:
                        estado["configuracoes"][chave] = valor
    except Exception as e:
        print(f"Aviso: nao carregou config user {user_id}: {e}")
    # Carregar agrupamentos persistidos
    _carregar_agrupamentos_usuario(user_id)
    # Carregar ultimo_resultado salvo em disco
    _carregar_resultado_usuario(user_id)


def _resultado_path(user_id):
    """Caminho do arquivo de resultado do usuario."""
    user = User.query.get(int(user_id))
    if not user:
        return None
    pasta = user.get_pasta_saida()
    return os.path.join(pasta, "_ultimo_resultado.json")


def _user_data_root(user_id):
    """Diretorio raiz de dados do usuario (irmao de entrada/saida)."""
    user = User.query.get(int(user_id))
    if not user:
        return None
    pasta_entrada = user.get_pasta_entrada()
    return os.path.dirname(pasta_entrada)


def _historico_geradas_dir(user_id):
    """Pasta persistente para snapshots de geracoes (ultimas 24h)."""
    root = _user_data_root(user_id)
    if not root:
        return None
    pasta = os.path.join(root, "historico_geradas")
    os.makedirs(pasta, exist_ok=True)
    return pasta


def _historico_geradas_index_path(user_id):
    pasta = _historico_geradas_dir(user_id)
    if not pasta:
        return None
    return os.path.join(pasta, "_index.json")


def _carregar_historico_geradas_raw(user_id):
    """Carrega indice bruto do historico de geracoes."""
    path = _historico_geradas_index_path(user_id)
    if not path or not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
    except Exception:
        pass
    return []


def _salvar_historico_geradas_raw(user_id, itens):
    """Salva indice bruto do historico de geracoes."""
    path = _historico_geradas_index_path(user_id)
    if not path:
        return
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(itens or [], f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Aviso: nao salvou historico de geradas user {user_id}: {e}")


def _parse_iso_dt(value):
    """Parse ISO date flexivel.
    Timestamps com 'Z' ou offset sao convertidos para horario de Brasilia.
    Timestamps naive sao assumidos como ja estando em Brasilia.
    """
    if not value:
        return None
    try:
        txt = str(value).strip()
        if txt.endswith("Z"):
            txt = txt[:-1] + "+00:00"
        dt = datetime.fromisoformat(txt)
        if getattr(dt, "tzinfo", None) is not None:
            # Converter para Brasilia e remover tzinfo
            dt = dt.astimezone(_FUSO_BRASILIA).replace(tzinfo=None)
        return dt
    except Exception:
        return None


def _parse_stamp_dt(value):
    """Extrai datetime a partir de padroes YYYYMMDD_HHMMSS em nomes."""
    if not value:
        return None
    try:
        m = _re.search(r"(\d{8})_(\d{6})", str(value))
        if not m:
            return None
        return datetime.strptime(f"{m.group(1)}_{m.group(2)}", "%Y%m%d_%H%M%S")
    except Exception:
        return None


def _zipar_pasta(origem, destino_zip, ignorar_underscore=True):
    """Compacta pasta inteira em ZIP; retorna quantidade de arquivos adicionados."""
    import zipfile

    arquivos = 0
    with zipfile.ZipFile(destino_zip, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, _dirs, files in os.walk(origem):
            for f in files:
                if ignorar_underscore and f.startswith("_"):
                    continue
                fp = os.path.join(root, f)
                if not os.path.isfile(fp):
                    continue
                arc = os.path.relpath(fp, origem)
                zf.write(fp, arc)
                arquivos += 1
    if arquivos <= 0:
        try:
            if os.path.exists(destino_zip):
                os.remove(destino_zip)
        except Exception:
            pass
    return arquivos


def _reconstruir_historico_geradas(user_id, horas=24):
    """
    Reconstroi entradas faltantes do historico com base em arquivos reais.

    Fontes:
    - ZIPs ja existentes em historico_geradas
    - Lotes em entrada/_upseller_lotes/*
    - Snapshot da pasta atual "Etiquetas prontas"
    """
    pasta_hist = _historico_geradas_dir(user_id)
    user = User.query.get(int(user_id))
    if not pasta_hist or not user:
        return 0

    horas = max(1, int(horas or 24))
    cutoff = _agora_brasil() - timedelta(hours=horas)
    itens = _carregar_historico_geradas_raw(user_id)
    by_arquivo = {}
    for it in itens:
        arq = os.path.basename(str((it or {}).get("arquivo", "")).strip())
        if arq:
            by_arquivo[arq] = it

    novos = 0

    def _registrar_zip_existente(caminho_zip, dt_hint=None, origem_hint="recuperado"):
        nonlocal novos, itens, by_arquivo

        if not caminho_zip or not os.path.exists(caminho_zip):
            return
        arquivo = os.path.basename(caminho_zip)
        if not arquivo.lower().endswith(".zip"):
            return
        if arquivo in by_arquivo:
            return

        dt = dt_hint or _parse_stamp_dt(arquivo)
        if dt is None:
            try:
                dt = datetime.utcfromtimestamp(os.path.getmtime(caminho_zip))
            except Exception:
                dt = None
        if dt is None or dt < cutoff:
            return

        try:
            size = int(os.path.getsize(caminho_zip) or 0)
        except Exception:
            size = 0
        if size <= 0:
            return

        stamp_token = dt.strftime("%Y%m%d_%H%M%S_%f")
        item = {
            "id": f"r{stamp_token}",
            "created_at": dt.isoformat() + "Z",
            "arquivo": arquivo,
            "origem": origem_hint,
            "total_lojas": 0,
            "total_etiquetas": 0,
            "size": size,
            "timestamp_resultado": "",
        }
        itens.append(item)
        by_arquivo[arquivo] = item
        novos += 1

    # 1) Reindexar ZIPs ja presentes na pasta de historico
    try:
        for nome in os.listdir(pasta_hist):
            if nome.lower().endswith(".zip"):
                _registrar_zip_existente(os.path.join(pasta_hist, nome), origem_hint="historico_orfao")
    except Exception:
        pass

    # 2) Recriar snapshots por lote (entrada/_upseller_lotes)
    try:
        pasta_lotes = os.path.join(user.get_pasta_entrada(), "_upseller_lotes")
        if os.path.isdir(pasta_lotes):
            for nome_dir in sorted(os.listdir(pasta_lotes), reverse=True):
                dir_lote = os.path.join(pasta_lotes, nome_dir)
                if not os.path.isdir(dir_lote):
                    continue
                dt_lote = _parse_stamp_dt(nome_dir)
                if dt_lote is None:
                    try:
                        dt_lote = datetime.utcfromtimestamp(os.path.getmtime(dir_lote))
                    except Exception:
                        dt_lote = None
                if dt_lote is None or dt_lote < cutoff:
                    continue

                zip_nome = f"recuperado_{nome_dir}.zip"
                zip_path = os.path.join(pasta_hist, zip_nome)
                if not os.path.exists(zip_path):
                    arquivos = _zipar_pasta(dir_lote, zip_path, ignorar_underscore=True)
                    if arquivos <= 0:
                        continue
                _registrar_zip_existente(zip_path, dt_hint=dt_lote, origem_hint="recuperado_lote")
    except Exception as e:
        print(f"Aviso: erro ao reconstruir historico por lotes user {user_id}: {e}")

    # 3) Snapshot da pasta de saida atual (caso tenha arquivos recentes ainda nao indexados)
    try:
        pasta_saida = user.get_pasta_saida()
        arquivos_saida = []
        if os.path.isdir(pasta_saida):
            for root, _dirs, files in os.walk(pasta_saida):
                for f in files:
                    if f.startswith("_"):
                        continue
                    fp = os.path.join(root, f)
                    if not os.path.isfile(fp):
                        continue
                    mt = datetime.utcfromtimestamp(os.path.getmtime(fp))
                    if mt >= cutoff:
                        arquivos_saida.append((fp, mt))
        if arquivos_saida:
            dt_saida = max(mt for _fp, mt in arquivos_saida)
            zip_nome = f"recuperado_saida_{dt_saida.strftime('%Y%m%d_%H%M%S')}.zip"
            zip_path = os.path.join(pasta_hist, zip_nome)
            if not os.path.exists(zip_path):
                arquivos = _zipar_pasta(pasta_saida, zip_path, ignorar_underscore=True)
                if arquivos > 0:
                    _registrar_zip_existente(zip_path, dt_hint=dt_saida, origem_hint="recuperado_saida")
            else:
                _registrar_zip_existente(zip_path, dt_hint=dt_saida, origem_hint="recuperado_saida")
    except Exception as e:
        print(f"Aviso: erro ao reconstruir historico da pasta de saida user {user_id}: {e}")

    if novos > 0:
        itens.sort(key=lambda x: str((x or {}).get("created_at", "")), reverse=True)
        _salvar_historico_geradas_raw(user_id, itens)
    return novos


def _limpar_historico_geradas_expirado(user_id, horas=24):
    """
    Remove registros/arquivos expirados do historico.

    Regra: manter somente ultimas `horas` (padrao 24h) e arquivos existentes.
    """
    pasta = _historico_geradas_dir(user_id)
    if not pasta:
        return []

    cutoff = _agora_brasil() - timedelta(hours=max(1, int(horas or 24)))
    itens = _carregar_historico_geradas_raw(user_id)
    kept = []

    for it in itens:
        arquivo = (it or {}).get("arquivo", "")
        if not arquivo:
            continue
        caminho = os.path.join(pasta, os.path.basename(arquivo))
        dt = _parse_iso_dt((it or {}).get("created_at"))

        # Expirado ou sem data => remover registro e arquivo.
        expired = (dt is None) or (dt < cutoff)
        missing = not os.path.exists(caminho)
        if expired or missing:
            if os.path.exists(caminho):
                try:
                    os.remove(caminho)
                except Exception:
                    pass
            continue

        kept.append(it)

    _salvar_historico_geradas_raw(user_id, kept)
    return kept


def _registrar_historico_gerada(user_id, resultado, pasta_saida, origem="processamento"):
    """
    Cria snapshot ZIP da geracao e registra no historico persistente (24h).
    """
    import zipfile

    if not resultado or not pasta_saida or not os.path.exists(pasta_saida):
        return None

    pasta_hist = _historico_geradas_dir(user_id)
    if not pasta_hist:
        return None

    _limpar_historico_geradas_expirado(user_id, horas=24)

    stamp = _agora_brasil()
    stamp_token = stamp.strftime("%Y%m%d_%H%M%S_%f")
    nome_zip = f"geradas_{stamp.strftime('%Y%m%d_%H%M%S')}.zip"
    caminho_zip = os.path.join(pasta_hist, nome_zip)

    arquivos = 0
    try:
        with zipfile.ZipFile(caminho_zip, "w", zipfile.ZIP_DEFLATED) as zf:
            for root, _dirs, files in os.walk(pasta_saida):
                for f in files:
                    if f.startswith("_"):
                        continue
                    fp = os.path.join(root, f)
                    if not os.path.isfile(fp):
                        continue
                    arc = os.path.relpath(fp, pasta_saida)
                    zf.write(fp, arc)
                    arquivos += 1
    except Exception as e:
        print(f"Aviso: nao criou ZIP de historico user {user_id}: {e}")
        return None

    if arquivos == 0:
        try:
            if os.path.exists(caminho_zip):
                os.remove(caminho_zip)
        except Exception:
            pass
        return None

    size = os.path.getsize(caminho_zip) if os.path.exists(caminho_zip) else 0
    entry_id = f"h{stamp_token}"
    item = {
        "id": entry_id,
        "created_at": stamp.isoformat(),
        "arquivo": nome_zip,
        "origem": origem,
        "total_lojas": int((resultado or {}).get("total_lojas", 0) or 0),
        "total_etiquetas": int((resultado or {}).get("total_etiquetas", 0) or 0),
        "size": int(size or 0),
        "timestamp_resultado": (resultado or {}).get("timestamp", ""),
    }

    itens = _carregar_historico_geradas_raw(user_id)
    itens.insert(0, item)
    _salvar_historico_geradas_raw(user_id, itens)
    _limpar_historico_geradas_expirado(user_id, horas=24)
    return item


def _listar_historico_geradas(user_id, horas=24):
    """Retorna historico pronto para UI (somente janela desejada)."""
    pasta_hist = _historico_geradas_dir(user_id)
    if not pasta_hist:
        return []

    horas = max(1, int(horas or 24))
    _reconstruir_historico_geradas(user_id, horas=horas)
    _limpar_historico_geradas_expirado(user_id, horas=horas)
    itens = _carregar_historico_geradas_raw(user_id)
    cutoff = _agora_brasil() - timedelta(hours=horas)

    out = []
    for it in itens:
        arquivo = (it or {}).get("arquivo", "")
        if not arquivo:
            continue
        caminho = os.path.join(pasta_hist, os.path.basename(arquivo))
        if not os.path.exists(caminho):
            continue

        dt = _parse_iso_dt((it or {}).get("created_at"))
        if dt is None or dt < cutoff:
            continue

        size = int((it or {}).get("size", 0) or 0)
        if size <= 0:
            try:
                size = os.path.getsize(caminho)
            except Exception:
                size = 0

        out.append({
            "id": (it or {}).get("id", ""),
            "created_at": (it or {}).get("created_at", ""),
            "created_at_fmt": dt.strftime("%d/%m/%Y %H:%M") if dt else "",
            "arquivo": os.path.basename(arquivo),
            "origem": (it or {}).get("origem", ""),
            "total_lojas": int((it or {}).get("total_lojas", 0) or 0),
            "total_etiquetas": int((it or {}).get("total_etiquetas", 0) or 0),
            "size": size,
            "size_fmt": _formatar_tamanho(size),
            "download_url": f"/api/historico-geradas/download/{(it or {}).get('id', '')}",
        })

    # Mais recente primeiro
    out.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return out


def _salvar_resultado_usuario(user_id):
    """Salva ultimo_resultado do usuario em JSON no volume persistente."""
    try:
        estado = estados.get(int(user_id))
        if not estado or not estado.get("ultimo_resultado"):
            return
        path = _resultado_path(user_id)
        if path:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(estado["ultimo_resultado"], f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Aviso: nao salvou resultado user {user_id}: {e}")


def _carregar_resultado_usuario(user_id):
    """Carrega ultimo_resultado do usuario se existir no disco."""
    try:
        path = _resultado_path(user_id)
        if path and os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                resultado_salvo = json.load(f)
            estado = estados.get(int(user_id))
            if estado and not estado["ultimo_resultado"]:
                estado["ultimo_resultado"] = resultado_salvo
    except Exception as e:
        print(f"Aviso: nao carregou resultado user {user_id}: {e}")
    # Carregar ultimo_lucro salvo em disco
    _carregar_lucro_usuario(user_id)


def _lucro_path(user_id):
    """Caminho do arquivo de lucro do usuario."""
    user = User.query.get(int(user_id))
    if not user:
        return None
    pasta = user.get_pasta_saida()
    return os.path.join(pasta, "_ultimo_lucro.json")


def _salvar_lucro_usuario(user_id):
    """Salva ultimo_lucro do usuario em JSON no volume persistente."""
    try:
        estado = estados.get(int(user_id))
        if not estado or not estado.get("ultimo_lucro"):
            return
        path = _lucro_path(user_id)
        if path:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(estado["ultimo_lucro"], f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Aviso: nao salvou lucro user {user_id}: {e}")


def _carregar_lucro_usuario(user_id):
    """Carrega ultimo_lucro do usuario se existir no disco."""
    try:
        path = _lucro_path(user_id)
        if path and os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                lucro_salvo = json.load(f)
            estado = estados.get(int(user_id))
            if estado and not estado.get("ultimo_lucro"):
                estado["ultimo_lucro"] = lucro_salvo
    except Exception as e:
        print(f"Aviso: nao carregou lucro user {user_id}: {e}")


def adicionar_log(estado, msg, tipo="info"):
    """Adiciona mensagem ao log do usuario."""
    estado["logs"].append({
        "timestamp": datetime.now().strftime("%H:%M:%S"),
        "mensagem": msg,
        "tipo": tipo
    })
    if len(estado["logs"]) > 500:
        estado["logs"] = estado["logs"][-500:]


# ----------------------------------------------------------------
# ROTAS PUBLICAS
# ----------------------------------------------------------------

@app.route('/version')
def version():
    """Retorna versão do código em execução."""
    _DEPLOY_ID = "2026-03-05-v3-db-pending"
    version_info = {
        'version': _DEPLOY_ID,
        'build': 'shopee-oauth-db-pending',
        'deploy_id': _DEPLOY_ID,
        'features': [
            'OAuth pending via DB (cross-worker)',
            'Callback timing logs (ms)',
            'Gunicorn --workers 1',
            'Etiquetas + Retirada + PyMuPDF 1.24.14',
        ],
        'startup_time': _agora_brasil().strftime("%Y-%m-%d %H:%M:%S BRT"),
    }
    # Tentar ler arquivo VERSION se existir
    try:
        with open('VERSION', 'r') as f:
            version_info['file'] = f.read().strip()
    except:
        pass
    return jsonify(version_info)

@app.route('/')
def index():
    """Serve o dashboard com auto-login (sem tela de login)."""
    import sys
    print(f"[ROTA /] method={request.method} args={dict(request.args)} url={request.url} remote={request.remote_addr} host={request.host}", flush=True, file=sys.stderr)

    # Fallback: se a Shopee redirecionar para / com code+shop_id (sem o path do callback),
    # encaminhar para o callback real.
    code = request.args.get("code", "")
    shop_id_arg = request.args.get("shop_id", "")
    if code and shop_id_arg:
        from urllib.parse import urlencode
        qs = urlencode(request.args)
        print(f"[ROTA /] FALLBACK SHOPEE -> redirecionando para /api/marketplace/shopee/callback?{qs}", flush=True, file=sys.stderr)
        return redirect(f"/api/marketplace/shopee/callback?{qs}")

    # Auto-login: gerar token para o primeiro usuario admin
    auto_token = ""
    auto_user_json = "{}"
    try:
        user = User.query.first()
        if user:
            token = create_access_token(identity=str(user.id))
            auto_token = token
            auto_user_json = json.dumps(user.to_dict())
    except Exception as e:
        print(f"[ROTA /] auto-login erro: {e}", flush=True, file=sys.stderr)

    # Ler index.html e injetar auto-login script antes do </head>
    import os as _os
    html_path = _os.path.join(app.static_folder, 'index.html')
    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()

    if auto_token:
        safe_user_json = auto_user_json.replace("'", "\\'")
        inject_script = (
            "<script>\n"
            "// Auto-login injetado pelo servidor\n"
            "localStorage.setItem('token', '" + auto_token + "');\n"
            "localStorage.setItem('user', '" + safe_user_json + "');\n"
            "</script>"
        )
        html = html.replace('</head>', inject_script + '\n</head>')

    resp = make_response(html)
    resp.content_type = 'text/html; charset=utf-8'
    resp.cache_control.no_store = True
    resp.cache_control.no_cache = True
    resp.cache_control.must_revalidate = True
    resp.cache_control.max_age = 0
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    resp.headers['X-Beka-Ui-Version'] = 'assinante-v2'
    return resp


@app.route('/login')
def login_page():
    # Redireciona para o dashboard — login automatico via /api/auth/local-token
    from flask import redirect
    return redirect('/')


@app.route('/auto-login')
def auto_login_page():
    """Auto-login para ambiente de desenvolvimento. Loga como admin@beka.com automaticamente."""
    DEV_EMAIL = 'admin@beka.com'
    user = User.query.filter_by(email=DEV_EMAIL).first()
    if not user:
        return redirect('/login')
    token = create_access_token(identity=str(user.id))
    user_data = user.to_dict() if hasattr(user, 'to_dict') else {"id": user.id, "email": user.email}
    return f"""<!DOCTYPE html>
<html><head><title>Auto Login</title></head><body>
<script>
localStorage.setItem('token', '{token}');
localStorage.setItem('user', JSON.stringify({json.dumps(user_data)}));
window.location.href = '/';
</script>
</body></html>"""


# ----------------------------------------------------------------
# ROTAS DA API (PROTEGIDAS COM JWT)
# ----------------------------------------------------------------

@app.route('/api/status')
@jwt_required()
def api_status():
    """Retorna o status atual do usuario."""
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404

    # Arquivos de etiquetas (pasta_entrada)
    pasta = estado["configuracoes"]["pasta_entrada"]
    arquivos = []
    if os.path.exists(pasta):
        for f in os.listdir(pasta):
            if f.startswith('_'):
                continue
            fp = os.path.join(pasta, f)
            if os.path.isfile(fp):
                ext = os.path.splitext(f)[1].lower()
                if ext in ('.pdf', '.xlsx', '.xls'):
                    tipo_arq = "PDF" if ext == '.pdf' else "XLSX"
                    tamanho = os.path.getsize(fp)
                    arquivos.append({
                        "nome": f,
                        "tipo": tipo_arq,
                        "tamanho": tamanho,
                        "tamanho_fmt": _formatar_tamanho(tamanho),
                    })

    # Arquivos de lucro (pasta_lucro) - separados
    pasta_lucro = estado["configuracoes"].get("pasta_lucro", "")
    arquivos_lucro = []
    if pasta_lucro and os.path.exists(pasta_lucro):
        for f in os.listdir(pasta_lucro):
            if f.startswith('_'):
                continue
            fp = os.path.join(pasta_lucro, f)
            if os.path.isfile(fp):
                ext = os.path.splitext(f)[1].lower()
                if ext in ('.zip', '.xml', '.xlsx', '.xls'):
                    if f == 'planilha_custos.xlsx':
                        tipo_arq = "CUSTOS"
                    elif ext == '.zip':
                        tipo_arq = "ZIP"
                    elif ext == '.xml':
                        tipo_arq = "XML"
                    else:
                        tipo_arq = "XLSX"
                    tamanho = os.path.getsize(fp)
                    arquivos_lucro.append({
                        "nome": f,
                        "tipo": tipo_arq,
                        "tamanho": tamanho,
                        "tamanho_fmt": _formatar_tamanho(tamanho),
                    })

    saidas = []
    pasta_saida = estado["configuracoes"]["pasta_saida"]
    if os.path.exists(pasta_saida):
        for loja in os.listdir(pasta_saida):
            pasta_loja = os.path.join(pasta_saida, loja)
            if os.path.isdir(pasta_loja):
                arquivos_loja = os.listdir(pasta_loja)
                saidas.append({
                    "loja": loja,
                    "arquivos": len(arquivos_loja),
                    "nomes": arquivos_loja,
                })

    # Info do usuario
    user = User.query.get(int(user_id))

    return jsonify({
        "processando": estado["processando"],
        "arquivos_entrada": arquivos,
        "arquivos_lucro": arquivos_lucro,
        "saidas": saidas,
        "ultimo_resultado": estado["ultimo_resultado"],
        "configuracoes": estado["configuracoes"],
        "agrupamentos": estado["agrupamentos"],
        "ultimo_lucro": estado.get("ultimo_lucro"),
        "user": user.to_dict() if user else None,
    })


@app.route('/api/logs')
@jwt_required()
def api_logs():
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    desde = request.args.get('desde', 0, type=int)
    return jsonify({
        "logs": estado["logs"][desde:],
        "total": len(estado["logs"]),
    })


@app.route('/api/escanear-lojas', methods=['POST'])
@jwt_required()
def api_escanear_lojas():
    """Escaneia PDFs enviados para identificar lojas sem processar tudo.

    Usa o processador real sem recorte de forma lightweight para
    extrair CNPJs e nomes de loja, sem gerar PDFs/XLSX de saida.
    """
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404

    pasta_entrada = estado["configuracoes"]["pasta_entrada"]
    if not os.path.exists(pasta_entrada):
        return jsonify({"erro": "Pasta de entrada nao encontrada"}), 404

    # Verificar se ha PDFs na pasta
    tem_pdf = any(f.lower().endswith('.pdf') and not f.startswith('_')
                  for f in os.listdir(pasta_entrada) if os.path.isfile(os.path.join(pasta_entrada, f)))
    if not tem_pdf:
        return jsonify({"erro": "Nenhum PDF encontrado na pasta de entrada"}), 400

    try:
        proc = ProcessadorEtiquetasShopee()
        proc.carregar_todos_xlsx(pasta_entrada)
        todas_etiquetas, cpf_auto, pdfs_shein = proc.carregar_todos_pdfs_sem_recorte(pasta_entrada)

        # Juntar CPF auto-detectadas
        etiquetas_cpf = proc.processar_cpf(pasta_entrada)
        etiquetas_cpf.extend(cpf_auto)
        if etiquetas_cpf:
            todas_etiquetas.extend(etiquetas_cpf)

        # Separar por loja
        lojas_dict = proc.separar_por_loja(todas_etiquetas)

        lojas = []
        for cnpj, etqs in lojas_dict.items():
            nome = proc.get_nome_loja(cnpj)
            lojas.append({"cnpj": cnpj, "nome": nome})

        # Adicionar Shein se houver
        if pdfs_shein:
            shein_etqs = proc.processar_shein(pasta_entrada, pdfs_extras=pdfs_shein)
            if shein_etqs:
                # Verificar se ha loja Shein no DB para consolidar
                _lojas_shein_db = Loja.query.filter(
                    Loja.user_id == int(user_id),
                    Loja.nome.ilike('%shein%')
                ).all()
                if _lojas_shein_db and len(_lojas_shein_db) == 1:
                    # 1 loja Shein no DB: todos os CNPJs (transportadoras) vao para ela
                    _nome_shein = _lojas_shein_db[0].nome
                    if not any(l['nome'] == _nome_shein for l in lojas):
                        lojas.append({"cnpj": "SHEIN", "nome": _nome_shein})
                else:
                    # Sem loja Shein no DB ou multiplas: listar por CNPJ
                    for etq in shein_etqs:
                        cnpj_s = etq.get('cnpj', 'SHEIN')
                        if not any(l['cnpj'] == cnpj_s for l in lojas):
                            nome_s = proc.get_nome_loja(cnpj_s)
                            lojas.append({"cnpj": cnpj_s, "nome": nome_s})

        lojas.sort(key=lambda x: x["nome"])
        return jsonify({"lojas": lojas})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"erro": f"Erro ao escanear: {str(e)}"}), 500


@app.route('/api/processar', methods=['POST'])
@jwt_required()
def api_processar():
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404

    if estado["processando"]:
        return jsonify({"erro": "Processamento ja em andamento"}), 409

    # Verificar limite do plano
    user = User.query.get(int(user_id))
    if not user.pode_processar():
        info = user.get_plano_info()
        return jsonify({
            "erro": f"Limite de {info['limite_proc']} processamentos/mes atingido. Faca upgrade do plano!"
        }), 403

    thread = threading.Thread(
        target=_executar_processamento,
        args=(int(user_id),),
        kwargs={"sem_recorte": True, "resumo_sku_somente": False},
    )
    thread.daemon = True
    thread.start()

    return jsonify({"mensagem": "Processamento iniciado"})


@app.route('/api/configuracoes', methods=['POST'])
@jwt_required()
def api_configuracoes():
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    dados = request.get_json()
    if dados:
        for chave, valor in dados.items():
            if chave in estado["configuracoes"] and chave not in ('pasta_entrada', 'pasta_saida'):
                estado["configuracoes"][chave] = valor
        _salvar_config_usuario(user_id)
        adicionar_log(estado, "Configuracoes atualizadas e salvas", "success")
    return jsonify(estado["configuracoes"])


@app.route('/api/configuracoes-lucro-lojas', methods=['GET', 'POST'])
@jwt_required()
def api_config_lucro_lojas():
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    if request.method == 'GET':
        return jsonify({
            "defaults": {
                "perc_declarado": estado["configuracoes"].get("perc_declarado", 100),
                "taxa_shopee": estado["configuracoes"].get("taxa_shopee", 18),
                "imposto_simples": estado["configuracoes"].get("imposto_simples", 4),
                "custo_fixo": estado["configuracoes"].get("custo_fixo", 3),
            },
            "por_loja": estado["configuracoes"].get("lucro_por_loja", {}),
        })
    dados = request.get_json()
    if dados:
        estado["configuracoes"]["lucro_por_loja"] = dados.get("por_loja", {})
        _salvar_config_usuario(user_id)
        adicionar_log(estado, "Config lucro por loja atualizada", "success")
    return jsonify({"ok": True})


@app.route('/api/historico')
@jwt_required()
def api_historico():
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    return jsonify({"historico": estado["historico"]})


@app.route('/api/historico-geradas', methods=['GET'])
@jwt_required()
def api_historico_geradas():
    """
    Lista snapshots de geracoes disponiveis para download (janela padrao: 24h).
    """
    user_id = int(get_jwt_identity())
    horas = request.args.get("horas", 24, type=int)
    horas = 24 if not horas else max(1, min(int(horas), 168))
    itens = _listar_historico_geradas(user_id, horas=horas)
    return jsonify({
        "historico": itens,
        "janela_horas": horas,
        "total": len(itens),
    })


@app.route('/api/historico-geradas/download/<item_id>', methods=['GET'])
@jwt_required()
def api_historico_geradas_download(item_id):
    """
    Download de um snapshot de geracao do historico (ultimas 24h).
    """
    user_id = int(get_jwt_identity())
    _limpar_historico_geradas_expirado(user_id, horas=24)
    itens = _carregar_historico_geradas_raw(user_id)
    match = next((it for it in itens if str((it or {}).get("id", "")) == str(item_id)), None)
    if not match:
        return jsonify({"erro": "Arquivo de historico nao encontrado ou expirado"}), 404

    dt = _parse_iso_dt((match or {}).get("created_at"))
    if dt is None or dt < (_agora_brasil() - timedelta(hours=24)):
        return jsonify({"erro": "Arquivo expirado (janela maxima: 24h)"}), 410

    pasta = _historico_geradas_dir(user_id)
    if not pasta:
        return jsonify({"erro": "Pasta de historico indisponivel"}), 404

    arquivo = os.path.basename((match or {}).get("arquivo", ""))
    if not arquivo:
        return jsonify({"erro": "Arquivo invalido no historico"}), 404

    caminho = os.path.realpath(os.path.join(pasta, arquivo))
    base = os.path.realpath(pasta)
    if not caminho.startswith(base):
        return jsonify({"erro": "Caminho invalido"}), 400
    if not os.path.exists(caminho):
        return jsonify({"erro": "Arquivo nao encontrado"}), 404

    return send_file(
        caminho,
        as_attachment=True,
        download_name=arquivo,
        mimetype="application/zip",
    )


@app.route('/api/abrir-pasta', methods=['POST'])
@jwt_required()
def api_abrir_pasta():
    dados = request.get_json()
    pasta = dados.get('pasta', '')
    if pasta and os.path.exists(pasta):
        try:
            os.startfile(pasta)
        except AttributeError:
            pass
        return jsonify({"ok": True})
    return jsonify({"erro": "Pasta nao encontrada"}), 404


@app.route('/api/upload', methods=['POST'])
@jwt_required()
def api_upload():
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404

    if 'arquivo' not in request.files:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400

    arquivo = request.files['arquivo']
    if arquivo.filename == '':
        return jsonify({"erro": "Nome de arquivo vazio"}), 400

    ext = os.path.splitext(arquivo.filename)[1].lower()
    if ext not in ('.pdf', '.xlsx', '.xls'):
        return jsonify({"erro": "Tipo de arquivo nao suportado. Use PDF, XLSX ou XLS."}), 400

    pasta = estado["configuracoes"]["pasta_entrada"]
    caminho = os.path.join(pasta, arquivo.filename)
    arquivo.save(caminho)
    adicionar_log(estado, f"Arquivo recebido: {arquivo.filename}", "success")
    return jsonify({"mensagem": f"Arquivo {arquivo.filename} salvo com sucesso"})


@app.route('/api/remover-arquivo', methods=['POST'])
@jwt_required()
def api_remover_arquivo():
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    dados = request.get_json()
    nome = dados.get('nome', '')
    if not nome:
        return jsonify({"erro": "Nome nao informado"}), 400
    caminho = os.path.join(estado["configuracoes"]["pasta_entrada"], nome)
    if os.path.exists(caminho):
        os.remove(caminho)
        adicionar_log(estado, f"Arquivo removido: {nome}", "warning")
        return jsonify({"ok": True})
    return jsonify({"erro": "Arquivo nao encontrado"}), 404


@app.route('/api/upload-lucro', methods=['POST'])
@jwt_required()
def api_upload_lucro():
    """Upload de arquivos para calculadora de lucro (ZIP/XML)."""
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404

    if 'arquivo' not in request.files:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400

    arquivo = request.files['arquivo']
    if arquivo.filename == '':
        return jsonify({"erro": "Nome de arquivo vazio"}), 400

    ext = os.path.splitext(arquivo.filename)[1].lower()
    if ext not in ('.zip', '.xml'):
        return jsonify({"erro": "Tipo nao suportado. Use ZIP ou XML."}), 400

    pasta_lucro = estado["configuracoes"].get("pasta_lucro", "")
    if not pasta_lucro:
        return jsonify({"erro": "Pasta de lucro nao configurada"}), 500
    os.makedirs(pasta_lucro, exist_ok=True)

    caminho = os.path.join(pasta_lucro, arquivo.filename)
    arquivo.save(caminho)
    adicionar_log(estado, f"Arquivo lucro recebido: {arquivo.filename}", "success")
    return jsonify({"mensagem": f"Arquivo {arquivo.filename} salvo com sucesso"})


@app.route('/api/remover-arquivo-lucro', methods=['POST'])
@jwt_required()
def api_remover_arquivo_lucro():
    """Remove arquivo da pasta de lucro."""
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    dados = request.get_json()
    nome = dados.get('nome', '')
    if not nome:
        return jsonify({"erro": "Nome nao informado"}), 400
    pasta_lucro = estado["configuracoes"].get("pasta_lucro", "")
    if not pasta_lucro:
        return jsonify({"erro": "Pasta de lucro nao configurada"}), 500
    caminho = os.path.join(pasta_lucro, nome)
    if os.path.exists(caminho):
        os.remove(caminho)
        adicionar_log(estado, f"Arquivo lucro removido: {nome}", "warning")
        return jsonify({"ok": True})
    return jsonify({"erro": "Arquivo nao encontrado"}), 404


@app.route('/api/limpar-lucro', methods=['POST'])
@jwt_required()
def api_limpar_lucro():
    """Remove todos os arquivos da pasta de lucro."""
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    pasta_lucro = estado["configuracoes"].get("pasta_lucro", "")
    if not pasta_lucro or not os.path.exists(pasta_lucro):
        return jsonify({"ok": True})
    removidos = 0
    for f in os.listdir(pasta_lucro):
        if f.startswith('_') or f == 'planilha_custos.xlsx':
            continue
        fp = os.path.join(pasta_lucro, f)
        if os.path.isfile(fp):
            os.remove(fp)
            removidos += 1
    estado["ultimo_lucro"] = None
    adicionar_log(estado, f"Pasta de lucro limpa ({removidos} arquivos removidos)", "warning")
    return jsonify({"ok": True, "removidos": removidos})


@app.route('/api/limpar-saida', methods=['POST'])
@jwt_required()
def api_limpar_saida():
    import shutil
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    pasta = estado["configuracoes"]["pasta_saida"]
    if os.path.exists(pasta):
        shutil.rmtree(pasta)
        os.makedirs(pasta, exist_ok=True)
        adicionar_log(estado, "Pasta de saida limpa", "warning")
    return jsonify({"ok": True})


@app.route('/api/novo-lote', methods=['POST'])
@jwt_required()
def api_novo_lote():
    """Limpa todos os arquivos de entrada, saida e reseta o estado para novo processamento."""
    import shutil
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404

    if estado["processando"]:
        return jsonify({"erro": "Aguarde o processamento atual terminar"}), 409

    pasta_entrada = estado["configuracoes"]["pasta_entrada"]
    pasta_saida = estado["configuracoes"]["pasta_saida"]

    # Limpar pasta de entrada de etiquetas (exceto _config.json)
    # Pasta de lucro NAO e limpa (independente)
    if os.path.exists(pasta_entrada):
        for f in os.listdir(pasta_entrada):
            if f.startswith('_'):
                continue  # Preservar _config.json etc
            fp = os.path.join(pasta_entrada, f)
            if os.path.isfile(fp):
                os.remove(fp)

    # Limpar pasta de saida
    if os.path.exists(pasta_saida):
        shutil.rmtree(pasta_saida)
        os.makedirs(pasta_saida, exist_ok=True)

    # Resetar estado
    estado["ultimo_resultado"] = None
    estado["ultimo_lucro"] = None
    estado["logs"] = []
    estado["_etiquetas_por_cnpj"] = {}
    estado["_proc_config"] = {}

    # Limpar arquivos de resultado persistidos
    rp = _resultado_path(user_id)
    if rp and os.path.exists(rp):
        os.remove(rp)
    lp = _lucro_path(user_id)
    if lp and os.path.exists(lp):
        os.remove(lp)

    adicionar_log(estado, "Novo lote iniciado - arquivos limpos", "success")
    return jsonify({"ok": True, "mensagem": "Pronto para novo lote"})


@app.route('/api/download-todos')
@jwt_required()
def api_download_todos():
    """Gera um ZIP com todos os arquivos de saida (PDFs + XLSXs de todas as lojas)."""
    import zipfile
    import io
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    pasta_saida = estado["configuracoes"]["pasta_saida"]
    if not os.path.exists(pasta_saida):
        return jsonify({"erro": "Nenhum resultado disponivel"}), 404

    buf = io.BytesIO()
    arquivos_adicionados = 0
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(pasta_saida):
            for f in files:
                if f.startswith('_'):
                    continue
                filepath = os.path.join(root, f)
                arcname = os.path.relpath(filepath, pasta_saida)
                zf.write(filepath, arcname)
                arquivos_adicionados += 1

    if arquivos_adicionados == 0:
        return jsonify({"erro": "Nenhum arquivo de resultado encontrado"}), 404

    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="Etiquetas_prontas.zip", mimetype="application/zip")


@app.route('/api/download/<loja>/<arquivo>')
@jwt_required()
def api_download(loja, arquivo):
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    pasta = os.path.join(estado["configuracoes"]["pasta_saida"], loja)
    caminho = os.path.join(pasta, arquivo)
    if os.path.exists(caminho):
        return send_file(caminho, as_attachment=True)
    return jsonify({"erro": "Arquivo nao encontrado"}), 404


@app.route('/api/download-resumo-geral')
@jwt_required()
def api_download_resumo_geral():
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    if not estado["ultimo_resultado"] or "resumo_geral" not in estado["ultimo_resultado"]:
        return jsonify({"erro": "Nenhum resumo geral disponivel"}), 404
    arquivo = estado["ultimo_resultado"]["resumo_geral"]["arquivo"]
    pasta = estado["configuracoes"]["pasta_saida"]
    caminho = os.path.join(pasta, arquivo)
    if os.path.exists(caminho):
        return send_file(caminho, as_attachment=True)
    return jsonify({"erro": "Arquivo nao encontrado"}), 404


@app.route('/api/exemplo-custos')
def api_exemplo_custos():
    """Gera e retorna um XLSX de exemplo para a planilha de custos."""
    import io
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Custos"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    ws.append(["SKU", "Custo Unitario (R$)"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    exemplos = [
        ("PROD001", 12.50),
        ("PROD002", 8.90),
        ("ABC123", 15.00),
        ("KIT-A", 25.00),
        ("CAMISETA-P", 18.50),
        ("CAMISETA-M", 18.50),
        ("CAMISETA-G", 19.00),
        ("CANECA01", 7.80),
    ]
    for sku, custo in exemplos:
        ws.append([sku, custo])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            if cell.col_idx == 2:
                cell.number_format = 'R$ #,##0.00'

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="exemplo_planilha_custos.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/api/upload-custos', methods=['POST'])
@jwt_required()
def api_upload_custos():
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404

    if 'arquivo' not in request.files:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400
    arquivo = request.files['arquivo']
    if arquivo.filename == '':
        return jsonify({"erro": "Nome de arquivo vazio"}), 400
    ext = os.path.splitext(arquivo.filename)[1].lower()
    if ext != '.xlsx':
        return jsonify({"erro": "Envie um arquivo .xlsx"}), 400

    pasta_lucro = estado["configuracoes"].get("pasta_lucro", "")
    if not pasta_lucro:
        return jsonify({"erro": "Pasta de lucro nao configurada"}), 500
    os.makedirs(pasta_lucro, exist_ok=True)
    caminho = os.path.join(pasta_lucro, "planilha_custos.xlsx")
    arquivo.save(caminho)
    estado["configuracoes"]["planilha_custos"] = caminho
    _salvar_config_usuario(user_id)
    adicionar_log(estado, f"Planilha de custos recebida: {arquivo.filename}", "success")
    return jsonify({"mensagem": "Planilha de custos salva", "caminho": caminho})


def _limpar_nome_loja(nome_raw):
    nome = _re.sub(r'^\d[\d.]+\s+', '', nome_raw)
    nome = _re.sub(r'\s+\d{11}$', '', nome)
    nome = _re.sub(r'\s+(LTDA|ME|MEI|EPP|EIRELI)\s*$', '', nome, flags=_re.IGNORECASE)
    nome = nome.strip().title()
    nome = _re.sub(r'[<>:"/\\|?*]', '', nome)
    return nome.strip() or 'Loja_Desconhecida'


def _formatar_cnpj_curto(cnpj):
    """Retorna a raiz do CNPJ formatada (8 primeiros digitos: XX.XXX.XXX)."""
    cnpj = str(cnpj).strip()
    if len(cnpj) >= 8:
        return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}"
    return cnpj


def _extrair_loja_nfe(nfe, cnpj_loja_map=None):
    emit = nfe.get("emit", {})
    if isinstance(emit, str):
        return "Desconhecida"
    # Tentar usar nome da loja Shopee (mesmo mapeamento do processamento principal)
    cnpj = str(emit.get("CNPJ", "")).strip()
    if cnpj_loja_map and cnpj in cnpj_loja_map:
        return cnpj_loja_map[cnpj]
    nome_raw = str(emit.get("xNome", "")).strip()
    nome = _limpar_nome_loja(nome_raw) if nome_raw else "Desconhecida"
    # Incluir CNPJ curto para diferenciar lojas com mesmo nome empresarial
    if cnpj:
        return f"{_formatar_cnpj_curto(cnpj)} {nome}"
    return nome


def _extrair_sku_principal(sku_completo):
    """Extrai o SKU principal (base) de um SKU completo.
    Estrategia: remover do FINAL APENAS a ultima parte se for claramente variacao.
    Variacao = tamanho roupa (P/M/G/GG/XG/PP/XL) OU numero de 2 digitos (34-50 = tamanho).
    Ex: 'TEN-BO-BR-38' -> 'TEN-BO-BR' (38 e tamanho roupa)
        'PROD-AZUL-M'  -> 'PROD-AZUL' (M e tamanho)
        'SKU-01-42'    -> 'SKU-01'    (42 e tamanho)
        'TEN-BO-BR'    -> 'TEN-BO-BR' (BR nao e tamanho, mantem)
    """
    if not sku_completo:
        return sku_completo
    partes = sku_completo.split('-')
    if len(partes) <= 1:
        return sku_completo

    # Tamanhos de roupa conhecidos
    tamanhos_letra = {'P', 'M', 'G', 'PP', 'GG', 'XG', 'XS', 'XL', 'XXL', 'XXG', 'EG', 'EGG'}

    # Remover do final apenas partes que sao claramente variacao
    i = len(partes) - 1
    while i > 0:
        p = partes[i].strip().upper()
        # Tamanho letra (P, M, G, GG, etc.)
        if p in tamanhos_letra:
            i -= 1
            continue
        # Numero 2 digitos no range tipico de tamanho (24-56)
        if _re.match(r'^\d{2}$', p) and 24 <= int(p) <= 56:
            i -= 1
            continue
        # Numero 1 digito solto (1-9) - pode ser variacao de quantidade
        if _re.match(r'^\d$', p):
            i -= 1
            continue
        break
    base = partes[:i + 1]
    return '-'.join(base) if base else sku_completo


def _buscar_custo_inteligente(sku_xml, dict_custos, chaves_ordenadas):
    """Busca inteligente de custo: SKU principal, match exato, prefixo."""
    sku_upper = sku_xml.upper().strip()
    if not sku_upper:
        return 0.0, False

    # 1. Match exato do SKU completo
    if sku_upper in dict_custos:
        return dict_custos[sku_upper], True

    # 2. Extrair SKU principal (sem sufixos de variacao) e tentar match exato
    sku_base = _extrair_sku_principal(sku_upper)
    if sku_base != sku_upper and sku_base in dict_custos:
        return dict_custos[sku_base], True

    # 3. SKU da planilha comeca com o SKU base (planilha tem chave mais longa)
    for chave in chaves_ordenadas:
        if chave.startswith(sku_base):
            return dict_custos[chave], True

    # 4. SKU do XML comeca com chave da planilha (planilha tem chave mais curta)
    for chave in chaves_ordenadas:
        if sku_upper.startswith(chave):
            return dict_custos[chave], True

    # 5. SKU base comeca com chave da planilha
    if sku_base != sku_upper:
        for chave in chaves_ordenadas:
            if sku_base.startswith(chave):
                return dict_custos[chave], True

    return 0.0, False


def _processar_nfe_lucro(nfe, dict_custos, cfg, cfg_por_loja, chaves_ordenadas=None, cnpj_loja_map=None):
    nome_loja = _extrair_loja_nfe(nfe, cnpj_loja_map)
    cfg_loja = cfg_por_loja.get(nome_loja, {})
    perc_declarado = float(cfg_loja.get("perc_declarado", cfg.get("perc_declarado", 100))) / 100
    taxa_shopee = float(cfg_loja.get("taxa_shopee", cfg.get("taxa_shopee", 18))) / 100
    taxa_imposto = float(cfg_loja.get("imposto_simples", cfg.get("imposto_simples", 4))) / 100
    custo_fixo = float(cfg_loja.get("custo_fixo", cfg.get("custo_fixo", 3)))

    if chaves_ordenadas is None:
        chaves_ordenadas = sorted(dict_custos.keys(), key=len, reverse=True)

    dets = nfe.get("det", [])
    if not isinstance(dets, list):
        dets = [dets]

    # Extrair SKU principal (primeiro item da NF) para busca de custo
    sku_principal = ""
    if dets:
        prod_primeiro = dets[0].get("prod", {})
        sku_principal = str(prod_primeiro.get("cProd", "")).strip()

    # Buscar custo usando apenas o SKU principal
    c_principal_unit, encontrou_principal = _buscar_custo_inteligente(sku_principal, dict_custos, chaves_ordenadas)

    # Somar todos os itens da NF como uma unica linha usando o SKU principal
    qtd_total = 0
    v_declarado_total = 0
    for item in dets:
        prod = item.get("prod", {})
        qtd_total += float(prod.get("qCom", 1))
        v_declarado_total += float(prod.get("vProd", 0))

    c_produto_total = c_principal_unit * qtd_total

    v_real = v_declarado_total / perc_declarado if perc_declarado > 0 else v_declarado_total
    c_imposto = v_declarado_total * taxa_imposto
    c_shopee = v_real * taxa_shopee
    c_fixo_total = custo_fixo * qtd_total
    lucro = v_real - c_imposto - c_shopee - c_fixo_total - c_produto_total

    itens = [{
        "SKU": sku_principal,
        "Qtd": qtd_total,
        "V. Real": round(v_real, 2),
        "V. Decl.": round(v_declarado_total, 2),
        "Custo": round(c_produto_total, 2),
        "Shopee": round(c_shopee, 2),
        "Imposto": round(c_imposto, 2),
        "Custo Fixo": round(c_fixo_total, 2),
        "LUCRO": round(lucro, 2),
    }]
    sem_custo = [0] if not encontrou_principal else []

    return nome_loja, itens, sem_custo


def _construir_mapa_cnpj_lojas(pasta_lucro):
    """Varre XMLs da pasta_lucro e retorna {cnpj_completo: 'XX.XXX.XXX nome_limpo'}.

    Usado por /api/lojas-lucro (config) e api_gerar_lucro (processamento)
    para garantir nomes de loja consistentes.
    """
    import zipfile
    lojas_dict = {}  # {cnpj: nome_limpo}

    def _extrair_emit(conteudo):
        try:
            doc = xmltodict.parse(conteudo)
            if "nfeProc" in doc:
                nfe = doc["nfeProc"]["NFe"]["infNFe"]
            elif "NFe" in doc:
                nfe = doc["NFe"]["infNFe"]
            else:
                return
            emit = nfe.get("emit", {})
            if isinstance(emit, str):
                return
            cnpj = str(emit.get("CNPJ", "")).strip()
            nome_raw = str(emit.get("xNome", "")).strip()
            nome = _limpar_nome_loja(nome_raw) if nome_raw else "Desconhecida"
            if cnpj:
                lojas_dict[cnpj] = nome
        except Exception:
            pass

    if not pasta_lucro or not os.path.exists(pasta_lucro):
        return {}

    for f in os.listdir(pasta_lucro):
        fp = os.path.join(pasta_lucro, f)
        if f.lower().endswith('.zip') and zipfile.is_zipfile(fp):
            try:
                with zipfile.ZipFile(fp) as zf:
                    for nome_arq in zf.namelist():
                        if nome_arq.lower().endswith('.xml'):
                            conteudo = zf.read(nome_arq).decode('utf-8', errors='ignore')
                            _extrair_emit(conteudo)
            except Exception:
                pass
        elif f.lower().endswith('.xml'):
            try:
                with open(fp, 'r', encoding='utf-8', errors='ignore') as xf:
                    _extrair_emit(xf.read())
            except Exception:
                pass

    return {c: f"{_formatar_cnpj_curto(c)} {n}" for c, n in lojas_dict.items()}


@app.route('/api/lojas-lucro')
@jwt_required()
def api_lojas_lucro():
    """Retorna lista de lojas encontradas nos XMLs da pasta_lucro, separadas por CNPJ."""
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404

    cfg = estado["configuracoes"]
    pasta_lucro = cfg.get("pasta_lucro", "")

    mapa = _construir_mapa_cnpj_lojas(pasta_lucro)
    lojas = [{"cnpj": c, "nome": n} for c, n in sorted(mapa.items(), key=lambda x: x[1])]

    return jsonify({"lojas": lojas})


@app.route('/api/gerar-lucro', methods=['POST'])
@jwt_required()
def api_gerar_lucro():
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404

    cfg = estado["configuracoes"]
    pasta_lucro = cfg.get("pasta_lucro", "")
    pasta_saida = cfg["pasta_saida"]
    caminho_custos = cfg.get("planilha_custos", "")

    if not caminho_custos or not os.path.exists(caminho_custos):
        return jsonify({"erro": "Planilha de custos nao encontrada. Faca upload primeiro."}), 400

    try:
        adicionar_log(estado, "Gerando relatorio de lucro...", "info")

        df_custos = pd.read_excel(caminho_custos)
        dict_custos = {}
        for _, row in df_custos.iterrows():
            sku_original = str(row.iloc[0]).strip().upper()
            custo = float(row.iloc[1]) if pd.notnull(row.iloc[1]) else 0.0
            if sku_original:
                dict_custos[sku_original] = custo
        # Pre-ordenar chaves por tamanho (maior primeiro) para busca inteligente
        _chaves_custos_ordenadas = sorted(dict_custos.keys(), key=len, reverse=True)

        cfg_por_loja = cfg.get("lucro_por_loja", {})

        import zipfile
        loja_dados = defaultdict(lambda: {"itens": [], "linhas_sem_custo": []})
        chaves_processadas = set()
        total_xmls_lidos = 0
        # Construir mapa CNPJ->nome inline durante processamento (evita ler XMLs 2x)
        _cnpj_loja_map = {}

        def _processar_doc(doc):
            nonlocal total_xmls_lidos
            if "nfeProc" in doc:
                nfe = doc["nfeProc"]["NFe"]["infNFe"]
            elif "NFe" in doc:
                nfe = doc["NFe"]["infNFe"]
            else:
                return
            total_xmls_lidos += 1

            # Construir mapa CNPJ->nome enquanto processa (mesma logica de _construir_mapa_cnpj_lojas)
            emit = nfe.get("emit", {})
            if isinstance(emit, dict):
                cnpj_emit = str(emit.get("CNPJ", "")).strip()
                if cnpj_emit and cnpj_emit not in _cnpj_loja_map:
                    nome_raw = str(emit.get("xNome", "")).strip()
                    nome = _limpar_nome_loja(nome_raw) if nome_raw else "Desconhecida"
                    _cnpj_loja_map[cnpj_emit] = f"{_formatar_cnpj_curto(cnpj_emit)} {nome}"

            # Deduplicar por chave NFe (evita contar mesma NF 2x)
            chave_nfe = str(nfe.get("@Id", "")).strip()
            if not chave_nfe:
                # Fallback: CNPJ + nNF
                cnpj_e = str(emit.get("CNPJ", "")).strip() if isinstance(emit, dict) else ""
                nf_num = str(nfe.get("ide", {}).get("nNF", "")).strip()
                chave_nfe = f"{cnpj_e}_{nf_num}" if cnpj_e and nf_num else ""
            if chave_nfe:
                if chave_nfe in chaves_processadas:
                    return  # Duplicada — ignorar
                chaves_processadas.add(chave_nfe)

            nome_loja, itens, sem_custo = _processar_nfe_lucro(nfe, dict_custos, cfg, cfg_por_loja, _chaves_custos_ordenadas, _cnpj_loja_map)
            offset = len(loja_dados[nome_loja]["itens"])
            loja_dados[nome_loja]["itens"].extend(itens)
            loja_dados[nome_loja]["linhas_sem_custo"].extend([i + offset for i in sem_custo])

        zips = [f for f in os.listdir(pasta_lucro) if f.lower().endswith('.zip')] if pasta_lucro and os.path.exists(pasta_lucro) else []
        for z in zips:
            caminho_zip = os.path.join(pasta_lucro, z)
            try:
                with zipfile.ZipFile(caminho_zip, 'r') as zf:
                    for nome_xml in zf.namelist():
                        if not nome_xml.lower().endswith('.xml'):
                            continue
                        try:
                            conteudo = zf.read(nome_xml).decode('utf-8')
                            doc = xmltodict.parse(conteudo)
                            _processar_doc(doc)
                        except Exception:
                            continue
            except Exception as e:
                adicionar_log(estado, f"Erro ao ler ZIP {z}: {e}", "warning")

        xmls_avulsos = [f for f in os.listdir(pasta_lucro) if f.lower().endswith('.xml')] if pasta_lucro and os.path.exists(pasta_lucro) else []
        for arq in xmls_avulsos:
            caminho_xml = os.path.join(pasta_lucro, arq)
            try:
                with open(caminho_xml, "r", encoding="utf-8") as f:
                    doc = xmltodict.parse(f.read())
                _processar_doc(doc)
            except Exception:
                continue

        # Log de duplicatas ignoradas
        duplicados = total_xmls_lidos - len(chaves_processadas)
        if duplicados > 0:
            adicionar_log(estado, f"{duplicados} NF(s) duplicada(s) ignorada(s) de {total_xmls_lidos} XMLs", "warning")
        adicionar_log(estado, f"{len(chaves_processadas)} NFs unicas processadas", "info")

        if not loja_dados:
            return jsonify({"erro": "Nenhum produto encontrado nos XMLs"}), 400

        # LOG DE SKUs PRINCIPAIS para conferencia do usuario
        adicionar_log(estado, "--- SKUs Principais extraidos (conferencia) ---", "info")
        skus_log = {}  # sku_principal -> set(sku_completos_originais)
        for nome_loja_l, dados_l in loja_dados.items():
            for item in dados_l["itens"]:
                sku_usado = item.get("SKU", "")
                if sku_usado:
                    sku_base = _extrair_sku_principal(sku_usado)
                    if sku_base not in skus_log:
                        skus_log[sku_base] = set()
                    skus_log[sku_base].add(sku_usado)
        for sku_base in sorted(skus_log.keys()):
            originais = sorted(skus_log[sku_base])
            if len(originais) == 1 and originais[0] == sku_base:
                adicionar_log(estado, f"  SKU: {sku_base}", "info")
            else:
                adicionar_log(estado, f"  SKU base: {sku_base} (de: {', '.join(originais)})", "info")
        # Mostrar chaves da planilha de custos para comparacao
        adicionar_log(estado, f"--- Planilha de custos: {len(dict_custos)} SKUs ---", "info")
        for sku_planilha in sorted(dict_custos.keys())[:20]:
            custo = dict_custos[sku_planilha]
            adicionar_log(estado, f"  Planilha: {sku_planilha} = R${custo:.2f}", "info")
        if len(dict_custos) > 20:
            adicionar_log(estado, f"  ... e mais {len(dict_custos) - 20} SKUs", "info")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        os.makedirs(pasta_saida, exist_ok=True)

        lojas_lucro = []
        lista_global = []
        linhas_sem_custo_global = []

        for nome_loja in sorted(loja_dados.keys()):
            dados_l = loja_dados[nome_loja]
            itens_l = dados_l["itens"]
            sem_custo_l = dados_l["linhas_sem_custo"]

            if not itens_l:
                continue

            pasta_loja = os.path.join(pasta_saida, nome_loja)
            os.makedirs(pasta_loja, exist_ok=True)
            caminho_loja_xlsx = os.path.join(pasta_loja, f"lucro_{nome_loja}_{timestamp}.xlsx")

            df_loja = pd.DataFrame(itens_l)
            df_loja = df_loja.sort_values('SKU', na_position='last').reset_index(drop=False)
            sem_custo_l_sorted = [i for i, orig in enumerate(df_loja['index']) if orig in sem_custo_l]
            df_loja = df_loja.drop(columns=['index'])
            totais_l = df_loja.sum(numeric_only=True)
            totais_l["SKU"] = "TOTAIS"
            df_loja = pd.concat([df_loja, pd.DataFrame([totais_l])], ignore_index=True)
            df_loja.to_excel(caminho_loja_xlsx, index=False)
            _formatar_excel_lucro(caminho_loja_xlsx, sem_custo_l_sorted)

            lucro_l = round(float(totais_l.get("LUCRO", 0)), 2)
            receita_l = round(float(totais_l.get("V. Real", 0)), 2)
            custo_l = round(float(totais_l.get("Custo", 0)), 2)

            lojas_lucro.append({
                "nome": nome_loja,
                "lucro": lucro_l,
                "receita": receita_l,
                "custo": custo_l,
                "total_itens": len(itens_l),
                "itens_sem_custo": len(sem_custo_l),
                "arquivo": f"lucro_{nome_loja}_{timestamp}.xlsx",
            })

            adicionar_log(estado, f"  Lucro {nome_loja}: {len(itens_l)} itens, R$ {lucro_l:.2f}", "info")

            offset_g = len(lista_global)
            for item in itens_l:
                lista_global.append(item)
            linhas_sem_custo_global.extend([i + offset_g for i in sem_custo_l])

        df_global = pd.DataFrame(lista_global)
        df_global = df_global.sort_values('SKU', na_position='last').reset_index(drop=False)
        sem_custo_global_sorted = [i for i, orig in enumerate(df_global['index']) if orig in linhas_sem_custo_global]
        df_global = df_global.drop(columns=['index'])
        totais_g = df_global.sum(numeric_only=True)
        totais_g["SKU"] = "TOTAIS"
        df_global = pd.concat([df_global, pd.DataFrame([totais_g])], ignore_index=True)

        caminho_xlsx = os.path.join(pasta_saida, f"relatorio_lucro_{timestamp}.xlsx")
        df_global.to_excel(caminho_xlsx, index=False)
        _formatar_excel_lucro(caminho_xlsx, sem_custo_global_sorted)

        lucro_total = round(float(totais_g.get("LUCRO", 0)), 2)
        receita_total = round(float(totais_g.get("V. Real", 0)), 2)
        custo_total = round(float(totais_g.get("Custo", 0)), 2)
        total_itens = len(lista_global)
        itens_sem_custo = len(linhas_sem_custo_global)

        estado["ultimo_lucro"] = {
            "arquivo": f"relatorio_lucro_{timestamp}.xlsx",
            "lucro_total": lucro_total,
            "receita_total": receita_total,
            "custo_total": custo_total,
            "total_itens": total_itens,
            "itens_sem_custo": itens_sem_custo,
            "timestamp": timestamp,
            "lojas": lojas_lucro,
        }

        # Salvar lucro em disco para persistir entre deploys
        _salvar_lucro_usuario(user_id)

        adicionar_log(estado, f"Relatorio de lucro gerado: {total_itens} itens, {len(lojas_lucro)} lojas, Lucro: R$ {lucro_total:.2f}", "success")

        return jsonify({
            "mensagem": "Relatorio gerado com sucesso",
            "lucro_total": lucro_total,
            "receita_total": receita_total,
            "custo_total": custo_total,
            "total_itens": total_itens,
            "itens_sem_custo": itens_sem_custo,
            "arquivo": f"relatorio_lucro_{timestamp}.xlsx",
            "lojas": lojas_lucro,
        })

    except Exception as e:
        import traceback
        adicionar_log(estado, f"Erro ao gerar lucro: {str(e)}", "error")
        adicionar_log(estado, traceback.format_exc(), "error")
        return jsonify({"erro": str(e)}), 500


@app.route('/api/download-lucro')
@jwt_required()
def api_download_lucro():
    """Gera ZIP com XLSX de lucro separado por loja + consolidado na raiz."""
    import zipfile as zf_mod
    import io as io_mod
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    lucro = estado.get("ultimo_lucro")
    if not lucro:
        return jsonify({"erro": "Nenhum relatorio de lucro disponivel"}), 404

    pasta_saida = estado["configuracoes"]["pasta_saida"]
    lojas = lucro.get("lojas", [])

    buf = io_mod.BytesIO()
    arquivos_adicionados = 0
    with zf_mod.ZipFile(buf, 'w', zf_mod.ZIP_DEFLATED) as zf:
        # Consolidado na raiz
        caminho_consolidado = os.path.join(pasta_saida, lucro["arquivo"])
        if os.path.exists(caminho_consolidado):
            zf.write(caminho_consolidado, lucro["arquivo"])
            arquivos_adicionados += 1

        # XLSX separado por loja em pastas
        for loja_info in lojas:
            nome_loja = loja_info.get("nome", "")
            arquivo_loja = loja_info.get("arquivo", "")
            if nome_loja and arquivo_loja:
                caminho_loja = os.path.join(pasta_saida, nome_loja, arquivo_loja)
                if os.path.exists(caminho_loja):
                    zf.write(caminho_loja, os.path.join(nome_loja, arquivo_loja))
                    arquivos_adicionados += 1

    if arquivos_adicionados == 0:
        return jsonify({"erro": "Nenhum arquivo de lucro encontrado"}), 404

    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="Relatorio_Lucro.zip", mimetype="application/zip")


@app.route('/api/download-lucro/<loja>')
@jwt_required()
def api_download_lucro_loja(loja):
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    pasta = os.path.join(estado["configuracoes"]["pasta_saida"], loja)
    if not os.path.exists(pasta):
        return jsonify({"erro": "Pasta da loja nao encontrada"}), 404
    arquivos = [f for f in os.listdir(pasta) if f.startswith("lucro_") and f.endswith(".xlsx")]
    if not arquivos:
        return jsonify({"erro": "Arquivo de lucro nao encontrado"}), 404
    arquivo = sorted(arquivos)[-1]
    return send_file(os.path.join(pasta, arquivo), as_attachment=True)


def _formatar_excel_lucro(caminho_arquivo, linhas_sem_custo):
    wb = openpyxl.load_workbook(caminho_arquivo)
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    alert_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    lucro_positivo = Font(color="006100", bold=True)
    lucro_negativo = Font(color="9C0006", bold=True)
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    max_row = ws.max_row
    max_col = ws.max_column
    sem_custo_set = set(linhas_sem_custo)  # Converter para set para O(1) lookup
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            if cell.col_idx >= 3:
                cell.number_format = 'R$ #,##0.00'
            idx_dados = cell.row - 2
            if idx_dados in sem_custo_set:
                cell.fill = alert_fill
            if cell.col_idx == max_col:
                if isinstance(cell.value, (int, float)) and cell.value >= 0:
                    cell.font = lucro_positivo
                elif isinstance(cell.value, (int, float)):
                    cell.font = lucro_negativo

    for cell in ws[max_row]:
        cell.fill = total_fill
        cell.font = Font(bold=True)

    for column in ws.columns:
        col_cells = [cell for cell in column]
        max_length = 0
        for cell in col_cells:
            try:
                val = cell.value
                if isinstance(val, (int, float)) and cell.col_idx >= 3:
                    display_len = len(f"R$ {val:,.2f}")
                else:
                    display_len = len(str(val)) if val is not None else 0
                if display_len > max_length:
                    max_length = display_len
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(max_length + 3, 12)

    wb.save(caminho_arquivo)


def _normalizar_nome_loja_saida(valor: str) -> str:
    txt = _re.sub(r"\s+", " ", str(valor or "")).strip().lower()
    if not txt:
        return ""
    return "".join(
        ch for ch in unicodedata.normalize("NFD", txt)
        if unicodedata.category(ch) != "Mn"
    )


def _resolver_nome_final_loja_saida(proc, cnpj: str, etiquetas_loja: list, estado: dict) -> str:
    """Resolve o nome final da loja, priorizando o nome UpSeller do PDF baixado."""
    pdf_loja_map = (estado or {}).get("_pdf_loja_map", {}) or {}
    for etq in (etiquetas_loja or []):
        pdf_path = str((etq or {}).get("caminho_pdf", "") or "").strip()
        if not pdf_path:
            continue
        nome_upseller = pdf_loja_map.get(os.path.basename(pdf_path))
        if nome_upseller:
            return str(nome_upseller).strip()
    return str(proc.get_nome_loja(cnpj) or cnpj or "Loja_Desconhecida").strip()


def _agrupar_lojas_para_saida(proc, lojas_por_cnpj: dict, estado: dict):
    """
    Consolida lojas que terminariam com o mesmo nome final na pasta de saida.

    Isso evita gerar dois PDFs/JPEGs para a mesma loja quando multiplos CNPJs
    ou nomes marketplace diferentes convergem para um unico nome UpSeller.
    """
    buckets = {}
    ordem = []

    for cnpj, etiquetas_loja in (lojas_por_cnpj or {}).items():
        nome_final = _resolver_nome_final_loja_saida(proc, cnpj, etiquetas_loja, estado)
        chave = _normalizar_nome_loja_saida(nome_final) or str(cnpj).strip().lower()
        bucket = buckets.get(chave)
        if bucket is None:
            bucket = {
                "nome": nome_final or str(cnpj).strip() or "Loja_Desconhecida",
                "cnpjs": [],
                "etiquetas": [],
            }
            buckets[chave] = bucket
            ordem.append(chave)

        cnpj_txt = str(cnpj or "").strip()
        if cnpj_txt and cnpj_txt not in bucket["cnpjs"]:
            bucket["cnpjs"].append(cnpj_txt)
        bucket["etiquetas"].extend(etiquetas_loja or [])

    lojas_saida = []
    for chave in ordem:
        bucket = buckets[chave]
        if len(bucket["cnpjs"]) > 1:
            antes = len(bucket["etiquetas"])
            try:
                bucket["etiquetas"], duplicadas = proc.remover_duplicatas(bucket["etiquetas"])
            except Exception:
                duplicadas = []
            removidas = max(0, antes - len(bucket["etiquetas"]))
            cnpjs_fmt = ", ".join(bucket["cnpjs"])
            msg = (
                f"Consolidando loja '{bucket['nome']}' em 1 saida "
                f"({len(bucket['cnpjs'])} origens: {cnpjs_fmt})"
            )
            if removidas > 0 or duplicadas:
                msg += f" | {removidas or len(duplicadas)} duplicata(s) removida(s)"
            adicionar_log(estado, msg, "warning")
        lojas_saida.append(bucket)

    return lojas_saida


@app.route('/api/agrupar', methods=['POST'])
@jwt_required()
def api_agrupar():
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404

    dados = request.get_json()
    if not dados:
        return jsonify({"erro": "Dados nao enviados"}), 400

    cnpjs = dados.get('cnpjs', [])
    nome_grupo = dados.get('nome', 'Agrupado').strip() or 'Agrupado'

    if len(cnpjs) < 2:
        return jsonify({"erro": "Selecione pelo menos 2 lojas"}), 400

    if not estado["ultimo_resultado"]:
        return jsonify({"erro": "Nenhum processamento realizado ainda"}), 400

    etiquetas_por_cnpj = estado.get("_etiquetas_por_cnpj", {})
    if not etiquetas_por_cnpj:
        return jsonify({"erro": "Dados de etiquetas nao disponiveis. Reprocesse."}), 400

    pasta_saida = estado["configuracoes"]["pasta_saida"]

    try:
        etiquetas_combinadas = []
        nomes_lojas = []
        for cnpj in cnpjs:
            etqs = etiquetas_por_cnpj.get(cnpj, [])
            if etqs:
                etiquetas_combinadas.extend(etqs)
                cfg = estado.get("_proc_config", {})
                nome = cfg.get("cnpj_loja", {}).get(cnpj) or cfg.get("cnpj_nome", {}).get(cnpj, cnpj)
                nomes_lojas.append(nome)

        if len(nomes_lojas) < 2:
            return jsonify({"erro": "Lojas selecionadas nao encontradas"}), 400

        proc = ProcessadorEtiquetasShopee()
        cfg = estado.get("_proc_config", {})
        proc.LARGURA_PT = cfg.get("largura_pt", proc.LARGURA_PT)
        proc.ALTURA_PT = cfg.get("altura_pt", proc.ALTURA_PT)
        proc.MARGEM_ESQUERDA = cfg.get("margem_esquerda", proc.MARGEM_ESQUERDA)
        proc.MARGEM_DIREITA = cfg.get("margem_direita", proc.MARGEM_DIREITA)
        proc.MARGEM_TOPO = cfg.get("margem_topo", proc.MARGEM_TOPO)
        proc.MARGEM_INFERIOR = cfg.get("margem_inferior", proc.MARGEM_INFERIOR)
        proc.fonte_produto = cfg.get("fonte_produto", proc.fonte_produto)
        proc.exibicao_produto = cfg.get("exibicao_produto", getattr(proc, 'exibicao_produto', 'sku'))
        proc.cnpj_loja = cfg.get("cnpj_loja", {})
        proc.cnpj_nome = cfg.get("cnpj_nome", {})

        etiquetas_combinadas, duplicadas = proc.remover_duplicatas(etiquetas_combinadas)
        if duplicadas:
            adicionar_log(estado, f"  Agrupamento: {len(duplicadas)} duplicatas removidas", "warning")

        # Todas as etiquetas juntas no mesmo PDF (regular, cpf, retirada)
        pasta_grupo = os.path.join(pasta_saida, nome_grupo)
        if not os.path.exists(pasta_grupo):
            os.makedirs(pasta_grupo)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        total_pags = 0

        if etiquetas_combinadas:
            caminho_pdf = os.path.join(pasta_grupo, f"agrupado_{nome_grupo}_{timestamp}.pdf")
            t, _, _, _, _ = proc.gerar_pdf_loja(etiquetas_combinadas, caminho_pdf)
            total_pags += t

        # Gerar XLSX temporario → JPEG do resumo → deletar XLSX
        caminho_xlsx = os.path.join(pasta_grupo, f"resumo_{nome_grupo}_{timestamp}.xlsx")
        n_skus, total_qtd = proc.gerar_resumo_xlsx(etiquetas_combinadas, caminho_xlsx, nome_grupo)
        try:
            caminho_jpeg = os.path.splitext(caminho_xlsx)[0] + ".jpeg"
            proc.gerar_imagem_resumo_xlsx(caminho_xlsx, caminho_jpeg, max_pedidos_por_pagina=150)
        except Exception as e_jpg:
            print(f"[Agrupamento {nome_grupo}] Aviso: falha ao gerar JPEG resumo: {e_jpg}")
        try:
            if os.path.exists(caminho_xlsx):
                os.remove(caminho_xlsx)
        except Exception:
            pass

        adicionar_log(estado, f"Agrupamento '{nome_grupo}': {', '.join(nomes_lojas)}", "success")
        adicionar_log(estado, f"  {total_pags} pags, {n_skus} SKUs, {total_qtd} un.", "info")

        return jsonify({
            "mensagem": f"Agrupamento '{nome_grupo}' gerado com sucesso",
            "lojas": nomes_lojas,
            "total_etiquetas": len(etiquetas_combinadas),
            "arquivo": f"agrupado_{nome_grupo}_{timestamp}.pdf",
            "pasta": nome_grupo,
        })

    except Exception as e:
        adicionar_log(estado, f"ERRO ao agrupar: {str(e)}", "error")
        import traceback
        adicionar_log(estado, traceback.format_exc(), "error")
        return jsonify({"erro": str(e)}), 500


@app.route('/api/agrupamentos', methods=['GET', 'POST'])
@jwt_required()
def api_agrupamentos():
    user_id = get_jwt_identity()
    estado = _get_estado(user_id)
    if not estado:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    if request.method == 'GET':
        return jsonify({"agrupamentos": estado["agrupamentos"]})
    dados = request.get_json() or {}
    estado["agrupamentos"] = dados.get("agrupamentos", [])
    _salvar_agrupamentos_usuario(user_id)
    adicionar_log(estado, f"Agrupamentos salvos: {len(estado['agrupamentos'])} grupo(s)", "success")
    return jsonify({"ok": True})


# ----------------------------------------------------------------
# PROCESSAMENTO EM BACKGROUND
# ----------------------------------------------------------------
def _executar_processamento(user_id, sem_recorte=True, resumo_sku_somente=False, pasta_entrada_override=None):
    """Executa o processamento completo em thread separada.

    Args:
        user_id: ID do usuario.
        sem_recorte: quando True, usa PDF pagina inteira (fluxo UpSeller novo).
        resumo_sku_somente: quando True, gera XLSX em SKU + quantidade.
            Padrão atual: False (SKU + Variação + Soma Quant., igual exemplo validado).
        pasta_entrada_override: quando informado, processa apenas essa pasta de lote.
    """
    with app.app_context():
        estado = _get_estado(user_id)
        if not estado:
            return {"ok": False, "erro": "Estado do usuario nao encontrado", "total_etiquetas": 0, "total_lojas": 0}

        # Regra global atual: etiqueta inteira, sem recorte.
        sem_recorte = True

        estado["processando"] = True
        estado["logs"] = []
        inicio = time.time()
        exec_info = {"ok": False, "erro": "", "total_etiquetas": 0, "total_lojas": 0}

        try:
            pasta_entrada_padrao = estado["configuracoes"]["pasta_entrada"]
            pasta_entrada = pasta_entrada_override or pasta_entrada_padrao
            pasta_saida = estado["configuracoes"]["pasta_saida"]
            os.makedirs(pasta_entrada, exist_ok=True)

            # Valida se o lote/pasta de entrada tem pelo menos 1 PDF real.
            # Isso evita processar "vazio" e sobrescrever o ultimo resultado.
            pdfs_entrada = []
            for raiz, _dirs, arquivos in os.walk(pasta_entrada):
                for nome_arq in arquivos:
                    low = str(nome_arq).lower()
                    if low.endswith(".pdf") and not low.startswith("._"):
                        pdfs_entrada.append(os.path.join(raiz, nome_arq))
            if not pdfs_entrada:
                raise RuntimeError(
                    "Nenhum PDF encontrado no lote atual para processar. "
                    "A saida anterior foi preservada."
                )

            adicionar_log(estado, "Iniciando processamento...", "info")
            if pasta_entrada_override:
                adicionar_log(estado, f"Lote isolado de entrada: {pasta_entrada}", "info")
            adicionar_log(estado, f"PDFs detectados no lote: {len(pdfs_entrada)}", "info")

            proc = ProcessadorEtiquetasShopee()

            proc.LARGURA_PT = estado["configuracoes"]["largura_mm"] * 2.835
            proc.ALTURA_PT = estado["configuracoes"]["altura_mm"] * 2.835
            proc.MARGEM_ESQUERDA = estado["configuracoes"]["margem_esq"]
            proc.MARGEM_DIREITA = estado["configuracoes"]["margem_dir"]
            proc.MARGEM_TOPO = estado["configuracoes"]["margem_topo"]
            proc.MARGEM_INFERIOR = estado["configuracoes"]["margem_inf"]
            proc.fonte_produto = estado["configuracoes"].get("fonte_produto", 7)
            proc.exibicao_produto = estado["configuracoes"].get("exibicao_produto", "sku")

            # Carregar dados dos XLSX de empacotamento (produtos, tracking, order_sn)
            adicionar_log(estado, "Carregando dados dos XLSX...", "info")
            proc.carregar_todos_xlsx(pasta_entrada)

            # Fallback para lotes isolados (ex.: /api/upseller/imprimir):
            # quando o lote atual tiver apenas PDF, tentar reaproveitar XLSX
            # de lotes gerar_* recentes do mesmo usuario.
            if not proc.dados_xlsx_global and pasta_entrada_override:
                try:
                    pasta_lotes_base = os.path.dirname(os.path.realpath(pasta_entrada))
                    if os.path.isdir(pasta_lotes_base):
                        candidatos = []
                        for nome in os.listdir(pasta_lotes_base):
                            if not str(nome).lower().startswith("gerar_"):
                                continue
                            d = os.path.join(pasta_lotes_base, nome)
                            if not os.path.isdir(d):
                                continue
                            try:
                                mt = os.path.getmtime(d)
                            except Exception:
                                mt = 0
                            candidatos.append((mt, d))

                        candidatos.sort(reverse=True)
                        pastas_testadas = 0
                        for _mt, d in candidatos[:10]:
                            pastas_testadas += 1
                            proc.carregar_todos_xlsx(d)
                            # Se ja encontrou bastante pedidos, nao precisa varrer tudo.
                            if len(proc.dados_xlsx_global) >= 300:
                                break

                        if proc.dados_xlsx_global:
                            adicionar_log(
                                estado,
                                f"XLSX fallback: {len(proc.dados_xlsx_global)} pedidos de lotes gerar_* recentes",
                                "success"
                            )
                        else:
                            adicionar_log(
                                estado,
                                f"XLSX fallback: nenhum pedido encontrado em {pastas_testadas} lote(s) gerar_*",
                                "warning"
                            )
                except Exception as e_xlsx_fb:
                    adicionar_log(estado, f"Falha no fallback de XLSX: {e_xlsx_fb}", "warning")

            if proc.dados_xlsx_global:
                adicionar_log(estado, f"XLSX: {len(proc.dados_xlsx_global)} pedidos, {len(proc.dados_xlsx_tracking)} trackings", "success")
            else:
                adicionar_log(estado, "Nenhum XLSX de empacotamento encontrado", "warning")

            if sem_recorte:
                adicionar_log(estado, "Modo: etiqueta inteira (sem recorte)", "info")
                adicionar_log(estado, "Carregando etiquetas dos PDFs (sem recorte)...", "info")
                todas_etiquetas, cpf_auto_detectadas, pdfs_shein_auto = proc.carregar_todos_pdfs_sem_recorte(pasta_entrada)
            else:
                adicionar_log(estado, "Carregando etiquetas dos PDFs...", "info")
                todas_etiquetas, cpf_auto_detectadas, pdfs_shein_auto = proc.carregar_todos_pdfs(pasta_entrada)
            adicionar_log(estado, f"Total: {len(todas_etiquetas)} etiquetas extraídas", "success")

            # Avisar sobre PDFs que podem ter quadrantes/páginas ignorados
            pdfs_na_pasta = [f for f in os.listdir(pasta_entrada) if f.lower().endswith('.pdf')]
            if len(pdfs_na_pasta) > len(todas_etiquetas) / 2:
                adicionar_log(estado, f"INFO: {len(pdfs_na_pasta)} PDFs encontrados, verifique se todos foram processados", "info")
            if cpf_auto_detectadas:
                adicionar_log(estado, f"CPF auto-detectadas: {len(cpf_auto_detectadas)} etiquetas", "info")
            if pdfs_shein_auto:
                adicionar_log(estado, f"Shein auto-detectados: {len(pdfs_shein_auto)} PDF(s)", "info")

            # Verificar quais etiquetas tem/nao tem dados de produto
            n_com_dados = sum(1 for e in todas_etiquetas if e.get('dados_xml', {}).get('produtos'))
            n_sem_dados = len(todas_etiquetas) - n_com_dados
            if n_sem_dados > 0:
                adicionar_log(estado, f"AVISO: {n_sem_dados} etiquetas sem dados de produto (de {len(todas_etiquetas)} total)", "warning")

            adicionar_log(estado, "Verificando etiquetas especiais...", "info")

            etiquetas_cpf_especial = proc.processar_cpf(pasta_entrada)
            # Juntar CPF do lanim*.pdf com CPF auto-detectadas de PDFs genericos
            etiquetas_cpf_especial.extend(cpf_auto_detectadas)
            if etiquetas_cpf_especial:
                todas_etiquetas.extend(etiquetas_cpf_especial)
                adicionar_log(estado, f"CPF: {len(etiquetas_cpf_especial)} etiquetas ({len(cpf_auto_detectadas)} auto-detectadas)", "success")

            etiquetas_shein = proc.processar_shein(pasta_entrada, pdfs_extras=pdfs_shein_auto)
            if etiquetas_shein:
                todas_etiquetas.extend(etiquetas_shein)
                adicionar_log(estado, f"Shein: {len(etiquetas_shein)} etiquetas", "success")

            if not etiquetas_cpf_especial and not etiquetas_shein:
                adicionar_log(estado, "Nenhuma etiqueta especial encontrada", "info")

            todas_etiquetas, duplicadas = proc.remover_duplicatas(todas_etiquetas)
            if duplicadas:
                adicionar_log(estado, f"AVISO: {len(duplicadas)} etiquetas duplicadas removidas:", "warning")
                for d in duplicadas:
                    adicionar_log(estado, f"  NF duplicada: {d.get('nf', '?')}", "warning")
            else:
                adicionar_log(estado, "Nenhuma duplicata encontrada", "info")

            adicionar_log(estado, "Separando etiquetas por loja...", "info")
            lojas_por_cnpj = proc.separar_por_loja(todas_etiquetas)
            lojas_saida = _agrupar_lojas_para_saida(proc, lojas_por_cnpj, estado)
            adicionar_log(estado, f"{len(lojas_saida)} lojas para processar", "info")
            # Avisos sobre tipos especiais de etiquetas
            n_retirada = sum(1 for e in todas_etiquetas if e.get('tipo_especial') == 'retirada')
            if n_retirada > 0:
                adicionar_log(estado, f"AVISO: {n_retirada} etiqueta(s) de RETIRADA (cliente retira na loja - sem endereço)", "warning")

            total_etiquetas_lojas = sum(len(item.get("etiquetas", []) or []) for item in lojas_saida)
            if total_etiquetas_lojas <= 0:
                raise RuntimeError(
                    "Nenhuma etiqueta foi extraida dos PDFs do lote atual. "
                    "A saida anterior foi preservada."
                )

            # Limpar pasta de saida somente apos confirmar que o lote e valido.
            import shutil
            if os.path.exists(pasta_saida):
                shutil.rmtree(pasta_saida)
            os.makedirs(pasta_saida, exist_ok=True)

            estado["_etiquetas_por_cnpj"] = dict(lojas_por_cnpj)
            estado["_proc_config"] = {
                "largura_pt": proc.LARGURA_PT,
                "altura_pt": proc.ALTURA_PT,
                "margem_esquerda": proc.MARGEM_ESQUERDA,
                "margem_direita": proc.MARGEM_DIREITA,
                "margem_topo": proc.MARGEM_TOPO,
                "margem_inferior": proc.MARGEM_INFERIOR,
                "fonte_produto": proc.fonte_produto,
                "cnpj_loja": dict(proc.cnpj_loja),
                "cnpj_nome": dict(proc.cnpj_nome),
                "resumo_sku_somente": bool(resumo_sku_somente),
            }

            # Contar etiquetas sem XML/declaracao para aviso
            etiquetas_sem_nf = []

            resultado_lojas = []
            for loja_saida_info in lojas_saida:
                nome_loja = str(loja_saida_info.get("nome", "") or "").strip() or "Loja_Desconhecida"
                etiquetas_loja = loja_saida_info.get("etiquetas", []) or []
                cnpjs_loja = [str(x).strip() for x in (loja_saida_info.get("cnpjs", []) or []) if str(x).strip()]
                n_etiquetas = len(etiquetas_loja)

                # Pular lojas sem etiquetas ou "Loja_Desconhecida" vazia
                if n_etiquetas == 0:
                    continue

                # Verificar etiquetas sem dados XML/declaracao
                for etq in etiquetas_loja:
                    dados = etq.get('dados_xml', {})
                    if not dados.get('chave') and not dados.get('produtos'):
                        etiquetas_sem_nf.append({
                            'nf': etq.get('nf', '?'),
                            'loja': nome_loja,
                        })

                try:
                    adicionar_log(estado, f"Processando: {nome_loja} ({n_etiquetas} etiquetas)...", "info")

                    pasta_loja = os.path.join(pasta_saida, nome_loja)
                    if not os.path.exists(pasta_loja):
                        os.makedirs(pasta_loja)

                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

                    # Todas as etiquetas vão no mesmo PDF (regular, cpf e retirada juntas)
                    total_pags = 0
                    n_simples = n_multi = com_xml = sem_xml = 0
                    pdf_nome = ''

                    if etiquetas_loja:
                        caminho_pdf = os.path.join(pasta_loja, f"etiquetas_{nome_loja}_{timestamp}.pdf")
                        t, ns, nm, cx, sx = proc.gerar_pdf_loja(etiquetas_loja, caminho_pdf)
                        total_pags += t
                        n_simples, n_multi, com_xml, sem_xml = ns, nm, cx, sx
                        pdf_nome = os.path.basename(caminho_pdf)

                    # Gerar XLSX temporario → JPEG do resumo → deletar XLSX
                    caminho_xlsx = os.path.join(pasta_loja, f"resumo_{nome_loja}_{timestamp}.xlsx")
                    n_skus, total_qtd = proc.gerar_resumo_xlsx(
                        etiquetas_loja,
                        caminho_xlsx,
                        nome_loja,
                        sku_somente=resumo_sku_somente
                    )
                    try:
                        caminho_jpeg = os.path.splitext(caminho_xlsx)[0] + ".jpeg"
                        proc.gerar_imagem_resumo_xlsx(caminho_xlsx, caminho_jpeg, max_pedidos_por_pagina=150)
                    except Exception as e_jpg:
                        print(f"[{nome_loja}] Aviso: falha ao gerar JPEG resumo: {e_jpg}")
                    # Remover XLSX — so o JPEG importa para envio
                    try:
                        if os.path.exists(caminho_xlsx):
                            os.remove(caminho_xlsx)
                    except Exception:
                        pass

                    info_loja = {
                        "nome": nome_loja,
                        "cnpj": ",".join(cnpjs_loja) if len(cnpjs_loja) > 1 else (cnpjs_loja[0] if cnpjs_loja else ""),
                        "etiquetas": n_etiquetas,
                        "paginas": total_pags,
                        "simples": n_simples,
                        "multi_produto": n_multi,
                        "com_xml": com_xml,
                        "sem_xml": sem_xml,
                        "skus": n_skus,
                        "total_qtd": total_qtd,
                        "pdf": pdf_nome,
                        "xlsx": "",
                    }
                    resultado_lojas.append(info_loja)

                    adicionar_log(estado, f"  {nome_loja}: {total_pags} pags, {n_skus} SKUs, {total_qtd} un.", "success")
                    if sem_xml > 0:
                        adicionar_log(estado, f"  AVISO: {sem_xml} etiquetas sem XML", "warning")

                except Exception as e_loja:
                    adicionar_log(estado, f"ERRO ao processar loja {nome_loja}: {str(e_loja)}", "error")
                    import traceback
                    adicionar_log(estado, traceback.format_exc(), "error")
                    continue

            # Acumular resumo diario e gerar JPEG consolidado
            _acumular_resumo_diario(user_id, resultado_lojas)
            adicionar_log(estado, f"Processamento concluido: {len(resultado_lojas)} lojas", "success")

            # Aviso de etiquetas sem NF/declaracao
            if etiquetas_sem_nf:
                adicionar_log(estado, f"AVISO: {len(etiquetas_sem_nf)} etiquetas sem nota fiscal ou declaracao de conteudo!", "warning")
                lojas_afetadas = set()
                for e in etiquetas_sem_nf:
                    lojas_afetadas.add(e['loja'])
                for loja_a in sorted(lojas_afetadas):
                    n_sem = sum(1 for e in etiquetas_sem_nf if e['loja'] == loja_a)
                    adicionar_log(estado, f"  {loja_a}: {n_sem} etiquetas sem NF/declaracao", "warning")

            if etiquetas_shein:
                adicionar_log(estado, "Gerando PDF Shein...", "info")
                from collections import defaultdict as dd
                shein_por_cnpj = dd(list)
                for etq in etiquetas_shein:
                    shein_por_cnpj[etq.get('cnpj', '')].append(etq)

                # Mapear CNPJs Shein para nomes de loja do DB/UpSeller
                # Shein usa varias transportadoras (cada uma com CNPJ diferente),
                # entao varios CNPJs pertencem a mesma loja.
                _shein_cnpj_nome_map = {}
                try:
                    lojas_db_shein = Loja.query.filter(
                        Loja.user_id == int(user_id),
                        Loja.nome.ilike('%shein%')
                    ).all()
                    if lojas_db_shein:
                        cnpjs_shein = list(shein_por_cnpj.keys())
                        if len(lojas_db_shein) == 1:
                            # 1 loja Shein → TODOS os CNPJs vao para ela
                            # (transportadoras diferentes = CNPJs diferentes, mas mesma loja)
                            for cnpj_sh in cnpjs_shein:
                                _shein_cnpj_nome_map[cnpj_sh] = lojas_db_shein[0].nome
                        else:
                            # Multiplas lojas Shein: tenta associar por ordem
                            for idx, cnpj_sh in enumerate(sorted(cnpjs_shein)):
                                if idx < len(lojas_db_shein):
                                    _shein_cnpj_nome_map[cnpj_sh] = lojas_db_shein[idx].nome
                except Exception as e_map:
                    print(f"[Shein] Aviso: falha ao mapear nomes Shein: {e_map}")

                # Consolidar por nome de loja (varios CNPJs podem pertencer a mesma loja)
                _shein_por_loja = dd(list)
                for cnpj_s, etqs_s in shein_por_cnpj.items():
                    nome_loja_s = _shein_cnpj_nome_map.get(cnpj_s) or proc.get_nome_loja(cnpj_s)
                    _shein_por_loja[nome_loja_s].extend(etqs_s)

                for nome_loja_s, etqs_s in _shein_por_loja.items():
                    pasta_loja_s = os.path.join(pasta_saida, nome_loja_s)
                    if not os.path.exists(pasta_loja_s):
                        os.makedirs(pasta_loja_s)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    caminho_shein = os.path.join(pasta_loja_s, f"shein_{nome_loja_s}_{timestamp}.pdf")
                    total_shein = proc.gerar_pdf_shein(etqs_s, caminho_shein)

                    # Gerar XLSX temporario → JPEG do resumo Shein → deletar XLSX
                    caminho_xlsx_shein = os.path.join(pasta_loja_s, f"resumo_shein_{nome_loja_s}_{timestamp}.xlsx")
                    n_skus_s, total_qtd_s = proc.gerar_resumo_xlsx_shein(etqs_s, caminho_xlsx_shein, nome_loja_s)
                    try:
                        caminho_jpeg_shein = os.path.splitext(caminho_xlsx_shein)[0] + ".jpeg"
                        proc.gerar_imagem_resumo_xlsx(caminho_xlsx_shein, caminho_jpeg_shein, max_pedidos_por_pagina=150)
                    except Exception as e_jpg:
                        print(f"[Shein] Aviso: falha ao gerar JPEG resumo: {e_jpg}")
                    # Remover XLSX — so o JPEG importa para envio
                    try:
                        if os.path.exists(caminho_xlsx_shein):
                            os.remove(caminho_xlsx_shein)
                    except Exception:
                        pass

                    # Coletar todos os CNPJs deste grupo consolidado
                    cnpjs_grupo = set(e.get('cnpj', '') for e in etqs_s)

                    # Adicionar Shein ao resultado para WhatsApp delivery
                    info_shein = {
                        "nome": nome_loja_s,
                        "cnpj": ",".join(cnpjs_grupo) if len(cnpjs_grupo) > 1 else next(iter(cnpjs_grupo), ""),
                        "etiquetas": len(etqs_s),
                        "paginas": total_shein,
                        "simples": 0,
                        "multi_produto": 0,
                        "com_xml": 0,
                        "sem_xml": 0,
                        "skus": n_skus_s,
                        "total_qtd": total_qtd_s,
                        "pdf": os.path.basename(caminho_shein),
                        "xlsx": "",
                    }
                    resultado_lojas.append(info_shein)

                    adicionar_log(estado, f"  Shein {nome_loja_s}: {total_shein} paginas, {n_skus_s} itens, {total_qtd_s} unidades", "success")

            if estado["agrupamentos"] and resultado_lojas:
                adicionar_log(estado, "Gerando agrupamentos pre-configurados...", "info")

                for grupo in estado["agrupamentos"]:
                    nome_grupo = grupo.get("nome", "Agrupado")
                    cnpjs_grupo = grupo.get("cnpjs", [])
                    if len(cnpjs_grupo) < 2:
                        continue

                    etiquetas_grupo = []
                    nomes_g = []
                    for c in cnpjs_grupo:
                        if c in lojas_por_cnpj:
                            etiquetas_grupo.extend(lojas_por_cnpj[c])
                            nomes_g.append(proc.get_nome_loja(c))

                    if len(nomes_g) < 2:
                        adicionar_log(estado, f"  Grupo '{nome_grupo}': lojas insuficientes, pulando", "warning")
                        continue

                    try:
                        etiquetas_grupo, _ = proc.remover_duplicatas(etiquetas_grupo)

                        # Todas as etiquetas juntas no mesmo PDF (regular, cpf, retirada)
                        pasta_grupo = os.path.join(pasta_saida, nome_grupo)
                        if not os.path.exists(pasta_grupo):
                            os.makedirs(pasta_grupo)
                        timestamp_g = datetime.now().strftime("%Y%m%d_%H%M%S")

                        total_pags_g = 0
                        if etiquetas_grupo:
                            caminho_agrup = os.path.join(pasta_grupo, f"etiquetas_{nome_grupo}_{timestamp_g}.pdf")
                            t_g, _, _, _, _ = proc.gerar_pdf_loja(etiquetas_grupo, caminho_agrup)
                            total_pags_g += t_g

                        # Gerar XLSX temporario → JPEG do resumo → deletar XLSX
                        caminho_xlsx_g = os.path.join(pasta_grupo, f"resumo_{nome_grupo}_{timestamp_g}.xlsx")
                        proc.gerar_resumo_xlsx(
                            etiquetas_grupo,
                            caminho_xlsx_g,
                            nome_grupo,
                            sku_somente=resumo_sku_somente
                        )
                        try:
                            caminho_jpeg_g = os.path.splitext(caminho_xlsx_g)[0] + ".jpeg"
                            proc.gerar_imagem_resumo_xlsx(caminho_xlsx_g, caminho_jpeg_g, max_pedidos_por_pagina=150)
                        except Exception as e_jpg_g:
                            print(f"[Grupo {nome_grupo}] Aviso: falha ao gerar JPEG resumo: {e_jpg_g}")
                        try:
                            if os.path.exists(caminho_xlsx_g):
                                os.remove(caminho_xlsx_g)
                        except Exception:
                            pass

                        adicionar_log(estado, f"  Grupo '{nome_grupo}': {', '.join(nomes_g)} ({total_pags_g} pags)", "success")
                    except Exception as e_g:
                        adicionar_log(estado, f"  ERRO grupo '{nome_grupo}': {str(e_g)}", "error")

            # Remover pastas vazias (ex: Loja_Desconhecida sem arquivos)
            if os.path.exists(pasta_saida):
                for d in os.listdir(pasta_saida):
                    dp = os.path.join(pasta_saida, d)
                    if os.path.isdir(dp) and not os.listdir(dp):
                        os.rmdir(dp)
                        adicionar_log(estado, f"  Pasta vazia removida: {d}", "info")

            duracao = round(time.time() - inicio, 1)

            resultado = {
                "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                "duracao_s": duracao,
                "total_xlsx": len(proc.dados_xlsx_global),
                "total_etiquetas": len(todas_etiquetas),
                "total_lojas": len(resultado_lojas),
                "lojas": resultado_lojas,
                "etiquetas_sem_nf": len(etiquetas_sem_nf) if etiquetas_sem_nf else 0,
                "resumo_geral": {},
            }

            estado["ultimo_resultado"] = resultado
            estado["historico"].insert(0, resultado)
            estado["historico"] = estado["historico"][:20]

            # Salvar resultado em disco para persistir entre deploys
            _salvar_resultado_usuario(user_id)

            # Registrar snapshot ZIP no historico persistente (ultimas 24h)
            try:
                hist_item = _registrar_historico_gerada(
                    user_id=user_id,
                    resultado=resultado,
                    pasta_saida=pasta_saida,
                    origem="processamento",
                )
                if hist_item:
                    resultado["historico_gerada_id"] = hist_item.get("id", "")
                    resultado["historico_gerada_arquivo"] = hist_item.get("arquivo", "")
                    adicionar_log(estado, "Historico 24h atualizado (download disponivel).", "info")
            except Exception as e_hist:
                print(f"Aviso: falha ao registrar historico 24h user {user_id}: {e_hist}")

            # Registrar processamento no contador do usuario
            user = User.query.get(user_id)
            if user:
                user.registrar_processamento()

            adicionar_log(estado, f"Processamento concluido em {duracao}s!", "success")
            exec_info = {
                "ok": True,
                "erro": "",
                "total_etiquetas": resultado.get("total_etiquetas", 0),
                "total_lojas": resultado.get("total_lojas", 0),
            }

        except Exception as e:
            adicionar_log(estado, f"ERRO: {str(e)}", "error")
            import traceback
            adicionar_log(estado, traceback.format_exc(), "error")
            exec_info = {
                "ok": False,
                "erro": str(e),
                "total_etiquetas": 0,
                "total_lojas": 0,
            }

        finally:
            estado["processando"] = False
        return exec_info


def _formatar_tamanho(bytes_val):
    if bytes_val < 1024:
        return f"{bytes_val} B"
    elif bytes_val < 1024 * 1024:
        return f"{bytes_val / 1024:.1f} KB"
    else:
        return f"{bytes_val / (1024 * 1024):.1f} MB"


def _normalizar_nome_loja_match(nome: str) -> str:
    """Normaliza nome de loja para comparacoes robustas (fallback de envio)."""
    if not nome:
        return ""
    try:
        nome = _re.sub(r"\s+", " ", str(nome)).strip().lower()
        # Remove acentos
        import unicodedata
        nome = ''.join(ch for ch in unicodedata.normalize('NFD', nome) if unicodedata.category(ch) != 'Mn')
        return nome
    except Exception:
        return (str(nome) or "").strip().lower()


# ----------------------------------------------------------------
# WHATSAPP - FILA PERSISTENTE + SUPERVISOR BAILEYS
# ----------------------------------------------------------------

_BAILEYS_PROC = None
_BAILEYS_PROC_LOCK = threading.Lock()
_BAILEYS_LAST_START_TS = 0.0
_BAILEYS_LOG_HANDLES = []
_BACKGROUND_WORKERS_STARTED = False
_BACKGROUND_WORKERS_LOCK = threading.Lock()
_WHATSAPP_ENQUEUE_LOCKS = {}
_WHATSAPP_ENQUEUE_LOCKS_LOCK = threading.Lock()
_WHATSAPP_WORKER_RUN_LOCK = threading.Lock()
_WHATSAPP_PHONE_SEND_LOCK = threading.Lock()
_WHATSAPP_LAST_SENT_TS = {}
_WHATSAPP_MIN_INTERVAL_SAME_PHONE = max(0, int(os.environ.get("WHATSAPP_MIN_INTERVAL_SAME_PHONE", "20") or 0))


def _whatsapp_api_base_url() -> str:
    return os.environ.get("WHATSAPP_API_URL", "http://localhost:3005").rstrip("/")


def _agora_utc() -> datetime:
    return _agora_brasil()


def _get_whatsapp_enqueue_lock(user_id: int) -> threading.Lock:
    uid = int(user_id)
    with _WHATSAPP_ENQUEUE_LOCKS_LOCK:
        lock = _WHATSAPP_ENQUEUE_LOCKS.get(uid)
        if lock is None:
            lock = threading.Lock()
            _WHATSAPP_ENQUEUE_LOCKS[uid] = lock
        return lock


def _normalizar_telefone_whatsapp(telefone: str) -> str:
    numero = "".join(ch for ch in str(telefone or "") if ch.isdigit())
    if len(numero) in (10, 11):
        numero = "55" + numero
    return numero


def _normalizar_chave_whatsapp_queue(telefone: str, file_path: str, loja_nome: str = ""):
    telefone_norm = _normalizar_telefone_whatsapp(telefone)

    caminho = str(file_path or "").strip()
    caminho_norm = os.path.normcase(os.path.abspath(caminho)) if caminho else ""
    arquivo_base = os.path.basename(caminho_norm).strip().lower() if caminho_norm else ""
    loja_norm = _re.sub(r"\s+", " ", str(loja_nome or "")).strip().casefold()
    return telefone_norm, caminho_norm, arquivo_base, loja_norm


def _whatsapp_worker_lock_path() -> str:
    return os.path.join(_VOLUME_PATH, "whatsapp_queue_worker.lock")


@contextmanager
def _try_acquire_whatsapp_worker_lock():
    os.makedirs(_VOLUME_PATH, exist_ok=True)
    lock_path = _whatsapp_worker_lock_path()
    handle = open(lock_path, "a+b")
    acquired = False
    try:
        handle.seek(0, os.SEEK_END)
        if handle.tell() == 0:
            handle.write(b"0")
            handle.flush()
        if os.name == "nt":
            import msvcrt
            handle.seek(0)
            try:
                msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
                acquired = True
            except OSError:
                acquired = False
        else:
            import fcntl
            try:
                fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                acquired = True
            except OSError:
                acquired = False
        yield handle if acquired else None
    finally:
        try:
            if acquired:
                if os.name == "nt":
                    import msvcrt
                    handle.seek(0)
                    msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
                else:
                    import fcntl
                    fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
        except Exception:
            pass
        try:
            handle.close()
        except Exception:
            pass


@contextmanager
def _try_acquire_whatsapp_worker_guard():
    with _try_acquire_whatsapp_worker_lock() as worker_lock:
        if worker_lock is None:
            yield None
            return
        acquired = _WHATSAPP_WORKER_RUN_LOCK.acquire(blocking=False)
        try:
            yield worker_lock if acquired else None
        finally:
            if acquired:
                _WHATSAPP_WORKER_RUN_LOCK.release()


def _chave_dedupe_batch_whatsapp(telefone: str, file_path: str, loja_nome: str = ""):
    telefone_norm, caminho_norm, arquivo_base, loja_norm = _normalizar_chave_whatsapp_queue(
        telefone, file_path, loja_nome
    )
    if not telefone_norm or not caminho_norm:
        return None
    return telefone_norm, caminho_norm, arquivo_base, loja_norm


def _to_bool(value, default=False):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    txt = str(value).strip().lower()
    if txt in ("1", "true", "yes", "sim", "on"):
        return True
    if txt in ("0", "false", "no", "nao", "off"):
        return False
        return default


def _to_int_list(value):
    """Normaliza listas de IDs aceitando lista Python ou JSON string."""
    if value is None:
        return None

    raw = value
    if isinstance(raw, str):
        txt = raw.strip()
        if not txt:
            return []
        try:
            raw = json.loads(txt)
        except Exception:
            raw = [txt]

    if not isinstance(raw, (list, tuple, set)):
        raw = [raw]

    items = []
    seen = set()
    for item in raw:
        try:
            num = int(item)
        except Exception:
            continue
        if num in seen or num <= 0:
            continue
        seen.add(num)
        items.append(num)
    return items


def _baileys_healthy(timeout=2.0) -> bool:
    from whatsapp_service import DEFAULT_PROVIDER
    if DEFAULT_PROVIDER == "uazapi":
        return True  # Baileys nao e usado com UAZAPI
    try:
        resp = requests.get(f"{_whatsapp_api_base_url()}/health", timeout=timeout)
        return resp.status_code == 200
    except Exception:
        return False


def _erro_sessao_whatsapp_desconectada(msg: str) -> bool:
    texto = str(msg or "").strip().lower()
    if not texto:
        return False
    sinais = (
        "sessao nao conectada",
        "sessão não conectada",
        "session not connected",
        "not connected",
    )
    return any(s in texto for s in sinais)


def _whatsapp_provider_pronto(wa: WhatsAppService = None) -> tuple:
    from whatsapp_service import DEFAULT_PROVIDER
    if DEFAULT_PROVIDER != "uazapi":
        return _baileys_healthy(), ""
    try:
        svc = wa or WhatsAppService()
        status = svc.verificar_conexao() or {}
        if status.get("connected"):
            return True, ""
        return False, str(status.get("error") or "Sessao nao conectada")
    except Exception as e:
        return False, str(e)


def _aplicar_cooldown_envio_whatsapp(telefone: str):
    numero = _normalizar_telefone_whatsapp(telefone)
    intervalo = int(_WHATSAPP_MIN_INTERVAL_SAME_PHONE or 0)
    if not numero or intervalo <= 0:
        return
    while True:
        with _WHATSAPP_PHONE_SEND_LOCK:
            agora_ts = time.time()
            ultimo = float(_WHATSAPP_LAST_SENT_TS.get(numero, 0.0) or 0.0)
            restante = intervalo - (agora_ts - ultimo)
            if restante <= 0:
                return
        time.sleep(min(restante, 1.0))


def _registrar_envio_whatsapp(telefone: str):
    numero = _normalizar_telefone_whatsapp(telefone)
    if not numero:
        return
    with _WHATSAPP_PHONE_SEND_LOCK:
        _WHATSAPP_LAST_SENT_TS[numero] = time.time()


def _garantir_baileys_rodando(motivo: str = "") -> bool:
    """Tenta manter o baileys-api ativo (sem derrubar fluxo principal).
    Quando o provider é UAZAPI, Baileys não é necessário."""
    from whatsapp_service import DEFAULT_PROVIDER
    if DEFAULT_PROVIDER == "uazapi":
        return True  # UAZAPI gerencia conexao na nuvem, Baileys desnecessario

    global _BAILEYS_PROC, _BAILEYS_LAST_START_TS
    if _baileys_healthy():
        return True

    with _BAILEYS_PROC_LOCK:
        if _baileys_healthy():
            return True

        now_ts = time.time()
        if now_ts - _BAILEYS_LAST_START_TS < 15:
            return False

        api_base = _whatsapp_api_base_url().lower()
        if "localhost" not in api_base and "127.0.0.1" not in api_base:
            # Em ambiente com API remota, nao tenta spawn local.
            return False

        baileys_dir = os.path.join(_BASE_DIR, "baileys-api")
        server_js = os.path.join(baileys_dir, "server.js")
        node_bin = os.environ.get("NODE_BIN", "node")
        if not os.path.exists(server_js):
            print(f"[BaileysSupervisor] server.js nao encontrado em {server_js}")
            return False
        if shutil.which(node_bin) is None:
            print(f"[BaileysSupervisor] executavel Node.js nao encontrado: {node_bin}")
            return False

        # Evitar spawn duplicado se processo conhecido ainda estiver vivo.
        if _BAILEYS_PROC is not None and _BAILEYS_PROC.poll() is None:
            _BAILEYS_LAST_START_TS = now_ts
            return False

        try:
            stdout_path = os.path.join(baileys_dir, "baileys_stdout.log")
            stderr_path = os.path.join(baileys_dir, "baileys_stderr.log")
            os.makedirs(baileys_dir, exist_ok=True)
            stdout_f = open(stdout_path, "a", encoding="utf-8", buffering=1)
            stderr_f = open(stderr_path, "a", encoding="utf-8", buffering=1)
            _BAILEYS_LOG_HANDLES.extend([stdout_f, stderr_f])

            kwargs = {}
            if os.name == "nt" and hasattr(subprocess, "CREATE_NO_WINDOW"):
                kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
            elif os.name != "nt":
                kwargs["start_new_session"] = True

            # Forcar PORT=3005 para o Baileys nao herdar PORT do Flask
            baileys_env = os.environ.copy()
            baileys_env["PORT"] = "3005"

            _BAILEYS_PROC = subprocess.Popen(
                [node_bin, "server.js"],
                cwd=baileys_dir,
                env=baileys_env,
                stdout=stdout_f,
                stderr=stderr_f,
                stdin=subprocess.DEVNULL,
                **kwargs,
            )
            _BAILEYS_LAST_START_TS = now_ts
            print(f"[BaileysSupervisor] iniciado (PID={_BAILEYS_PROC.pid}) motivo={motivo or 'n/a'}")
        except Exception as e:
            print(f"[BaileysSupervisor] erro ao iniciar: {e}")
            return False

    # Aguarda subir (fora do lock).
    for _ in range(20):
        if _baileys_healthy(timeout=1.5):
            return True
        time.sleep(0.5)
    return False


def _calc_backoff_seconds(tentativas: int) -> int:
    # 8s, 16s, 32s... ate 15 min
    base = 8
    return min(900, base * (2 ** max((tentativas or 1) - 1, 0)))


# ================================================================
# RESUMO GERAL DIARIO: acumula lojas do dia e gera JPEG consolidado
# ================================================================

def _caminho_resumo_diario(user_id: int) -> str:
    """Retorna caminho do arquivo JSON para persistir resumo diario."""
    return os.path.join("data", f"resumo_diario_{user_id}.json")


def _carregar_resumo_diario_disco(user_id: int) -> dict:
    """Carrega resumo diario do disco (sobrevive reinicializacoes)."""
    caminho = _caminho_resumo_diario(user_id)
    if os.path.exists(caminho):
        try:
            with open(caminho, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _salvar_resumo_diario_disco(user_id: int, dados: dict):
    """Salva resumo diario em disco."""
    caminho = _caminho_resumo_diario(user_id)
    os.makedirs(os.path.dirname(caminho), exist_ok=True)
    try:
        with open(caminho, 'w', encoding='utf-8') as f:
            json.dump(dados, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[ResumoGeral] Erro ao salvar disco: {e}")


def _acumular_resumo_diario(user_id: int, resultado_lojas: list):
    """Acumula lojas processadas no dia para gerar resumo geral consolidado.
    Persiste em disco (JSON) para sobreviver reinicializacoes do servidor."""
    estado = _get_estado(user_id)
    if not estado:
        return
    hoje = datetime.now().strftime("%Y-%m-%d")

    # Carregar do disco (mais confiavel que so memoria)
    dados_disco = _carregar_resumo_diario_disco(user_id)

    # Limpar dias antigos (manter so hoje)
    for k in list(dados_disco.keys()):
        if k != hoje:
            del dados_disco[k]
    if hoje not in dados_disco:
        dados_disco[hoje] = {}
    diario = dados_disco[hoje]

    for loja in (resultado_lojas or []):
        nome = str(loja.get("nome", "") or "").strip()
        if not nome:
            continue
        # Acumula: soma etiquetas/skus/unidades se ja existia
        if nome in diario:
            diario[nome]["etiquetas"] += int(loja.get("etiquetas", 0) or 0)
            diario[nome]["skus"] = max(diario[nome]["skus"], int(loja.get("skus", 0) or 0))
            diario[nome]["unidades"] += int(loja.get("total_qtd", 0) or 0)
        else:
            diario[nome] = {
                "etiquetas": int(loja.get("etiquetas", 0) or 0),
                "skus": int(loja.get("skus", 0) or 0),
                "unidades": int(loja.get("total_qtd", 0) or 0),
            }

    # Persistir em disco
    _salvar_resumo_diario_disco(user_id, dados_disco)
    # Manter em memoria tambem
    estado["resumo_diario"] = dados_disco


def _obter_diario_hoje(user_id: int) -> dict:
    """Retorna o dict do resumo diario de hoje (disco + memoria)."""
    hoje = datetime.now().strftime("%Y-%m-%d")
    # Tentar disco primeiro (mais confiavel)
    dados_disco = _carregar_resumo_diario_disco(user_id)
    diario = dados_disco.get(hoje, {})
    if diario:
        return diario
    # Fallback: memoria
    estado = _get_estado(user_id)
    if estado:
        return (estado.get("resumo_diario") or {}).get(hoje, {})
    return {}


def _gerar_xlsx_resumo_geral(user_id: int, pasta_saida: str) -> str:
    """Gera XLSX consolidado com todas as lojas do dia.
    Formato identico ao resumo_geral individual (Loja|Etiquetas|SKUs|Unidades|TOTAL).
    Retorna caminho do XLSX gerado ou string vazia."""
    diario = _obter_diario_hoje(user_id)
    if not diario:
        return ""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
    except Exception:
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo Geral"
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    borda = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['Loja', 'Etiquetas', 'SKUs', 'Unidades']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = borda

    row = 2
    sum_etiq = sum_skus = sum_un = 0
    for nome in sorted(diario.keys()):
        d = diario[nome]
        etiq = d.get("etiquetas", 0)
        skus = d.get("skus", 0)
        un = d.get("unidades", 0)
        ws.cell(row=row, column=1, value=nome).border = borda
        ws.cell(row=row, column=2, value=etiq).border = borda
        ws.cell(row=row, column=3, value=skus).border = borda
        ws.cell(row=row, column=4, value=un).border = borda
        sum_etiq += etiq
        sum_skus += skus
        sum_un += un
        row += 1

    for col, val in enumerate(['TOTAL', sum_etiq, sum_skus, sum_un], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True, size=11)
        cell.fill = total_fill
        cell.border = borda

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15

    os.makedirs(pasta_saida, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho = os.path.join(pasta_saida, f"resumo_geral_{timestamp}.xlsx")
    try:
        wb.save(caminho)
        wb.close()
        return caminho
    except Exception as e:
        print(f"[ResumoGeral] Erro ao salvar XLSX: {e}")
        return ""


def _gerar_jpeg_resumo_geral(user_id: int, pasta_saida: str) -> str:
    """Gera JPEG com tabela consolidada de todas as lojas do dia.
    Estilo: header escuro, linhas claras, coluna Loja|Etiquetas|SKUs|Unidades, TOTAL.
    Retorna caminho do JPEG gerado ou string vazia."""
    diario = _obter_diario_hoje(user_id)
    if not diario:
        return ""

    try:
        from PIL import Image, ImageDraw, ImageFont
    except Exception:
        return ""

    # Montar dados da tabela
    header = ["Loja", "Etiquetas", "SKUs", "Unidades"]
    rows = []
    total_etiq = 0
    total_skus = 0
    total_un = 0
    for nome in sorted(diario.keys()):
        d = diario[nome]
        etiq = d.get("etiquetas", 0)
        skus = d.get("skus", 0)
        un = d.get("unidades", 0)
        rows.append([nome, str(etiq), str(skus), str(un)])
        total_etiq += etiq
        total_skus += skus
        total_un += un
    total_row = ["TOTAL", str(total_etiq), str(total_skus), str(total_un)]

    if not rows:
        return ""

    # Fontes
    try:
        font = ImageFont.truetype(r"C:\Windows\Fonts\arial.ttf", 15)
        font_bold = ImageFont.truetype(r"C:\Windows\Fonts\arialbd.ttf", 15)
    except Exception:
        font = ImageFont.load_default()
        font_bold = font

    # Calcular larguras
    all_rows = [header] + rows + [total_row]
    col_widths = [0] * 4
    for r in all_rows:
        for i, v in enumerate(r):
            try:
                tw = font_bold.getlength(str(v))
            except Exception:
                tw = len(str(v)) * 8
            col_widths[i] = max(col_widths[i], int(tw) + 24)
    col_widths[0] = max(col_widths[0], 160)  # Loja col min width

    row_h = 28
    margin = 2
    img_w = sum(col_widths) + margin * 2 + 1
    img_h = (len(rows) + 2) * row_h + margin * 2 + 1  # +2 = header + total

    # Cores estilo planilha
    cor_header_bg = (68, 84, 106)      # Azul escuro
    cor_header_text = (255, 255, 255)  # Branco
    cor_total_bg = (226, 239, 218)     # Verde claro
    cor_borda = (180, 180, 180)
    cor_texto = (20, 20, 20)
    cor_alt = (245, 245, 245)          # Linhas alternadas

    img = Image.new('RGB', (img_w, img_h), 'white')
    draw = ImageDraw.Draw(img)

    y = margin
    all_render = [header] + rows + [total_row]
    for ridx, r in enumerate(all_render):
        x = margin
        is_header = (ridx == 0)
        is_total = (ridx == len(all_render) - 1)
        if is_header:
            bg = cor_header_bg
            txt_color = cor_header_text
            use_font = font_bold
        elif is_total:
            bg = cor_total_bg
            txt_color = cor_texto
            use_font = font_bold
        else:
            bg = cor_alt if (ridx % 2 == 0) else (255, 255, 255)
            txt_color = cor_texto
            use_font = font

        for cidx, text in enumerate(r):
            w = col_widths[cidx]
            draw.rectangle([(x, y), (x + w, y + row_h)], fill=bg, outline=cor_borda, width=1)
            txt = str(text or '')
            tx = x + 6
            # Alinhar numeros a direita
            if cidx > 0:
                try:
                    tw = draw.textlength(txt, font=use_font)
                except Exception:
                    tw = len(txt) * 8
                tx = x + w - int(tw) - 8
            ty = y + 6
            draw.text((tx, ty), txt, fill=txt_color, font=use_font)
            x += w
        y += row_h

    # Salvar
    os.makedirs(pasta_saida, exist_ok=True)
    caminho = os.path.join(pasta_saida, f"resumo_geral_{hoje}.jpeg")
    try:
        img.save(caminho, format='JPEG', quality=92, optimize=True)
        return caminho
    except Exception as e:
        print(f"[ResumoGeral] Erro ao salvar JPEG: {e}")
        return ""


def _enviar_resumo_geral_whatsapp(user_id: int, pasta_saida: str, consolidado: bool = False) -> dict:
    """Gera e enfileira resumo geral para contatos com flag resumo_geral=True.
    consolidado=False: envia so JPEG parcial (apos cada processamento — comportamento original)
    consolidado=True:  envia XLSX + JPEG completo (fim do dia, resumo consolidado)"""
    uid = int(user_id)
    contatos = WhatsAppContact.query.filter_by(user_id=uid, ativo=True, resumo_geral=True).all()
    if not contatos:
        return {"ok": False, "motivo": "sem_contatos_resumo"}

    caminho_jpeg = _gerar_jpeg_resumo_geral(uid, pasta_saida)
    if not caminho_jpeg or not os.path.exists(caminho_jpeg):
        return {"ok": False, "motivo": "falha_gerar_jpeg"}

    # XLSX so no consolidado do fim do dia
    caminho_xlsx = ""
    if consolidado:
        caminho_xlsx = _gerar_xlsx_resumo_geral(uid, pasta_saida)

    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    agora = _agora_utc()
    enfileirados = 0
    hoje_label = datetime.now().strftime("%d/%m/%Y")
    label = "Resumo Geral do Dia" if consolidado else "Resumo Geral"
    with _get_whatsapp_enqueue_lock(uid):
        for c in contatos:
            telefone = str(getattr(c, "telefone", "") or "").strip()
            if not telefone:
                continue
            # Enviar JPEG (imagem do resumo)
            if not _ja_na_fila_whatsapp(uid, telefone, caminho_jpeg, label):
                db.session.add(WhatsAppQueueItem(
                    user_id=uid,
                    batch_id=batch_id,
                    origem="resumo_geral",
                    loja_nome=label,
                    telefone=telefone,
                    pdf_path=caminho_jpeg,
                    caption=f"{label} - {hoje_label}",
                    status="pending",
                    tentativas=0,
                    max_tentativas=5,
                    next_attempt_at=agora,
                ))
                enfileirados += 1
            # Enviar XLSX (so no consolidado do fim do dia)
            if caminho_xlsx and os.path.exists(caminho_xlsx) and not _ja_na_fila_whatsapp(uid, telefone, caminho_xlsx, label):
                db.session.add(WhatsAppQueueItem(
                    user_id=uid,
                    batch_id=batch_id,
                    origem="resumo_geral",
                    loja_nome=label,
                    telefone=telefone,
                    pdf_path=caminho_xlsx,
                    caption=f"{label} - {hoje_label}",
                    status="pending",
                    tentativas=0,
                    max_tentativas=5,
                    next_attempt_at=agora,
                ))
                enfileirados += 1

        if enfileirados:
            db.session.commit()

    if enfileirados:
        _garantir_baileys_rodando(motivo="resumo_geral")
        tipo = "consolidado XLSX+JPEG" if consolidado else "parcial JPEG"
        print(f"[ResumoGeral] Enfileirados {enfileirados} item(ns) ({tipo})")
    return {"ok": enfileirados > 0, "enfileirados": enfileirados}


# ================================================================
# PIPELINE DIRETO: Imprimir + Processar (para envio automatico)
# ================================================================

def _gerar_jpeg_de_xlsx(pasta_saida: str) -> int:
    """
    Percorre subpastas de pasta_saida, encontra XLSX e gera JPEG do resumo
    usando o mesmo metodo do ProcessadorEtiquetasShopee (visual padrao).
    Divide em paginas quando excede 150 linhas para manter legibilidade.
    Retorna quantidade de JPEGs gerados.
    """
    from etiquetas_shopee import ProcessadorEtiquetasShopee

    gerados = 0
    if not os.path.isdir(pasta_saida):
        return 0

    proc = ProcessadorEtiquetasShopee()
    for loja_dir in os.listdir(pasta_saida):
        loja_path = os.path.join(pasta_saida, loja_dir)
        if not os.path.isdir(loja_path):
            continue
        for arq in os.listdir(loja_path):
            if not arq.lower().endswith('.xlsx'):
                continue
            xlsx_path = os.path.join(loja_path, arq)
            jpg_name = os.path.splitext(arq)[0] + ".jpeg"
            jpg_path = os.path.join(loja_path, jpg_name)
            if os.path.exists(jpg_path):
                continue
            try:
                result = proc.gerar_imagem_resumo_xlsx(
                    caminho_xlsx=xlsx_path,
                    caminho_imagem=jpg_path,
                    max_pedidos_por_pagina=150,
                )
                if result:
                    gerados += 1
                    print(f"[JPEG] Gerado: {result}")
            except Exception as e:
                print(f"[JPEG] Erro ao converter {xlsx_path}: {e}")
    return gerados


def _mesclar_lojas_agrupadas(pasta_saida: str, estado: dict, user_id: int):
    """
    Mescla lojas do mesmo grupo em UMA pasta, reprocessando as etiquetas juntas
    como se fossem uma unica loja (ordenadas por SKU, com rodape correto).
    Ex: grupo "Leone" = [HEITOR, LEONE] -> pasta "Leone/" com PDF+XLSX gerados do zero.
    """
    import shutil
    from etiquetas_shopee import ProcessadorEtiquetasShopee

    agrupamentos = (estado or {}).get("agrupamentos", []) or []
    if not agrupamentos:
        return

    ultimo_resultado = estado.get("ultimo_resultado", {})
    lojas_resultado = ultimo_resultado.get("lojas", []) if ultimo_resultado else []
    if not lojas_resultado:
        return

    etiq_por_cnpj = estado.get("_etiquetas_por_cnpj", {})
    if not etiq_por_cnpj:
        print("[Agrupamento] Sem dados de etiquetas em memoria, pulando")
        return

    # Construir mapa: nome_upseller (lower) -> cnpj_marketplace
    # Usando _pdf_loja_map + _etiquetas_por_cnpj + _proc_config
    pdf_loja_map = estado.get("_pdf_loja_map", {})  # {pdf_basename: upseller_name}
    proc_config = estado.get("_proc_config", {})
    cnpj_loja = proc_config.get("cnpj_loja", {})

    # Mapear: upseller_lower -> cnpj
    upseller_to_cnpj = {}
    for cnpj, etiquetas in etiq_por_cnpj.items():
        if not etiquetas:
            continue
        # Descobrir nome UpSeller para este CNPJ via pdf_loja_map
        for etq in etiquetas:
            pdf_path = etq.get("caminho_pdf", "")
            if not pdf_path:
                continue
            pdf_base = os.path.basename(pdf_path)
            upseller_nome = pdf_loja_map.get(pdf_base)
            if upseller_nome:
                upseller_to_cnpj[upseller_nome.strip().lower()] = cnpj
                break

    if not upseller_to_cnpj:
        print("[Agrupamento] Sem mapeamento UpSeller->CNPJ, pulando")
        return

    print(f"[Agrupamento] Mapa UpSeller->CNPJ: {upseller_to_cnpj}")

    # Pastas existentes na saida
    pastas_existentes = {}
    for d in os.listdir(pasta_saida):
        if os.path.isdir(os.path.join(pasta_saida, d)):
            pastas_existentes[d.strip().lower()] = d

    proc = ProcessadorEtiquetasShopee()
    mesclou = False

    for grupo in agrupamentos:
        nome_grupo = (grupo.get("nome") or "").strip()
        nomes_lojas = grupo.get("nomes_lojas") or []
        if not nome_grupo or len(nomes_lojas) < 2:
            continue

        # Coletar etiquetas de todas as lojas do grupo
        etiquetas_grupo = []
        pastas_remover = []
        nomes_encontrados = []

        for nl in nomes_lojas:
            nl_lower = nl.strip().lower()
            cnpj = upseller_to_cnpj.get(nl_lower)
            if cnpj and cnpj in etiq_por_cnpj:
                etiquetas_grupo.extend(etiq_por_cnpj[cnpj])
                nomes_encontrados.append(nl.strip())
            # Coletar pasta para remover
            if nl_lower in pastas_existentes:
                nome_real = pastas_existentes[nl_lower]
                pastas_remover.append((nome_real, os.path.join(pasta_saida, nome_real)))

        if len(nomes_encontrados) < 2 or not etiquetas_grupo:
            continue

        print(f"[Agrupamento] Grupo '{nome_grupo}': {nomes_encontrados} -> {len(etiquetas_grupo)} etiquetas")

        # Pasta temporaria (evita conflito Windows case-insensitive)
        pasta_temp = os.path.join(pasta_saida, f"_temp_grupo_{nome_grupo}")
        if os.path.exists(pasta_temp):
            shutil.rmtree(pasta_temp, ignore_errors=True)
        os.makedirs(pasta_temp, exist_ok=True)
        timestamp_g = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Todas as etiquetas juntas no mesmo PDF (regular, cpf, retirada)
        total_pags = 0
        pdf_name = ""

        if etiquetas_grupo:
            caminho_pdf = os.path.join(pasta_temp, f"etiquetas_{nome_grupo}_{timestamp_g}.pdf")
            t, _, _, _, _ = proc.gerar_pdf_loja(etiquetas_grupo, caminho_pdf)
            total_pags += t
            pdf_name = os.path.basename(caminho_pdf)
            print(f"[Agrupamento]   PDF etiquetas: {t} paginas (ordenado por SKU)")

        # Gerar XLSX temporario → JPEG do resumo → deletar XLSX
        caminho_xlsx = os.path.join(pasta_temp, f"resumo_{nome_grupo}_{timestamp_g}.xlsx")
        proc.gerar_resumo_xlsx(etiquetas_grupo, caminho_xlsx, nome_grupo)
        try:
            caminho_jpeg_g = os.path.splitext(caminho_xlsx)[0] + ".jpeg"
            proc.gerar_imagem_resumo_xlsx(caminho_xlsx, caminho_jpeg_g, max_pedidos_por_pagina=150)
        except Exception as e_jpg:
            print(f"[Agrupamento]   Aviso: falha ao gerar JPEG resumo: {e_jpg}")
        try:
            if os.path.exists(caminho_xlsx):
                os.remove(caminho_xlsx)
        except Exception:
            pass
        print(f"[Agrupamento]   JPEG resumo gerado (XLSX removido)")

        # Remover pastas individuais
        for nome_real, pasta_path in pastas_remover:
            shutil.rmtree(pasta_path, ignore_errors=True)
            print(f"[Agrupamento]   Removida: {nome_real}/")

        # Renomear temp -> nome final
        pasta_destino = os.path.join(pasta_saida, nome_grupo)
        if os.path.exists(pasta_destino):
            shutil.rmtree(pasta_destino, ignore_errors=True)
        os.rename(pasta_temp, pasta_destino)

        # Atualizar resultado["lojas"]
        total_etiq = len(etiquetas_grupo)
        total_qtd = sum(e.get('num_produtos', 1) or 1 for e in etiquetas_grupo)
        nomes_removidos = set(n.strip().lower() for n in nomes_encontrados)

        novas_lojas = []
        for lj in lojas_resultado:
            nome_lj = (lj.get("nome") or "").strip()
            if nome_lj.lower() not in nomes_removidos:
                novas_lojas.append(lj)

        novas_lojas.append({
            "nome": nome_grupo,
            "cnpj": f"GRUPO_{nome_grupo}",
            "total_paginas": total_pags,
            "total_etiquetas": total_etiq,
            "total_qtd": total_qtd,
            "pdf": pdf_name,
            "xlsx": "",
        })
        lojas_resultado[:] = novas_lojas
        ultimo_resultado["total_lojas"] = len(lojas_resultado)
        mesclou = True
        print(f"[Agrupamento] Grupo '{nome_grupo}' OK: {total_pags} pags, {total_etiq} etiq, {total_qtd} un")

    # Salvar resultado atualizado
    if mesclou and ultimo_resultado:
        resultado_path = os.path.join(pasta_saida, "_ultimo_resultado.json")
        try:
            with open(resultado_path, 'w', encoding='utf-8') as f:
                json.dump(ultimo_resultado, f, ensure_ascii=False, indent=2)
        except Exception:
            pass


def _renomear_pastas_saida_para_upseller(
    pasta_saida: str, resultado_proc: dict, estado: dict,
    pdf_loja_map: dict, pasta_lote: str
):
    """
    Renomeia pastas de saida de nomes do marketplace (REMETENTE na etiqueta)
    para nomes do UpSeller, usando o mapeamento PDF→loja do download por-loja.

    Logica:
    - Para cada CNPJ no estado["_etiquetas_por_cnpj"], pegar o caminho_pdf da 1a etiqueta
    - Mapear o basename do PDF para o nome UpSeller via pdf_loja_map
    - Se o nome atual da pasta != nome UpSeller, renomear pasta e atualizar resultado
    """
    import shutil

    etiq_por_cnpj = estado.get("_etiquetas_por_cnpj", {})
    if not etiq_por_cnpj:
        return

    ultimo_resultado = estado.get("ultimo_resultado", {})
    lojas_resultado = ultimo_resultado.get("lojas", []) if ultimo_resultado else []

    # Mapear PDF (que pode estar no lote) → nome UpSeller
    # O mover_para_pasta_entrada copia/move PDFs para pasta_lote, mantendo basename
    lote_pdf_map = {}
    for pdf_base, upseller_nome in pdf_loja_map.items():
        lote_pdf_map[pdf_base.lower()] = upseller_nome
        # Tambem verificar se o PDF foi renomeado no lote
        lote_file = os.path.join(pasta_lote, pdf_base) if pasta_lote else ""
        if lote_file and os.path.exists(lote_file):
            lote_pdf_map[os.path.basename(lote_file).lower()] = upseller_nome

    # Mapear: nome_marketplace_atual → nome_upseller
    renomear = {}  # {nome_marketplace: nome_upseller}

    for cnpj, etiquetas in etiq_por_cnpj.items():
        if not etiquetas:
            continue
        # Pegar o caminho_pdf da primeira etiqueta deste CNPJ
        pdf_path = ""
        for etq in etiquetas:
            pdf_path = etq.get("caminho_pdf", "")
            if pdf_path:
                break
        if not pdf_path:
            continue

        pdf_base = os.path.basename(pdf_path).lower()
        upseller_nome = lote_pdf_map.get(pdf_base)
        if not upseller_nome:
            continue

        # Nome atual da pasta (marketplace) — vem de get_nome_loja(cnpj)
        proc_config = estado.get("_proc_config", {})
        cnpj_loja = proc_config.get("cnpj_loja", {})
        cnpj_nome = proc_config.get("cnpj_nome", {})
        nome_marketplace = cnpj_loja.get(cnpj) or cnpj_nome.get(cnpj, "")
        if not nome_marketplace:
            continue

        # Se os nomes ja sao iguais (case-insensitive), pular
        if nome_marketplace.strip().lower() == upseller_nome.strip().lower():
            continue

        renomear[nome_marketplace] = upseller_nome

    if not renomear:
        print("[ImprDireto] Renomear pastas: nenhuma pasta precisa ser renomeada")
        return

    print(f"[ImprDireto] Renomeando {len(renomear)} pasta(s): {renomear}")

    for nome_antigo, nome_novo in renomear.items():
        pasta_antiga = os.path.join(pasta_saida, nome_antigo)
        pasta_nova = os.path.join(pasta_saida, nome_novo)

        if not os.path.isdir(pasta_antiga):
            print(f"[ImprDireto]   Pasta '{nome_antigo}' nao encontrada, pulando")
            continue
        if os.path.exists(pasta_nova) and pasta_antiga != pasta_nova:
            # Se ja existe pasta com o nome UpSeller, mesclar
            for arq in os.listdir(pasta_antiga):
                src = os.path.join(pasta_antiga, arq)
                dst = os.path.join(pasta_nova, arq)
                if os.path.isfile(src) and not os.path.exists(dst):
                    shutil.move(src, dst)
            shutil.rmtree(pasta_antiga, ignore_errors=True)
        else:
            try:
                os.rename(pasta_antiga, pasta_nova)
            except Exception:
                shutil.move(pasta_antiga, pasta_nova)

        print(f"[ImprDireto]   Renomeado: '{nome_antigo}' -> '{nome_novo}'")

        # Renomear arquivos dentro da pasta para usar o nome UpSeller
        if os.path.isdir(pasta_nova):
            for arq in os.listdir(pasta_nova):
                if nome_antigo in arq:
                    novo_arq = arq.replace(nome_antigo, nome_novo)
                    try:
                        os.rename(
                            os.path.join(pasta_nova, arq),
                            os.path.join(pasta_nova, novo_arq)
                        )
                    except Exception:
                        pass

        # Atualizar resultado em memoria
        for loja_info in lojas_resultado:
            if loja_info.get("nome") == nome_antigo:
                loja_info["nome"] = nome_novo
                loja_info["cnpj"] = f"LOJA_{nome_novo.replace(' ', '_')}"
                # Atualizar pdf
                pdf_old = loja_info.get("pdf", "")
                if pdf_old and nome_antigo in pdf_old:
                    loja_info["pdf"] = pdf_old.replace(nome_antigo, nome_novo)
                # Atualizar xlsx
                xlsx_old = loja_info.get("xlsx", "")
                if xlsx_old and nome_antigo in xlsx_old:
                    loja_info["xlsx"] = xlsx_old.replace(nome_antigo, nome_novo)

    # Atualizar ultimo_resultado no estado e salvar em disco
    if ultimo_resultado:
        ultimo_resultado["total_lojas"] = len(lojas_resultado)
        estado["ultimo_resultado"] = ultimo_resultado
        resultado_path = os.path.join(pasta_saida, "_ultimo_resultado.json")
        try:
            with open(resultado_path, 'w', encoding='utf-8') as f:
                json.dump(ultimo_resultado, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    # Atualizar resumo geral XLSX
    resumo_geral = ultimo_resultado.get("resumo_geral", {}) if ultimo_resultado else {}
    resumo_geral_arq = resumo_geral.get("arquivo", "")
    if resumo_geral_arq:
        resumo_path = os.path.join(pasta_saida, resumo_geral_arq)
        if os.path.exists(resumo_path):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(resumo_path)
                ws = wb.active
                for row in ws.iter_rows(min_row=2):
                    cell_nome = row[0]
                    if cell_nome.value and str(cell_nome.value) in renomear:
                        cell_nome.value = renomear[str(cell_nome.value)]
                wb.save(resumo_path)
                wb.close()
            except Exception:
                pass


def _executar_imprimir_direto(user_id: int, lojas_alvo: list = None) -> dict:
    """
    Pipeline direto para scheduler/execucao manual:
    1. Baixa etiquetas do UpSeller (SEM emitir notas, SEM gerar pedidos)
    2. Baixa XLSX de resumo
    3. Move para pasta de lote
    4. Processa com Beka MKT
    Retorna: {"ok": True/False, "pdfs_movidos": int, "total_etiquetas": int, "total_lojas": int}
    """
    import asyncio

    user = User.query.get(user_id)
    if not user:
        return {"ok": False, "erro": "Usuario nao encontrado"}

    pasta_entrada = user.get_pasta_entrada()
    pasta_lote = _criar_pasta_lote_upseller(pasta_entrada, prefixo="envio_auto")
    # Pasta temp isolada por execucao (uuid) — evita conflito quando 2 contatos rodam proximo
    import uuid as _uuid
    download_dir = os.path.join(pasta_entrada, f'_upseller_temp_{_uuid.uuid4().hex[:8]}')
    os.makedirs(download_dir, exist_ok=True)

    config = _get_or_create_upseller_config(user_id)
    if not config:
        return {"ok": False, "erro": "UpSeller nao configurado"}

    # Criar/reutilizar scraper
    # IMPORTANTE: usar headless=False pois UpSeller bloqueia headless.
    # Se ja existir um scraper vivo (ex: usuario conectou via UI), reutiliza.
    print(f"[ImprDireto] Verificando scraper para user {user_id}...")
    scraper_vivo = _upseller_mgr.is_alive(user_id)
    print(f"[ImprDireto] scraper_vivo={scraper_vivo}")
    if not scraper_vivo:
        try:
            print(f"[ImprDireto] Criando scraper (headless=False)...")
            scraper = _upseller_mgr.criar_scraper(
                user_id, config.session_dir, download_dir, headless=False
            )
            print(f"[ImprDireto] Scraper criado, verificando login...")
            logado = _upseller_mgr._run_async(user_id, scraper._esta_logado())
            print(f"[ImprDireto] logado={logado}")
            if not logado:
                return {"ok": False, "erro": "Sessao UpSeller expirada. Reconecte manualmente."}
            config.status_conexao = "ok"
            db.session.commit()
            # Minimizar janela para nao atrapalhar
            try:
                _upseller_mgr._run_async(user_id, scraper._page.evaluate("""
                    (() => { try { window.resizeTo(1,1); window.moveTo(-2000,-2000); } catch(e){} })()
                """))
            except Exception:
                pass
        except Exception as e:
            print(f"[ImprDireto] ERRO ao conectar: {e}")
            import traceback; traceback.print_exc()
            return {"ok": False, "erro": f"Erro ao conectar UpSeller: {e}"}

    scraper = _upseller_mgr.get_scraper(user_id)
    if not scraper:
        return {"ok": False, "erro": "Scraper nao disponivel"}
    scraper.download_dir = download_dir

    filtro = lojas_alvo if lojas_alvo else None

    # Baixar XLSX de resumo
    xlsx_paths = []
    try:
        print(f"[ImprDireto] Baixando XLSX de resumo (filtro={filtro})...")
        xlsx_lote = _upseller_mgr._run_async(
            user_id, scraper.baixar_lista_resumo(filtro_loja=filtro)
        ) or []
        if isinstance(xlsx_lote, str):
            xlsx_lote = [xlsx_lote]
        for xp in xlsx_lote:
            if xp and os.path.exists(xp):
                xlsx_paths.append(xp)
        print(f"[ImprDireto] {len(xlsx_paths)} XLSX(s) baixado(s)")
    except Exception as e:
        print(f"[ImprDireto] Aviso: falha ao baixar XLSX: {e}")

    # Baixar etiquetas — UMA LOJA POR VEZ quando ha multiplas lojas.
    # O filtro multi-loja do UpSeller pode selecionar lojas erradas;
    # baixar por-loja e mais confiavel e permite rastrear qual PDF veio de qual loja.
    pdfs = []
    _pdf_loja_map = {}  # {pdf_basename -> upseller_name} para renomear depois

    if filtro and len(filtro) > 1:
        # Multiplas lojas: baixar uma por vez
        print(f"[ImprDireto] Baixando etiquetas por loja ({len(filtro)} lojas)...")
        for loja_nome in filtro:
            try:
                print(f"[ImprDireto]   Baixando loja '{loja_nome}'...")
                pdfs_loja = _upseller_mgr._run_async(
                    user_id, scraper.baixar_etiquetas(filtro_loja=loja_nome)
                ) or []
                if isinstance(pdfs_loja, str):
                    pdfs_loja = [pdfs_loja]
                for p in pdfs_loja:
                    if p:
                        pdfs.append(p)
                        _pdf_loja_map[os.path.basename(p)] = loja_nome
                if pdfs_loja:
                    print(f"[ImprDireto]   '{loja_nome}': {len(pdfs_loja)} PDF(s)")
                else:
                    print(f"[ImprDireto]   '{loja_nome}': nenhuma etiqueta pendente")
            except Exception as e_loja:
                print(f"[ImprDireto]   ERRO ao baixar '{loja_nome}': {e_loja}")
        print(f"[ImprDireto] Total: {len(pdfs)} PDF(s) de {len(filtro)} lojas")
    else:
        # Loja unica ou sem filtro: baixar normalmente
        try:
            print(f"[ImprDireto] Baixando etiquetas (filtro={filtro})...")
            pdfs_lote = _upseller_mgr._run_async(
                user_id, scraper.baixar_etiquetas(filtro_loja=filtro)
            ) or []
            if isinstance(pdfs_lote, list):
                pdfs = pdfs_lote
            elif pdfs_lote:
                pdfs = [pdfs_lote]
            # Se foi filtro de 1 loja, mapear
            if filtro and len(filtro) == 1:
                for p in pdfs:
                    _pdf_loja_map[os.path.basename(p)] = filtro[0]
            print(f"[ImprDireto] {len(pdfs)} PDF(s) baixado(s)")
        except Exception as e:
            return {"ok": False, "erro": f"Erro ao baixar etiquetas: {e}"}

    if not pdfs:
        return {
            "ok": False,
            "erro": (
                "Nenhuma etiqueta baixada. "
                "Se houver etiquetas pendentes, verifique se as configuracoes de impressao "
                "estao definidas no UpSeller (Imprimir Etiquetas > Ir para Configurar)."
            )
        }

    # Mover para lote isolado
    xlsx_path = xlsx_paths[0] if xlsx_paths else ""
    xlsx_extra = xlsx_paths[1:] if len(xlsx_paths) > 1 else []
    resultado_dl = {"pdfs": pdfs, "xmls": [], "xlsx": xlsx_path, "xlsx_extra": xlsx_extra}
    resumo = scraper.mover_para_pasta_entrada(resultado_dl, pasta_lote)

    if resumo.get("pdfs_movidos", 0) <= 0:
        return {"ok": False, "erro": "Nenhum PDF foi movido para o lote"}

    # Processar com Beka MKT
    estado = _get_estado(user_id)
    if estado is not None:
        estado["_pdf_loja_map"] = dict(_pdf_loja_map)
    proc_result = _executar_processamento(
        user_id,
        sem_recorte=True,
        resumo_sku_somente=False,
        pasta_entrada_override=pasta_lote,
    )

    # Aguardar conclusao do processamento (poll a cada 2s, max 10min)
    estado = _get_estado(user_id)
    elapsed = 0
    while estado.get("processando") and elapsed < 600:
        time.sleep(2)
        elapsed += 2

    resultado_proc = estado.get("ultimo_resultado", {})

    # Limpar pasta temp
    try:
        import shutil
        shutil.rmtree(download_dir, ignore_errors=True)
    except Exception:
        pass

    if not proc_result or not proc_result.get("ok"):
        return {"ok": False, "erro": (proc_result or {}).get("erro", "Erro no processamento")}

    # ---- Renomear pastas de saida: marketplace → UpSeller ----
    # O processador usa nomes do REMETENTE na etiqueta (marketplace), mas o usuario
    # conhece suas lojas pelo nome do UpSeller. Quando temos o mapeamento
    # PDF→loja (download por-loja), podemos renomear.
    pasta_saida = user.get_pasta_saida()

    # Salvar mapa PDF→UpSeller no estado para uso no agrupamento
    if estado is not None:
        estado["_pdf_loja_map"] = dict(_pdf_loja_map)

    if _pdf_loja_map and resultado_proc:
        try:
            _renomear_pastas_saida_para_upseller(
                pasta_saida, resultado_proc, estado, _pdf_loja_map, pasta_lote
            )
        except Exception as e_ren:
            print(f"[ImprDireto] Aviso: falha ao renomear pastas: {e_ren}")
            import traceback; traceback.print_exc()

    # ---- Mesclar lojas agrupadas (ex: HEITOR+LEONE → Leone/) ----
    try:
        _mesclar_lojas_agrupadas(pasta_saida, estado, user_id)
    except Exception as e_merge:
        print(f"[ImprDireto] Aviso: falha ao mesclar agrupamentos: {e_merge}")
        import traceback; traceback.print_exc()

    # JPEG de resumo desativado — pasta de saida contem apenas PDFs

    return {
        "ok": True,
        "pdfs_movidos": resumo.get("pdfs_movidos", 0),
        "total_etiquetas": resultado_proc.get("total_etiquetas", proc_result.get("total_etiquetas", 0)),
        "total_lojas": resultado_proc.get("total_lojas", proc_result.get("total_lojas", 0)),
    }


def _enviar_email_resultado_agendado(user_id: int, resultado: dict = None) -> dict:
    """Envia ultimo resultado via email para contatos configurados (uso do scheduler)."""
    user = User.query.get(user_id)
    if not user:
        return {"ok": False, "erro": "Usuario nao encontrado"}

    estado = _get_estado(user_id)
    resultado_ref = resultado or (estado.get("ultimo_resultado", {}) if estado else {})
    if not resultado_ref or not resultado_ref.get("lojas"):
        return {"ok": False, "erro": "Nenhum resultado para enviar"}

    from_addr = (getattr(user, "email_remetente", "") or "").strip()
    if not from_addr or "@" not in from_addr:
        return {"ok": False, "erro": "Email remetente nao configurado"}

    smtp_cfg, _ = _smtp_config_resolver(user)
    if not smtp_cfg:
        return {"ok": False, "erro": "SMTP nao configurado"}

    contatos = EmailContact.query.filter_by(user_id=user_id, ativo=True).all()
    if not contatos:
        return {"ok": False, "erro": "Nenhum contato email cadastrado"}

    pasta_saida = user.get_pasta_saida()
    agrupamentos = (estado or {}).get("agrupamentos", []) if estado else []
    envios, diagnostico = montar_destinos_por_resultado(
        resultado=resultado_ref,
        pasta_saida=pasta_saida,
        contatos=contatos,
        destino_attr="email",
        agrupamentos_usuario=agrupamentos,
    )

    if not envios:
        return {"ok": False, "erro": "Nenhuma entrega email valida", "diagnostico": diagnostico}

    timestamp = resultado_ref.get("timestamp", "")
    from_name = (getattr(user, "nome_remetente", "") or "").strip()

    # Agrupar por (destino, loja)
    envios_agrupados = {}
    for envio in envios:
        destino = str(envio.get("destino", "") or "").strip()
        loja = str(envio.get("loja", "") or "").strip()
        file_path = str(envio.get("file_path", envio.get("pdf_path", "")) or "").strip()
        if not destino or not file_path:
            continue
        chave = (destino.lower(), loja)
        if chave not in envios_agrupados:
            envios_agrupados[chave] = {"destino": destino, "loja": loja, "arquivos": []}
        if file_path not in envios_agrupados[chave]["arquivos"]:
            envios_agrupados[chave]["arquivos"].append(file_path)

    total_ok = 0
    for grupo in envios_agrupados.values():
        try:
            res = enviar_email_com_anexos(
                email_destino=grupo["destino"],
                assunto=f"Arquivos {grupo['loja']} - {timestamp}",
                loja_nome=grupo["loja"],
                timestamp=timestamp,
                anexos_paths=grupo["arquivos"],
                from_addr_override=from_addr,
                from_name_override=from_name,
                smtp_override=smtp_cfg,
            )
            if res.get("success"):
                total_ok += 1
        except Exception as e:
            print(f"[EmailAgendado] Erro ao enviar para {grupo['destino']}: {e}")
        time.sleep(2)

    return {"ok": True, "total": total_ok, "total_envios": len(envios_agrupados)}


def _enviar_email_para_contato(user_id: int, contato_id: int) -> dict:
    """Envia resultado por email para UM contato especifico."""
    uid = int(user_id)
    contato = EmailContact.query.filter_by(id=contato_id, user_id=uid, ativo=True).first()
    if not contato:
        return {"ok": False, "erro": "Contato email nao encontrado"}

    estado = _get_estado(uid)
    resultado_ref = (estado.get("ultimo_resultado", {}) if estado else {})
    if not resultado_ref:
        return {"ok": False, "erro": "Nenhum resultado"}

    user = User.query.get(uid)
    pasta_saida = user.get_pasta_saida()
    agrupamentos_usuario = (estado or {}).get("agrupamentos", []) if estado else []

    envios, _ = montar_destinos_por_resultado(
        resultado=resultado_ref,
        pasta_saida=pasta_saida,
        contatos=[contato],
        destino_attr="email",
        agrupamentos_usuario=agrupamentos_usuario,
    )

    enviados = 0
    for envio in envios:
        email_dest = envio.get("destino", "")
        file_path = envio.get("file_path", "")
        if not email_dest or not file_path:
            continue
        try:
            enviar_email_com_anexo(uid, email_dest, f"Etiquetas {envio.get('loja','')}", file_path)
            enviados += 1
        except Exception as e:
            print(f"[EmailIndividual] Erro: {e}")

    return {"ok": enviados > 0, "enviados": enviados}


def _enviar_email_para_contato_direto(user_id: int, contato_id: int) -> dict:
    """
    Envia TODOS os arquivos do ultimo resultado por email para UM contato,
    SEM depender de matching de nomes de lojas.
    Usado na execucao individual.
    """
    uid = int(user_id)
    contato = EmailContact.query.filter_by(id=contato_id, user_id=uid, ativo=True).first()
    if not contato:
        return {"ok": False, "erro": "Contato email nao encontrado"}

    email_dest = str(getattr(contato, "email", "") or "").strip()
    if not email_dest:
        return {"ok": False, "erro": "Contato sem email"}

    estado = _get_estado(uid)
    resultado_ref = (estado.get("ultimo_resultado", {}) if estado else {})
    if not resultado_ref or not resultado_ref.get("lojas"):
        return {"ok": False, "erro": "Nenhum resultado"}

    user = User.query.get(uid)
    pasta_saida = user.get_pasta_saida()
    timestamp = resultado_ref.get("timestamp", "")
    lojas_resultado = resultado_ref.get("lojas", []) or []

    from_addr = (getattr(user, "email_remetente", "") or "").strip()
    if not from_addr or "@" not in from_addr:
        return {"ok": False, "erro": "Email remetente nao configurado"}

    smtp_cfg, _ = _smtp_config_resolver(user)
    if not smtp_cfg:
        return {"ok": False, "erro": "SMTP nao configurado"}

    from_name = (getattr(user, "nome_remetente", "") or "").strip()
    from whatsapp_delivery import _listar_arquivos_loja
    total_ok = 0

    # Para cada loja no resultado, enviar todos os arquivos
    for loja_info in lojas_resultado:
        nome = str((loja_info or {}).get("nome", "") or "")
        pdf_nome = str((loja_info or {}).get("pdf", "") or "")
        arquivos = _listar_arquivos_loja(pasta_saida, nome, pdf_hint=pdf_nome)
        if not arquivos:
            continue
        try:
            res = enviar_email_com_anexos(
                email_destino=email_dest,
                assunto=f"Etiquetas {nome} - {timestamp}",
                loja_nome=nome,
                timestamp=timestamp,
                anexos_paths=arquivos,
                from_addr_override=from_addr,
                from_name_override=from_name,
                smtp_override=smtp_cfg,
            )
            if res.get("success"):
                total_ok += 1
        except Exception as e:
            print(f"[EmailIndividualDireto] Erro ao enviar para {email_dest}: {e}")
        time.sleep(2)

    return {"ok": total_ok > 0, "enviados": total_ok}


def _ja_na_fila_whatsapp(uid: int, telefone: str, file_path: str, loja_nome: str = "") -> bool:
    """Verifica se um item identico ja esta pendente ou foi enviado recentemente (2h).
    Evita duplicacao tanto cross-batch (pendentes) quanto re-envio (recentes)."""
    telefone_norm, caminho_norm, arquivo_base, loja_norm = _normalizar_chave_whatsapp_queue(
        telefone, file_path, loja_nome
    )
    if not telefone_norm or not caminho_norm:
        return False

    limite = _agora_utc() - timedelta(hours=2)
    candidatos = []
    candidatos.extend(
        WhatsAppQueueItem.query.filter(
            WhatsAppQueueItem.user_id == uid,
            WhatsAppQueueItem.status.in_(["pending", "retry", "sending"]),
        ).all()
    )
    candidatos.extend(
        WhatsAppQueueItem.query.filter(
            WhatsAppQueueItem.user_id == uid,
            WhatsAppQueueItem.status == "sent",
            WhatsAppQueueItem.sent_at >= limite,
        ).all()
    )

    for item in candidatos:
        tel_item, caminho_item, arquivo_item, loja_item = _normalizar_chave_whatsapp_queue(
            getattr(item, "telefone", ""),
            getattr(item, "pdf_path", ""),
            getattr(item, "loja_nome", ""),
        )
        if tel_item != telefone_norm:
            continue
        if caminho_item == caminho_norm:
            return True
        if loja_norm and loja_item == loja_norm and arquivo_base and arquivo_item == arquivo_base:
            return True
    return False


def _enfileirar_envio_whatsapp_resultado(
    user_id: int,
    resultado: dict = None,
    origem: str = "manual",
    respeitar_toggle_auto: bool = False,
) -> dict:
    """Enfileira envios WhatsApp a partir do ultimo resultado processado."""
    uid = int(user_id)
    user = User.query.get(uid)
    if not user:
        return {"ok": False, "erro": "Usuario nao encontrado"}

    if respeitar_toggle_auto and not _to_bool(getattr(user, "auto_send_whatsapp", False), False):
        return {"ok": False, "erro": "Auto-envio desabilitado", "ignorado": True}

    estado = _get_estado(uid)
    resultado_ref = resultado or (estado.get("ultimo_resultado", {}) if estado else {})
    if not resultado_ref or not resultado_ref.get("lojas"):
        return {"ok": False, "erro": "Nenhum resultado para enviar"}

    contatos = WhatsAppContact.query.filter_by(user_id=uid, ativo=True).all()
    if not contatos:
        return {"ok": False, "erro": "Nenhum contato WhatsApp cadastrado"}

    pasta_saida = user.get_pasta_saida()
    agrupamentos_usuario = (estado or {}).get("agrupamentos", []) if estado else []
    entregas, diagnostico = montar_entregas_por_resultado(
        resultado=resultado_ref,
        pasta_saida=pasta_saida,
        contatos=contatos,
        agrupamentos_usuario=agrupamentos_usuario,
    )
    if not entregas:
        return {"ok": False, "erro": "Nenhuma entrega valida para enviar", "diagnostico": diagnostico}

    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    agora = _agora_utc()
    enfileirados = 0
    duplicados_batch = 0
    duplicados_cross = 0
    dedupe_batch = set()
    with _get_whatsapp_enqueue_lock(uid):
        for ent in entregas:
            file_path = ent.get("file_path", ent.get("pdf_path", ""))
            if not file_path:
                continue
            telefone_ent = ent.get("telefone", "")
            loja_ent = ent.get("loja", "")
            chave_batch = _chave_dedupe_batch_whatsapp(telefone_ent, file_path, loja_ent)
            if chave_batch and chave_batch in dedupe_batch:
                duplicados_batch += 1
                print(f"[WhatsApp] Dedupe skip (batch): loja={loja_ent} tel={telefone_ent} arquivo={os.path.basename(file_path)}")
                continue
            if chave_batch:
                dedupe_batch.add(chave_batch)
            # Cross-batch dedup: pular se pendente ou enviado recentemente
            if _ja_na_fila_whatsapp(uid, telefone_ent, file_path, loja_ent):
                duplicados_cross += 1
                print(f"[WhatsApp] Dedupe skip: loja={loja_ent} tel={telefone_ent} arquivo={os.path.basename(file_path)}")
                continue
            db.session.add(WhatsAppQueueItem(
                user_id=uid,
                batch_id=batch_id,
                origem=origem,
                loja_nome=loja_ent,
                telefone=telefone_ent,
                pdf_path=file_path,
                caption=ent.get("caption", ""),
                status="pending",
                tentativas=0,
                max_tentativas=5,
                next_attempt_at=agora,
            ))
            enfileirados += 1
        if enfileirados > 0:
            db.session.commit()
    if duplicados_batch:
        print(f"[WhatsApp] {duplicados_batch} item(ns) ignorado(s) por dedup no batch atual")
    if duplicados_cross:
        print(f"[WhatsApp] {duplicados_cross} item(ns) ignorado(s) por cross-batch dedup")
    if enfileirados <= 0:
        db.session.rollback()
        if duplicados_batch > 0 or duplicados_cross > 0:
            return {
                "ok": False,
                "erro": "Todos os arquivos dessa selecao ja foram enviados recentemente no WhatsApp.",
                "diagnostico": diagnostico,
                "duplicados_batch": duplicados_batch,
                "duplicados_ignorados": duplicados_cross,
            }
        return {"ok": False, "erro": "Nenhum arquivo valido para enfileirar", "diagnostico": diagnostico}

    _garantir_baileys_rodando(motivo=f"queue:{origem}")
    return {
        "ok": True,
        "batch_id": batch_id,
        "total_entregas": enfileirados,
        "diagnostico": diagnostico,
    }


def _enfileirar_envio_whatsapp_para_contato(user_id: int, contato_id: int) -> dict:
    """Enfileira envios WhatsApp para UM contato especifico."""
    uid = int(user_id)
    user = User.query.get(uid)
    if not user:
        return {"ok": False, "erro": "Usuario nao encontrado"}

    estado = _get_estado(uid)
    resultado_ref = (estado.get("ultimo_resultado", {}) if estado else {})
    if not resultado_ref or not resultado_ref.get("lojas"):
        return {"ok": False, "erro": "Nenhum resultado para enviar"}

    contato = WhatsAppContact.query.filter_by(id=contato_id, user_id=uid, ativo=True).first()
    if not contato:
        return {"ok": False, "erro": "Contato nao encontrado ou inativo"}

    contatos = [contato]
    pasta_saida = user.get_pasta_saida()
    agrupamentos_usuario = (estado or {}).get("agrupamentos", []) if estado else []
    entregas, diagnostico = montar_entregas_por_resultado(
        resultado=resultado_ref,
        pasta_saida=pasta_saida,
        contatos=contatos,
        agrupamentos_usuario=agrupamentos_usuario,
    )
    if not entregas:
        return {"ok": False, "erro": "Nenhuma entrega valida", "diagnostico": diagnostico}

    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    agora = _agora_utc()
    enfileirados = 0
    duplicados_batch = 0
    duplicados_cross = 0
    dedupe_batch = set()
    with _get_whatsapp_enqueue_lock(uid):
        for ent in entregas:
            file_path = ent.get("file_path", ent.get("pdf_path", ""))
            if not file_path:
                continue
            telefone_ent = ent.get("telefone", "")
            loja_ent = ent.get("loja", "")
            chave_batch = _chave_dedupe_batch_whatsapp(telefone_ent, file_path, loja_ent)
            if chave_batch and chave_batch in dedupe_batch:
                duplicados_batch += 1
                print(f"[WhatsApp] Dedupe skip (batch contato): loja={loja_ent} tel={telefone_ent} arquivo={os.path.basename(file_path)}")
                continue
            if chave_batch:
                dedupe_batch.add(chave_batch)
            # Cross-batch dedup: pular se pendente ou enviado recentemente
            if _ja_na_fila_whatsapp(uid, telefone_ent, file_path, loja_ent):
                duplicados_cross += 1
                print(f"[WhatsApp] Dedupe skip (contato): loja={loja_ent} tel={telefone_ent} arquivo={os.path.basename(file_path)}")
                continue
            db.session.add(WhatsAppQueueItem(
                user_id=uid, batch_id=batch_id, origem="individual",
                loja_nome=loja_ent, telefone=telefone_ent,
                pdf_path=file_path, caption=ent.get("caption", ""),
                status="pending", tentativas=0, max_tentativas=5,
                next_attempt_at=agora,
            ))
            enfileirados += 1
        if enfileirados > 0:
            db.session.commit()
    if duplicados_batch:
        print(f"[WhatsApp] {duplicados_batch} item(ns) ignorado(s) por dedup no batch atual (contato)")
    if duplicados_cross:
        print(f"[WhatsApp] {duplicados_cross} item(ns) ignorado(s) por cross-batch dedup (contato)")
    if enfileirados <= 0:
        db.session.rollback()
        if duplicados_batch > 0 or duplicados_cross > 0:
            return {
                "ok": False,
                "erro": "Os arquivos desse contato ja foram enviados recentemente no WhatsApp.",
                "duplicados_batch": duplicados_batch,
                "duplicados_ignorados": duplicados_cross,
            }
        return {"ok": False, "erro": "Nenhum arquivo valido para enfileirar"}
    _garantir_baileys_rodando(motivo="individual")
    return {"ok": True, "batch_id": batch_id, "total_entregas": enfileirados, "diagnostico": diagnostico}


def _enfileirar_whatsapp_todos_arquivos_para_contato(user_id: int, contato_id: int, lojas_filtro=None) -> dict:
    """
    Enfileira TODOS os arquivos do ultimo resultado para UM contato,
    SEM depender de matching de nomes de lojas.
    Usado na execucao individual: como ja filtramos no download,
    todo conteudo pertence a esse contato.

    Args:
        lojas_filtro: lista opcional de nomes de lojas. Quando fornecido,
                      enfileira apenas arquivos dessas lojas especificas.
    """
    uid = int(user_id)
    user = User.query.get(uid)
    if not user:
        return {"ok": False, "erro": "Usuario nao encontrado"}

    estado = _get_estado(uid)
    resultado_ref = (estado.get("ultimo_resultado", {}) if estado else {})
    if not resultado_ref or not resultado_ref.get("lojas"):
        return {"ok": False, "erro": "Nenhum resultado para enviar"}

    contato = WhatsAppContact.query.filter_by(id=contato_id, user_id=uid, ativo=True).first()
    if not contato:
        return {"ok": False, "erro": "Contato nao encontrado ou inativo"}

    telefone = str(getattr(contato, "telefone", "") or "").strip()
    if not telefone:
        return {"ok": False, "erro": "Contato sem telefone"}

    pasta_saida = user.get_pasta_saida()
    timestamp = resultado_ref.get("timestamp", "")
    lojas_resultado = resultado_ref.get("lojas", []) or []

    # Normalizar lojas_filtro para comparacao case-insensitive
    lojas_filtro_norm = None
    if lojas_filtro:
        lojas_filtro_norm = {str(x).strip().lower() for x in lojas_filtro if str(x).strip()}

    # Coletar TODOS os arquivos de TODAS as lojas do resultado
    from whatsapp_delivery import _listar_arquivos_loja, _deduplicar_entregas_logicas
    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    agora = _agora_utc()
    enfileirados = 0
    duplicados_batch = 0
    duplicados_cross = 0
    duplicados_logicos = 0
    dedupe_arquivos = set()
    dedupe_batch = set()
    candidatos = []

    for loja_info in lojas_resultado:
        nome = str((loja_info or {}).get("nome", "") or "")
        # Filtrar por lojas_filtro se fornecido
        if lojas_filtro_norm and nome.strip().lower() not in lojas_filtro_norm:
            continue
        pdf_nome = str((loja_info or {}).get("pdf", "") or "")
        arquivos = _listar_arquivos_loja(pasta_saida, nome, pdf_hint=pdf_nome)
        caption_base = f"Etiquetas {nome} - {timestamp}"

        for file_path in arquivos:
            abs_path = os.path.abspath(file_path)
            if abs_path in dedupe_arquivos:
                continue
            dedupe_arquivos.add(abs_path)
            candidatos.append({
                "telefone": telefone,
                "file_path": file_path,
                "pdf_path": file_path,
                "loja": nome,
                "caption": f"{caption_base} - {os.path.basename(file_path)}",
            })

    candidatos, duplicados_logicos = _deduplicar_entregas_logicas(candidatos, "telefone")

    with _get_whatsapp_enqueue_lock(uid):
        for ent in candidatos:
            file_path = ent.get("file_path", ent.get("pdf_path", ""))
            nome = ent.get("loja", "")
            chave_batch = _chave_dedupe_batch_whatsapp(telefone, file_path, nome)
            if chave_batch and chave_batch in dedupe_batch:
                duplicados_batch += 1
                print(f"[ExecIndividual] Dedupe skip (batch): loja={nome} tel={telefone} arquivo={os.path.basename(file_path)}")
                continue
            if chave_batch:
                dedupe_batch.add(chave_batch)
            # Cross-batch dedup: pular se pendente ou enviado recentemente
            if _ja_na_fila_whatsapp(uid, telefone, file_path, nome):
                duplicados_cross += 1
                print(f"[ExecIndividual] Dedupe skip: loja={nome} tel={telefone} arquivo={os.path.basename(file_path)}")
                continue
            db.session.add(WhatsAppQueueItem(
                user_id=uid, batch_id=batch_id, origem="individual",
                loja_nome=nome, telefone=telefone,
                pdf_path=file_path,
                caption=ent.get("caption", ""),
                status="pending", tentativas=0, max_tentativas=5,
                next_attempt_at=agora,
            ))
            enfileirados += 1
        if enfileirados > 0:
            db.session.commit()

    if duplicados_logicos:
        print(f"[ExecIndividual] {duplicados_logicos} item(ns) logico(s) ignorado(s), mantendo apenas a versao mais recente")
    if duplicados_batch:
        print(f"[ExecIndividual] {duplicados_batch} item(ns) ignorado(s) por dedup no batch atual")
    if duplicados_cross:
        print(f"[ExecIndividual] {duplicados_cross} item(ns) ignorado(s) por cross-batch dedup")
    if enfileirados <= 0:
        db.session.rollback()
        if duplicados_logicos > 0 or duplicados_batch > 0 or duplicados_cross > 0:
            return {
                "ok": False,
                "erro": "Os arquivos desse contato ja foram enviados recentemente no WhatsApp.",
                "duplicados_logicos": duplicados_logicos,
                "duplicados_batch": duplicados_batch,
                "duplicados_ignorados": duplicados_cross,
            }
        return {"ok": False, "erro": "Nenhum arquivo encontrado no resultado"}
    _garantir_baileys_rodando(motivo="individual_direto")
    print(f"[ExecIndividual] Enfileirados {enfileirados} arquivo(s) direto para tel={telefone}")
    return {"ok": True, "batch_id": batch_id, "total_entregas": enfileirados}


def _status_batch_fila_whatsapp(user_id: int, batch_id: str) -> dict:
    itens = WhatsAppQueueItem.query.filter_by(user_id=int(user_id), batch_id=batch_id).all()
    total = len(itens)
    counts = defaultdict(int)
    for it in itens:
        counts[it.status or "pending"] += 1
    sent = counts.get("sent", 0)
    dead = counts.get("dead", 0)
    pending = counts.get("pending", 0) + counts.get("retry", 0) + counts.get("sending", 0)
    done = sent + dead
    em_andamento = total > 0 and done < total
    if total == 0:
        etapa = "idle"
        detalhes = "Fila vazia"
    elif em_andamento:
        etapa = "enviando"
        detalhes = f"Fila em andamento: {sent}/{total} enviado(s), {counts.get('retry', 0)} em retry"
    else:
        etapa = "concluido" if dead == 0 else "parcial"
        detalhes = f"Fila finalizada: {sent}/{total} enviado(s), {dead} falha(s)"
    progresso = int((done / total) * 100) if total else 0
    return {
        "batch_id": batch_id,
        "etapa": etapa,
        "em_andamento": em_andamento,
        "progresso": progresso,
        "detalhes": detalhes,
        "total": total,
        "enviados": sent,
        "erros": dead,
        "retry": counts.get("retry", 0),
        "pendentes": pending,
    }


def _processar_fila_whatsapp_once() -> int:
    """Processa ate 8 itens da fila. Retorna quantos itens foram pegos."""
    with _try_acquire_whatsapp_worker_guard() as worker_lock:
        if worker_lock is None:
            return 0

        agora = _agora_utc()

        # Rearm de itens presos em "sending".
        limite_stale = agora - timedelta(minutes=10)
        stale = WhatsAppQueueItem.query.filter(
            WhatsAppQueueItem.status == "sending",
            WhatsAppQueueItem.updated_at < limite_stale
        ).all()
        for item in stale:
            item.status = "retry"
            item.next_attempt_at = agora
            item.last_error = "Timeout interno: item rearmado automaticamente."
        if stale:
            db.session.commit()

        itens = WhatsAppQueueItem.query.filter(
            WhatsAppQueueItem.status.in_(["pending", "retry"]),
            WhatsAppQueueItem.next_attempt_at <= agora
        ).order_by(
            WhatsAppQueueItem.next_attempt_at.asc(),
            WhatsAppQueueItem.id.asc()
        ).limit(8).all()
        if not itens:
            return 0

        if not _baileys_healthy():
            _garantir_baileys_rodando(motivo="worker")
            return len(itens)

        wa = WhatsAppService()
        prov_pronto, prov_erro = _whatsapp_provider_pronto(wa)
        if not prov_pronto:
            msg = (prov_erro or "Sessao nao conectada")[:500]
            for item in itens:
                fresh = WhatsAppQueueItem.query.get(item.id)
                if not fresh or fresh.status not in ("pending", "retry"):
                    continue
                fresh.status = "retry"
                fresh.next_attempt_at = _agora_utc() + timedelta(minutes=2)
                fresh.last_error = msg
                fresh.updated_at = _agora_utc()
            db.session.commit()
            print(f"[WhatsAppWorker] Provedor indisponivel, itens rearmados para retry: {msg}")
            return len(itens)

        for item in itens:
            try:
                _aplicar_cooldown_envio_whatsapp(item.telefone)

                # Re-verificar status atomicamente (protecao contra duplo worker)
                fresh = WhatsAppQueueItem.query.get(item.id)
                if not fresh or fresh.status not in ("pending", "retry"):
                    continue  # Ja foi processado por outro worker
                fresh.status = "sending"
                fresh.tentativas = (fresh.tentativas or 0) + 1
                fresh.updated_at = _agora_utc()
                db.session.commit()
                item = fresh  # usar referencia atualizada

                if not os.path.exists(item.pdf_path or ""):
                    raise FileNotFoundError(f"Arquivo nao encontrado: {item.pdf_path}")

                ext = os.path.splitext(item.pdf_path or "")[1].lower()
                if ext in (".jpg", ".jpeg", ".png", ".webp"):
                    res = wa.enviar_imagem(item.telefone, item.pdf_path, item.caption or "")
                else:
                    res = wa.enviar_arquivo(item.telefone, item.pdf_path, item.caption or "")
                if res.get("success"):
                    item.status = "sent"
                    item.sent_at = _agora_utc()
                    item.message_id = (res.get("messageId") or "")[:190]
                    item.last_error = ""
                else:
                    msg = (res.get("error") or "Falha desconhecida")[:500]
                    if _erro_sessao_whatsapp_desconectada(msg):
                        item.status = "retry"
                        item.tentativas = max(0, (item.tentativas or 1) - 1)
                        item.next_attempt_at = _agora_utc() + timedelta(minutes=2)
                        item.last_error = msg
                    elif (item.tentativas or 0) >= (item.max_tentativas or 5):
                        item.status = "dead"
                        item.last_error = msg
                    else:
                        item.status = "retry"
                        item.next_attempt_at = _agora_utc() + timedelta(
                            seconds=_calc_backoff_seconds(item.tentativas or 1)
                        )
                        item.last_error = msg
                _registrar_envio_whatsapp(item.telefone)
                item.updated_at = _agora_utc()
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                try:
                    item_ref = WhatsAppQueueItem.query.get(item.id)
                    if item_ref:
                        msg = str(e)[:500]
                        if _erro_sessao_whatsapp_desconectada(msg):
                            item_ref.status = "retry"
                            item_ref.next_attempt_at = _agora_utc() + timedelta(minutes=2)
                            item_ref.last_error = msg
                        else:
                            item_ref.tentativas = (item_ref.tentativas or 0) + 1
                            if (item_ref.tentativas or 0) >= (item_ref.max_tentativas or 5):
                                item_ref.status = "dead"
                                item_ref.last_error = msg
                            else:
                                item_ref.status = "retry"
                                item_ref.next_attempt_at = _agora_utc() + timedelta(
                                    seconds=_calc_backoff_seconds(item_ref.tentativas or 1)
                                )
                                item_ref.last_error = msg
                        _registrar_envio_whatsapp(item_ref.telefone)
                        if _erro_sessao_whatsapp_desconectada(msg):
                            item_ref.tentativas = max(0, item_ref.tentativas or 0)
                        item_ref.updated_at = _agora_utc()
                        db.session.commit()
                except Exception:
                    db.session.rollback()
        return len(itens)


def _whatsapp_queue_worker_loop():
    print("[WhatsAppQueueWorker] iniciado")
    while True:
        try:
            with app.app_context():
                picked = _processar_fila_whatsapp_once()
            time.sleep(1.0 if picked else 3.0)
        except Exception as e:
            print(f"[WhatsAppQueueWorker] erro: {e}")
            time.sleep(5)


def _baileys_supervisor_loop():
    print("[BaileysSupervisor] iniciado")
    while True:
        try:
            _garantir_baileys_rodando(motivo="supervisor")
        except Exception as e:
            print(f"[BaileysSupervisor] erro no loop: {e}")
        time.sleep(20)


def _iniciar_background_workers():
    global _BACKGROUND_WORKERS_STARTED
    with _BACKGROUND_WORKERS_LOCK:
        if _BACKGROUND_WORKERS_STARTED:
            return
        _BACKGROUND_WORKERS_STARTED = True
        threading.Thread(target=_baileys_supervisor_loop, daemon=True).start()
        threading.Thread(target=_whatsapp_queue_worker_loop, daemon=True).start()


# ================================================================
# AUTOMACAO: UpSeller + WhatsApp + Agendamentos
# ================================================================

# ----------------------------------------------------------------
# ENDPOINTS - UPSELLER (sessao persistente, sem credenciais)
# ----------------------------------------------------------------

def _get_upseller_session_dir(user_id):
    """Retorna diretorio de sessao do UpSeller para o usuario."""
    d = os.path.join(os.path.expanduser("~"), ".upseller_sessions", str(user_id))
    os.makedirs(d, exist_ok=True)
    return d


def _get_or_create_upseller_config(user_id):
    """Retorna ou cria config do UpSeller (sem precisar de email/senha)."""
    config = UpSellerConfig.query.filter_by(user_id=user_id).first()
    if not config:
        config = UpSellerConfig(
            user_id=user_id,
            email="",
            session_dir=_get_upseller_session_dir(user_id),
        )
        config.password_encrypted = ""
        db.session.add(config)
        db.session.commit()
    # Corrigir session_dir antigo se necessario
    if not config.session_dir or '/tmp/' in config.session_dir or '\\tmp\\' in config.session_dir:
        config.session_dir = _get_upseller_session_dir(user_id)
        db.session.commit()
    return config


# ----------------------------------------------------------------
# MARKETPLACE API DIRETA (Shopee - fase 1)
# ----------------------------------------------------------------

def _get_or_create_marketplace_api_config(user_id, marketplace: str = "shopee"):
    mp = (marketplace or "shopee").strip().lower()
    cfg = MarketplaceApiConfig.query.filter_by(user_id=user_id, marketplace=mp).first()
    if not cfg:
        cfg = MarketplaceApiConfig(
            user_id=user_id,
            marketplace=mp,
            api_base_url="https://openplatform.sandbox.test-stable.shopee.sg",
            status_conexao="nao_configurado",
            ativo=False,
        )
        db.session.add(cfg)
        db.session.commit()
    return cfg


def _set_marketplace_sidebar_cache(user_id, sidebar_info):
    if not hasattr(app, "_marketplace_sidebar_cache"):
        app._marketplace_sidebar_cache = {}
    app._marketplace_sidebar_cache[int(user_id)] = sidebar_info or {}


def _get_marketplace_sidebar_cache(user_id):
    if hasattr(app, "_marketplace_sidebar_cache"):
        return app._marketplace_sidebar_cache.get(int(user_id), {}) or {}
    return {}


def _persistir_lojas_marketplace_api(user_id, lojas):
    """
    Persiste snapshot de lojas da API direta em tabela separada.
    Nao interfere na tabela `lojas` usada pela automacao UpSeller.
    """
    agora = _agora_brasil()
    lojas = lojas or []

    existentes = MarketplaceLoja.query.filter_by(user_id=user_id, ativo=True).all()
    mapa = {}
    for l in existentes:
        key = f"{(l.marketplace or '').strip().casefold()}::{_norm_loja_nome(l.nome)}"
        mapa[key] = l

    keys_snapshot = set()
    for lj in lojas:
        nome = str(lj.get("nome", "") or "").strip()
        if not nome:
            continue
        marketplace = str(lj.get("marketplace", "Shopee") or "Shopee").strip() or "Shopee"
        key = f"{marketplace.casefold()}::{_norm_loja_nome(nome)}"
        keys_snapshot.add(key)

        try:
            pedidos = max(0, int(lj.get("pedidos", 0) or 0))
        except Exception:
            pedidos = 0
        try:
            notas = max(0, int(lj.get("notas_pendentes", 0) or 0))
        except Exception:
            notas = 0
        try:
            etiquetas = max(0, int(lj.get("etiquetas_pendentes", 0) or 0))
        except Exception:
            etiquetas = 0

        row = mapa.get(key)
        if row:
            row.nome = nome
            row.marketplace = marketplace
            row.pedidos_pendentes = pedidos
            row.notas_pendentes = notas
            row.etiquetas_pendentes = etiquetas
            row.ultima_atualizacao = agora
            row.ativo = True
        else:
            row = MarketplaceLoja(
                user_id=user_id,
                marketplace=marketplace,
                nome=nome,
                pedidos_pendentes=pedidos,
                notas_pendentes=notas,
                etiquetas_pendentes=etiquetas,
                ultima_atualizacao=agora,
                ativo=True,
            )
            db.session.add(row)
            mapa[key] = row

    # Zera as que nao vieram no snapshot atual, mantendo historico da lista.
    for key, row in mapa.items():
        if key not in keys_snapshot:
            row.pedidos_pendentes = 0
            row.notas_pendentes = 0
            row.etiquetas_pendentes = 0
            row.ultima_atualizacao = agora

    db.session.commit()


class _ShopeeOpenApiClient:
    """
    Cliente minimo Shopee Open API v2 para sincronizacao de contagens.
    Fase 1: leitura de loja e pedidos (READY_TO_SHIP).
    """

    def __init__(self, cfg: MarketplaceApiConfig):
        self.cfg = cfg
        self.base_url = (cfg.api_base_url or "https://openplatform.sandbox.test-stable.shopee.sg").strip().rstrip("/")
        self.partner_id = str(cfg.partner_id or "").strip()
        self.partner_key = str(cfg.get_partner_key() or "").strip()
        self.shop_id = str(cfg.shop_id or "").strip()
        self.access_token = str(cfg.get_access_token() or "").strip()
        self.refresh_token = str(cfg.get_refresh_token() or "").strip()

    def _timestamp(self):
        return int(time.time())

    def _sign(self, path: str, timestamp: int, access_token: str = "", shop_id: str = ""):
        base = f"{self.partner_id}{path}{timestamp}{access_token}{shop_id}"
        return hmac.new(
            self.partner_key.encode("utf-8"),
            base.encode("utf-8"),
            hashlib.sha256
        ).hexdigest()

    def _request(self, method: str, path: str, params=None, body=None, with_auth=True, timeout=30):
        ts = self._timestamp()
        params = dict(params or {})
        sign = self._sign(path, ts, self.access_token if with_auth else "", self.shop_id if with_auth else "")
        params.update({
            "partner_id": self.partner_id,
            "timestamp": ts,
            "sign": sign,
        })
        if with_auth:
            params["access_token"] = self.access_token
            params["shop_id"] = self.shop_id

        url = f"{self.base_url}{path}"
        if method.upper() == "GET":
            resp = requests.get(url, params=params, timeout=timeout)
        else:
            resp = requests.post(url, params=params, json=body or {}, timeout=timeout)

        try:
            data = resp.json()
        except Exception:
            data = {"error": f"http_{resp.status_code}", "message": resp.text[:500]}

        if resp.status_code >= 400:
            return {"ok": False, "http_status": resp.status_code, "data": data}
        err = str((data or {}).get("error") or "").strip()
        if err:
            return {"ok": False, "http_status": resp.status_code, "data": data}
        return {"ok": True, "http_status": resp.status_code, "data": data}

    def refresh_access_token(self):
        """
        Refresh de access token Shopee.
        """
        path = "/api/v2/auth/access_token/get"
        ts = self._timestamp()
        sign = self._sign(path, ts, "", "")
        params = {
            "partner_id": self.partner_id,
            "timestamp": ts,
            "sign": sign,
        }
        body = {
            "shop_id": int(self.shop_id) if str(self.shop_id).isdigit() else self.shop_id,
            "refresh_token": self.refresh_token,
            "partner_id": int(self.partner_id) if str(self.partner_id).isdigit() else self.partner_id,
        }
        url = f"{self.base_url}{path}"
        resp = requests.post(url, params=params, json=body, timeout=30)
        try:
            data = resp.json()
        except Exception:
            data = {"error": f"http_{resp.status_code}", "message": resp.text[:500]}
        if resp.status_code >= 400 or str((data or {}).get("error") or "").strip():
            return {"ok": False, "data": data, "http_status": resp.status_code}

        payload = (data or {}).get("response") or data or {}
        shop_from = payload.get("shop_id") or ""
        if not shop_from and isinstance(payload.get("shop_id_list"), list) and payload["shop_id_list"]:
            shop_from = payload["shop_id_list"][0]
        return {
            "ok": True,
            "access_token": str(payload.get("access_token") or "").strip(),
            "refresh_token": str(payload.get("refresh_token") or "").strip(),
            "expire_in": int(payload.get("expire_in") or 0),
            "shop_id": str(shop_from or self.shop_id).strip(),
            "data": data,
        }

    def get_shop_info(self):
        return self._request("GET", "/api/v2/shop/get_shop_info", with_auth=True)

    def get_order_count(self, order_status: str, days: int = 15, page_size: int = 100):
        """
        Conta pedidos por status usando paginacao do get_order_list.
        Obs: Shopee limita janela de tempo; usamos ultimos `days` dias.
        """
        now_ts = int(time.time())
        from_ts = now_ts - max(1, int(days or 15)) * 24 * 3600
        cursor = ""
        total = 0
        loops = 0

        while loops < 80:
            loops += 1
            params = {
                "time_range_field": "create_time",
                "time_from": from_ts,
                "time_to": now_ts,
                "page_size": min(max(int(page_size or 100), 1), 100),
                "cursor": cursor,
                "order_status": order_status,
            }
            ret = self._request("GET", "/api/v2/order/get_order_list", params=params, with_auth=True)
            if not ret.get("ok"):
                return {"ok": False, "erro": (ret.get("data") or {}).get("message") or (ret.get("data") or {}).get("error") or "falha_get_order_list"}

            resp = ((ret.get("data") or {}).get("response") or {})
            order_list = resp.get("order_list") or []
            total += len(order_list)
            has_next = bool(resp.get("more"))
            cursor = str(resp.get("next_cursor") or "").strip()
            if not has_next or not cursor:
                break

        return {"ok": True, "total": total}

    # ---- Product APIs (sandbox test) ----

    def get_item_list(self, offset=0, page_size=50, item_status="NORMAL"):
        params = {"offset": offset, "page_size": page_size, "item_status": item_status}
        return self._request("GET", "/api/v2/product/get_item_list", params=params, with_auth=True)

    def get_category_list(self, language="pt-BR"):
        """Retorna arvore de categorias da Shopee."""
        params = {"language": language}
        return self._request("GET", "/api/v2/product/get_category", params=params, with_auth=True)

    def upload_image_url(self, image_url="https://via.placeholder.com/600x600.png?text=TestProduct"):
        """Faz upload de imagem para Shopee a partir de URL."""
        body = {"image_url_list": [image_url]}
        return self._request("POST", "/api/v2/media_space/upload_image", body=body, with_auth=True)

    def add_item(self, item_name, category_id, description, price, stock, image_id_list=None, weight=0.5):
        """Cria um produto na loja Shopee sandbox."""
        body = {
            "original_price": float(price),
            "description": description,
            "item_name": item_name,
            "normal_stock": int(stock),
            "weight": float(weight),
            "category_id": int(category_id),
            "image": {"image_id_list": image_id_list or []},
            "logistic_info": [{"logistic_id": 0, "enabled": True}],
            "item_status": "NORMAL",
            "condition": "NEW",
        }
        return self._request("POST", "/api/v2/product/add_item", body=body, with_auth=True)

    def get_item_base_info(self, item_id_list):
        """Retorna info basica de itens."""
        params = {"item_id_list": ",".join(str(i) for i in item_id_list)}
        return self._request("GET", "/api/v2/product/get_item_base_info", params=params, with_auth=True)

    def get_logistics_channel(self):
        """Lista canais de logistica disponiveis."""
        return self._request("GET", "/api/v2/logistics/get_channel_list", with_auth=True)


def _marketplace_cfg_to_client(cfg: MarketplaceApiConfig):
    if not cfg:
        return None, "config_nao_encontrada"
    c = _ShopeeOpenApiClient(cfg)
    if not c.partner_id or not c.partner_key or not c.shop_id or not c.access_token:
        return None, "credenciais_incompletas"
    return c, ""


def _marketplace_shopee_sync_snapshot(user_id: int):
    """
    Faz sincronizacao Shopee API e retorna snapshot de lojas do modo API.
    """
    cfg = _get_or_create_marketplace_api_config(user_id, "shopee")
    cli, err = _marketplace_cfg_to_client(cfg)
    if not cli:
        return {"sucesso": False, "erro": err}

    # Tentativa de refresh preventivo quando token estiver perto do vencimento.
    try:
        if cfg.token_expires_at and cfg.token_expires_at <= (_agora_brasil() + timedelta(minutes=10)):
            ref = cli.refresh_access_token()
            if ref.get("ok"):
                cfg.set_access_token(ref.get("access_token", ""))
                if ref.get("refresh_token"):
                    cfg.set_refresh_token(ref.get("refresh_token", ""))
                cfg.token_expires_at = _agora_brasil() + timedelta(seconds=max(1, int(ref.get("expire_in") or 0)))
                db.session.commit()
                cli = _ShopeeOpenApiClient(cfg)
    except Exception:
        pass

    shop = cli.get_shop_info()
    if not shop.get("ok"):
        # Se erro de token, tenta refresh e repete uma vez.
        msg = str(((shop.get("data") or {}).get("message") or "")).lower()
        err_code = str(((shop.get("data") or {}).get("error") or "")).lower()
        if "token" in msg or "token" in err_code:
            ref = cli.refresh_access_token()
            if ref.get("ok"):
                cfg.set_access_token(ref.get("access_token", ""))
                if ref.get("refresh_token"):
                    cfg.set_refresh_token(ref.get("refresh_token", ""))
                cfg.token_expires_at = _agora_brasil() + timedelta(seconds=max(1, int(ref.get("expire_in") or 0)))
                db.session.commit()
                cli = _ShopeeOpenApiClient(cfg)
                shop = cli.get_shop_info()
        if not shop.get("ok"):
            return {
                "sucesso": False,
                "erro": (shop.get("data") or {}).get("message") or (shop.get("data") or {}).get("error") or "falha_shop_info"
            }

    shop_info = ((shop.get("data") or {}).get("response") or {})
    loja_nome = (cfg.loja_nome or "").strip() or str(shop_info.get("shop_name") or "").strip() or f"Shopee Shop {cfg.shop_id}"

    # Fase 1: usamos READY_TO_SHIP como base de pedidos pendentes.
    cnt_ready = cli.get_order_count("READY_TO_SHIP", days=15)
    if not cnt_ready.get("ok"):
        return {"sucesso": False, "erro": cnt_ready.get("erro", "falha_contagem_ready_to_ship")}
    pedidos = int(cnt_ready.get("total") or 0)

    # Shopee API nao expoe diretamente "nota fiscal pendente" no mesmo formato do UpSeller.
    notas = 0
    etiquetas = pedidos

    lojas = [{
        "nome": loja_nome,
        "marketplace": "Shopee",
        "pedidos": pedidos,
        "notas_pendentes": notas,
        "etiquetas_pendentes": etiquetas,
    }]
    sidebar = {
        "Para Enviar": pedidos,
        "Para Emitir": notas,
        "Para Imprimir": etiquetas,
    }

    _persistir_lojas_marketplace_api(user_id, lojas)
    _set_marketplace_sidebar_cache(user_id, sidebar)

    cfg.status_conexao = "ok"
    cfg.ultima_sincronizacao = _agora_brasil()
    cfg.ativo = True
    db.session.commit()

    return {
        "sucesso": True,
        "lojas": lojas,
        "total_pedidos": pedidos,
        "sidebar_info": sidebar,
        "detalhes": f"Shopee sincronizado: {pedidos} pedido(s) READY_TO_SHIP",
    }


# ----------------------------------------------------------------
# GERENCIADOR GLOBAL DE SCRAPERS UPSELLER (mantém instância viva)
# ----------------------------------------------------------------
# Problema: UpSeller usa cookies de sessão (não-persistentes) que são
# perdidos quando o Playwright fecha. Solução: manter o Playwright
# rodando em background com o navegador minimizado.

import asyncio as _asyncio_global
import threading as _threading_global

class _UpSellerManager:
    """Gerencia instâncias do UpSellerScraper por usuário, mantendo-as vivas."""

    def __init__(self):
        self._scrapers = {}      # user_id -> scraper instance
        self._loops = {}         # user_id -> asyncio event loop
        self._threads = {}       # user_id -> thread running the loop
        self._locks = {}         # user_id -> threading.Lock
        self._global_lock = _threading_global.Lock()

    def _get_lock(self, user_id):
        with self._global_lock:
            if user_id not in self._locks:
                self._locks[user_id] = _threading_global.RLock()
            return self._locks[user_id]

    def _ensure_loop(self, user_id):
        """Garante que existe um event loop rodando para o user_id."""
        if user_id in self._loops and self._loops[user_id].is_running():
            return self._loops[user_id]

        loop = _asyncio_global.new_event_loop()

        def _run_loop():
            _asyncio_global.set_event_loop(loop)
            loop.run_forever()

        t = _threading_global.Thread(target=_run_loop, daemon=True, name=f"upseller-loop-{user_id}")
        t.start()
        self._loops[user_id] = loop
        self._threads[user_id] = t
        return loop

    def _run_async(self, user_id, coro, timeout=300):
        """Executa coroutine no loop do user_id e retorna resultado."""
        loop = self._ensure_loop(user_id)
        future = _asyncio_global.run_coroutine_threadsafe(coro, loop)
        return future.result(timeout=timeout)

    def get_scraper(self, user_id):
        """Retorna scraper existente ou None."""
        return self._scrapers.get(user_id)

    def is_alive(self, user_id):
        """Verifica se o scraper do user_id está vivo e com página aberta."""
        scraper = self._scrapers.get(user_id)
        if not scraper or not scraper._page:
            return False
        try:
            # Testa se a página ainda responde
            self._run_async(user_id, scraper._page.evaluate("1+1"))
            return True
        except Exception:
            return False

    def is_logged_in(self, user_id):
        """Verifica se o scraper está vivo E logado no UpSeller."""
        scraper = self._scrapers.get(user_id)
        if not scraper or not scraper._page:
            return False
        try:
            return self._run_async(user_id, scraper._esta_logado())
        except Exception:
            return False

    def criar_scraper(self, user_id, config_session_dir, download_dir, headless=False):
        """Cria novo scraper, fechando o anterior se existir."""
        from upseller_scraper import UpSellerScraper

        lock = self._get_lock(user_id)
        with lock:
            # Fechar anterior se existir
            self._fechar_interno(user_id)

            scraper = UpSellerScraper({
                "email": "",
                "password": "",
                "profile_dir": config_session_dir,
                "headless": headless,
                "download_dir": download_dir,
            })
            self._scrapers[user_id] = scraper

            # Iniciar navegador
            self._run_async(user_id, scraper._iniciar_navegador())
            return scraper

    def login_manual(self, user_id, timeout=180):
        """Executa login_manual no scraper do user_id."""
        scraper = self._scrapers.get(user_id)
        if not scraper:
            raise RuntimeError("Scraper não inicializado. Chame criar_scraper primeiro.")
        return self._run_async(user_id, scraper.login_manual(timeout_seconds=timeout))

    def listar_lojas(self, user_id):
        """Executa listar_lojas_pendentes no scraper do user_id."""
        scraper = self._scrapers.get(user_id)
        if not scraper:
            raise RuntimeError("Scraper não inicializado")
        return self._run_async(user_id, scraper.listar_lojas_pendentes())

    def contar_loja(self, user_id, nome_loja, pedidos_fallback=0, marketplace_fallback=""):
        """Executa contagem precisa de uma loja no scraper do user_id."""
        scraper = self._scrapers.get(user_id)
        if not scraper:
            raise RuntimeError("Scraper nao inicializado")
        return self._run_async(
            user_id,
            scraper.contar_pedidos_loja(
                nome_loja=nome_loja,
                pedidos_fallback=pedidos_fallback,
                marketplace_fallback=marketplace_fallback,
            ),
        )

    def esta_logado(self, user_id):
        """Verifica login no scraper do user_id."""
        scraper = self._scrapers.get(user_id)
        if not scraper:
            return False
        try:
            return self._run_async(user_id, scraper._esta_logado())
        except Exception:
            return False

    def reconectar(self, user_id, config_session_dir, download_dir):
        """
        Tenta reconectar automaticamente:
        1. Se scraper está vivo e logado → retorna True
        2. Se scraper está vivo mas deslogado → tenta navegar de volta
        3. Se scraper morreu → cria novo com headless=False para login manual
        """
        lock = self._get_lock(user_id)
        with lock:
            # 1. Scraper vivo e logado?
            if self.is_alive(user_id):
                try:
                    logado = self._run_async(user_id, self._scrapers[user_id]._esta_logado())
                    if logado:
                        print(f"[UpSellerManager] User {user_id}: já está logado!")
                        return True
                except Exception:
                    pass

            # 2. Criar novo scraper headless e testar sessão persistente
            try:
                scraper = self.criar_scraper(user_id, config_session_dir, download_dir, headless=True)
                logado = self._run_async(user_id, scraper._esta_logado())
                if logado:
                    print(f"[UpSellerManager] User {user_id}: sessão persistente reutilizada (headless)!")
                    return True
            except Exception as e:
                print(f"[UpSellerManager] Erro tentando headless: {e}")

            # 3. Sessão expirou → abrir visível para login manual
            try:
                self._fechar_interno(user_id)
                scraper = self.criar_scraper(user_id, config_session_dir, download_dir, headless=False)
                logado = self._run_async(user_id, scraper.login_manual(timeout_seconds=180))
                if logado:
                    # Minimizar janela após login bem-sucedido
                    try:
                        self._run_async(user_id, scraper._page.evaluate("""
                            (() => { try { window.resizeTo(1,1); window.moveTo(-2000,-2000); } catch(e){} })()
                        """))
                    except Exception:
                        pass
                    print(f"[UpSellerManager] User {user_id}: login manual concluído, scraper mantido vivo!")
                    return True
                else:
                    return False
            except Exception as e:
                print(f"[UpSellerManager] Erro no login manual: {e}")
                return False

    def _fechar_interno(self, user_id):
        """Fecha scraper e loop sem lock (chamado internamente)."""
        scraper = self._scrapers.pop(user_id, None)
        loop = self._loops.get(user_id)
        if scraper and loop and loop.is_running():
            try:
                future = _asyncio_global.run_coroutine_threadsafe(scraper.fechar(), loop)
                future.result(timeout=10)
            except Exception:
                pass
        # Parar o loop
        if loop and loop.is_running():
            loop.call_soon_threadsafe(loop.stop)
        self._loops.pop(user_id, None)
        self._threads.pop(user_id, None)

    def fechar(self, user_id):
        """Fecha scraper do user_id (com lock)."""
        lock = self._get_lock(user_id)
        with lock:
            self._fechar_interno(user_id)

    def fechar_todos(self):
        """Fecha todos os scrapers (para shutdown)."""
        for uid in list(self._scrapers.keys()):
            self.fechar(uid)


# Instância global do gerenciador
_upseller_mgr = _UpSellerManager()


def _norm_loja_nome(nome):
    txt = (nome or "").strip().casefold()
    txt = ''.join(ch for ch in unicodedata.normalize('NFD', txt) if unicodedata.category(ch) != 'Mn')
    txt = _re.sub(r"\s+", " ", txt).strip()
    return txt


def _sanitizar_lojas_selecionadas(user_id, lojas_raw):
    """
    Normaliza e valida lista de lojas recebida do frontend.
    Retorna (lojas_validas, lojas_ignoradas) preservando a ordem do usuario.
    """
    lojas_raw = lojas_raw if isinstance(lojas_raw, list) else []
    if not lojas_raw:
        return [], []

    lojas_db = Loja.query.filter_by(user_id=user_id, ativo=True).all()
    mapa_por_key = {}
    for l in lojas_db:
        nome = (l.nome or "").strip()
        key = _norm_loja_nome(nome)
        if key and key not in mapa_por_key:
            mapa_por_key[key] = nome

    validas = []
    ignoradas = []
    vistos = set()
    for item in lojas_raw:
        nome_in = str(item or "").strip()
        if not nome_in:
            continue
        key = _norm_loja_nome(nome_in)
        nome_canon = mapa_por_key.get(key)
        if not nome_canon:
            ignoradas.append(nome_in)
            continue
        if key in vistos:
            continue
        vistos.add(key)
        validas.append(nome_canon)

    return validas, ignoradas


def _agrupar_lojas_por_marketplace(user_id: int, lojas: list) -> list:
    """
    Agrupa lojas por marketplace preservando a ordem de selecao do usuario.
    Retorna lista [{marketplace: str, lojas: [nomes...]}].
    """
    lojas = lojas if isinstance(lojas, list) else []
    if not lojas:
        return []

    lojas_db = Loja.query.filter_by(user_id=user_id, ativo=True).all()
    mapa_db = {_norm_loja_nome(l.nome): l for l in lojas_db if l.nome}

    grupos = {}
    ordem_marketplaces = []
    for nome in lojas:
        nome_txt = str(nome or "").strip()
        if not nome_txt:
            continue
        key = _norm_loja_nome(nome_txt)
        loja_db = mapa_db.get(key)
        nome_canon = (loja_db.nome if loja_db and loja_db.nome else nome_txt).strip()
        marketplace = (loja_db.marketplace if loja_db else "").strip() or "Shopee"
        if marketplace not in grupos:
            grupos[marketplace] = []
            ordem_marketplaces.append(marketplace)
        if nome_canon not in grupos[marketplace]:
            grupos[marketplace].append(nome_canon)

    return [{"marketplace": mp, "lojas": grupos.get(mp, [])} for mp in ordem_marketplaces if grupos.get(mp)]


def _set_sidebar_cache(user_id, sidebar_info):
    if not hasattr(app, '_upseller_sidebar_cache'):
        app._upseller_sidebar_cache = {}
    app._upseller_sidebar_cache[user_id] = sidebar_info or {}


def _get_sidebar_cache(user_id):
    if hasattr(app, '_upseller_sidebar_cache'):
        return app._upseller_sidebar_cache.get(user_id, {}) or {}
    return {}


def _criar_pasta_lote_upseller(pasta_entrada: str, prefixo: str = "lote") -> str:
    """
    Cria pasta de lote isolada para cada execucao do UpSeller.
    Evita que uma geracao reutilize arquivos de execucoes anteriores.
    """
    lotes_base = os.path.join(pasta_entrada, "_upseller_lotes")
    os.makedirs(lotes_base, exist_ok=True)
    nome_lote = f"{prefixo}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
    pasta_lote = os.path.join(lotes_base, nome_lote)
    os.makedirs(pasta_lote, exist_ok=True)
    return pasta_lote


def _persistir_lojas_upseller(user_id, lojas):
    """
    Upsert da lista de lojas no banco:
    - mantem todas as lojas ativas visiveis (inclusive com 0 pedidos)
    - atualiza marketplace, pedidos e timestamp
    - lojas nao retornadas no snapshot atual ficam com 0 pedidos
    """
    agora = _agora_brasil()
    lojas = lojas or []

    existentes = Loja.query.filter_by(user_id=user_id).all()
    mapa_existentes = {_norm_loja_nome(l.nome): l for l in existentes if l.nome}

    nomes_snapshot = set()
    for lj in lojas:
        nome = (lj.get("nome") or "").strip()
        if not nome:
            continue
        key = _norm_loja_nome(nome)
        pedidos = lj.get("pedidos", 0)
        try:
            pedidos = max(0, int(pedidos or 0))
        except Exception:
            pedidos = 0
        marketplace = (lj.get("marketplace") or "Shopee").strip() or "Shopee"

        notas = lj.get("notas_pendentes", 0)
        etiquetas = lj.get("etiquetas_pendentes", 0)
        try:
            notas = max(0, int(notas or 0))
        except Exception:
            notas = 0
        try:
            etiquetas = max(0, int(etiquetas or 0))
        except Exception:
            etiquetas = 0

        nomes_snapshot.add(key)
        loja_db = mapa_existentes.get(key)
        if loja_db:
            loja_db.nome = nome  # atualiza capitalizacao/nome exibido
            loja_db.marketplace = marketplace
            loja_db.pedidos_pendentes = pedidos
            loja_db.notas_pendentes = notas
            loja_db.etiquetas_pendentes = etiquetas
            loja_db.ultima_atualizacao = agora
            loja_db.ativo = True
        else:
            nova = Loja(
                user_id=user_id,
                nome=nome,
                marketplace=marketplace,
                pedidos_pendentes=pedidos,
                notas_pendentes=notas,
                etiquetas_pendentes=etiquetas,
                ultima_atualizacao=agora,
                ativo=True,
            )
            db.session.add(nova)
            mapa_existentes[key] = nova

    # Lojas fora do snapshot atual permanecem salvas, com 0 pedido.
    for key, loja_db in mapa_existentes.items():
        if loja_db.ativo and key not in nomes_snapshot:
            loja_db.pedidos_pendentes = 0
            loja_db.notas_pendentes = 0
            loja_db.etiquetas_pendentes = 0
            loja_db.ultima_atualizacao = agora

    db.session.commit()


def _contagens_lojas_suspeitas(lojas):
    """
    Heuristica para detectar snapshot claramente anomalo:
    muitas lojas com o MESMO valor alto de pedidos.
    """
    try:
        lojas = lojas or []
        if len(lojas) < 4:
            return False
        vals = []
        for lj in lojas:
            try:
                v = int((lj.get("pedidos", 0) if isinstance(lj, dict) else getattr(lj, "pedidos_pendentes", 0)) or 0)
            except Exception:
                v = 0
            if v > 0:
                vals.append(v)
        if len(vals) < 4:
            return False
        freq = {}
        for v in vals:
            freq[v] = freq.get(v, 0) + 1
        val_top, qtd_top = max(freq.items(), key=lambda x: x[1])
        ratio = qtd_top / max(len(vals), 1)
        # Ex.: 393 repetido em quase todas as lojas.
        return val_top >= 50 and ratio >= 0.75
    except Exception:
        return False


def _atualizar_lojas_apos_acao(user_id):
    """
    Rele snapshot atual de lojas no UpSeller e atualiza o banco.
    Usado apos programar/emitir/imprimir/gerar para manter contagens corretas.
    """
    try:
        user = db.session.get(User, user_id)
        config = _get_or_create_upseller_config(user_id)
        download_dir = user.get_pasta_entrada() if user else os.path.join(os.path.expanduser("~"), ".upseller_downloads")

        scraper_vivo = _upseller_mgr.is_alive(user_id)
        scraper_atual = _upseller_mgr.get_scraper(user_id) if scraper_vivo else None
        headless_atual = getattr(scraper_atual, 'headless', True) if scraper_atual else True

        # Para leitura confiavel do SPA (filtro loja + contadores), usar browser visivel.
        if (not scraper_vivo) or headless_atual:
            try:
                _upseller_mgr.criar_scraper(user_id, config.session_dir, download_dir, headless=False)
            except Exception as e:
                return {"sucesso": False, "erro": f"scraper_inativo: {e}"}

        if not _upseller_mgr.esta_logado(user_id):
            return {"sucesso": False, "erro": "sessao_expirada"}

        resultado = _upseller_mgr.listar_lojas(user_id)
        if not resultado or not resultado.get("sucesso"):
            return {"sucesso": False, "erro": (resultado or {}).get("erro", "falha_listar_lojas")}
        lojas_snap = resultado.get("lojas", []) or []
        if _contagens_lojas_suspeitas(lojas_snap):
            return {
                "sucesso": False,
                "erro": "Contagens suspeitas detectadas no UpSeller. Reconecte e tente sincronizar novamente.",
            }
        _persistir_lojas_upseller(user_id, lojas_snap)
        _set_sidebar_cache(user_id, resultado.get("sidebar_info", {}))
        return resultado
    except Exception as e:
        print(f"[Lojas] Falha ao atualizar snapshot apos acao: {e}")
        return {"sucesso": False, "erro": str(e)}


def _set_atualizacao_lojas_em_andamento(user_id, valor: bool):
    if not hasattr(app, "_lojas_atualizando"):
        app._lojas_atualizando = {}
    if not hasattr(app, "_lojas_atualizando_ts"):
        app._lojas_atualizando_ts = {}
    uid = int(user_id)
    ativo = bool(valor)
    app._lojas_atualizando[uid] = ativo
    if ativo:
        app._lojas_atualizando_ts[uid] = _agora_brasil()
    else:
        app._lojas_atualizando_ts.pop(uid, None)


def _esta_atualizando_lojas(user_id) -> bool:
    if not hasattr(app, "_lojas_atualizando"):
        return False
    uid = int(user_id)
    ativo = bool(app._lojas_atualizando.get(uid, False))
    if not ativo:
        return False
    # Blindagem: se a thread travar, libera automaticamente apos TTL.
    ttl_segundos = 900  # 15 min
    ts = None
    if hasattr(app, "_lojas_atualizando_ts"):
        ts = app._lojas_atualizando_ts.get(uid)
    if isinstance(ts, datetime):
        if (_agora_brasil() - ts).total_seconds() > ttl_segundos:
            app._lojas_atualizando[uid] = False
            if hasattr(app, "_lojas_atualizando_ts"):
                app._lojas_atualizando_ts.pop(uid, None)
            return False
    return True


def _obter_acao_massa_em_andamento(user_id):
    if not hasattr(app, "_acoes_massa"):
        app._acoes_massa = {}
    return app._acoes_massa.get(int(user_id))


def _iniciar_acao_massa(user_id, acao: str):
    """
    Trava de exclusao mutua para processos em massa.
    Garante 1 processo por usuario por vez (emitir/programar/imprimir/gerar).
    """
    if not hasattr(app, "_acoes_massa"):
        app._acoes_massa = {}
    uid = int(user_id)
    atual = app._acoes_massa.get(uid)
    if atual:
        return False, atual
    novo = {
        "acao": str(acao or "").strip() or "processo",
        "inicio": _agora_brasil().isoformat(),
    }
    app._acoes_massa[uid] = novo
    return True, novo


def _finalizar_acao_massa(user_id, acao: str = ""):
    if not hasattr(app, "_acoes_massa"):
        return
    uid = int(user_id)
    atual = app._acoes_massa.get(uid)
    if not atual:
        return
    if acao and atual.get("acao") != acao:
        return
    app._acoes_massa.pop(uid, None)


def _disparar_atualizacao_lojas_background(user_id, status_attr: str = ""):
    """
    Atualiza snapshot de lojas em background sem bloquear retorno da acao principal.
    """
    user_id = int(user_id)
    if _esta_atualizando_lojas(user_id):
        return False

    _set_atualizacao_lojas_em_andamento(user_id, True)

    if status_attr and hasattr(app, status_attr):
        status_map = getattr(app, status_attr)
        st = status_map.get(user_id)
        if isinstance(st, dict):
            st["atualizando_pedidos"] = True
            st["aviso_atualizacao"] = "Atualizando pedidos em segundo plano..."

    def _runner():
        with app.app_context():
            try:
                ret = _atualizar_lojas_apos_acao(user_id)
                if status_attr and hasattr(app, status_attr):
                    status_map = getattr(app, status_attr)
                    st = status_map.get(user_id)
                    if isinstance(st, dict):
                        st["atualizacao_lojas_ok"] = bool((ret or {}).get("sucesso"))
                        if not bool((ret or {}).get("sucesso")):
                            st["aviso_atualizacao"] = (ret or {}).get("erro", "Falha ao atualizar pedidos em segundo plano")
            except Exception as e:
                if status_attr and hasattr(app, status_attr):
                    status_map = getattr(app, status_attr)
                    st = status_map.get(user_id)
                    if isinstance(st, dict):
                        st["atualizacao_lojas_ok"] = False
                        st["aviso_atualizacao"] = str(e)
            finally:
                _set_atualizacao_lojas_em_andamento(user_id, False)
                if status_attr and hasattr(app, status_attr):
                    status_map = getattr(app, status_attr)
                    st = status_map.get(user_id)
                    if isinstance(st, dict):
                        st["atualizando_pedidos"] = False

    threading.Thread(target=_runner, daemon=True).start()
    return True


def _atualizar_loja_individual(user_id, nome_loja):
    """
    Atualiza somente UMA loja no banco via filtro dedicado do UpSeller.
    Mantem as demais lojas intactas.
    """
    nome = (nome_loja or "").strip()
    if not nome:
        return {"sucesso": False, "erro": "loja_obrigatoria"}

    try:
        user = db.session.get(User, user_id)
        config = _get_or_create_upseller_config(user_id)
        download_dir = user.get_pasta_entrada() if user else os.path.join(os.path.expanduser("~"), ".upseller_downloads")

        scraper_vivo = _upseller_mgr.is_alive(user_id)
        scraper_atual = _upseller_mgr.get_scraper(user_id) if scraper_vivo else None
        headless_atual = getattr(scraper_atual, 'headless', True) if scraper_atual else True

        # Para leitura confiavel do SPA usar browser visivel.
        if (not scraper_vivo) or headless_atual:
            try:
                _upseller_mgr.criar_scraper(user_id, config.session_dir, download_dir, headless=False)
            except Exception as e:
                return {"sucesso": False, "erro": f"scraper_inativo: {e}"}

        if not _upseller_mgr.esta_logado(user_id):
            return {"sucesso": False, "erro": "sessao_expirada"}

        loja_db = Loja.query.filter_by(user_id=user_id, nome=nome).first()
        if not loja_db:
            # Busca case-insensitive para evitar duplicidade por capitalizacao.
            loja_db = next(
                (l for l in Loja.query.filter_by(user_id=user_id, ativo=True).all()
                 if _norm_loja_nome(l.nome) == _norm_loja_nome(nome)),
                None
            )

        pedidos_fb = int((loja_db.pedidos_pendentes if loja_db else 0) or 0)
        marketplace_fb = (loja_db.marketplace if loja_db else "Shopee") or "Shopee"

        resultado = _upseller_mgr.contar_loja(
            user_id,
            nome_loja=nome,
            pedidos_fallback=pedidos_fb,
            marketplace_fallback=marketplace_fb,
        )
        if not resultado or not resultado.get("sucesso"):
            return {"sucesso": False, "erro": (resultado or {}).get("erro", "falha_contagem_loja")}

        item = (resultado.get("loja") or {})
        nome_final = (item.get("nome") or nome).strip() or nome
        marketplace = (item.get("marketplace") or marketplace_fb).strip() or "Shopee"
        try:
            pedidos = max(0, int(item.get("pedidos", pedidos_fb) or 0))
        except Exception:
            pedidos = max(0, int(pedidos_fb or 0))

        agora = _agora_brasil()
        if loja_db:
            loja_db.nome = nome_final
            loja_db.marketplace = marketplace
            loja_db.pedidos_pendentes = pedidos
            loja_db.ultima_atualizacao = agora
            loja_db.ativo = True
        else:
            loja_db = Loja(
                user_id=user_id,
                nome=nome_final,
                marketplace=marketplace,
                pedidos_pendentes=pedidos,
                ultima_atualizacao=agora,
                ativo=True,
            )
            db.session.add(loja_db)

        db.session.commit()

        lojas = Loja.query.filter_by(user_id=user_id, ativo=True).all()
        lojas.sort(key=lambda l: (-(l.pedidos_pendentes or 0), (l.nome or "").casefold()))
        total = sum(int(l.pedidos_pendentes or 0) for l in lojas)

        return {
            "sucesso": True,
            "loja": loja_db.to_dict(),
            "lojas": [l.to_dict() for l in lojas],
            "total_pedidos": total,
            "sidebar_info": _get_sidebar_cache(user_id),
            "fonte": item.get("_src", ""),
        }
    except Exception as e:
        print(f"[Lojas] Falha ao atualizar loja individual '{nome}': {e}")
        return {"sucesso": False, "erro": str(e)}


def _atualizar_todas_lojas_preciso(user_id):
    """
    Atualiza contagem de TODAS as lojas lendo diretamente do UpSeller.
    Usa listar_lojas() que le sidebar + tabela (rapido, sem navegar pagina por pagina).
    """
    try:
        user = db.session.get(User, user_id)
        config = _get_or_create_upseller_config(user_id)
        download_dir = user.get_pasta_entrada() if user else os.path.join(os.path.expanduser("~"), ".upseller_downloads")

        scraper_vivo = _upseller_mgr.is_alive(user_id)
        scraper_atual = _upseller_mgr.get_scraper(user_id) if scraper_vivo else None
        headless_atual = getattr(scraper_atual, 'headless', True) if scraper_atual else True

        if (not scraper_vivo) or headless_atual:
            try:
                _upseller_mgr.criar_scraper(user_id, config.session_dir, download_dir, headless=False)
            except Exception as e:
                return {"sucesso": False, "erro": f"scraper_inativo: {e}"}

        if not _upseller_mgr.esta_logado(user_id):
            return {"sucesso": False, "erro": "sessao_expirada"}

        # Leitura rapida: sidebar + tabela (sem paginacao loja-a-loja)
        snap = _upseller_mgr.listar_lojas(user_id)
        if not snap or not snap.get("sucesso"):
            return {"sucesso": False, "erro": (snap or {}).get("erro", "falha_snapshot_base")}

        lojas_snapshot = snap.get("lojas", []) or []
        _set_sidebar_cache(user_id, snap.get("sidebar_info", {}) or {})

        # Mesclar com lojas do banco para nao "sumirem"
        lojas_db = Loja.query.filter_by(user_id=user_id, ativo=True).all()
        snap_keys = set()
        for lj in lojas_snapshot:
            key = _norm_loja_nome((lj.get("nome") or "").strip())
            if key:
                snap_keys.add(key)
        # Incluir lojas do banco que nao apareceram no snapshot
        for ldb in lojas_db:
            key = _norm_loja_nome((ldb.nome or "").strip())
            if key and key not in snap_keys:
                lojas_snapshot.append({
                    "nome": ldb.nome,
                    "marketplace": ldb.marketplace or "Shopee",
                    "pedidos": 0,
                    "notas_pendentes": 0,
                    "etiquetas_pendentes": 0,
                })

        # Persistir todas as lojas (lista vazia se nao encontrou nada)
        _persistir_lojas_upseller(user_id, lojas_snapshot)

        lojas = Loja.query.filter_by(user_id=user_id, ativo=True).all()
        lojas.sort(key=lambda l: (-(l.pedidos_pendentes or 0), (l.nome or "").casefold()))
        total = sum(int(l.pedidos_pendentes or 0) for l in lojas)
        return {
            "sucesso": True,
            "lojas": [l.to_dict() for l in lojas],
            "total_pedidos": total,
            "sidebar_info": _get_sidebar_cache(user_id),
        }
    except Exception as e:
        err_txt = str(e).strip() or e.__class__.__name__
        if err_txt == "TimeoutError" or e.__class__.__name__.lower() == "timeouterror":
            err_txt = (
                "Timeout na atualizacao de lojas. "
                "Tente novamente com menos abas abertas no UpSeller."
            )
        print(f"[Lojas] Falha na atualizacao precisa de todas as lojas user {user_id}: {err_txt}")
        return {"sucesso": False, "erro": err_txt}


@app.route('/api/upseller/status', methods=['GET'])
@jwt_required()
def api_upseller_status():
    """Retorna status da conexao UpSeller."""
    user_id = int(get_jwt_identity())
    config = UpSellerConfig.query.filter_by(user_id=user_id).first()
    if config and config.status_conexao == "ok":
        return jsonify({
            "status": "ok",
            "ultima_sincronizacao": config.ultima_sincronizacao.strftime("%d/%m/%Y %H:%M") if config.ultima_sincronizacao else ""
        })
    return jsonify({"status": "desconectado", "ultima_sincronizacao": ""})


@app.route('/api/marketplace/status', methods=['GET'])
@jwt_required()
def api_marketplace_status():
    """Status da conexao API direta (fase 1: Shopee)."""
    user_id = int(get_jwt_identity())
    cfg = _get_or_create_marketplace_api_config(user_id, "shopee")
    data = cfg.to_dict()
    status = "ok" if data.get("configurado") and cfg.status_conexao == "ok" else "desconectado"
    return jsonify({
        "status": status,
        "marketplace": "shopee",
        "configurado": bool(data.get("configurado")),
        "loja_nome": data.get("loja_nome", ""),
        "shop_id": data.get("shop_id", ""),
        "ultima_sincronizacao": data.get("ultima_sincronizacao", ""),
        "status_conexao": data.get("status_conexao", "nao_configurado"),
        "api_base_url": data.get("api_base_url", ""),
        "has_partner_key": bool(data.get("has_partner_key")),
        "has_access_token": bool(data.get("has_access_token")),
        "has_refresh_token": bool(data.get("has_refresh_token")),
        "redirect_url": _get_shopee_redirect_url(),
        "redirect_domain": _get_shopee_redirect_domain(),
    })


@app.route('/api/marketplace/config', methods=['GET'])
@jwt_required()
def api_marketplace_config_get():
    """Retorna configuracao da API direta (Shopee)."""
    user_id = int(get_jwt_identity())
    cfg = _get_or_create_marketplace_api_config(user_id, "shopee")
    data = cfg.to_dict()
    data["redirect_url"] = _get_shopee_redirect_url()
    data["redirect_domain"] = _get_shopee_redirect_domain()
    return jsonify(data)


@app.route('/api/marketplace/shopee/redirect-info', methods=['GET'])
@jwt_required()
def api_marketplace_shopee_redirect_info():
    """Info de redirect URL/domain para configurar no painel Shopee."""
    return jsonify({
        "redirect_url": _get_shopee_redirect_url(),
        "redirect_domain": _get_shopee_redirect_domain(),
        "base_publica": _detectar_base_publica(),
    })


@app.route('/api/marketplace/shopee/debug-sign', methods=['POST'])
@jwt_required()
def api_marketplace_shopee_debug_sign():
    """
    Debug de assinatura Shopee v2 para comparacao 1:1.
    Retorna base_string, timestamp, sign e resposta bruta do endpoint auth_partner.
    """
    user_id = int(get_jwt_identity())
    cfg = _get_or_create_marketplace_api_config(user_id, "shopee")
    data = request.get_json(force=True, silent=True) or {}

    partner_id = str(cfg.partner_id or "").strip()
    partner_key = str(cfg.get_partner_key() or "").strip()
    if not partner_id:
        return jsonify({"status": "erro", "erro": "Partner ID nao configurado."}), 400
    if not partner_key:
        return jsonify({"status": "erro", "erro": "Partner Key nao configurada."}), 400

    ts_in = data.get("timestamp")
    try:
        ts = int(ts_in) if ts_in is not None and str(ts_in).strip() else int(time.time())
    except Exception:
        ts = int(time.time())

    path = "/api/v2/shop/auth_partner"
    base_string = f"{partner_id}{path}{ts}"
    sign = hmac.new(
        partner_key.encode("utf-8"),
        base_string.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()

    redirect_url = _get_shopee_redirect_url()
    state = _build_shopee_oauth_state(user_id)
    auth_base = (cfg.api_base_url or "https://openplatform.sandbox.test-stable.shopee.sg").strip().rstrip("/")
    query = {
        "partner_id": partner_id,
        "timestamp": ts,
        "sign": sign,
        "redirect": redirect_url,
        "state": state,
    }
    auth_url_preview = f"{auth_base}{path}?{urlencode(query)}"

    test_result = {
        "ok": False,
        "http_status": 0,
        "error": "",
        "message": "",
        "request_id": "",
        "location": "",
        "body_excerpt": "",
    }
    try:
        resp = requests.get(f"{auth_base}{path}", params=query, timeout=20, allow_redirects=False)
        test_result["http_status"] = int(resp.status_code)
        test_result["location"] = str(resp.headers.get("Location") or "")
        ctype = str(resp.headers.get("content-type") or "").lower()
        if "json" in ctype:
            try:
                j = resp.json() or {}
            except Exception:
                j = {}
            test_result["error"] = str(j.get("error") or "").strip()
            test_result["message"] = str(j.get("message") or "").strip()
            test_result["request_id"] = str(j.get("request_id") or "").strip()
            test_result["ok"] = not test_result["error"]
            test_result["body_excerpt"] = str(resp.text or "")[:500]
        else:
            txt = str(resp.text or "")
            test_result["body_excerpt"] = txt[:500]
            test_result["ok"] = resp.status_code in (200, 301, 302, 303, 307, 308)
    except Exception as e:
        test_result["error"] = "network_error"
        test_result["message"] = str(e)

    # Diagnostico detalhado se error_sign
    diagnostico = {}
    if test_result.get("error") == "error_sign":
        key_txt = partner_key
        diagnostico = {
            "causa": "Partner Key NAO corresponde ao Partner ID neste ambiente.",
            "key_format": f"{'v2 (shpk...)' if key_txt.startswith('shpk') else 'desconhecido'}, {len(key_txt)} chars",
            "ambiente": "Test/Sandbox" if "test-stable" in auth_base else "Producao",
            "solucao": (
                "Acesse open.shopee.com > App Management > seu app > "
                "copie App ID e App Key do MESMO app e ambiente."
            ),
        }

    return jsonify({
        "status": "ok",
        "debug": {
            "partner_id": partner_id,
            "api_base_url": auth_base,
            "path": path,
            "timestamp": ts,
            "base_string": base_string,
            "sign": sign,
            "redirect_url": redirect_url,
            "state": state,
            "auth_url_preview": auth_url_preview,
            "test_result": test_result,
            "diagnostico": diagnostico,
        }
    })


def _shopee_exchange_code_for_tokens_fast(
    partner_id: str, partner_key: str, base_url: str,
    code: str, shop_id: str,
):
    """
    Troca authorization code por access/refresh token na Shopee Open API.
    Versao otimizada: recebe credenciais ja prontas (sem DB/Fernet no caminho critico).
    O code da Shopee sandbox expira em ~30s, entao esta funcao deve ser chamada
    o mais rapido possivel apos receber o callback.
    """
    t0 = time.time()
    code_txt = str(code or "").strip()
    shop_txt = str(shop_id or "").strip()
    base_url = (base_url or "https://openplatform.sandbox.test-stable.shopee.sg").strip().rstrip("/")

    if not partner_id:
        return {"ok": False, "erro": "Partner ID nao configurado."}
    if not partner_key:
        return {"ok": False, "erro": "Partner Key nao configurada."}
    if not code_txt:
        return {"ok": False, "erro": "Authorization code ausente no callback."}
    if not shop_txt:
        return {"ok": False, "erro": "shop_id ausente no callback."}

    path = "/api/v2/auth/token/get"
    ts = int(time.time())
    base_string = f"{partner_id}{path}{ts}"
    sign = hmac.new(
        partner_key.encode("utf-8"),
        base_string.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()

    pid_int = int(partner_id)
    sid_int = int(shop_txt) if shop_txt.isdigit() else shop_txt
    params = {
        "partner_id": pid_int,
        "timestamp": ts,
        "sign": sign,
    }
    body = {
        "code": code_txt,
        "shop_id": sid_int,
        "partner_id": pid_int,
    }

    url = f"{base_url}{path}"
    t_pre = time.time()
    print(f"[Shopee token/get] POST {url} pid={pid_int} sid={sid_int} code={code_txt[:12]}... ts={ts} prep={int((t_pre-t0)*1000)}ms", flush=True, file=sys.stderr)
    try:
        resp = requests.post(url, params=params, json=body, timeout=30)
    except Exception as e:
        t_err = time.time()
        print(f"[Shopee token/get] REDE FALHOU em {int((t_err-t0)*1000)}ms: {e}", flush=True, file=sys.stderr)
        return {"ok": False, "erro": f"Falha de rede na troca de token: {e}"}

    t_resp = time.time()
    try:
        data = resp.json()
    except Exception:
        data = {"error": f"http_{resp.status_code}", "message": (resp.text or "")[:400]}

    print(f"[Shopee token/get] Status={resp.status_code} error={data.get('error','')} msg={data.get('message','')[:80]} total={int((t_resp-t0)*1000)}ms", flush=True, file=sys.stderr)
    print(f"[Shopee token/get] FULL RESPONSE: {json.dumps(data, default=str)[:1000]}", flush=True, file=sys.stderr)

    if resp.status_code >= 400:
        msg = (data or {}).get("message") or (data or {}).get("error") or f"http_{resp.status_code}"
        return {"ok": False, "erro": f"Falha Shopee token/get: {msg}", "data": data}

    err = str((data or {}).get("error") or "").strip()
    if err:
        msg = (data or {}).get("message") or err
        return {"ok": False, "erro": f"Erro Shopee token/get: {msg}", "data": data}

    # Shopee sandbox retorna tokens no nivel raiz; producao usa "response" nested.
    payload = (data or {}).get("response") or data or {}
    access_token = str(payload.get("access_token") or "").strip()
    refresh_token = str(payload.get("refresh_token") or "").strip()
    # shop_id pode estar em shop_id ou shop_id_list[0]
    shop_from_payload = payload.get("shop_id") or ""
    if not shop_from_payload and isinstance(payload.get("shop_id_list"), list) and payload["shop_id_list"]:
        shop_from_payload = payload["shop_id_list"][0]
    shop_final = str(shop_from_payload or shop_txt).strip()
    expire_in = int(payload.get("expire_in") or 0)

    if not access_token:
        return {"ok": False, "erro": "Shopee nao retornou access_token.", "data": data}

    return {
        "ok": True,
        "access_token": access_token,
        "refresh_token": refresh_token,
        "shop_id": shop_final,
        "expire_in": expire_in,
        "data": data,
    }


def _diagnosticar_partner_key_v2(cfg: MarketplaceApiConfig, partner_id: str, ts: int, sign_v2: str, redirect_url: str, state: str):
    """
    Diagnostica erro de assinatura antes de abrir o login.
    Faz um GET simples no auth_partner para checar se a Shopee aceita o sign.
    Sem logica v1/v2 — apenas testa e reporta.
    """
    base_url = (cfg.api_base_url or "https://openplatform.sandbox.test-stable.shopee.sg").strip().rstrip("/")
    path = "/api/v2/shop/auth_partner"
    try:
        r = requests.get(
            f"{base_url}{path}",
            params={
                "partner_id": partner_id,
                "timestamp": ts,
                "sign": sign_v2,
                "redirect": redirect_url,
                "state": state,
            },
            timeout=18,
            allow_redirects=False,
        )
        data = {}
        try:
            data = r.json() if "json" in (r.headers.get("content-type") or "") else {}
        except Exception:
            data = {}
        err = str((data or {}).get("error") or "").strip().lower()
        msg = str((data or {}).get("message") or "").strip()

        if err == "error_sign":
            return {
                "ok": False,
                "erro": (
                    f"Assinatura rejeitada pela Shopee (Wrong sign). "
                    f"Verifique Partner ID ({partner_id}), Partner Key e api_base_url ({base_url})."
                ),
            }

        if err == "invalid_partner_id":
            return {
                "ok": False,
                "erro": f"Partner ID {partner_id} nao existe no ambiente {base_url}.",
            }

        # Qualquer outro erro (ex: error_param sobre redirect) ou redirect 302 = sign OK
        return {"ok": True}

    except Exception:
        # Falha de rede/timeout nao bloqueia o fluxo de login.
        return {"ok": True}


@app.route('/api/marketplace/shopee/login-url', methods=['GET'])
@jwt_required()
def api_marketplace_shopee_login_url():
    """
    Gera URL de login/autorizacao Shopee para o assinante.
    Campos tecnicos devem estar configurados na aba ADM.
    """
    user_id = int(get_jwt_identity())
    cfg = _get_or_create_marketplace_api_config(user_id, "shopee")

    partner_id = str(cfg.partner_id or "").strip()
    partner_key = str(cfg.get_partner_key() or "").strip()
    redirect_url = _get_shopee_redirect_url()

    if not partner_id:
        return jsonify({"status": "erro", "erro": "Partner ID nao configurado na aba Admin."}), 400
    if not partner_key:
        return jsonify({"status": "erro", "erro": "Partner Key nao configurada na aba Admin."}), 400
    if not redirect_url:
        return jsonify({
            "status": "erro",
            "erro": "Redirect URL nao disponivel. Configure NGROK_URL ou SHOPEE_REDIRECT_BASE_URL."
        }), 400

    ts = int(time.time())
    path = "/api/v2/shop/auth_partner"
    base_string = f"{partner_id}{path}{ts}"
    sign = hmac.new(
        partner_key.encode("utf-8"),
        base_string.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()
    state = _build_shopee_oauth_state(user_id)
    diag = _diagnosticar_partner_key_v2(
        cfg=cfg,
        partner_id=partner_id,
        ts=ts,
        sign_v2=sign,
        redirect_url=redirect_url,
        state=state,
    )
    if not diag.get("ok"):
        return jsonify({
            "status": "erro",
            "erro": diag.get("erro") or "Falha de assinatura na Shopee API.",
        }), 400

    # Registrar OAuth pendente no DB (funciona cross-worker no Railway).
    auth_base = (cfg.api_base_url or "https://openplatform.sandbox.test-stable.shopee.sg").strip().rstrip("/")
    _register_pending_oauth(user_id)
    auth_url = (
        f"{auth_base}{path}"
        f"?partner_id={quote_plus(partner_id)}"
        f"&timestamp={ts}"
        f"&sign={sign}"
        f"&redirect={quote_plus(redirect_url)}"
        f"&state={quote_plus(state)}"
    )

    return jsonify({
        "status": "ok",
        "auth_url": auth_url,
        "api_base_url": auth_base,
        "redirect_url": redirect_url,
        "redirect_domain": _get_shopee_redirect_domain(),
    })


@app.route('/api/marketplace/shopee/callback', methods=['GET'])
def api_marketplace_shopee_callback():
    """
    Endpoint de callback OAuth Shopee.
    Fluxo otimizado para evitar code expirado (~30s sandbox):
    1. Identifica usuario (state ou DB pending)
    2. Carrega cfg do DB (1 query, funciona cross-worker)
    3. Decrypt partner_key + POST token/get IMEDIATAMENTE
    4. Salva tokens no DB
    """
    t_start = time.time()
    code = str(request.args.get("code", "") or "").strip()
    shop_id = str(request.args.get("shop_id", "") or "").strip()
    state = str(request.args.get("state", "") or "").strip()
    print(f"[CALLBACK] t=0ms code={code[:16] if code else 'VAZIO'} shop_id={shop_id or 'VAZIO'} state={'SET' if state else 'VAZIO'}", flush=True, file=sys.stderr)

    if not code:
        return redirect('/?shopee_login=erro&msg=' + quote_plus("Shopee callback sem code."))

    # --- FASE 1: Identificar usuario + carregar cfg do DB ---
    cfg = None
    user_id = None

    # Tentar state assinado primeiro
    user_id, err_state = _parse_shopee_oauth_state(state)
    if user_id:
        cfg = _get_or_create_marketplace_api_config(int(user_id), "shopee")
        print(f"[CALLBACK] t={int((time.time()-t_start)*1000)}ms state OK user_id={user_id}", flush=True, file=sys.stderr)

    # Fallback: buscar no DB quem tem oauth_pending_at recente (cross-worker safe)
    if not cfg:
        cfg = _find_pending_oauth_cfg()
        if cfg:
            user_id = cfg.user_id
            print(f"[CALLBACK] t={int((time.time()-t_start)*1000)}ms DB pending user_id={user_id}", flush=True, file=sys.stderr)
        else:
            print(f"[CALLBACK] t={int((time.time()-t_start)*1000)}ms NENHUM usuario encontrado!", flush=True, file=sys.stderr)
            return redirect('/?shopee_login=erro&msg=' + quote_plus(f"State invalido: {err_state}"))

    # --- FASE 2: Decrypt + Exchange IMEDIATO ---
    partner_id = str(cfg.partner_id or "").strip()
    partner_key = str(cfg.get_partner_key() or "").strip()
    base_url = (cfg.api_base_url or "https://openplatform.sandbox.test-stable.shopee.sg").strip()
    print(f"[CALLBACK] t={int((time.time()-t_start)*1000)}ms creds OK pid={partner_id} key={'SET' if partner_key else 'VAZIO'}", flush=True, file=sys.stderr)

    troca = _shopee_exchange_code_for_tokens_fast(
        partner_id=partner_id,
        partner_key=partner_key,
        base_url=base_url,
        code=code,
        shop_id=shop_id,
    )
    t_exchange = time.time()
    print(f"[CALLBACK] t={int((t_exchange-t_start)*1000)}ms exchange ok={troca.get('ok')} erro={troca.get('erro','')[:60]}", flush=True, file=sys.stderr)

    # --- FASE 3: Salvar resultado no DB ---
    if not troca.get("ok"):
        try:
            cfg.status_conexao = "erro"
            cfg.oauth_pending_at = None
            db.session.commit()
        except Exception:
            pass
        detail = json.dumps(troca.get("data") or {}, default=str)[:300]
        msg_full = (troca.get("erro") or "Falha ao obter token Shopee.") + " | response: " + detail
        return redirect('/?shopee_login=erro&msg=' + quote_plus(msg_full))

    cfg.shop_id = str(troca.get("shop_id") or shop_id or cfg.shop_id or "").strip()
    cfg.set_access_token(str(troca.get("access_token") or "").strip())
    refresh_token = str(troca.get("refresh_token") or "").strip()
    if refresh_token:
        cfg.set_refresh_token(refresh_token)
    expire_in = int(troca.get("expire_in") or 0)
    if expire_in > 0:
        cfg.token_expires_at = _agora_brasil() + timedelta(seconds=expire_in)
    cfg.status_conexao = "ok"
    cfg.ativo = bool(cfg.configurado())
    cfg.oauth_pending_at = None  # Limpar flag de pending
    t_save = time.time()
    print(f"[CALLBACK] t={int((t_save-t_start)*1000)}ms tokens salvos, status=ok", flush=True, file=sys.stderr)

    # Tenta preencher nome da loja automaticamente (nao critico).
    try:
        cli = _ShopeeOpenApiClient(cfg)
        info = cli.get_shop_info()
        if info.get("ok"):
            shop_name = str((((info.get("data") or {}).get("response") or {}).get("shop_name")) or "").strip()
            if shop_name:
                cfg.loja_nome = shop_name
    except Exception:
        pass

    db.session.commit()
    t_end = time.time()
    print(f"[CALLBACK] t={int((t_end-t_start)*1000)}ms COMPLETO - redirect /?shopee_login=ok", flush=True, file=sys.stderr)
    return redirect('/?shopee_login=ok')


@app.route('/api/marketplace/config', methods=['POST'])
@jwt_required()
def api_marketplace_config_set():
    """Salva configuracao da API Shopee direta."""
    user_id = int(get_jwt_identity())
    cfg = _get_or_create_marketplace_api_config(user_id, "shopee")
    data = request.get_json(force=True, silent=True) or {}

    cfg.loja_nome = str(data.get("loja_nome", cfg.loja_nome or "") or "").strip()
    cfg.api_base_url = str(data.get("api_base_url", cfg.api_base_url or "https://openplatform.sandbox.test-stable.shopee.sg") or "").strip() or "https://openplatform.sandbox.test-stable.shopee.sg"
    cfg.partner_id = str(data.get("partner_id", cfg.partner_id or "") or "").strip()
    cfg.shop_id = str(data.get("shop_id", cfg.shop_id or "") or "").strip()

    partner_key = data.get("partner_key", None)
    access_token = data.get("access_token", None)
    refresh_token = data.get("refresh_token", None)
    limpar_tokens = _to_bool(data.get("limpar_tokens"), False)

    if partner_key is not None:
        key_txt = str(partner_key or "").strip()
        if key_txt:
            cfg.set_partner_key(key_txt)
    if access_token is not None:
        at_txt = str(access_token or "").strip()
        if at_txt:
            cfg.set_access_token(at_txt)
        elif limpar_tokens:
            cfg.access_token_enc = ""
    if refresh_token is not None:
        rt_txt = str(refresh_token or "").strip()
        if rt_txt:
            cfg.set_refresh_token(rt_txt)
        elif limpar_tokens:
            cfg.refresh_token_enc = ""

    # Atualizar expiração opcional se vier do frontend.
    token_expires_in = data.get("token_expires_in")
    if token_expires_in is not None:
        try:
            sec = max(1, int(token_expires_in))
            cfg.token_expires_at = _agora_brasil() + timedelta(seconds=sec)
        except Exception:
            pass

    cfg.ativo = bool(cfg.configurado())
    if not cfg.configurado():
        cfg.status_conexao = "nao_configurado"
    db.session.commit()
    data_out = cfg.to_dict()
    data_out["redirect_url"] = _get_shopee_redirect_url()
    data_out["redirect_domain"] = _get_shopee_redirect_domain()
    return jsonify(data_out)


@app.route('/api/marketplace/testar-conexao', methods=['POST'])
@jwt_required()
def api_marketplace_testar():
    """Testa credenciais da API Shopee (shop/get_shop_info)."""
    user_id = int(get_jwt_identity())
    cfg = _get_or_create_marketplace_api_config(user_id, "shopee")
    cli, err = _marketplace_cfg_to_client(cfg)
    if not cli:
        return jsonify({"status": "erro", "erro": err}), 400
    try:
        ret = cli.get_shop_info()
        if not ret.get("ok"):
            cfg.status_conexao = "erro"
            db.session.commit()
            return jsonify({
                "status": "erro",
                "erro": (ret.get("data") or {}).get("message") or (ret.get("data") or {}).get("error") or "falha_shop_info",
            }), 400
        info = ((ret.get("data") or {}).get("response") or {})
        shop_name = str(info.get("shop_name") or "").strip()
        if shop_name and not (cfg.loja_nome or "").strip():
            cfg.loja_nome = shop_name
        cfg.status_conexao = "ok"
        cfg.ativo = True
        db.session.commit()
        return jsonify({
            "status": "ok",
            "mensagem": "Conexao Shopee API validada",
            "shop_name": shop_name,
            "shop_id": cfg.shop_id,
        })
    except Exception as e:
        cfg.status_conexao = "erro"
        db.session.commit()
        return jsonify({"status": "erro", "erro": str(e)}), 500


@app.route('/api/marketplace/reconectar', methods=['POST'])
@jwt_required()
def api_marketplace_reconectar():
    """Alias de teste/revalidacao para API direta."""
    return api_marketplace_testar()


@app.route('/api/marketplace/shopee/diagnostico-itens', methods=['GET'])
@jwt_required()
def api_marketplace_shopee_diagnostico_itens():
    """Diagnostico: lista itens e seu status na loja sandbox."""
    user_id = int(get_jwt_identity())
    cfg = MarketplaceApiConfig.query.filter_by(user_id=user_id, marketplace="shopee").first()
    if not cfg:
        return jsonify({"status": "erro", "mensagem": "Shopee nao configurada"}), 400
    cli, err = _marketplace_cfg_to_client(cfg)
    if not cli:
        return jsonify({"status": "erro", "mensagem": err or "Token expirado"}), 400
    try:
        # Get item list
        item_list_resp = cli.get_item_list(item_status="NORMAL")
        items = []
        item_ids = []
        for it in ((item_list_resp.get("response") or {}).get("item") or []):
            items.append(it)
            item_ids.append(it.get("item_id"))
        # Get base info for those items
        base_info = {}
        if item_ids:
            try:
                bi_resp = cli.get_item_base_info(item_ids[:20])
                for it in ((bi_resp.get("response") or {}).get("item_list") or []):
                    base_info[it.get("item_id")] = it
            except Exception as e2:
                base_info = {"error": str(e2)}
        return jsonify({"status": "ok", "items": items, "base_info": base_info, "total": len(items), "raw_item_list": item_list_resp})
    except Exception as e:
        return jsonify({"status": "erro", "mensagem": str(e)}), 500

@app.route('/api/marketplace/shopee/criar-produto-teste', methods=['POST'])
@jwt_required()
def api_marketplace_shopee_criar_produto_teste():
    """Cria produtos de teste na loja Shopee sandbox via API."""
    user_id = int(get_jwt_identity())
    cfg = _get_or_create_marketplace_api_config(user_id, "shopee")
    cli, err = _marketplace_cfg_to_client(cfg)
    if not cli:
        return jsonify({"status": "erro", "erro": err}), 400

    try:
        # 1. Buscar categorias disponiveis
        cat_ret = cli.get_category_list(language="en")
        if not cat_ret.get("ok"):
            cat_ret = cli.get_category_list(language="")
        cats = ((cat_ret.get("data") or {}).get("response") or {}).get("category_list") or []
        leaf_cats = [c for c in cats if not c.get("has_children", True)]
        if not leaf_cats:
            leaf_cats = cats[:5]
        if not leaf_cats:
            return jsonify({"status": "erro", "erro": "Nenhuma categoria disponivel na Shopee sandbox"}), 400
        cat_id = leaf_cats[0].get("category_id", 0)
        cat_name = leaf_cats[0].get("display_category_name") or leaf_cats[0].get("category_name") or "Categoria"

        # 1b. Buscar marcas disponiveis para a categoria
        brand_ret = cli._request("GET", "/api/v2/product/get_brand_list", params={
            "offset": 0, "page_size": 10, "category_id": cat_id, "status": 1
        }, with_auth=True)
        brand_list = ((brand_ret.get("data") or {}).get("response") or {}).get("brand_list") or []
        # Usar "No Brand" (id=0) ou a primeira marca disponivel
        brand_id = 0
        brand_name = "NoBrand"
        for b in brand_list:
            bn = str(b.get("original_brand_name") or b.get("display_brand_name") or "").lower()
            if "no brand" in bn or bn == "nobrand":
                brand_id = b.get("brand_id", 0)
                brand_name = b.get("original_brand_name") or "NoBrand"
                break
        if brand_id == 0 and brand_list:
            brand_id = brand_list[0].get("brand_id", 0)
            brand_name = brand_list[0].get("original_brand_name") or "Brand"

        # 2. Buscar canais de logistica
        log_ret = cli.get_logistics_channel()
        logistics = ((log_ret.get("data") or {}).get("response") or {}).get("logistics_channel_list") or []
        logistic_info = []
        for lg in logistics[:3]:
            logistic_info.append({"logistic_id": lg.get("logistics_channel_id", 0), "enabled": True})
        if not logistic_info:
            logistic_info = [{"logistic_id": 0, "enabled": True}]

        # 3. Upload de imagem placeholder (multipart/form-data obrigatorio pela Shopee)
        image_ids = []
        try:
            import struct, zlib
            # Gerar PNG 200x200 laranja
            w, h = 200, 200
            raw = b''
            for _ in range(h):
                raw += b'\x00' + bytes([255, 140, 0]) * w
            def _chunk(ct, d):
                c = ct + d
                return struct.pack('>I', len(d)) + c + struct.pack('>I', zlib.crc32(c) & 0xffffffff)
            ihdr = struct.pack('>IIBBBBB', w, h, 8, 2, 0, 0, 0)
            png_bytes = b'\x89PNG\r\n\x1a\n' + _chunk(b'IHDR', ihdr) + _chunk(b'IDAT', zlib.compress(raw)) + _chunk(b'IEND', b'')

            # Upload via multipart/form-data (nao usa _request pois precisa de files=)
            path = "/api/v2/media_space/upload_image"
            ts = cli._timestamp()
            sign = cli._sign(path, ts, cli.access_token, cli.shop_id)
            params = {
                "partner_id": cli.partner_id, "timestamp": ts, "sign": sign,
                "access_token": cli.access_token, "shop_id": cli.shop_id,
            }
            url = f"{cli.base_url}{path}"
            resp_img = requests.post(url, params=params, files={"image": ("test.png", png_bytes, "image/png")}, timeout=30)
            img_data = resp_img.json() if resp_img.status_code < 400 else {}
            img_info = (img_data.get("response") or {}).get("image_info") or {}
            if isinstance(img_info, dict) and img_info.get("image_id"):
                image_ids = [img_info["image_id"]]
            elif isinstance(img_info, list):
                for ii in img_info:
                    if ii.get("image_id"):
                        image_ids.append(ii["image_id"])
        except Exception as img_err:
            print(f"[SHOPEE] Image upload error: {img_err}", flush=True, file=sys.stderr)

        # 4. Criar 3 produtos de teste
        produtos_teste = [
            {"name": "Camiseta Teste Beka - Preta M", "price": 59.90, "stock": 50, "desc": "Camiseta preta tamanho M para teste de integracao"},
            {"name": "Caneca Personalizada Beka 300ml", "price": 29.90, "stock": 100, "desc": "Caneca ceramica 300ml para teste de integracao"},
            {"name": "Adesivo Beka Kit com 10 unidades", "price": 14.90, "stock": 200, "desc": "Kit 10 adesivos variados para teste de integracao"},
        ]

        criados = []
        erros = []
        for p in produtos_teste:
            body = {
                "original_price": p["price"],
                "description": p["desc"],
                "item_name": p["name"],
                "seller_stock": [{"stock": p["stock"]}],
                "weight": 0.3,
                "category_id": cat_id,
                "image": {"image_id_list": image_ids} if image_ids else {},
                "logistic_info": logistic_info,
                "item_status": "NORMAL",
                "condition": "NEW",
                "brand": {"brand_id": brand_id, "original_brand_name": brand_name},
                "dimension": {"package_length": 20, "package_width": 15, "package_height": 5},
            }
            ret = cli._request("POST", "/api/v2/product/add_item", body=body, with_auth=True)
            if ret.get("ok"):
                item_id = ((ret.get("data") or {}).get("response") or {}).get("item_id", "?")
                criados.append({"name": p["name"], "item_id": item_id})
            else:
                err_msg = (ret.get("data") or {}).get("message") or (ret.get("data") or {}).get("error") or str(ret)
                erros.append({"name": p["name"], "erro": err_msg, "data": ret.get("data")})

        return jsonify({
            "status": "ok" if criados else "erro",
            "mensagem": f"{len(criados)} produto(s) criado(s), {len(erros)} erro(s)",
            "criados": criados,
            "erros": erros,
            "categoria_usada": {"id": cat_id, "nome": cat_name},
            "brand_usada": {"id": brand_id, "nome": brand_name},
            "logistica": logistic_info,
            "imagens": image_ids,
        })

    except Exception as e:
        import traceback
        return jsonify({"status": "erro", "erro": str(e), "trace": traceback.format_exc()}), 500


@app.route('/api/marketplace/desconectar', methods=['POST'])
@jwt_required()
def api_marketplace_desconectar():
    """Desativa conexao API direta, com opcao de limpar tokens."""
    user_id = int(get_jwt_identity())
    cfg = _get_or_create_marketplace_api_config(user_id, "shopee")
    data = request.get_json(silent=True) or {}
    limpar = _to_bool(data.get("limpar_tokens"), False)
    cfg.ativo = False
    cfg.status_conexao = "nao_configurado"
    if limpar:
        cfg.access_token_enc = ""
        cfg.refresh_token_enc = ""
    db.session.commit()
    return jsonify({"status": "desconectado", "mensagem": "API direta desativada"})


@app.route('/api/marketplace/lojas', methods=['GET'])
@jwt_required()
def api_marketplace_lojas_listar():
    """Lista lojas sincronizadas via API direta (separadas do UpSeller)."""
    user_id = int(get_jwt_identity())
    lojas = MarketplaceLoja.query.filter_by(user_id=user_id, ativo=True).all()
    lojas.sort(key=lambda l: (-(l.pedidos_pendentes or 0), (l.nome or "").casefold()))
    total = sum(int(l.pedidos_pendentes or 0) for l in lojas)
    return jsonify({
        "lojas": [l.to_dict() for l in lojas],
        "total_pedidos": total,
        "sidebar_info": _get_marketplace_sidebar_cache(user_id),
    })


@app.route('/api/marketplace/sincronizar', methods=['POST'])
@jwt_required()
def api_marketplace_sincronizar():
    """Sincroniza contagens via API direta (Shopee)."""
    user_id = int(get_jwt_identity())
    if not hasattr(app, "_marketplace_sync_em_andamento"):
        app._marketplace_sync_em_andamento = {}
    if not hasattr(app, "_marketplace_sync_status"):
        app._marketplace_sync_status = {}

    if app._marketplace_sync_em_andamento.get(user_id):
        return jsonify({"erro": "Sincronizacao API ja em andamento"}), 409

    def _worker():
        app._marketplace_sync_em_andamento[user_id] = True
        app._marketplace_sync_status[user_id] = {
            "etapa": "iniciando",
            "progresso": 5,
            "detalhes": "Iniciando sincronizacao Shopee API...",
            "em_andamento": True,
        }
        try:
            app._marketplace_sync_status[user_id].update({
                "etapa": "consultando",
                "progresso": 35,
                "detalhes": "Consultando pedidos na Shopee API...",
            })
            ret = _marketplace_shopee_sync_snapshot(user_id)
            if not ret.get("sucesso"):
                app._marketplace_sync_status[user_id].update({
                    "etapa": "erro",
                    "progresso": 0,
                    "detalhes": ret.get("erro", "Falha na sincronizacao API"),
                    "em_andamento": False,
                })
                return

            lojas = ret.get("lojas", []) or []
            total = int(ret.get("total_pedidos") or 0)
            app._marketplace_sync_status[user_id].update({
                "etapa": "concluido",
                "progresso": 100,
                "detalhes": ret.get("detalhes") or f"Concluido: {len(lojas)} loja(s), {total} pedido(s).",
                "em_andamento": False,
                "lojas_encontradas": lojas,
                "total_pedidos": total,
                "sidebar_info": ret.get("sidebar_info", {}),
            })
        except Exception as e:
            app._marketplace_sync_status[user_id].update({
                "etapa": "erro",
                "progresso": 0,
                "detalhes": str(e),
                "em_andamento": False,
            })
        finally:
            app._marketplace_sync_em_andamento[user_id] = False

    threading.Thread(target=_worker, daemon=True).start()
    return jsonify({"sucesso": True, "mensagem": "Sincronizacao API iniciada"})


@app.route('/api/marketplace/sincronizar/status', methods=['GET'])
@jwt_required()
def api_marketplace_sync_status():
    """Status da sincronizacao API direta."""
    user_id = int(get_jwt_identity())
    if hasattr(app, "_marketplace_sync_status") and user_id in app._marketplace_sync_status:
        st = dict(app._marketplace_sync_status[user_id])
        st["em_andamento"] = bool(hasattr(app, "_marketplace_sync_em_andamento") and app._marketplace_sync_em_andamento.get(user_id, False))
        return jsonify(st)
    return jsonify({
        "etapa": "idle",
        "progresso": 0,
        "detalhes": "",
        "em_andamento": False,
    })


@app.route('/api/lojas', methods=['GET'])
@jwt_required()
def api_lojas_listar():
    """Retorna todas as lojas persistidas do user (inclusive com 0 pedidos)."""
    user_id = int(get_jwt_identity())
    lojas = Loja.query.filter_by(user_id=user_id, ativo=True).all()
    lojas.sort(key=lambda l: (-(l.pedidos_pendentes or 0), (l.nome or "").casefold()))
    total = sum(l.pedidos_pendentes for l in lojas)
    return jsonify({
        "lojas": [l.to_dict() for l in lojas],
        "total_pedidos": total,
        "sidebar_info": _get_sidebar_cache(user_id),
    })


@app.route('/api/lojas/atualizar-todas', methods=['POST'])
@jwt_required()
def api_lojas_atualizar_todas():
    """Atualiza snapshot de TODAS as lojas (somente contagens), sem pipeline de geracao."""
    user_id = int(get_jwt_identity())

    # Evitar concorrencia com pipelines que usam o mesmo scraper.
    if (hasattr(app, '_gerar_em_andamento') and app._gerar_em_andamento.get(user_id)) or \
       (hasattr(app, '_imprimir_em_andamento') and app._imprimir_em_andamento.get(user_id)) or \
       (hasattr(app, '_sync_em_andamento') and app._sync_em_andamento.get(user_id)):
        return jsonify({"sucesso": False, "erro": "Aguarde a automacao atual finalizar"}), 409

    ret = _atualizar_todas_lojas_preciso(user_id)
    if not ret.get("sucesso"):
        return jsonify({"sucesso": False, "erro": ret.get("erro", "falha_atualizar_lojas")}), 400
    return jsonify(ret)


@app.route('/api/lojas/atualizar-individual', methods=['POST'])
@jwt_required()
def api_loja_atualizar_individual():
    """Atualiza contagem de uma loja especifica sem reprocessar a lista inteira."""
    user_id = int(get_jwt_identity())
    data = request.get_json(silent=True) or {}
    nome_loja = (data.get("loja") or "").strip()
    if not nome_loja:
        return jsonify({"sucesso": False, "erro": "Nome da loja e obrigatorio"}), 400

    # Evitar concorrencia com pipelines de automacao que usam o mesmo scraper.
    if (hasattr(app, '_gerar_em_andamento') and app._gerar_em_andamento.get(user_id)) or \
       (hasattr(app, '_imprimir_em_andamento') and app._imprimir_em_andamento.get(user_id)) or \
       (hasattr(app, '_sync_em_andamento') and app._sync_em_andamento.get(user_id)):
        return jsonify({"sucesso": False, "erro": "Aguarde a automacao atual finalizar"}), 409

    ret = _atualizar_loja_individual(user_id, nome_loja)
    if not ret.get("sucesso"):
        return jsonify(ret), 400
    return jsonify(ret)


@app.route('/api/upseller/conectar', methods=['POST'])
@jwt_required()
def api_upseller_conectar():
    """
    Abre navegador VISIVEL para login manual no UpSeller.
    MANTÉM o scraper vivo em background após login (cookies de sessão ficam na memória).
    """
    user_id = int(get_jwt_identity())
    config = _get_or_create_upseller_config(user_id)

    try:
        user = db.session.get(User, user_id)
        download_dir = user.get_pasta_entrada() if user else os.path.join(os.path.expanduser("~"), ".upseller_downloads")

        # Verificar se já está logado (scraper vivo)
        if _upseller_mgr.is_alive(user_id):
            try:
                logado = _upseller_mgr.esta_logado(user_id)
                if logado:
                    config.status_conexao = "ok"
                    db.session.commit()
                    return jsonify({"mensagem": "Já conectado ao UpSeller!", "status": "ok"})
            except Exception:
                pass

        # Criar scraper VISÍVEL e fazer login manual
        scraper = _upseller_mgr.criar_scraper(user_id, config.session_dir, download_dir, headless=False)
        logado = _upseller_mgr.login_manual(user_id, timeout=180)

        config.status_conexao = "ok" if logado else "erro"
        db.session.commit()

        if logado:
            # NÃO fecha o scraper! Ele fica vivo em background.
            return jsonify({"mensagem": "Conectado ao UpSeller! Navegador mantido em background.", "status": "ok"})

        # Se não logou, fechar scraper
        _upseller_mgr.fechar(user_id)
        return jsonify({"erro": "Login nao concluido. Tente novamente.", "status": "erro"}), 400

    except Exception as e:
        print(f"[UpSeller] Erro ao conectar: {e}")
        import traceback
        traceback.print_exc()
        _upseller_mgr.fechar(user_id)
        return jsonify({"erro": str(e), "status": "erro"}), 500


@app.route('/api/upseller/testar-conexao', methods=['POST'])
@jwt_required()
def api_upseller_testar():
    """Verifica se scraper está vivo e logado no UpSeller."""
    user_id = int(get_jwt_identity())
    config = _get_or_create_upseller_config(user_id)

    try:
        # Verificar scraper vivo primeiro (sem criar novo)
        if _upseller_mgr.is_alive(user_id):
            logado = _upseller_mgr.esta_logado(user_id)
            config.status_conexao = "ok" if logado else "erro"
            db.session.commit()
            if logado:
                return jsonify({"mensagem": "Sessao ativa!", "status": "ok"})
            return jsonify({"erro": "Sessao expirada. Clique em Reconectar.", "status": "erro"}), 400

        # Scraper não está vivo
        config.status_conexao = "desconectado"
        db.session.commit()
        return jsonify({"erro": "Navegador não está ativo. Clique em Conectar.", "status": "desconectado"}), 400

    except Exception as e:
        return jsonify({"erro": str(e), "status": "erro"}), 500


@app.route('/api/upseller/reconectar', methods=['POST'])
@jwt_required()
def api_upseller_reconectar():
    """
    Reconexão automática ao UpSeller:
    1. Se scraper vivo e logado → retorna OK imediatamente
    2. Se scraper vivo mas deslogado → tenta reconectar
    3. Se scraper morto → tenta criar headless (sessão persistente)
    4. Se tudo falhou → abre navegador visível para login manual
    """
    user_id = int(get_jwt_identity())
    config = _get_or_create_upseller_config(user_id)

    try:
        user = db.session.get(User, user_id)
        download_dir = user.get_pasta_entrada() if user else os.path.join(os.path.expanduser("~"), ".upseller_downloads")

        logado = _upseller_mgr.reconectar(user_id, config.session_dir, download_dir)

        config.status_conexao = "ok" if logado else "erro"
        db.session.commit()

        if logado:
            return jsonify({"mensagem": "Reconectado ao UpSeller!", "status": "ok"})
        return jsonify({"erro": "Não foi possível reconectar. Tente Conectar novamente.", "status": "erro"}), 400

    except Exception as e:
        print(f"[UpSeller] Erro ao reconectar: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"erro": str(e), "status": "erro"}), 500


@app.route('/api/upseller/desconectar', methods=['POST'])
@jwt_required()
def api_upseller_desconectar():
    """
    Desconecta do UpSeller: fecha o navegador/scraper, limpa sessao
    e reseta status. Permite trocar de conta.
    """
    user_id = int(get_jwt_identity())
    config = _get_or_create_upseller_config(user_id)

    try:
        # Fechar scraper (navegador)
        try:
            _upseller_mgr.fechar(user_id)
        except Exception:
            pass

        # Resetar status no banco
        config.status_conexao = "desconectado"
        db.session.commit()

        # Se pediu para limpar sessao, apagar diretorio de cookies
        data = request.get_json(silent=True) or {}
        limpar_sessao = data.get("limpar_sessao", False)
        if limpar_sessao and config.session_dir:
            import shutil
            try:
                shutil.rmtree(config.session_dir, ignore_errors=True)
                os.makedirs(config.session_dir, exist_ok=True)
            except Exception:
                pass

        return jsonify({
            "mensagem": "Desconectado do UpSeller. Clique 'Conectar' para fazer login com outra conta.",
            "status": "desconectado"
        })
    except Exception as e:
        print(f"[UpSeller] Erro ao desconectar: {e}")
        return jsonify({"erro": str(e), "status": "erro"}), 500


@app.route('/api/upseller/sincronizar', methods=['POST'])
@jwt_required()
def api_upseller_sincronizar():
    """
    Sincronizar = APENAS ler pedidos pendentes do UpSeller e mostrar lojas.
    NAO baixa nada, NAO processa nada. Rapido e leve.
    """
    user_id = int(get_jwt_identity())
    config = _get_or_create_upseller_config(user_id)

    if hasattr(app, '_sync_em_andamento') and app._sync_em_andamento.get(user_id):
        return jsonify({"erro": "Sincronizacao ja em andamento"}), 409

    def _sync_leitura():
        if not hasattr(app, '_sync_em_andamento'):
            app._sync_em_andamento = {}
        app._sync_em_andamento[user_id] = True
        if not hasattr(app, '_sync_status'):
            app._sync_status = {}
        app._sync_status[user_id] = {"etapa": "iniciando", "progresso": 0, "detalhes": "Conectando ao UpSeller..."}

        with app.app_context():
            try:
                # Etapa 1: Verificar se scraper está vivo e logado
                app._sync_status[user_id] = {"etapa": "login", "progresso": 20, "detalhes": "Verificando sessao UpSeller..."}

                user = db.session.get(User, user_id)
                download_dir = user.get_pasta_entrada() if user else os.path.join(os.path.expanduser("~"), ".upseller_downloads")

                scraper_vivo = _upseller_mgr.is_alive(user_id)
                scraper_atual = _upseller_mgr.get_scraper(user_id) if scraper_vivo else None
                headless_atual = getattr(scraper_atual, 'headless', True) if scraper_atual else True

                # Para leitura confiavel de lojas/filtros no UpSeller SPA, usar browser visivel.
                precisa_visivel = (not scraper_vivo) or headless_atual

                if precisa_visivel:
                    app._sync_status[user_id] = {"etapa": "reconectando", "progresso": 30, "detalhes": "Abrindo navegador visivel para sincronizacao precisa..."}
                    try:
                        scraper = _upseller_mgr.criar_scraper(user_id, config.session_dir, download_dir, headless=False)
                        logado = _upseller_mgr._run_async(user_id, scraper._esta_logado())
                    except Exception as e:
                        print(f"[Sync] Erro ao criar scraper visivel: {e}")
                        logado = False
                else:
                    # Scraper visivel e vivo: apenas validar sessao
                    logado = _upseller_mgr.esta_logado(user_id)

                if not logado:
                    app._sync_status[user_id] = {
                        "etapa": "erro", "progresso": 0,
                        "detalhes": "Sessao expirada. Clique 'Conectar ao UpSeller' para fazer login."
                    }
                    config.status_conexao = "desconectado"
                    db.session.commit()
                    return

                # Etapa 2: Ler pedidos pendentes usando scraper vivo
                app._sync_status[user_id] = {"etapa": "lendo_pedidos", "progresso": 60, "detalhes": "Lendo pedidos pendentes..."}
                resultado = _upseller_mgr.listar_lojas(user_id)

                # NÃO fecha o scraper! Ele continua vivo para próximas operações.

                # Atualizar config
                config.ultima_sincronizacao = _agora_brasil()
                config.status_conexao = "ok"
                db.session.commit()

                # Concluir
                if resultado.get("sucesso"):
                    lojas = resultado.get("lojas", [])
                    total = resultado.get("total_pedidos", 0)
                    sidebar = resultado.get("sidebar_info", {})

                    # Persistir snapshot completo de lojas (incluindo lojas com 0 pedido)
                    _persistir_lojas_upseller(user_id, lojas)
                    _set_sidebar_cache(user_id, sidebar)

                    app._sync_status[user_id] = {
                        "etapa": "concluido", "progresso": 100,
                        "detalhes": f"{len(lojas)} lojas, {total} pedidos pendentes",
                        "lojas_encontradas": lojas,
                        "total_pedidos": total,
                        "sidebar_info": sidebar,
                    }
                else:
                    app._sync_status[user_id] = {
                        "etapa": "erro", "progresso": 0,
                        "detalhes": resultado.get("erro", "Erro ao ler pedidos")
                    }

            except Exception as e:
                print(f"[Sync] Erro: {e}")
                import traceback
                traceback.print_exc()
                app._sync_status[user_id] = {"etapa": "erro", "progresso": 0, "detalhes": str(e)}
            finally:
                app._sync_em_andamento[user_id] = False

    thread = threading.Thread(target=_sync_leitura, daemon=True)
    thread.start()
    return jsonify({"mensagem": "Sincronizacao iniciada", "status": "iniciando"})


@app.route('/api/upseller/gerar', methods=['POST'])
@jwt_required()
def api_upseller_gerar():
    """
    Pipeline completo de geracao (atualizado 2026-02-28):

    Fluxo correto POR LOJA:
    1. Verificar/reconectar sessao UpSeller
    2. Atualizar contagens de lojas
    3. Emitir NF-e (com retry de falhas)
    4. Programar envio dos pedidos ("Para Programar")
    5. Aguardar tracking numbers
    6. Baixar etiquetas com DDC (Etiqueta Casada + Declaracao de Conteudo)
    7. Extrair dados de pedidos → XLSX
    8. Mover tudo para pasta_entrada
    9. Processar etiquetas (DDC dados no rodape, organizar por SKU)

    Body JSON (opcional):
    {
        "lojas": ["DAHIANE", "LOJA_X"],  // Lojas para processar (vazio = todas)
    }
    """
    user_id = int(get_jwt_identity())
    config = _get_or_create_upseller_config(user_id)

    if _esta_atualizando_lojas(user_id):
        return jsonify({
            "sucesso": False,
            "mensagem": "Atualizando pedidos, aguarde."
        }), 409

    acao_em_execucao = _obter_acao_massa_em_andamento(user_id)
    if acao_em_execucao:
        return jsonify({
            "sucesso": False,
            "mensagem": (
                f"Processo em massa em andamento ({acao_em_execucao.get('acao', 'processando')}). "
                "Aguarde finalizar."
            )
        }), 409

    if hasattr(app, '_gerar_em_andamento') and app._gerar_em_andamento.get(user_id):
        return jsonify({"erro": "Geracao ja em andamento"}), 409

    # Extrair lojas selecionadas do request
    data = request.get_json(silent=True) or {}
    lojas_raw = data.get("lojas", [])
    lojas_selecionadas, lojas_ignoradas = _sanitizar_lojas_selecionadas(user_id, lojas_raw)
    if isinstance(lojas_raw, list) and lojas_raw and not lojas_selecionadas:
        return jsonify({"erro": "Nenhuma loja valida selecionada para o pipeline."}), 400

    # Pre-setar status ANTES da thread para evitar race condition no polling
    if not hasattr(app, '_gerar_em_andamento'):
        app._gerar_em_andamento = {}
    if not hasattr(app, '_gerar_status'):
        app._gerar_status = {}
    lock_ok, lock_atual = _iniciar_acao_massa(user_id, "gerar")
    if not lock_ok:
        return jsonify({
            "sucesso": False,
            "mensagem": (
                f"Processo em massa em andamento ({(lock_atual or {}).get('acao', 'processando')}). "
                "Aguarde finalizar."
            )
        }), 409
    app._gerar_em_andamento[user_id] = True
    app._gerar_status[user_id] = {"etapa": "iniciando", "progresso": 0, "detalhes": "Iniciando geracao..."}

    def _gerar_pipeline():
        with app.app_context():
            try:
                user = db.session.get(User, user_id)
                if not user:
                    app._gerar_status[user_id] = {"etapa": "erro", "progresso": 0, "detalhes": "Usuario nao encontrado"}
                    return

                pasta_entrada = user.get_pasta_entrada()
                pasta_lote = _criar_pasta_lote_upseller(pasta_entrada, prefixo="gerar")
                download_dir = os.path.join(pasta_entrada, '_upseller_temp')
                os.makedirs(download_dir, exist_ok=True)

                resultado = {"pdfs": [], "xmls": [], "xlsx": "", "sucesso": False}
                lojas_falha_pipeline = []
                grupos_marketplace = []
                if lojas_selecionadas:
                    grupos_marketplace = _agrupar_lojas_por_marketplace(user_id, lojas_selecionadas)
                    if not grupos_marketplace:
                        grupos_marketplace = [{"marketplace": "Shopee", "lojas": list(lojas_selecionadas)}]

                # === Etapa 1: Verificar/reconectar scraper (headless=False para SPA) ===
                app._gerar_status[user_id] = {"etapa": "login", "progresso": 5, "detalhes": "Verificando sessao UpSeller..."}

                scraper_vivo = _upseller_mgr.is_alive(user_id)
                precisa_recriar = False
                if not scraper_vivo:
                    precisa_recriar = True
                else:
                    scraper_atual = _upseller_mgr.get_scraper(user_id)
                    if scraper_atual and getattr(scraper_atual, 'headless', True):
                        precisa_recriar = True
                    elif not _upseller_mgr.esta_logado(user_id):
                        precisa_recriar = True

                if precisa_recriar:
                    app._gerar_status[user_id] = {"etapa": "reconectando", "progresso": 8, "detalhes": "Abrindo navegador para UpSeller..."}
                    try:
                        scraper = _upseller_mgr.criar_scraper(user_id, config.session_dir, download_dir, headless=False)
                        logado = _upseller_mgr._run_async(user_id, scraper._esta_logado())
                        if not logado:
                            app._gerar_status[user_id] = {
                                "etapa": "erro", "progresso": 0,
                                "detalhes": "Sessao expirada. Clique 'Reconectar' para login manual."
                            }
                            config.status_conexao = "desconectado"
                            db.session.commit()
                            return
                        config.status_conexao = "ok"
                        db.session.commit()
                    except Exception as e:
                        print(f"[Gerar] Erro ao criar scraper visivel: {e}")
                        app._gerar_status[user_id] = {"etapa": "erro", "progresso": 0, "detalhes": f"Erro ao abrir navegador: {e}"}
                        return

                scraper = _upseller_mgr.get_scraper(user_id)
                if not scraper:
                    app._gerar_status[user_id] = {"etapa": "erro", "progresso": 0, "detalhes": "Scraper nao disponivel"}
                    return

                # Atualizar download_dir do scraper
                scraper.download_dir = download_dir
                if download_dir:
                    os.makedirs(download_dir, exist_ok=True)

                def _xlsx_download_valido(path):
                    try:
                        if not path or not os.path.exists(path):
                            return False
                        checker = getattr(scraper, "_arquivo_tabulado_valido", None)
                        if callable(checker):
                            return bool(checker(path))
                        with open(path, "rb") as f:
                            head = (f.read(512) or b"").lstrip().lower()
                        if head.startswith(b"<!doctype html") or head.startswith(b"<html"):
                            return False
                        return True
                    except Exception:
                        return False

                # === Etapa 2: Atualizar contagens de lojas ===
                app._gerar_status[user_id] = {
                    "etapa": "atualizando_lojas", "progresso": 7,
                    "detalhes": "Atualizando contagens de lojas..."
                }
                try:
                    if lojas_selecionadas:
                        # Otimizacao: evita varredura completa inicial quando o usuario
                        # ja escolheu lojas especificas para processar.
                        app._gerar_status[user_id]["detalhes"] = (
                            f"Lojas selecionadas ({len(lojas_selecionadas)}), "
                            "pulando varredura completa inicial"
                        )
                    else:
                        _atualizar_lojas_apos_acao(user_id)
                except Exception as e:
                    print(f"[Gerar] Aviso ao atualizar lojas: {e}")

                # === Etapa 3: Emitir NF-e (por loja ou todas) ===
                total_emitidos = 0
                aviso_falhas_nfe = ""
                motivos_falhas_nfe = []
                if lojas_selecionadas:
                    total_grupos = max(1, len(grupos_marketplace))
                    for idx_gp, gp in enumerate(grupos_marketplace, start=1):
                        lojas_gp = list(gp.get("lojas") or [])
                        mp_nome = (gp.get("marketplace") or "Shopee").strip() or "Shopee"
                        if not lojas_gp:
                            continue
                        app._gerar_status[user_id] = {
                            "etapa": "emitindo_nfe",
                            "progresso": 10,
                            "detalhes": (
                                f"Emitindo NF-e ({idx_gp}/{total_grupos}) - "
                                f"{mp_nome}: {len(lojas_gp)} loja(s)"
                            )
                        }
                        try:
                            nfe_result = _upseller_mgr._run_async(
                                user_id, scraper.emitir_nfe(filtro_loja=lojas_gp)
                            )
                            if isinstance(nfe_result, dict):
                                total_emitidos += int(nfe_result.get("total_emitidos", 0) or 0)
                                if nfe_result.get("aviso_falhas"):
                                    aviso_falhas_nfe = (
                                        (aviso_falhas_nfe + " | ") if aviso_falhas_nfe else ""
                                    ) + str(nfe_result.get("aviso_falhas") or "")
                                for m in (nfe_result.get("motivos_falhas") or []):
                                    mt = str(m or "").strip()
                                    if mt and mt not in motivos_falhas_nfe:
                                        motivos_falhas_nfe.append(mt)
                        except Exception as e:
                            print(f"[Gerar] Erro emitir NF-e ({mp_nome}): {e}")
                else:
                    app._gerar_status[user_id] = {
                        "etapa": "emitindo_nfe", "progresso": 10,
                        "detalhes": "Emitindo NF-e dos pedidos..."
                    }
                    try:
                        nfe_result = _upseller_mgr._run_async(user_id, scraper.emitir_nfe())
                        if isinstance(nfe_result, dict):
                            total_emitidos = int(nfe_result.get("total_emitidos", 0) or 0)
                            if nfe_result.get("aviso_falhas"):
                                aviso_falhas_nfe = nfe_result["aviso_falhas"]
                            for m in (nfe_result.get("motivos_falhas") or []):
                                mt = str(m or "").strip()
                                if mt and mt not in motivos_falhas_nfe:
                                    motivos_falhas_nfe.append(mt)
                    except Exception as e:
                        print(f"[Gerar] Erro emitir NF-e: {e}")

                if total_emitidos > 0:
                    app._gerar_status[user_id]["detalhes"] = f"{total_emitidos} NF-e emitidas"
                else:
                    app._gerar_status[user_id]["detalhes"] = "Nenhuma NF-e para emitir"

                # === Etapa 4: Programar envio (por loja ou todas) ===
                total_programados = 0
                if lojas_selecionadas:
                    total_grupos = max(1, len(grupos_marketplace))
                    for idx_gp, gp in enumerate(grupos_marketplace, start=1):
                        lojas_gp = list(gp.get("lojas") or [])
                        mp_nome = (gp.get("marketplace") or "Shopee").strip() or "Shopee"
                        if not lojas_gp:
                            continue
                        app._gerar_status[user_id] = {
                            "etapa": "programando",
                            "progresso": 20,
                            "detalhes": (
                                f"Programando envio ({idx_gp}/{total_grupos}) - "
                                f"{mp_nome}: {len(lojas_gp)} loja(s)"
                            )
                        }
                        try:
                            prog_result = _upseller_mgr._run_async(
                                user_id, scraper.programar_envio(filtro_loja=lojas_gp)
                            )
                            total_programados += int((prog_result or {}).get("total_programados", 0) or 0)
                        except Exception as e:
                            print(f"[Gerar] Erro programar envio ({mp_nome}): {e}")
                else:
                    # Programar todas as lojas de uma vez (sem filtro)
                    app._gerar_status[user_id] = {"etapa": "programando", "progresso": 20, "detalhes": "Programando envio dos pedidos..."}
                    try:
                        prog_result = _upseller_mgr._run_async(user_id, scraper.programar_envio())
                        total_programados = prog_result.get("total_programados", 0)
                    except Exception as e:
                        print(f"[Gerar] Erro programar envio: {e}")

                if total_programados > 0:
                    app._gerar_status[user_id]["detalhes"] = f"{total_programados} pedidos programados"
                else:
                    app._gerar_status[user_id]["detalhes"] = "Nenhum pedido novo para programar"

                # === Etapa 5: Aguardar tracking numbers ===
                app._gerar_status[user_id] = {
                    "etapa": "aguardando_tracking",
                    "progresso": 32,
                    "detalhes": "Aguardando tracking numbers do UpSeller..."
                }

                if total_programados > 0:
                    # Aguardar tracking (poll a cada 10s, max 2 min)
                    try:
                        tracking_ok = _upseller_mgr._run_async(
                            user_id, scraper._aguardar_tracking(timeout_segundos=180)
                        )
                        if tracking_ok:
                            app._gerar_status[user_id]["detalhes"] = "Tracking numbers recebidos!"
                        else:
                            app._gerar_status[user_id]["detalhes"] = "Timeout aguardando tracking, tentando baixar mesmo assim..."
                    except Exception as e:
                        print(f"[Gerar] Erro ao aguardar tracking: {e}")
                        app._gerar_status[user_id]["detalhes"] = "Continuando sem aguardar tracking..."
                else:
                    # Se nao programou nada, esperar um pouco apenas
                    import time
                    time.sleep(3)

                # === Etapa 5.5: Baixar Lista de Resumo (XLSX) antes das etiquetas ===
                app._gerar_status[user_id] = {
                    "etapa": "lista_resumo", "progresso": 38,
                    "detalhes": "Baixando Lista de Resumo (XLSX)..."
                }

                xlsx_paths = []
                try:
                    if lojas_selecionadas:
                        total_grupos = max(1, len(grupos_marketplace))
                        for idx_gp, gp in enumerate(grupos_marketplace, start=1):
                            lojas_gp = list(gp.get("lojas") or [])
                            mp_nome = (gp.get("marketplace") or "Shopee").strip() or "Shopee"
                            if not lojas_gp:
                                continue
                            app._gerar_status[user_id] = {
                                "etapa": "lista_resumo", "progresso": 40,
                                "detalhes": (
                                    f"Lista de Resumo ({idx_gp}/{total_grupos}) - "
                                    f"{mp_nome}: {len(lojas_gp)} loja(s)"
                                )
                            }
                            lista_xlsx = _upseller_mgr._run_async(
                                user_id, scraper.baixar_lista_resumo(filtro_loja=lojas_gp)
                            ) or []
                            if isinstance(lista_xlsx, str):
                                lista_xlsx = [lista_xlsx]
                            for xp in lista_xlsx:
                                if _xlsx_download_valido(xp):
                                    xlsx_paths.append(xp)
                                elif xp:
                                    print(f"[Gerar] Aviso: arquivo de resumo invalido (ignorado): {xp}")
                    else:
                        lista_xlsx = _upseller_mgr._run_async(
                            user_id, scraper.baixar_lista_resumo()
                        ) or []
                        if isinstance(lista_xlsx, str):
                            lista_xlsx = [lista_xlsx]
                        for xp in lista_xlsx:
                            if _xlsx_download_valido(xp):
                                xlsx_paths.append(xp)
                            elif xp:
                                print(f"[Gerar] Aviso: arquivo de resumo invalido (ignorado): {xp}")
                except Exception as e:
                    print(f"[Gerar] Erro lista resumo: {e}")

                # Deduplicar mantendo ordem.
                if xlsx_paths:
                    vistos_xlsx = set()
                    dedup_xlsx = []
                    for xp in xlsx_paths:
                        if not xp or xp in vistos_xlsx:
                            continue
                        vistos_xlsx.add(xp)
                        dedup_xlsx.append(xp)
                    xlsx_paths = dedup_xlsx

                if xlsx_paths:
                    app._gerar_status[user_id]["detalhes"] = f"{len(xlsx_paths)} arquivo(s) de Lista de Resumo baixado(s)"
                else:
                    app._gerar_status[user_id]["detalhes"] = "Lista de Resumo indisponivel, tentando fallback depois das etiquetas"

                # === Etapa 6: Baixar etiquetas com DDC ===
                app._gerar_status[user_id] = {
                    "etapa": "baixando_etiquetas",
                    "progresso": 45,
                    "detalhes": "Baixando etiquetas (Casada + DDC)..."
                }

                if lojas_selecionadas:
                    total_grupos = max(1, len(grupos_marketplace))
                    for idx_gp, gp in enumerate(grupos_marketplace, start=1):
                        lojas_gp = list(gp.get("lojas") or [])
                        mp_nome = (gp.get("marketplace") or "Shopee").strip() or "Shopee"
                        if not lojas_gp:
                            continue
                        app._gerar_status[user_id] = {
                            "etapa": "baixando_etiquetas",
                            "progresso": 48,
                            "detalhes": (
                                f"Baixando etiquetas ({idx_gp}/{total_grupos}) - "
                                f"{mp_nome}: {len(lojas_gp)} loja(s)"
                            )
                        }
                        try:
                            pdfs_lote = _upseller_mgr._run_async(
                                user_id, scraper.baixar_etiquetas(filtro_loja=lojas_gp)
                            ) or []
                            if pdfs_lote:
                                resultado["pdfs"].extend(list(pdfs_lote))
                                continue
                        except Exception as e:
                            print(f"[Gerar] Erro baixar etiquetas ({mp_nome}): {e}")
                            lojas_falha_pipeline.append({
                                "loja": f"lote_{mp_nome}",
                                "etapa": "gerar_etiquetas",
                                "motivo": str(e),
                            })

                        # Resiliencia: se grupo vier vazio/erro, tentar loja a loja.
                        for loja_nome in lojas_gp:
                            try:
                                pdfs_loja = _upseller_mgr._run_async(
                                    user_id, scraper.baixar_etiquetas(filtro_loja=loja_nome)
                                ) or []
                                if pdfs_loja:
                                    resultado["pdfs"].extend(list(pdfs_loja))
                                else:
                                    lojas_falha_pipeline.append({
                                        "loja": loja_nome,
                                        "marketplace": mp_nome,
                                        "etapa": "gerar_etiquetas",
                                        "motivo": "sem_pdf_gerado",
                                    })
                            except Exception as e_loja_pdf:
                                lojas_falha_pipeline.append({
                                    "loja": loja_nome,
                                    "marketplace": mp_nome,
                                    "etapa": "gerar_etiquetas",
                                    "motivo": str(e_loja_pdf),
                                })
                else:
                    # Baixar todas (sem filtro)
                    try:
                        resultado["pdfs"] = _upseller_mgr._run_async(user_id, scraper.baixar_etiquetas())
                    except Exception as e:
                        print(f"[Gerar] Erro baixar etiquetas: {e}")

                app._gerar_status[user_id]["detalhes"] = f"{len(resultado['pdfs'])} PDFs baixados"

                # Sem PDF = nada para processar. Nao continuar para evitar zerar saida/resultado.
                if not resultado["pdfs"]:
                    detalhe_falhas = ""
                    if lojas_falha_pipeline:
                        amostra = " | ".join(
                            f"{x.get('loja', '?')}: {x.get('motivo', 'erro')}"
                            for x in lojas_falha_pipeline[:5]
                        )
                        detalhe_falhas = f" Falhas: {amostra}"
                    app._gerar_status[user_id] = {
                        "etapa": "erro",
                        "progresso": 0,
                        "detalhes": (
                            "Nenhuma etiqueta (PDF) foi baixada do UpSeller. "
                            "Nada para processar." + detalhe_falhas
                        ),
                        "lojas_falha": lojas_falha_pipeline[:20],
                    }
                    # Atualizar sync de lojas para refletir estado atual do UpSeller
                    _atualizar_lojas_apos_acao(user_id)
                    config.ultima_sincronizacao = _agora_brasil()
                    config.status_conexao = "ok"
                    db.session.commit()
                    return

                # === Etapa 7: Consolidar XLSX (fallback legado apenas se necessario) ===
                app._gerar_status[user_id] = {
                    "etapa": "extraindo_pedidos",
                    "progresso": 60,
                    "detalhes": "Consolidando dados de pedidos (XLSX)..."
                }
                try:
                    if not xlsx_paths:
                        # Fallback legado: extracao via tela apenas se Lista de Resumo falhar.
                        if lojas_selecionadas:
                            app._gerar_status[user_id] = {
                                "etapa": "extraindo_pedidos",
                                "progresso": 64,
                                "detalhes": f"Fallback XLSX em lote ({len(lojas_selecionadas)} lojas)..."
                            }
                            xp = _upseller_mgr._run_async(
                                user_id,
                                scraper.extrair_dados_pedidos(
                                    status_filter="para_imprimir",
                                    filtro_loja=lojas_selecionadas
                                )
                            )
                            if _xlsx_download_valido(xp):
                                xlsx_paths.append(xp)
                        else:
                            xp = _upseller_mgr._run_async(
                                user_id,
                                scraper.extrair_dados_pedidos(status_filter="para_imprimir")
                            )
                            if _xlsx_download_valido(xp):
                                xlsx_paths.append(xp)

                    if xlsx_paths:
                        resultado["xlsx"] = xlsx_paths[0]
                        if len(xlsx_paths) > 1:
                            resultado["xlsx_extra"] = xlsx_paths[1:]
                except Exception as e:
                    print(f"[Gerar] Erro consolidar/extrair XLSX: {e}")

                resultado["sucesso"] = True

                # NÃO fecha o scraper - ele fica vivo!

                # === Etapa 8: Mover arquivos para lote isolado da pasta_entrada ===
                app._gerar_status[user_id] = {"etapa": "movendo", "progresso": 72, "detalhes": "Movendo arquivos para lote atual..."}
                resumo = scraper.mover_para_pasta_entrada(resultado, pasta_lote)
                detalhes_mov = f"{resumo['pdfs_movidos']} PDFs"
                if resumo['xlsx_copiado']:
                    detalhes_mov += ", XLSX"
                if resumo.get("pdfs_movidos", 0) <= 0:
                    app._gerar_status[user_id] = {
                        "etapa": "erro",
                        "progresso": 0,
                        "detalhes": (
                            "Nenhum PDF foi movido para o lote atual. "
                            "Processamento cancelado para preservar o ultimo resultado valido."
                        ),
                        "processado": False,
                    }
                    _atualizar_lojas_apos_acao(user_id)
                    config.ultima_sincronizacao = _agora_brasil()
                    config.status_conexao = "ok"
                    db.session.commit()
                    return

                # === Etapa 9: Processar etiquetas ===
                app._gerar_status[user_id] = {"etapa": "processando", "progresso": 86, "detalhes": "Processando etiquetas + DDC + produtos..."}
                try:
                    proc_result = _executar_processamento(
                        user_id,
                        sem_recorte=True,
                        resumo_sku_somente=False,
                        pasta_entrada_override=pasta_lote
                    )
                    if not proc_result or not proc_result.get("ok"):
                        raise RuntimeError(
                            (proc_result or {}).get("erro") or
                            "Processamento concluido sem gerar resultado valido."
                        )
                    status_final = {
                        "etapa": "concluido", "progresso": 100,
                        "detalhes": (
                            f"Concluido! {total_emitidos} NF-e | {detalhes_mov} | "
                            f"{proc_result.get('total_etiquetas', 0)} etiquetas em "
                            f"{proc_result.get('total_lojas', 0)} loja(s)"
                        ),
                        "processado": True,
                    }
                    if aviso_falhas_nfe:
                        status_final["aviso_falhas"] = aviso_falhas_nfe
                    if motivos_falhas_nfe:
                        status_final["motivos_falhas"] = motivos_falhas_nfe[:8]
                    if lojas_falha_pipeline:
                        status_final["lojas_falha"] = lojas_falha_pipeline[:20]
                        status_final["aviso_falhas"] = (
                            (status_final.get("aviso_falhas", "") + " | " if status_final.get("aviso_falhas") else "")
                            + f"{len(lojas_falha_pipeline)} loja(s) com erro foram ignoradas."
                        )

                    # Auto-disparo opcional WhatsApp (fila persistente)
                    try:
                        auto_res = _enfileirar_envio_whatsapp_resultado(
                            user_id=user_id,
                            origem="auto",
                            respeitar_toggle_auto=True,
                        )
                        if auto_res.get("ok"):
                            status_final["auto_whatsapp"] = {
                                "enfileirados": auto_res.get("total_entregas", 0),
                                "batch_id": auto_res.get("batch_id", ""),
                            }
                            status_final["detalhes"] += f" | WhatsApp fila: {auto_res.get('total_entregas', 0)}"
                        elif not auto_res.get("ignorado"):
                            status_final["aviso_whatsapp"] = auto_res.get("erro", "Falha ao enfileirar WhatsApp")
                    except Exception as e_auto:
                        status_final["aviso_whatsapp"] = f"Falha no auto-envio WhatsApp: {e_auto}"
                    # Resumo geral diario: enviar JPEG parcial + acumular para consolidado
                    try:
                        _enviar_resumo_geral_whatsapp(user_id, user.get_pasta_saida())
                    except Exception as e_rg:
                        print(f"[ResumoGeral] Erro: {e_rg}")
                    app._gerar_status[user_id] = status_final
                except Exception as e:
                    print(f"[Gerar] Erro processamento: {e}")
                    import traceback
                    traceback.print_exc()
                    app._gerar_status[user_id] = {
                        "etapa": "parcial", "progresso": 90,
                        "detalhes": f"Download OK ({detalhes_mov}), erro no processamento: {e}",
                        "processado": False,
                    }

                # Re-sincronizar contagens em background para liberar resultado imediatamente.
                _disparar_atualizacao_lojas_background(user_id, status_attr="_gerar_status")

                # Atualizar config
                config.ultima_sincronizacao = _agora_brasil()
                config.status_conexao = "ok"
                db.session.commit()

                # Limpar pasta temp
                try:
                    import shutil
                    shutil.rmtree(download_dir, ignore_errors=True)
                except:
                    pass

            except Exception as e:
                print(f"[Gerar] Erro geral: {e}")
                import traceback
                traceback.print_exc()
                app._gerar_status[user_id] = {"etapa": "erro", "progresso": 0, "detalhes": str(e)}
            finally:
                app._gerar_em_andamento[user_id] = False
                _finalizar_acao_massa(user_id, "gerar")

    thread = threading.Thread(target=_gerar_pipeline, daemon=True)
    thread.start()
    payload = {"mensagem": "Geracao iniciada", "status": "iniciando"}
    if lojas_ignoradas:
        payload["lojas_ignoradas"] = lojas_ignoradas
    return jsonify(payload)


@app.route('/api/upseller/sincronizar/status', methods=['GET'])
@jwt_required()
def api_upseller_sync_status():
    """Retorna status da sincronizacao em andamento, incluindo lojas escaneadas quando concluido."""
    user_id = int(get_jwt_identity())
    if hasattr(app, '_sync_status') and user_id in app._sync_status:
        status = dict(app._sync_status[user_id])
        em_andamento = hasattr(app, '_sync_em_andamento') and app._sync_em_andamento.get(user_id, False)
        status["em_andamento"] = em_andamento

        # Quando concluido, incluir lojas escaneadas para o usuario escolher
        if status.get("etapa") == "concluido":
            # lojas_encontradas ja estao no status (set pelo pipeline)
            if "lojas_encontradas" not in status:
                status["lojas_encontradas"] = []
            if "total_etiquetas_scan" not in status:
                status["total_etiquetas_scan"] = 0

        return jsonify(status)
    return jsonify({"etapa": "idle", "progresso": 0, "detalhes": "", "em_andamento": False})


@app.route('/api/upseller/gerar/status', methods=['GET'])
@jwt_required()
def api_upseller_gerar_status():
    """Retorna status da geracao em andamento (separado do status de sincronizacao)."""
    user_id = int(get_jwt_identity())
    if hasattr(app, '_gerar_status') and user_id in app._gerar_status:
        status = dict(app._gerar_status[user_id])
        em_andamento = hasattr(app, '_gerar_em_andamento') and app._gerar_em_andamento.get(user_id, False)
        status["em_andamento"] = em_andamento
        status["atualizando_pedidos"] = _esta_atualizando_lojas(user_id) or bool(status.get("atualizando_pedidos"))
        return jsonify(status)
    return jsonify({
        "etapa": "idle",
        "progresso": 0,
        "detalhes": "",
        "em_andamento": False,
        "atualizando_pedidos": _esta_atualizando_lojas(user_id),
    })


# ----------------------------------------------------------------
# ENDPOINTS - ATALHOS INDIVIDUAIS (Emitir, Programar, Imprimir)
# ----------------------------------------------------------------

@app.route('/api/upseller/emitir', methods=['POST'])
@jwt_required()
def api_upseller_emitir():
    """Emite NF-e dos pedidos pendentes no UpSeller (atalho 'Para Emitir')."""
    user_id = int(get_jwt_identity())

    if _esta_atualizando_lojas(user_id):
        return jsonify({"sucesso": False, "mensagem": "Atualizando pedidos, aguarde."}), 409

    acao_em_execucao = _obter_acao_massa_em_andamento(user_id)
    if acao_em_execucao:
        return jsonify({
            "sucesso": False,
            "mensagem": (
                f"Processo em massa em andamento ({acao_em_execucao.get('acao', 'processando')}). "
                "Aguarde finalizar."
            )
        }), 409

    if not _upseller_mgr.is_alive(user_id):
        return jsonify({"sucesso": False, "mensagem": "UpSeller nao conectado. Conecte primeiro."}), 400

    data = request.get_json() or {}
    filtro_loja = (data.get("loja") or "").strip() or None
    lojas_lista_raw = data.get("lojas", [])
    lojas_lista, lojas_ignoradas = _sanitizar_lojas_selecionadas(user_id, lojas_lista_raw)

    if isinstance(lojas_lista_raw, list) and lojas_lista_raw and not lojas_lista:
        return jsonify({
            "sucesso": False,
            "mensagem": "Nenhuma loja valida selecionada para emitir NF-e."
        }), 400
    # Se vier lista de lojas, prioriza SEMPRE execucao em lote
    # (ignora campo legado "loja" para evitar cair em fluxo individual).
    if lojas_lista:
        filtro_loja = None

    if filtro_loja:
        filtro_lista, _ = _sanitizar_lojas_selecionadas(user_id, [filtro_loja])
        if not filtro_lista:
            return jsonify({
                "sucesso": False,
                "mensagem": f"Loja invalida para emissao: {filtro_loja}"
            }), 400
        filtro_loja = filtro_lista[0]

    lock_ok, lock_atual = _iniciar_acao_massa(user_id, "emitir")
    if not lock_ok:
        return jsonify({
            "sucesso": False,
            "mensagem": (
                f"Processo em massa em andamento ({(lock_atual or {}).get('acao', 'processando')}). "
                "Aguarde finalizar."
            )
        }), 409

    try:
        scraper = _upseller_mgr._scrapers.get(user_id)
        if not scraper:
            return jsonify({"sucesso": False, "mensagem": "Scraper nao encontrado"}), 400

        # Se recebeu lista de lojas, executar em lotes separados por marketplace
        # para evitar erro de plataforma mista no UpSeller.
        if lojas_lista:
            grupos_mp = _agrupar_lojas_por_marketplace(user_id, lojas_lista)
            if not grupos_mp:
                grupos_mp = [{"marketplace": "Shopee", "lojas": list(lojas_lista)}]

            total_emitidos_all = 0
            lojas_ok = []
            lojas_falha = []
            motivos_falhas = []
            grupos_processados = []

            for idx_gp, gp in enumerate(grupos_mp, start=1):
                lojas_gp = list(gp.get("lojas") or [])
                mp_nome = (gp.get("marketplace") or "Shopee").strip() or "Shopee"
                if not lojas_gp:
                    continue

                grupos_processados.append({"marketplace": mp_nome, "lojas": lojas_gp})
                try:
                    r_gp = _upseller_mgr._run_async(user_id, scraper.emitir_nfe(filtro_loja=lojas_gp)) or {}
                    if not isinstance(r_gp, dict):
                        r_gp = {"sucesso": False, "mensagem": "retorno_invalido"}

                    if bool(r_gp.get("sucesso")):
                        lojas_ok.extend(lojas_gp)
                        total_emitidos_all += int((r_gp.get("total_emitidos", 0) or 0))
                        for m in (r_gp.get("motivos_falhas") or []):
                            mt = str(m or "").strip()
                            if mt and mt not in motivos_falhas:
                                motivos_falhas.append(mt)
                        continue

                    # Fallback por loja (somente dentro do marketplace atual)
                    if len(lojas_gp) > 1:
                        for loja_nome in lojas_gp:
                            try:
                                r_loja = _upseller_mgr._run_async(user_id, scraper.emitir_nfe(filtro_loja=loja_nome)) or {}
                                if not isinstance(r_loja, dict):
                                    r_loja = {"sucesso": False, "mensagem": "retorno_invalido"}
                                if bool(r_loja.get("sucesso")):
                                    lojas_ok.append(loja_nome)
                                    total_emitidos_all += int((r_loja.get("total_emitidos", 0) or 0))
                                    for m in (r_loja.get("motivos_falhas") or []):
                                        mt = str(m or "").strip()
                                        if mt and mt not in motivos_falhas:
                                            motivos_falhas.append(mt)
                                else:
                                    lojas_falha.append({
                                        "loja": loja_nome,
                                        "marketplace": mp_nome,
                                        "motivo": (r_loja.get("mensagem") or "erro_na_emissao"),
                                    })
                            except Exception as e_loja:
                                lojas_falha.append({
                                    "loja": loja_nome,
                                    "marketplace": mp_nome,
                                    "motivo": str(e_loja),
                                })
                    else:
                        lojas_falha.append({
                            "loja": lojas_gp[0],
                            "marketplace": mp_nome,
                            "motivo": (r_gp.get("mensagem") or "erro_na_emissao"),
                        })
                except Exception as e_gp:
                    for loja_nome in lojas_gp:
                        lojas_falha.append({
                            "loja": loja_nome,
                            "marketplace": mp_nome,
                            "motivo": str(e_gp),
                        })

            resultado = {
                "sucesso": bool(lojas_ok),
                "mensagem": (
                    f"Emissao concluida por marketplace: {len(lojas_ok)} loja(s) ok, "
                    f"{len(lojas_falha)} com erro"
                ),
                "total_emitidos": int(total_emitidos_all),
                "modo_execucao": "lote_marketplace",
                "lojas_processadas": lojas_ok,
                "grupos_marketplace": grupos_processados,
            }
            if motivos_falhas:
                resultado["motivos_falhas"] = motivos_falhas[:8]
            if lojas_falha:
                resultado["lojas_falha"] = lojas_falha[:20]
                resultado["aviso_falhas"] = (
                    f"{len(lojas_falha)} loja(s) com erro foram ignoradas."
                )

            snapshot = _atualizar_lojas_apos_acao(user_id)
            resultado["lojas_atualizadas"] = bool(snapshot.get("sucesso"))
            if lojas_ignoradas:
                resultado["lojas_ignoradas"] = lojas_ignoradas
            return jsonify(resultado)

        resultado = _upseller_mgr._run_async(user_id, scraper.emitir_nfe(filtro_loja=filtro_loja))
        snapshot = _atualizar_lojas_apos_acao(user_id)
        if isinstance(resultado, dict):
            resultado["lojas_atualizadas"] = bool(snapshot.get("sucesso"))
        return jsonify(resultado)
    except Exception as e:
        print(f"[Emitir] Erro: {e}")
        return jsonify({"sucesso": False, "mensagem": str(e)}), 500
    finally:
        _finalizar_acao_massa(user_id, "emitir")


@app.route('/api/upseller/programar', methods=['POST'])
@jwt_required()
def api_upseller_programar():
    """Programa envio dos pedidos pendentes no UpSeller (atalho 'Para Enviar')."""
    user_id = int(get_jwt_identity())

    if _esta_atualizando_lojas(user_id):
        return jsonify({"sucesso": False, "mensagem": "Atualizando pedidos, aguarde."}), 409

    acao_em_execucao = _obter_acao_massa_em_andamento(user_id)
    if acao_em_execucao:
        return jsonify({
            "sucesso": False,
            "mensagem": (
                f"Processo em massa em andamento ({acao_em_execucao.get('acao', 'processando')}). "
                "Aguarde finalizar."
            )
        }), 409

    if not _upseller_mgr.is_alive(user_id):
        return jsonify({"sucesso": False, "mensagem": "UpSeller nao conectado. Conecte primeiro."}), 400

    data = request.get_json() or {}
    filtro_loja = (data.get("loja") or "").strip() or None
    lojas_lista_raw = data.get("lojas", [])
    lojas_lista, lojas_ignoradas = _sanitizar_lojas_selecionadas(user_id, lojas_lista_raw)

    if isinstance(lojas_lista_raw, list) and lojas_lista_raw and not lojas_lista:
        return jsonify({
            "sucesso": False,
            "mensagem": "Nenhuma loja valida selecionada para programar envio."
        }), 400
    # Se vier lista de lojas, prioriza SEMPRE execucao em lote
    # (ignora campo legado "loja" para evitar cair em fluxo individual).
    if lojas_lista:
        filtro_loja = None

    if filtro_loja:
        filtro_lista, _ = _sanitizar_lojas_selecionadas(user_id, [filtro_loja])
        if not filtro_lista:
            return jsonify({
                "sucesso": False,
                "mensagem": f"Loja invalida para programar envio: {filtro_loja}"
            }), 400
        filtro_loja = filtro_lista[0]

    lock_ok, lock_atual = _iniciar_acao_massa(user_id, "programar")
    if not lock_ok:
        return jsonify({
            "sucesso": False,
            "mensagem": (
                f"Processo em massa em andamento ({(lock_atual or {}).get('acao', 'processando')}). "
                "Aguarde finalizar."
            )
        }), 409

    try:
        scraper = _upseller_mgr._scrapers.get(user_id)
        if not scraper:
            return jsonify({"sucesso": False, "mensagem": "Scraper nao encontrado"}), 400

        if lojas_lista:
            grupos_mp = _agrupar_lojas_por_marketplace(user_id, lojas_lista)
            if not grupos_mp:
                grupos_mp = [{"marketplace": "Shopee", "lojas": list(lojas_lista)}]

            total_programados_all = 0
            lojas_ok = []
            lojas_falha = []
            grupos_processados = []

            for idx_gp, gp in enumerate(grupos_mp, start=1):
                lojas_gp = list(gp.get("lojas") or [])
                mp_nome = (gp.get("marketplace") or "Shopee").strip() or "Shopee"
                if not lojas_gp:
                    continue

                grupos_processados.append({"marketplace": mp_nome, "lojas": lojas_gp})
                try:
                    r_gp = _upseller_mgr._run_async(user_id, scraper.programar_envio(filtro_loja=lojas_gp)) or {}
                    if not isinstance(r_gp, dict):
                        r_gp = {"sucesso": False, "mensagem": "retorno_invalido"}

                    if bool(r_gp.get("sucesso")):
                        lojas_ok.extend(lojas_gp)
                        total_programados_all += int((r_gp.get("total_programados", 0) or 0))
                        continue

                    # Fallback por loja dentro do marketplace atual.
                    if len(lojas_gp) > 1:
                        for loja_nome in lojas_gp:
                            try:
                                r_loja = _upseller_mgr._run_async(user_id, scraper.programar_envio(filtro_loja=loja_nome)) or {}
                                if not isinstance(r_loja, dict):
                                    r_loja = {"sucesso": False, "mensagem": "retorno_invalido"}
                                if bool(r_loja.get("sucesso")):
                                    lojas_ok.append(loja_nome)
                                    total_programados_all += int((r_loja.get("total_programados", 0) or 0))
                                else:
                                    lojas_falha.append({
                                        "loja": loja_nome,
                                        "marketplace": mp_nome,
                                        "motivo": (r_loja.get("mensagem") or "erro_na_programacao"),
                                    })
                            except Exception as e_loja:
                                lojas_falha.append({
                                    "loja": loja_nome,
                                    "marketplace": mp_nome,
                                    "motivo": str(e_loja),
                                })
                    else:
                        lojas_falha.append({
                            "loja": lojas_gp[0],
                            "marketplace": mp_nome,
                            "motivo": (r_gp.get("mensagem") or "erro_na_programacao"),
                        })
                except Exception as e_gp:
                    for loja_nome in lojas_gp:
                        lojas_falha.append({
                            "loja": loja_nome,
                            "marketplace": mp_nome,
                            "motivo": str(e_gp),
                        })

            retorno = {
                "sucesso": bool(lojas_ok),
                "mensagem": (
                    f"Programacao concluida por marketplace: {len(lojas_ok)} loja(s) ok, "
                    f"{len(lojas_falha)} com erro"
                ),
                "total_programados": int(total_programados_all),
                "modo_execucao": "lote_marketplace",
                "lojas_processadas": lojas_ok,
                "grupos_marketplace": grupos_processados,
            }
            if lojas_falha:
                retorno["lojas_falha"] = lojas_falha[:20]
                retorno["aviso_falhas"] = (
                    f"{len(lojas_falha)} loja(s) com erro foram ignoradas."
                )

            snapshot = _atualizar_lojas_apos_acao(user_id)
            retorno["lojas_atualizadas"] = bool(snapshot.get("sucesso"))
            if lojas_ignoradas:
                retorno["lojas_ignoradas"] = lojas_ignoradas
            return jsonify(retorno)

        resultado = _upseller_mgr._run_async(user_id, scraper.programar_envio(filtro_loja=filtro_loja))
        snapshot = _atualizar_lojas_apos_acao(user_id)
        if isinstance(resultado, dict):
            resultado["lojas_atualizadas"] = bool(snapshot.get("sucesso"))
        return jsonify(resultado)
    except Exception as e:
        print(f"[Programar] Erro: {e}")
        return jsonify({"sucesso": False, "mensagem": str(e)}), 500
    finally:
        _finalizar_acao_massa(user_id, "programar")


@app.route('/api/upseller/imprimir', methods=['POST'])
@jwt_required()
def api_upseller_imprimir():
    """
    Pipeline 'Gerar Etiquetas' (atalho do UpSeller):
    1. Baixa PDFs de etiquetas do UpSeller (Etiqueta para Impressao)
    2. Move para pasta_entrada do usuario
    3. Processa com regras do Beka MKT (organizar SKU, agrupar loja, numerar, etc.)
    Executa em background com status consultavel via /api/upseller/imprimir/status.
    """
    user_id = int(get_jwt_identity())

    if _esta_atualizando_lojas(user_id):
        return jsonify({"sucesso": False, "mensagem": "Atualizando pedidos, aguarde."}), 409

    acao_em_execucao = _obter_acao_massa_em_andamento(user_id)
    if acao_em_execucao:
        return jsonify({
            "sucesso": False,
            "mensagem": (
                f"Processo em massa em andamento ({acao_em_execucao.get('acao', 'processando')}). "
                "Aguarde finalizar."
            )
        }), 409

    if hasattr(app, '_imprimir_em_andamento') and app._imprimir_em_andamento.get(user_id):
        return jsonify({"sucesso": False, "mensagem": "Geracao de etiquetas ja em andamento"}), 409

    data = request.get_json() or {}
    filtro_loja = (data.get("loja") or "").strip() or None
    lojas_lista_raw = data.get("lojas", [])
    lojas_lista, lojas_ignoradas = _sanitizar_lojas_selecionadas(user_id, lojas_lista_raw)

    if isinstance(lojas_lista_raw, list) and lojas_lista_raw and not lojas_lista:
        return jsonify({
            "sucesso": False,
            "mensagem": "Nenhuma loja valida selecionada para gerar etiquetas."
        }), 400
    # Se vier lista de lojas, prioriza SEMPRE execucao em lote
    # (ignora campo legado "loja" para evitar cair em fluxo individual).
    if lojas_lista:
        filtro_loja = None

    if filtro_loja:
        filtro_lista, _ = _sanitizar_lojas_selecionadas(user_id, [filtro_loja])
        if not filtro_lista:
            return jsonify({
                "sucesso": False,
                "mensagem": f"Loja invalida para gerar etiquetas: {filtro_loja}"
            }), 400
        filtro_loja = filtro_lista[0]

    # Pre-setar status ANTES da thread para evitar race condition no polling
    if not hasattr(app, '_imprimir_em_andamento'):
        app._imprimir_em_andamento = {}
    if not hasattr(app, '_imprimir_status'):
        app._imprimir_status = {}
    lock_ok, lock_atual = _iniciar_acao_massa(user_id, "imprimir")
    if not lock_ok:
        return jsonify({
            "sucesso": False,
            "mensagem": (
                f"Processo em massa em andamento ({(lock_atual or {}).get('acao', 'processando')}). "
                "Aguarde finalizar."
            )
        }), 409
    app._imprimir_em_andamento[user_id] = True
    app._imprimir_status[user_id] = {"etapa": "iniciando", "progresso": 5, "detalhes": "Iniciando..."}

    def _imprimir_pipeline():
        with app.app_context():
            try:
                user = db.session.get(User, user_id)
                if not user:
                    app._imprimir_status[user_id] = {"etapa": "erro", "progresso": 0, "detalhes": "Usuario nao encontrado"}
                    return

                pasta_entrada = user.get_pasta_entrada()
                pasta_lote = _criar_pasta_lote_upseller(pasta_entrada, prefixo="imprimir")
                download_dir = os.path.join(pasta_entrada, '_upseller_temp')
                os.makedirs(download_dir, exist_ok=True)

                # === Auto-reconexao com headless=False (SPA precisa de browser visivel) ===
                config = _get_or_create_upseller_config(user_id)
                scraper_vivo = _upseller_mgr.is_alive(user_id)

                # Para imprimir etiquetas, o UpSeller SPA precisa de browser VISIVEL.
                # Se o scraper esta headless ou morto, recriar com headless=False.
                precisa_recriar = False
                if not scraper_vivo:
                    precisa_recriar = True
                else:
                    scraper_atual = _upseller_mgr.get_scraper(user_id)
                    if scraper_atual and getattr(scraper_atual, 'headless', True):
                        precisa_recriar = True

                if precisa_recriar:
                    app._imprimir_status[user_id] = {"etapa": "reconectando", "progresso": 8, "detalhes": "Abrindo navegador para UpSeller..."}
                    try:
                        scraper = _upseller_mgr.criar_scraper(user_id, config.session_dir, download_dir, headless=False)
                        logado = _upseller_mgr._run_async(user_id, scraper._esta_logado())
                        if not logado:
                            app._imprimir_status[user_id] = {"etapa": "erro", "progresso": 0, "detalhes": "Sessao expirada. Clique 'Reconectar' para login manual."}
                            config.status_conexao = "desconectado"
                            db.session.commit()
                            return
                        config.status_conexao = "ok"
                        db.session.commit()
                    except Exception as e:
                        print(f"[Imprimir] Erro ao criar scraper visivel: {e}")
                        app._imprimir_status[user_id] = {"etapa": "erro", "progresso": 0, "detalhes": f"Erro ao abrir navegador: {e}"}
                        return

                scraper = _upseller_mgr.get_scraper(user_id)
                if not scraper:
                    app._imprimir_status[user_id] = {"etapa": "erro", "progresso": 0, "detalhes": "Scraper nao disponivel apos reconexao"}
                    return

                # Atualizar download_dir do scraper
                scraper.download_dir = download_dir

                # === Etapa 1: Baixar "Lista de Resumo" (XLSX) antes das etiquetas ===
                app._imprimir_status[user_id] = {
                    "etapa": "resumo",
                    "progresso": 12,
                    "detalhes": "Baixando Lista de Resumo (XLSX)..."
                }
                xlsx_paths = []
                lojas_falha = []
                grupos_marketplace = []
                if lojas_lista and not filtro_loja:
                    grupos_marketplace = _agrupar_lojas_por_marketplace(user_id, lojas_lista)
                    if not grupos_marketplace:
                        grupos_marketplace = [{"marketplace": "Shopee", "lojas": list(lojas_lista)}]
                def _xlsx_download_valido(path):
                    try:
                        if not path or not os.path.exists(path):
                            return False
                        checker = getattr(scraper, "_arquivo_tabulado_valido", None)
                        if callable(checker):
                            return bool(checker(path))
                        # Fallback simples: rejeita HTML.
                        with open(path, "rb") as f:
                            head = (f.read(512) or b"").lstrip().lower()
                        if head.startswith(b"<!doctype html") or head.startswith(b"<html"):
                            return False
                        return True
                    except Exception:
                        return False
                try:
                    if lojas_lista and not filtro_loja:
                        total_grupos = max(1, len(grupos_marketplace))
                        for idx_gp, gp in enumerate(grupos_marketplace, start=1):
                            lojas_gp = list(gp.get("lojas") or [])
                            mp_nome = (gp.get("marketplace") or "Shopee").strip() or "Shopee"
                            if not lojas_gp:
                                continue
                            app._imprimir_status[user_id] = {
                                "etapa": "resumo",
                                "progresso": 18,
                                "detalhes": (
                                    f"Lista de resumo ({idx_gp}/{total_grupos}) - "
                                    f"{mp_nome}: {len(lojas_gp)} loja(s)"
                                )
                            }
                            xlsx_lote = _upseller_mgr._run_async(
                                user_id, scraper.baixar_lista_resumo(filtro_loja=lojas_gp)
                            ) or []
                            if isinstance(xlsx_lote, str):
                                xlsx_lote = [xlsx_lote]
                            salvou_gp = False
                            for xp in xlsx_lote:
                                if _xlsx_download_valido(xp):
                                    xlsx_paths.append(xp)
                                    salvou_gp = True
                                elif xp:
                                    print(f"[Imprimir] Aviso: arquivo de resumo invalido (ignorado): {xp}")

                            if (not salvou_gp) and len(lojas_gp) > 1:
                                for loja_nome in lojas_gp:
                                    try:
                                        xlsx_loja = _upseller_mgr._run_async(
                                            user_id, scraper.baixar_lista_resumo(filtro_loja=loja_nome)
                                        ) or []
                                        if isinstance(xlsx_loja, str):
                                            xlsx_loja = [xlsx_loja]
                                        salvou_loja = False
                                        for xp in xlsx_loja:
                                            if _xlsx_download_valido(xp):
                                                xlsx_paths.append(xp)
                                                salvou_loja = True
                                        if not salvou_loja:
                                            lojas_falha.append({
                                                "loja": loja_nome,
                                                "marketplace": mp_nome,
                                                "etapa": "lista_resumo",
                                                "motivo": "sem_xlsx_valido",
                                            })
                                    except Exception as e_xlsx_loja:
                                        lojas_falha.append({
                                            "loja": loja_nome,
                                            "marketplace": mp_nome,
                                            "etapa": "lista_resumo",
                                            "motivo": str(e_xlsx_loja),
                                        })
                    else:
                        xlsx_unico = _upseller_mgr._run_async(
                            user_id, scraper.baixar_lista_resumo(filtro_loja=filtro_loja)
                        ) or []
                        if isinstance(xlsx_unico, str):
                            xlsx_unico = [xlsx_unico]
                        for xp in xlsx_unico:
                            if _xlsx_download_valido(xp):
                                xlsx_paths.append(xp)
                            elif xp:
                                print(f"[Imprimir] Aviso: arquivo de resumo invalido (ignorado): {xp}")
                except Exception as e_resumo:
                    print(f"[Imprimir] Aviso: falha ao baixar Lista de Resumo (XLSX): {e_resumo}")

                # Deduplicar mantendo ordem.
                vistos_xlsx = set()
                xlsx_paths_filtrados = []
                for xp in xlsx_paths:
                    if not xp or xp in vistos_xlsx:
                        continue
                    vistos_xlsx.add(xp)
                    xlsx_paths_filtrados.append(xp)
                xlsx_paths = xlsx_paths_filtrados

                # === Etapa 2: Baixar etiquetas do UpSeller ===
                app._imprimir_status[user_id] = {
                    "etapa": "baixando",
                    "progresso": 28,
                    "detalhes": "Baixando etiquetas do UpSeller..."
                }
                pdfs = []
                try:
                    if lojas_lista and not filtro_loja:
                        total_grupos = max(1, len(grupos_marketplace))
                        for idx_gp, gp in enumerate(grupos_marketplace, start=1):
                            lojas_gp = list(gp.get("lojas") or [])
                            mp_nome = (gp.get("marketplace") or "Shopee").strip() or "Shopee"
                            if not lojas_gp:
                                continue
                            app._imprimir_status[user_id] = {
                                "etapa": "baixando", "progresso": 32,
                                "detalhes": (
                                    f"Baixando etiquetas ({idx_gp}/{total_grupos}) - "
                                    f"{mp_nome}: {len(lojas_gp)} loja(s)"
                                )
                            }
                            pdfs_lote = _upseller_mgr._run_async(
                                user_id, scraper.baixar_etiquetas(filtro_loja=lojas_gp)
                            ) or []
                            if pdfs_lote:
                                pdfs.extend(list(pdfs_lote))
                                continue

                            if len(lojas_gp) > 1:
                                for loja_nome in lojas_gp:
                                    try:
                                        pdfs_loja = _upseller_mgr._run_async(
                                            user_id, scraper.baixar_etiquetas(filtro_loja=loja_nome)
                                        ) or []
                                        if pdfs_loja:
                                            pdfs.extend(list(pdfs_loja))
                                        else:
                                            lojas_falha.append({
                                                "loja": loja_nome,
                                                "marketplace": mp_nome,
                                                "etapa": "gerar_etiquetas",
                                                "motivo": "sem_pdf_gerado",
                                            })
                                    except Exception as e_loja_pdf:
                                        lojas_falha.append({
                                            "loja": loja_nome,
                                            "marketplace": mp_nome,
                                            "etapa": "gerar_etiquetas",
                                            "motivo": str(e_loja_pdf),
                                        })
                            else:
                                lojas_falha.append({
                                    "loja": lojas_gp[0],
                                    "marketplace": mp_nome,
                                    "etapa": "gerar_etiquetas",
                                    "motivo": "sem_pdf_gerado",
                                })
                    else:
                        pdfs = _upseller_mgr._run_async(user_id, scraper.baixar_etiquetas(filtro_loja=filtro_loja))
                except Exception as e:
                    print(f"[Imprimir] Erro ao baixar etiquetas: {e}")
                    if lojas_lista and not filtro_loja:
                        # Fallback por loja mesmo com erro de lote, respeitando marketplace.
                        for gp in grupos_marketplace:
                            lojas_gp = list(gp.get("lojas") or [])
                            mp_nome = (gp.get("marketplace") or "Shopee").strip() or "Shopee"
                            for loja_nome in lojas_gp:
                                try:
                                    pdfs_loja = _upseller_mgr._run_async(
                                        user_id, scraper.baixar_etiquetas(filtro_loja=loja_nome)
                                    ) or []
                                    if pdfs_loja:
                                        pdfs.extend(list(pdfs_loja))
                                    else:
                                        lojas_falha.append({
                                            "loja": loja_nome,
                                            "marketplace": mp_nome,
                                            "etapa": "gerar_etiquetas",
                                            "motivo": "sem_pdf_gerado",
                                        })
                                except Exception as e_loja_pdf:
                                    lojas_falha.append({
                                        "loja": loja_nome,
                                        "marketplace": mp_nome,
                                        "etapa": "gerar_etiquetas",
                                        "motivo": str(e_loja_pdf),
                                    })
                    else:
                        app._imprimir_status[user_id] = {"etapa": "erro", "progresso": 0, "detalhes": f"Erro ao baixar: {e}"}
                        return

                if not pdfs:
                    detalhe_falhas = ""
                    if lojas_falha:
                        amostra = " | ".join(
                            f"{x.get('loja', '?')}: {x.get('motivo', 'erro')}"
                            for x in lojas_falha[:5]
                        )
                        detalhe_falhas = f" Falhas: {amostra}"
                    app._imprimir_status[user_id] = {
                        "etapa": "erro", "progresso": 0,
                        "detalhes": (
                            "Nenhuma etiqueta pendente no UpSeller. "
                            "Verifique a aba 'Etiqueta para Impressão' no UpSeller." + detalhe_falhas
                        ),
                        "lojas_falha": lojas_falha[:20],
                    }
                    return

                # === Etapa 2.5: Fallback antigo de extracao caso Lista de Resumo falhe ===
                if not xlsx_paths:
                    app._imprimir_status[user_id] = {
                        "etapa": "extraindo",
                        "progresso": 50,
                        "detalhes": "Lista de Resumo nao retornou XLSX. Tentando extracao de fallback..."
                    }
                    try:
                        if lojas_lista and not filtro_loja:
                            app._imprimir_status[user_id] = {
                                "etapa": "extraindo", "progresso": 54,
                                "detalhes": f"Fallback XLSX em lote ({len(lojas_lista)} lojas)..."
                            }
                            xp = _upseller_mgr._run_async(
                                user_id,
                                scraper.extrair_dados_pedidos_em_impressao(filtro_loja=lojas_lista)
                            ) or ""
                            if not xp:
                                xp = _upseller_mgr._run_async(
                                    user_id,
                                    scraper.extrair_dados_pedidos(status_filter="para_imprimir", filtro_loja=lojas_lista)
                                ) or ""
                            if not xp:
                                xp = _upseller_mgr._run_async(
                                    user_id,
                                    scraper.extrair_dados_pedidos(status_filter="para_enviar", filtro_loja=lojas_lista)
                                ) or ""
                            if _xlsx_download_valido(xp):
                                xlsx_paths.append(xp)
                        else:
                            xp = _upseller_mgr._run_async(
                                user_id,
                                scraper.extrair_dados_pedidos_em_impressao(filtro_loja=filtro_loja)
                            ) or ""
                            if not xp:
                                xp = _upseller_mgr._run_async(
                                    user_id,
                                    scraper.extrair_dados_pedidos(status_filter="para_imprimir", filtro_loja=filtro_loja)
                                ) or ""
                            if not xp:
                                xp = _upseller_mgr._run_async(
                                    user_id,
                                    scraper.extrair_dados_pedidos(status_filter="para_enviar", filtro_loja=filtro_loja)
                                ) or ""
                            if _xlsx_download_valido(xp):
                                xlsx_paths.append(xp)
                    except Exception as e_xlsx:
                        print(f"[Imprimir] Aviso: falha no fallback de XLSX: {e_xlsx}")

                # Deduplicar novamente apos fallback.
                if xlsx_paths:
                    vistos_xlsx2 = set()
                    dedup_xlsx = []
                    for xp in xlsx_paths:
                        if not xp or xp in vistos_xlsx2:
                            continue
                        vistos_xlsx2.add(xp)
                        dedup_xlsx.append(xp)
                    xlsx_paths = dedup_xlsx

                xlsx_path = xlsx_paths[0] if xlsx_paths else ""
                xlsx_extra = xlsx_paths[1:] if len(xlsx_paths) > 1 else []

                if xlsx_path and os.path.exists(xlsx_path):
                    app._imprimir_status[user_id] = {
                        "etapa": "movendo",
                        "progresso": 58,
                        "detalhes": f"{len(pdfs)} PDF(s) + {len(xlsx_paths)} XLSX, movendo..."
                    }
                else:
                    app._imprimir_status[user_id] = {
                        "etapa": "movendo",
                        "progresso": 58,
                        "detalhes": f"{len(pdfs)} PDF(s) baixado(s), movendo..."
                    }

                # === Etapa 3: Mover arquivos para lote isolado da pasta_entrada ===
                resultado = {"pdfs": pdfs, "xmls": [], "xlsx": xlsx_path or "", "xlsx_extra": xlsx_extra}
                resumo = scraper.mover_para_pasta_entrada(resultado, pasta_lote)
                if resumo.get("pdfs_movidos", 0) <= 0:
                    app._imprimir_status[user_id] = {
                        "etapa": "erro",
                        "progresso": 0,
                        "detalhes": (
                            "Nenhum PDF foi movido para o lote atual. "
                            "Processamento cancelado para preservar o ultimo resultado valido."
                        ),
                        "processado": False,
                    }
                    return

                app._imprimir_status[user_id] = {
                    "etapa": "processando", "progresso": 65,
                    "detalhes": f"{resumo['pdfs_movidos']} PDF(s) na pasta, processando etiquetas..."
                }

                # === Etapa 3: Processar etiquetas com regras do Beka MKT ===
                try:
                    proc_result = _executar_processamento(
                        user_id,
                        sem_recorte=True,
                        resumo_sku_somente=False,
                        pasta_entrada_override=pasta_lote
                    )
                    if not proc_result or not proc_result.get("ok"):
                        raise RuntimeError(
                            (proc_result or {}).get("erro") or
                            "Processamento concluido sem gerar resultado valido."
                        )
                    app._imprimir_status[user_id] = {
                        "etapa": "concluido", "progresso": 100,
                        "detalhes": (
                            f"Concluido! {resumo['pdfs_movidos']} PDF(s) processado(s) | "
                            f"{proc_result.get('total_etiquetas', 0)} etiquetas em "
                            f"{proc_result.get('total_lojas', 0)} loja(s)"
                        ),
                        "processado": True,
                    }

                    # Auto-disparo opcional WhatsApp (fila persistente)
                    try:
                        auto_res = _enfileirar_envio_whatsapp_resultado(
                            user_id=user_id,
                            origem="auto",
                            respeitar_toggle_auto=True,
                        )
                        if auto_res.get("ok"):
                            app._imprimir_status[user_id]["auto_whatsapp"] = {
                                "enfileirados": auto_res.get("total_entregas", 0),
                                "batch_id": auto_res.get("batch_id", ""),
                            }
                            app._imprimir_status[user_id]["detalhes"] += f" | WhatsApp fila: {auto_res.get('total_entregas', 0)}"
                        elif not auto_res.get("ignorado"):
                            app._imprimir_status[user_id]["aviso_whatsapp"] = auto_res.get("erro", "Falha ao enfileirar WhatsApp")
                    except Exception as e_auto:
                        app._imprimir_status[user_id]["aviso_whatsapp"] = f"Falha no auto-envio WhatsApp: {e_auto}"
                    # Resumo geral diario: enviar JPEG parcial + acumular
                    try:
                        _enviar_resumo_geral_whatsapp(user_id, user.get_pasta_saida())
                    except Exception as e_rg:
                        print(f"[ResumoGeral] Erro: {e_rg}")

                    if lojas_falha:
                        app._imprimir_status[user_id]["aviso_falhas"] = (
                            f"{len(lojas_falha)} loja(s) com erro foram ignoradas e o restante foi processado."
                        )
                        app._imprimir_status[user_id]["lojas_falha"] = lojas_falha[:20]
                except Exception as e:
                    print(f"[Imprimir] Erro no processamento: {e}")
                    import traceback
                    traceback.print_exc()
                    app._imprimir_status[user_id] = {
                        "etapa": "parcial", "progresso": 80,
                        "detalhes": f"Download OK ({resumo['pdfs_movidos']} PDFs), erro no processamento: {e}",
                        "processado": False,
                    }
                    if lojas_falha:
                        app._imprimir_status[user_id]["aviso_falhas"] = (
                            f"{len(lojas_falha)} loja(s) com erro foram ignoradas durante o download."
                        )
                        app._imprimir_status[user_id]["lojas_falha"] = lojas_falha[:20]

                # Atualizar snapshot em background para liberar resultado imediatamente.
                _disparar_atualizacao_lojas_background(user_id, status_attr="_imprimir_status")

                # Limpar pasta temp
                try:
                    import shutil
                    shutil.rmtree(download_dir, ignore_errors=True)
                except:
                    pass

            except Exception as e:
                print(f"[Imprimir] Erro geral: {e}")
                import traceback
                traceback.print_exc()
                app._imprimir_status[user_id] = {"etapa": "erro", "progresso": 0, "detalhes": str(e)}
            finally:
                app._imprimir_em_andamento[user_id] = False
                _finalizar_acao_massa(user_id, "imprimir")

    thread = threading.Thread(target=_imprimir_pipeline, daemon=True)
    thread.start()
    payload = {"sucesso": True, "mensagem": "Geracao de etiquetas iniciada", "async": True}
    if lojas_ignoradas:
        payload["lojas_ignoradas"] = lojas_ignoradas
    return jsonify(payload)


@app.route('/api/upseller/imprimir/status', methods=['GET'])
@jwt_required()
def api_upseller_imprimir_status():
    """Retorna status do pipeline 'Gerar Etiquetas' em andamento."""
    user_id = int(get_jwt_identity())
    if hasattr(app, '_imprimir_status') and user_id in app._imprimir_status:
        status = dict(app._imprimir_status[user_id])
        em_andamento = hasattr(app, '_imprimir_em_andamento') and app._imprimir_em_andamento.get(user_id, False)
        status["em_andamento"] = em_andamento
        status["atualizando_pedidos"] = _esta_atualizando_lojas(user_id) or bool(status.get("atualizando_pedidos"))
        return jsonify(status)
    return jsonify({
        "etapa": "idle",
        "progresso": 0,
        "detalhes": "",
        "em_andamento": False,
        "atualizando_pedidos": _esta_atualizando_lojas(user_id),
    })


# ----------------------------------------------------------------
# ENDPOINTS AIOS: movidos para aios_routes.py (Blueprint separado)
# ----------------------------------------------------------------


# ----------------------------------------------------------------
# ENDPOINTS - WHATSAPP
# ----------------------------------------------------------------

@app.route('/api/whatsapp/status', methods=['GET'])
@jwt_required()
def api_whatsapp_status():
    """Verifica status da conexao WhatsApp."""
    _garantir_baileys_rodando(motivo="status")
    wa = WhatsAppService()
    return jsonify(wa.verificar_conexao())


@app.route('/api/whatsapp/qr', methods=['GET'])
@jwt_required()
def api_whatsapp_qr():
    """Retorna QR code para escanear."""
    _garantir_baileys_rodando(motivo="qr")
    wa = WhatsAppService()
    # Iniciar sessao se necessario
    wa.iniciar_sessao()
    return jsonify(wa.get_qr_code())


@app.route('/api/whatsapp/config', methods=['GET'])
@jwt_required()
def api_whatsapp_config_get():
    """Retorna configuracoes do modulo WhatsApp do usuario."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    return jsonify({
        "auto_send_whatsapp": bool(getattr(user, "auto_send_whatsapp", False)) if user else False
    })


@app.route('/api/whatsapp/config', methods=['POST'])
@jwt_required()
def api_whatsapp_config_set():
    """Atualiza configuracoes do modulo WhatsApp do usuario."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    data = request.get_json(force=True, silent=True) or {}
    user.auto_send_whatsapp = _to_bool(data.get("auto_send_whatsapp"), False)
    db.session.commit()
    return jsonify({
        "mensagem": "Configuracao salva",
        "auto_send_whatsapp": bool(user.auto_send_whatsapp),
    })


@app.route('/api/whatsapp/contatos', methods=['GET'])
@jwt_required()
def api_whatsapp_contatos_listar():
    """Lista contatos WhatsApp do usuario."""
    user_id = int(get_jwt_identity())
    contatos = WhatsAppContact.query.filter_by(user_id=user_id).order_by(WhatsAppContact.loja_nome).all()
    return jsonify([c.to_dict() for c in contatos])


@app.route('/api/whatsapp/contatos', methods=['POST'])
@jwt_required()
def api_whatsapp_contatos_salvar():
    """Cadastra ou atualiza contato WhatsApp (com selecao de lojas/grupos)."""
    user_id = int(get_jwt_identity())
    data = request.get_json(force=True, silent=True) or {}

    telefone = str(data.get("telefone", "") or "").strip()

    # Se tem id, buscar contato existente primeiro (permite atualizar so horarios)
    contato_id = data.get("id")
    contato = None
    if contato_id:
        contato = WhatsAppContact.query.filter_by(id=int(contato_id), user_id=user_id).first()

    if not telefone and not contato:
        return jsonify({"erro": "Telefone e obrigatorio"}), 400

    if not telefone and contato:
        telefone = contato.telefone

    lojas_enviado = "lojas" in data
    grupos_enviado = "grupos" in data
    lote_ids_enviado = "lote_ids" in data or "lote_ids_json" in data

    lojas = data.get("lojas", []) or []
    grupos = data.get("grupos", []) or []
    lojas = [str(x).strip() for x in lojas if str(x).strip()]
    grupos = [str(x).strip() for x in grupos if str(x).strip()]
    lote_ids = _to_int_list(data.get("lote_ids"))
    if lote_ids is None and "lote_ids_json" in data:
        lote_ids = _to_int_list(data.get("lote_ids_json"))

    loja_cnpj = str(data.get("loja_cnpj", "") or "").strip()
    loja_nome = str(data.get("loja_nome", "") or "").strip()

    if not loja_nome and lojas:
        loja_nome = lojas[0]
    if not loja_cnpj:
        # Mantem compatibilidade com chave legada e evita campo vazio.
        loja_cnpj = "ALVO_CUSTOM"

    # Buscar duplicata APENAS quando editando (id fornecido mas nao encontrado).
    # Quando criando novo contato (sem id), SEMPRE criar — permite mesmo
    # telefone com lojas/horarios diferentes (ex: Shopee + Shein separados).
    if not contato and data.get('id'):
        contato = WhatsAppContact.query.filter_by(
            user_id=user_id,
            loja_cnpj=loja_cnpj,
            telefone=telefone
        ).first()

    # Horarios multiplos: [{"dias":["seg","ter"],"horas":["07:00","11:30"]}, ...]
    horarios = data.get("horarios", None)

    if contato:
        # Atualizar existente — so atualiza campos que foram explicitamente enviados
        if loja_cnpj and loja_cnpj != "ALVO_CUSTOM":
            contato.loja_cnpj = loja_cnpj
        if loja_nome:
            contato.loja_nome = loja_nome
        if "nome_contato" in data:
            contato.nome_contato = data.get("nome_contato", contato.nome_contato)
        if "ativo" in data:
            contato.ativo = _to_bool(data.get("ativo"), contato.ativo)
        if lojas_enviado:
            contato.lojas_json = json.dumps(lojas, ensure_ascii=False)
        if grupos_enviado:
            contato.grupos_json = json.dumps(grupos, ensure_ascii=False)
        if horarios is not None:
            contato.horarios_json = json.dumps(horarios, ensure_ascii=False)
        if lote_ids_enviado:
            contato.lote_ids_json = json.dumps(lote_ids or [], ensure_ascii=False)
        if "resumo_geral" in data:
            contato.resumo_geral = _to_bool(data.get("resumo_geral"), False)
        if "agendamento_ativo" in data:
            contato.agendamento_ativo = _to_bool(data.get("agendamento_ativo"), True)
    else:
        # Criar novo
        contato = WhatsAppContact(
            user_id=user_id,
            loja_cnpj=loja_cnpj,
            loja_nome=loja_nome,
            telefone=telefone,
            nome_contato=data.get("nome_contato", ""),
            lojas_json=json.dumps(lojas, ensure_ascii=False),
            grupos_json=json.dumps(grupos, ensure_ascii=False),
            horarios_json=json.dumps(horarios or [], ensure_ascii=False),
            lote_ids_json=json.dumps(lote_ids or [], ensure_ascii=False),
            ativo=_to_bool(data.get("ativo"), True),
            resumo_geral=_to_bool(data.get("resumo_geral"), False),
        )
        db.session.add(contato)

    db.session.commit()

    # Sincronizar job APScheduler do contato
    try:
        beka_scheduler.registrar_job_contato(contato, 'whatsapp')
    except Exception:
        pass

    return jsonify({"mensagem": "Contato salvo", **contato.to_dict()})


@app.route('/api/whatsapp/contatos/<int:contato_id>', methods=['DELETE'])
@jwt_required()
def api_whatsapp_contatos_remover(contato_id):
    """Remove um contato WhatsApp."""
    user_id = int(get_jwt_identity())
    contato = WhatsAppContact.query.filter_by(id=contato_id, user_id=user_id).first()
    if not contato:
        return jsonify({"erro": "Contato nao encontrado"}), 404
    try:
        beka_scheduler.remover_job_contato(contato_id, 'whatsapp')
    except Exception:
        pass
    db.session.delete(contato)
    db.session.commit()
    return jsonify({"mensagem": "Contato removido"})


@app.route('/api/contatos/<string:tipo>/<int:contato_id>/toggle-agendamento', methods=['POST'])
@jwt_required()
def api_toggle_agendamento(tipo, contato_id):
    """Ativa/desativa agendamento individual de um contato (sem apagar horarios)."""
    user_id = int(get_jwt_identity())
    if tipo == 'whatsapp':
        contato = WhatsAppContact.query.filter_by(id=contato_id, user_id=user_id).first()
    elif tipo == 'email':
        contato = EmailContact.query.filter_by(id=contato_id, user_id=user_id).first()
    else:
        return jsonify({"erro": "Tipo invalido"}), 400
    if not contato:
        return jsonify({"erro": "Contato nao encontrado"}), 404

    atual = getattr(contato, 'agendamento_ativo', True)
    if atual is None:
        atual = True
    contato.agendamento_ativo = not atual
    db.session.commit()

    # Sincronizar scheduler
    try:
        if contato.agendamento_ativo:
            beka_scheduler.registrar_job_contato(contato, tipo)
        else:
            beka_scheduler.remover_job_contato(contato_id, tipo)
    except Exception:
        pass

    status = "ativado" if contato.agendamento_ativo else "desativado"
    return jsonify({"mensagem": f"Agendamento individual {status}", "agendamento_ativo": contato.agendamento_ativo})


@app.route('/api/whatsapp/enviar-teste', methods=['POST'])
@jwt_required()
def api_whatsapp_enviar_teste():
    """Envia mensagem de teste para um numero."""
    data = request.get_json(force=True, silent=True) or {}
    telefone = data.get("telefone", "")
    if not telefone:
        return jsonify({"success": False, "error": "Telefone obrigatorio"}), 400
    try:
        _garantir_baileys_rodando(motivo="teste")
        wa = WhatsAppService()
        status = wa.verificar_conexao()
        if not status.get("connected"):
            return jsonify({
                "success": False,
                "error": "WhatsApp desconectado. Escaneie o QR code para conectar.",
                "status": status,
            }), 400
        resultado = wa.enviar_mensagem(telefone, "Teste Beka MKT - Conexao WhatsApp OK!")
        if resultado.get("success"):
            return jsonify({
                "success": True,
                "message": "Mensagem de teste enviada com sucesso",
                "telefone": telefone,
                "messageId": resultado.get("messageId", ""),
            })
        return jsonify({
            "success": False,
            "error": resultado.get("error", "Falha ao enviar mensagem de teste"),
            "telefone": telefone,
        }), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/whatsapp/enviar-lote', methods=['POST'])
@jwt_required()
def api_whatsapp_enviar_lote():
    """Enfileira envio de arquivos (PDF/IMG/XLSX) para contatos cadastrados manualmente."""
    user_id = int(get_jwt_identity())
    enq = _enfileirar_envio_whatsapp_resultado(
        user_id=user_id,
        origem="manual",
        respeitar_toggle_auto=False,
    )
    if not enq.get("ok"):
        return jsonify({
            "erro": enq.get("erro", "Falha ao enfileirar envio"),
            "diagnostico": enq.get("diagnostico", {}),
            "ignorado": bool(enq.get("ignorado")),
        }), 400

    if not hasattr(app, "_whatsapp_send_status"):
        app._whatsapp_send_status = {}
    app._whatsapp_send_status[user_id] = {
        "batch_id": enq.get("batch_id"),
        "etapa": "enfileirado",
        "em_andamento": True,
        "progresso": 0,
        "detalhes": f"Fila criada com {enq.get('total_entregas', 0)} envio(s).",
        "total": enq.get("total_entregas", 0),
        "enviados": 0,
        "erros": 0,
        "diagnostico": enq.get("diagnostico", {}),
    }
    return jsonify({
        "mensagem": f"Fila WhatsApp criada com {enq.get('total_entregas', 0)} envio(s).",
        "batch_id": enq.get("batch_id"),
        "total_entregas": enq.get("total_entregas", 0),
        "diagnostico": enq.get("diagnostico", {}),
    })


@app.route('/api/whatsapp/enviar-lote/status', methods=['GET'])
@jwt_required()
def api_whatsapp_enviar_lote_status():
    """Status do envio manual de WhatsApp em background."""
    user_id = int(get_jwt_identity())
    status = {}
    if hasattr(app, "_whatsapp_send_status"):
        status = app._whatsapp_send_status.get(user_id, {}) or {}
    batch_id = status.get("batch_id", "") if status else ""
    if batch_id:
        fila = _status_batch_fila_whatsapp(user_id, batch_id)
        fila["diagnostico"] = status.get("diagnostico", {})
        app._whatsapp_send_status[user_id] = {**status, **fila}
        return jsonify({**status, **fila})

    # Sem batch ativo: retorna panorama rapido da fila do usuario.
    totais = defaultdict(int)
    for st in db.session.query(WhatsAppQueueItem.status, db.func.count(WhatsAppQueueItem.id)).filter(
        WhatsAppQueueItem.user_id == user_id
    ).group_by(WhatsAppQueueItem.status):
        totais[st[0] or "pending"] = st[1]
    total_fila = sum(totais.values())
    return jsonify({
        "etapa": "idle",
        "em_andamento": False,
        "progresso": 0,
        "detalhes": "Sem envio manual ativo",
        "total": total_fila,
        "enviados": totais.get("sent", 0),
        "erros": totais.get("dead", 0),
        "retry": totais.get("retry", 0),
        "pendentes": totais.get("pending", 0) + totais.get("sending", 0),
    })


@app.route('/api/whatsapp/fila/status', methods=['GET'])
@jwt_required()
def api_whatsapp_fila_status():
    """Resumo da fila persistente de WhatsApp."""
    user_id = int(get_jwt_identity())
    rows = db.session.query(WhatsAppQueueItem.status, db.func.count(WhatsAppQueueItem.id)).filter(
        WhatsAppQueueItem.user_id == user_id
    ).group_by(WhatsAppQueueItem.status).all()
    counts = defaultdict(int)
    for st, qtd in rows:
        counts[st or "pending"] = qtd

    ultimo_batch = db.session.query(WhatsAppQueueItem.batch_id).filter(
        WhatsAppQueueItem.user_id == user_id
    ).order_by(WhatsAppQueueItem.id.desc()).first()
    batch_id = ultimo_batch[0] if ultimo_batch else ""
    batch = _status_batch_fila_whatsapp(user_id, batch_id) if batch_id else {}

    return jsonify({
        "counts": {
            "pending": counts.get("pending", 0),
            "retry": counts.get("retry", 0),
            "sending": counts.get("sending", 0),
            "sent": counts.get("sent", 0),
            "dead": counts.get("dead", 0),
        },
        "batch_atual": batch,
    })


# ----------------------------------------------------------------
# ENDPOINTS - EMAIL (contatos + envio de etiquetas)
# ----------------------------------------------------------------

def _smtp_campos_usuario(user):
    if not user:
        return {
            "smtp_host": "",
            "smtp_port": 587,
            "smtp_user": "",
            "smtp_from": "",
            "smtp_pass_salva": False,
        }
    try:
        smtp_port = int(getattr(user, "smtp_port", 587) or 587)
    except Exception:
        smtp_port = 587
    return {
        "smtp_host": (getattr(user, "smtp_host", "") or "").strip(),
        "smtp_port": smtp_port,
        "smtp_user": (getattr(user, "smtp_user", "") or "").strip(),
        "smtp_from": (getattr(user, "smtp_from", "") or "").strip(),
        "smtp_pass_salva": bool((getattr(user, "smtp_pass_enc", "") or "").strip()),
    }


def _smtp_config_usuario(user):
    campos = _smtp_campos_usuario(user)
    smtp_pass = ""
    pass_enc = (getattr(user, "smtp_pass_enc", "") or "").strip() if user else ""
    if pass_enc:
        try:
            smtp_pass = decrypt_value(pass_enc)
        except Exception:
            smtp_pass = ""
    cfg = {
        "host": campos["smtp_host"],
        "port": campos["smtp_port"],
        "user": campos["smtp_user"],
        "password": smtp_pass,
        "from_addr": campos["smtp_from"] or campos["smtp_user"],
    }
    if smtp_configurado(cfg):
        return cfg
    return None


def _smtp_config_resolver(user):
    cfg_user = _smtp_config_usuario(user)
    if cfg_user:
        return cfg_user, "usuario"
    cfg_env = get_smtp_config()
    if cfg_env:
        return cfg_env, "ambiente"
    return None, "nao_configurado"


@app.route('/api/email/status', methods=['GET'])
@jwt_required()
def api_email_status():
    """Retorna se SMTP esta configurado."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    smtp_cfg, smtp_origem = _smtp_config_resolver(user)
    smtp_campos = _smtp_campos_usuario(user)
    return jsonify({
        "configurado": bool(smtp_cfg),
        "smtp_origem": smtp_origem,
        "email_remetente": (getattr(user, "email_remetente", "") or "").strip() if user else "",
        "nome_remetente": (getattr(user, "nome_remetente", "") or "").strip() if user else "",
        **smtp_campos,
    })


@app.route('/api/email/config', methods=['GET'])
@jwt_required()
def api_email_config_get():
    """Retorna configuracao de remetente do modulo de email."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    smtp_cfg, smtp_origem = _smtp_config_resolver(user)
    smtp_campos = _smtp_campos_usuario(user)
    return jsonify({
        "configurado": bool(smtp_cfg),
        "smtp_origem": smtp_origem,
        "email_remetente": (getattr(user, "email_remetente", "") or "").strip() if user else "",
        "nome_remetente": (getattr(user, "nome_remetente", "") or "").strip() if user else "",
        **smtp_campos,
    })


@app.route('/api/email/config', methods=['POST'])
@jwt_required()
def api_email_config_set():
    """Atualiza configuracao de remetente do modulo de email."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user:
        return jsonify({"erro": "Usuario nao encontrado"}), 404
    data = request.get_json(force=True, silent=True) or {}
    email_rem = str(data.get("email_remetente", "") or "").strip()
    nome_rem = str(data.get("nome_remetente", "") or "").strip()
    if not email_rem:
        return jsonify({"erro": "Email do remetente e obrigatorio"}), 400
    if "@" not in email_rem:
        return jsonify({"erro": "Email do remetente invalido"}), 400
    # Se o email remetente mudou, resetar TODA a config SMTP
    old_rem = (getattr(user, "email_remetente", "") or "").strip().lower()
    email_mudou = old_rem and old_rem != email_rem.lower()

    user.email_remetente = email_rem
    user.nome_remetente = nome_rem

    # SMTP por usuario (opcional no cadastro, obrigatorio para envio real)
    smtp_host_in = data.get("smtp_host", None)
    smtp_port_in = data.get("smtp_port", None)
    smtp_user_in = data.get("smtp_user", None)
    smtp_from_in = data.get("smtp_from", None)
    smtp_pass_in = data.get("smtp_pass", None)
    smtp_limpar_senha = _to_bool(data.get("smtp_limpar_senha"), False)

    # Se email mudou, auto-detectar SMTP pelo dominio e limpar senha antiga
    if email_mudou:
        domain = email_rem.split("@")[-1].lower() if "@" in email_rem else ""
        smtp_hosts = {
            "gmail.com": "smtp.gmail.com",
            "hotmail.com": "smtp-mail.outlook.com",
            "outlook.com": "smtp-mail.outlook.com",
            "live.com": "smtp-mail.outlook.com",
            "yahoo.com": "smtp.mail.yahoo.com",
            "yahoo.com.br": "smtp.mail.yahoo.com",
        }
        user.smtp_host = smtp_hosts.get(domain, f"smtp.{domain}")
        user.smtp_port = 587
        user.smtp_user = email_rem
        user.smtp_from = email_rem
        user.smtp_pass_enc = ""  # Limpar senha antiga — nao serve mais
    else:
        # Manter logica normal se email nao mudou
        if smtp_host_in is not None:
            user.smtp_host = str(smtp_host_in or "").strip()
        elif not (getattr(user, "smtp_host", "") or "").strip():
            user.smtp_host = "smtp.gmail.com"

        if smtp_port_in is not None:
            try:
                smtp_port = int(str(smtp_port_in).strip() or "587")
                if smtp_port <= 0 or smtp_port > 65535:
                    raise ValueError()
                user.smtp_port = smtp_port
            except Exception:
                return jsonify({"erro": "Porta SMTP invalida"}), 400
        elif not getattr(user, "smtp_port", None):
            user.smtp_port = 587

        if smtp_user_in is not None:
            user.smtp_user = str(smtp_user_in or "").strip()
        elif not (getattr(user, "smtp_user", "") or "").strip():
            user.smtp_user = email_rem

        if smtp_from_in is not None:
            user.smtp_from = str(smtp_from_in or "").strip()
        elif not (getattr(user, "smtp_from", "") or "").strip():
            user.smtp_from = email_rem

    if smtp_pass_in is not None:
        smtp_pass_txt = str(smtp_pass_in or "").strip()
        if smtp_pass_txt:
            try:
                user.smtp_pass_enc = encrypt_value(smtp_pass_txt)
            except Exception:
                return jsonify({"erro": "Nao foi possivel salvar a senha SMTP"}), 500
        elif smtp_limpar_senha:
            user.smtp_pass_enc = ""

    db.session.commit()
    smtp_cfg, smtp_origem = _smtp_config_resolver(user)
    smtp_campos = _smtp_campos_usuario(user)
    return jsonify({
        "mensagem": "Configuracao de email salva",
        "email_remetente": user.email_remetente or "",
        "nome_remetente": user.nome_remetente or "",
        "configurado": bool(smtp_cfg),
        "smtp_origem": smtp_origem,
        **smtp_campos,
    })


@app.route('/api/email/validar', methods=['POST'])
@jwt_required()
def api_email_validar():
    """Valida login SMTP atual (usuario ou ambiente)."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user:
        return jsonify({"sucesso": False, "erro": "Usuario nao encontrado"}), 404

    smtp_cfg, smtp_origem = _smtp_config_resolver(user)
    if not smtp_cfg:
        return jsonify({"sucesso": False, "erro": "SMTP nao configurado"}), 400

    host = str(smtp_cfg.get("host") or "").strip()
    port = int(smtp_cfg.get("port") or 587)
    user_smtp = str(smtp_cfg.get("user") or "").strip()
    pass_smtp = str(smtp_cfg.get("password") or "").strip()

    if not host or not user_smtp or not pass_smtp:
        return jsonify({"sucesso": False, "erro": "Configuracao SMTP incompleta"}), 400

    try:
        if port == 465:
            server = smtplib.SMTP_SSL(host, port, timeout=20)
            server.login(user_smtp, pass_smtp)
            server.quit()
        else:
            server = smtplib.SMTP(host, port, timeout=20)
            server.ehlo()
            if port in (587, 25, 2525):
                server.starttls()
                server.ehlo()
            server.login(user_smtp, pass_smtp)
            server.quit()
    except Exception as e:
        from email_utils import _traduzir_erro_smtp
        msg = _traduzir_erro_smtp(e, host, port)
        return jsonify({"sucesso": False, "erro": msg}), 400

    return jsonify({
        "sucesso": True,
        "mensagem": f"Remetente validado com sucesso ({smtp_origem}).",
    })


@app.route('/api/email/contatos', methods=['GET'])
@jwt_required()
def api_email_contatos_listar():
    """Lista contatos de email do usuario."""
    user_id = int(get_jwt_identity())
    contatos = EmailContact.query.filter_by(user_id=user_id).all()
    # Debug: logar se nao encontrar contatos
    if not contatos:
        all_contacts = EmailContact.query.all()
        print(f"[DEBUG email/contatos] user_id={user_id} retornou 0 contatos. "
              f"Total no banco: {len(all_contacts)} "
              f"(user_ids: {[c.user_id for c in all_contacts]})")
    return jsonify([c.to_dict() for c in contatos])


@app.route('/api/email/contatos', methods=['POST'])
@jwt_required()
def api_email_contatos_criar():
    """Cria/atualiza contato de email (com selecao de lojas/grupos)."""
    user_id = int(get_jwt_identity())
    data = request.get_json(force=True, silent=True) or {}
    email_addr = data.get("email", "").strip()

    # Se tem id, buscar contato existente primeiro (permite atualizar so horarios)
    contato_id = data.get("id")
    contato = None
    if contato_id:
        contato = EmailContact.query.filter_by(id=int(contato_id), user_id=user_id).first()

    if not email_addr and contato:
        email_addr = contato.email
    if (not email_addr or "@" not in email_addr) and not contato:
        return jsonify({"erro": "Email invalido"}), 400

    lojas_enviado = "lojas" in data
    grupos_enviado = "grupos" in data
    lote_ids_enviado = "lote_ids" in data or "lote_ids_json" in data

    lojas = data.get("lojas", []) or []
    grupos = data.get("grupos", []) or []
    lojas = [str(x).strip() for x in lojas if str(x).strip()]
    grupos = [str(x).strip() for x in grupos if str(x).strip()]
    lote_ids = _to_int_list(data.get("lote_ids"))
    if lote_ids is None and "lote_ids_json" in data:
        lote_ids = _to_int_list(data.get("lote_ids_json"))

    loja_cnpj = str(data.get("loja_cnpj", "") or "").strip() or "ALVO_CUSTOM"
    # Horarios multiplos: [{"dias":["seg","ter"],"horas":["07:00","11:30"]}, ...]
    horarios = data.get("horarios", None)

    if not contato:
        contato = EmailContact.query.filter_by(user_id=user_id, email=email_addr, loja_cnpj=loja_cnpj).first()

    if contato:
        if "nome_contato" in data:
            contato.nome_contato = data.get("nome_contato", contato.nome_contato)
        if "ativo" in data:
            contato.ativo = _to_bool(data.get("ativo"), contato.ativo)
        if loja_cnpj and loja_cnpj != "ALVO_CUSTOM":
            contato.loja_cnpj = loja_cnpj
        if lojas_enviado:
            contato.lojas_json = json.dumps(lojas, ensure_ascii=False)
        if grupos_enviado:
            contato.grupos_json = json.dumps(grupos, ensure_ascii=False)
        if horarios is not None:
            contato.horarios_json = json.dumps(horarios, ensure_ascii=False)
        if lote_ids_enviado:
            contato.lote_ids_json = json.dumps(lote_ids or [], ensure_ascii=False)
    else:
        contato = EmailContact(
            user_id=user_id,
            email=email_addr,
            loja_cnpj=loja_cnpj,
            nome_contato=data.get("nome_contato", ""),
            lojas_json=json.dumps(lojas, ensure_ascii=False),
            grupos_json=json.dumps(grupos, ensure_ascii=False),
            horarios_json=json.dumps(horarios or [], ensure_ascii=False),
            lote_ids_json=json.dumps(lote_ids or [], ensure_ascii=False),
            ativo=True,
        )
        db.session.add(contato)
    db.session.commit()

    # Sincronizar job APScheduler do contato
    try:
        beka_scheduler.registrar_job_contato(contato, 'email')
    except Exception:
        pass

    return jsonify(contato.to_dict()), 201


@app.route('/api/email/contatos/<int:contato_id>', methods=['DELETE'])
@jwt_required()
def api_email_contatos_deletar(contato_id):
    """Remove contato de email."""
    user_id = int(get_jwt_identity())
    contato = EmailContact.query.filter_by(id=contato_id, user_id=user_id).first()
    if not contato:
        return jsonify({"erro": "Contato nao encontrado"}), 404
    try:
        beka_scheduler.remover_job_contato(contato_id, 'email')
    except Exception:
        pass
    db.session.delete(contato)
    db.session.commit()
    return jsonify({"ok": True})


# ============================================================
# API - Lotes (TimeLote) — CRUD + execucao manual
# ============================================================

@app.route('/api/lotes', methods=['GET'])
@jwt_required()
def api_lotes_listar():
    user_id = int(get_jwt_identity())
    lotes = TimeLote.query.filter_by(user_id=user_id).order_by(TimeLote.hora).all()
    result = []
    for l in lotes:
        d = l.to_dict()
        # Contar contatos atribuidos a este lote
        total = 0
        for c in WhatsAppContact.query.filter_by(user_id=user_id, ativo=True).all():
            try:
                ids = json.loads(c.lote_ids_json or '[]')
            except Exception:
                ids = []
            if l.id in ids:
                total += 1
        for c in EmailContact.query.filter_by(user_id=user_id, ativo=True).all():
            try:
                ids = json.loads(c.lote_ids_json or '[]')
            except Exception:
                ids = []
            if l.id in ids:
                total += 1
        d['total_contatos'] = total
        result.append(d)
    return jsonify(result)


@app.route('/api/lotes', methods=['POST'])
@jwt_required()
def api_lotes_salvar():
    user_id = int(get_jwt_identity())
    data = request.get_json()
    lote_id = data.get('id')
    if lote_id:
        lote = TimeLote.query.get(lote_id)
        if not lote or lote.user_id != user_id:
            return jsonify({"erro": "Lote nao encontrado"}), 404
    else:
        lote = TimeLote(user_id=user_id)
        db.session.add(lote)

    hora_val = (data.get('hora') or data.get('horario') or '').strip()
    lote.nome = data.get('nome', '').strip() or f"Lote {hora_val}"
    lote.hora = hora_val
    lote.dias_semana = json.dumps(data.get('dias_semana', data.get('dias', ['seg', 'ter', 'qua', 'qui', 'sex'])))
    lote.ativo = data.get('ativo', True)
    db.session.commit()

    # Atualizar contatos atribuidos ao lote (se enviado)
    contatos_raw = data.get('contatos')  # ["whatsapp_7", "whatsapp_9", "email_1"]
    if contatos_raw is not None:
        # Parse selecionados
        wa_ids_selecionados = set()
        email_ids_selecionados = set()
        for item in contatos_raw:
            parts = str(item).split('_', 1)
            if len(parts) == 2:
                tipo, cid = parts[0], int(parts[1])
                if tipo == 'whatsapp':
                    wa_ids_selecionados.add(cid)
                elif tipo == 'email':
                    email_ids_selecionados.add(cid)

        # Atualizar WhatsApp contacts
        for c in WhatsAppContact.query.filter_by(user_id=user_id).all():
            try:
                ids = json.loads(c.lote_ids_json or '[]')
            except Exception:
                ids = []
            tinha = lote.id in ids
            deve_ter = c.id in wa_ids_selecionados
            if deve_ter and not tinha:
                ids.append(lote.id)
            elif not deve_ter and tinha:
                ids = [x for x in ids if x != lote.id]
            else:
                continue
            c.lote_ids_json = json.dumps(ids)

        # Atualizar Email contacts
        for c in EmailContact.query.filter_by(user_id=user_id).all():
            try:
                ids = json.loads(c.lote_ids_json or '[]')
            except Exception:
                ids = []
            tinha = lote.id in ids
            deve_ter = c.id in email_ids_selecionados
            if deve_ter and not tinha:
                ids.append(lote.id)
            elif not deve_ter and tinha:
                ids = [x for x in ids if x != lote.id]
            else:
                continue
            c.lote_ids_json = json.dumps(ids)

        db.session.commit()

    # Re-register scheduler jobs
    if hasattr(app, '_beka_scheduler') and app._beka_scheduler:
        app._beka_scheduler.registrar_jobs_lotes(user_id)

    return jsonify(lote.to_dict())


@app.route('/api/lotes/<int:lote_id>', methods=['DELETE'])
@jwt_required()
def api_lotes_excluir(lote_id):
    user_id = int(get_jwt_identity())
    lote = TimeLote.query.get(lote_id)
    if not lote or lote.user_id != user_id:
        return jsonify({"erro": "Nao encontrado"}), 404

    contatos_wa_afetados = []
    for c in WhatsAppContact.query.filter_by(user_id=user_id).all():
        try:
            ids = json.loads(c.lote_ids_json or '[]')
        except Exception:
            ids = []
        novos_ids = [x for x in ids if x != lote_id]
        if novos_ids != ids:
            c.lote_ids_json = json.dumps(novos_ids, ensure_ascii=False)
            contatos_wa_afetados.append(c)

    contatos_email_afetados = []
    for c in EmailContact.query.filter_by(user_id=user_id).all():
        try:
            ids = json.loads(c.lote_ids_json or '[]')
        except Exception:
            ids = []
        novos_ids = [x for x in ids if x != lote_id]
        if novos_ids != ids:
            c.lote_ids_json = json.dumps(novos_ids, ensure_ascii=False)
            contatos_email_afetados.append(c)

    db.session.delete(lote)
    db.session.commit()

    # Re-register scheduler jobs (remove the deleted lote's job)
    if hasattr(app, '_beka_scheduler') and app._beka_scheduler:
        try:
            app._beka_scheduler.registrar_jobs_lotes(user_id)
            for c in contatos_wa_afetados:
                try:
                    ids_restantes = json.loads(c.lote_ids_json or '[]')
                except Exception:
                    ids_restantes = []
                if ids_restantes:
                    app._beka_scheduler.remover_job_contato(c.id, 'whatsapp')
                else:
                    app._beka_scheduler.registrar_job_contato(c, 'whatsapp')
            for c in contatos_email_afetados:
                try:
                    ids_restantes = json.loads(c.lote_ids_json or '[]')
                except Exception:
                    ids_restantes = []
                if ids_restantes:
                    app._beka_scheduler.remover_job_contato(c.id, 'email')
                else:
                    app._beka_scheduler.registrar_job_contato(c, 'email')
        except Exception as e:
            print(f"[LoteExcluir] Erro ao sincronizar jobs apos excluir lote #{lote_id}: {e}")

    return jsonify({"ok": True})


@app.route('/api/lotes/<int:lote_id>/executar', methods=['POST'])
@jwt_required()
def api_lotes_executar(lote_id):
    user_id = int(get_jwt_identity())
    lote = TimeLote.query.get(lote_id)
    if not lote or lote.user_id != user_id:
        return jsonify({"erro": "Nao encontrado"}), 404
    # Execute in background thread
    import threading
    if hasattr(app, '_beka_scheduler') and app._beka_scheduler:
        t = threading.Thread(target=app._beka_scheduler._executar_lote, args=(user_id, lote_id), daemon=True)
        t.start()
        return jsonify({"ok": True, "msg": f"Lote '{lote.nome}' iniciado"})
    return jsonify({"erro": "Scheduler nao disponivel"}), 500


# ─── ETIQUETAS AVULSAS (independente) ────────────────────────────────

@app.route('/api/avulsas/config', methods=['GET'])
@jwt_required()
def api_avulsas_config():
    """Retorna a pasta salva para etiquetas avulsas."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    pasta = getattr(user, 'pasta_avulsas', '') or ''
    return jsonify({"pasta": pasta})


@app.route('/api/avulsas/config', methods=['POST'])
@jwt_required()
def api_avulsas_config_salvar():
    """Salva a pasta para etiquetas avulsas."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    data = request.get_json(silent=True) or {}
    pasta = data.get('pasta', '').strip()
    if not pasta:
        return jsonify({"erro": "Pasta nao informada"}), 400
    if not os.path.isdir(pasta):
        return jsonify({"erro": f"Pasta nao encontrada: {pasta}"}), 400
    user.pasta_avulsas = pasta
    db.session.commit()
    return jsonify({"ok": True, "pasta": pasta})


@app.route('/api/avulsas/processar', methods=['POST'])
@jwt_required()
def api_avulsas_processar():
    """Processa etiquetas avulsas da pasta configurada."""
    import threading
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    pasta = getattr(user, 'pasta_avulsas', '') or ''
    if not pasta or not os.path.isdir(pasta):
        return jsonify({"erro": "Pasta nao configurada ou nao existe"}), 400

    def _run():
        try:
            # Importar o processador avulso
            script_dir = os.path.dirname(os.path.abspath(__file__))
            import importlib.util
            spec = importlib.util.spec_from_file_location("recortar_etiquetas",
                os.path.join(script_dir, "recortar_etiquetas.py"))
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)

            etiquetas = mod.ler_pdfs(pasta)
            if not etiquetas:
                app._avulsas_resultado = {"erro": "Nenhum PDF valido encontrado na pasta."}
                return

            from collections import defaultdict
            lojas = defaultdict(list)
            for etq in etiquetas:
                lojas[etq['loja']].append(etq)

            output_base = os.path.join(pasta, "etiquetas_prontas")
            os.makedirs(output_base, exist_ok=True)

            import re as _re
            resultado_lojas = []
            for loja, etqs in sorted(lojas.items()):
                loja_safe = _re.sub(r'[<>:"/\\|?*]', '_', loja)
                loja_dir = os.path.join(output_base, loja_safe)
                os.makedirs(loja_dir, exist_ok=True)

                pdf_path = os.path.join(loja_dir, f"etiquetas_{loja_safe}.pdf")
                n = mod.gerar_pdf_loja(etqs, pdf_path)

                xlsx_path = os.path.join(loja_dir, f"resumo_{loja_safe}.xlsx")
                res = mod.gerar_resumo_xlsx(etqs, xlsx_path, loja)
                n_skus = res[0] if res else 0
                n_units = res[1] if res else 0

                resultado_lojas.append({
                    "loja": loja,
                    "etiquetas": len(etqs),
                    "skus": n_skus,
                    "unidades": n_units,
                    "pdf": pdf_path,
                })

            # Resumo geral
            lojas_info = {r['loja']: {'etiquetas': r['etiquetas'], 'skus': r['skus'], 'unidades': r['unidades']} for r in resultado_lojas}
            mod.gerar_resumo_geral(lojas_info, os.path.join(output_base, "resumo_geral.xlsx"))

            app._avulsas_resultado = {
                "ok": True,
                "total_etiquetas": len(etiquetas),
                "total_lojas": len(lojas),
                "lojas": resultado_lojas,
                "pasta_saida": output_base,
            }
        except Exception as e:
            import traceback
            app._avulsas_resultado = {"erro": str(e), "trace": traceback.format_exc()}

    app._avulsas_resultado = {"processando": True}
    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return jsonify({"ok": True, "msg": "Processamento iniciado"})


@app.route('/api/avulsas/status', methods=['GET'])
@jwt_required()
def api_avulsas_status():
    """Retorna status do processamento avulso."""
    resultado = getattr(app, '_avulsas_resultado', None)
    if resultado is None:
        return jsonify({"idle": True})
    return jsonify(resultado)


@app.route('/api/avulsas/abrir-pasta', methods=['POST'])
@jwt_required()
def api_avulsas_abrir_pasta():
    """Abre a pasta de saida no explorer."""
    data = request.get_json(silent=True) or {}
    pasta = data.get('pasta', '')
    if pasta and os.path.isdir(pasta):
        import subprocess
        subprocess.Popen(f'explorer "{pasta}"', shell=True)
        return jsonify({"ok": True})
    return jsonify({"erro": "Pasta nao encontrada"}), 400


@app.route('/api/email/enviar-lote', methods=['POST'])
@jwt_required()
def api_email_enviar_lote():
    """Envia arquivos de etiquetas (PDF/IMG/XLSX) por email para contatos cadastrados."""
    user_id = int(get_jwt_identity())
    estado = _get_estado(user_id)
    resultado = estado.get("ultimo_resultado", {})
    user = User.query.get(user_id)
    pasta_saida = user.get_pasta_saida()

    if not resultado or not resultado.get("lojas"):
        return jsonify({"erro": "Nenhum resultado para enviar. Processe as etiquetas primeiro."}), 400

    from_addr_override = (getattr(user, "email_remetente", "") or "").strip()
    if not from_addr_override or "@" not in from_addr_override:
        return jsonify({
            "erro": "Email remetente obrigatorio. Configure em Automacao > Contatos Email > Configurar Remetente."
        }), 400

    smtp_cfg, smtp_origem = _smtp_config_resolver(user)
    if not smtp_cfg:
        return jsonify({
            "erro": "SMTP nao configurado. Preencha host, usuario e senha SMTP em Automacao > Contatos Email > Configurar Remetente."
        }), 400

    contatos = EmailContact.query.filter_by(user_id=user_id, ativo=True).all()
    if not contatos:
        return jsonify({"erro": "Nenhum contato de email cadastrado"}), 400

    envios, diagnostico = montar_destinos_por_resultado(
        resultado=resultado,
        pasta_saida=pasta_saida,
        contatos=contatos,
        destino_attr="email",
        agrupamentos_usuario=(estado or {}).get("agrupamentos", []),
    )

    if not envios:
        return jsonify({
            "erro": "Nenhuma loja com contato email e arquivos encontrados",
            "diagnostico": diagnostico,
        }), 400

    timestamp = resultado.get("timestamp", "")
    from_name_override = (getattr(user, "nome_remetente", "") or "").strip()

    envios_agrupados = {}
    for envio in envios:
        destino = str(envio.get("destino", "") or "").strip()
        loja_nome = str(envio.get("loja", "") or "").strip()
        file_path = str(envio.get("file_path", envio.get("pdf_path", "")) or "").strip()
        if not destino or not file_path:
            continue
        chave = (destino.lower(), loja_nome)
        if chave not in envios_agrupados:
            envios_agrupados[chave] = {
                "destino": destino,
                "loja": loja_nome,
                "arquivos": [],
            }
        if file_path not in envios_agrupados[chave]["arquivos"]:
            envios_agrupados[chave]["arquivos"].append(file_path)

    grupos_envio = list(envios_agrupados.values())
    total_arquivos = sum(len(g.get("arquivos", [])) for g in grupos_envio)
    if not grupos_envio:
        return jsonify({
            "erro": "Nenhum arquivo valido encontrado para envio",
            "diagnostico": diagnostico,
        }), 400

    # Enviar em background
    def _enviar_emails():
        with app.app_context():
            resultados = []
            for envio in grupos_envio:
                res = enviar_email_com_anexos(
                    email_destino=envio["destino"],
                    assunto=f"Arquivos {envio['loja']} - {timestamp}",
                    loja_nome=envio["loja"],
                    timestamp=timestamp,
                    anexos_paths=envio.get("arquivos", []),
                    from_addr_override=from_addr_override,
                    from_name_override=from_name_override,
                    smtp_override=smtp_cfg,
                )
                resultados.append(res)
                import time
                time.sleep(2)  # anti-spam

            # Log
            log_exec = ExecutionLog(
                user_id=user_id,
                tipo="email",
                inicio=_agora_brasil(),
                fim=_agora_brasil(),
                status="sucesso" if all(r.get("success") for r in resultados) else "parcial",
                whatsapp_enviados=sum(1 for r in resultados if r.get("success")),
                whatsapp_erros=sum(1 for r in resultados if not r.get("success")),
                detalhes=json.dumps({
                    "envios": len(grupos_envio),
                    "arquivos": total_arquivos,
                    "resultados": resultados,
                    "diagnostico": diagnostico,
                }, ensure_ascii=False),
            )
            db.session.add(log_exec)
            db.session.commit()

    thread = threading.Thread(target=_enviar_emails, daemon=True)
    thread.start()
    return jsonify({
        "mensagem": f"Enviando {len(grupos_envio)} email(s) com {total_arquivos} arquivo(s) em background...",
        "total_envios": len(grupos_envio),
        "total_arquivos": total_arquivos,
        "smtp_origem": smtp_origem,
        "diagnostico": diagnostico,
    })


# ----------------------------------------------------------------
# ENDPOINTS - AGENDAMENTOS
# ----------------------------------------------------------------

@app.route('/api/agendamentos', methods=['GET'])
@jwt_required()
def api_agendamentos_listar():
    """Lista agendamentos do usuario."""
    user_id = int(get_jwt_identity())
    return jsonify(beka_scheduler.listar_agendamentos(user_id))


@app.route('/api/agendamentos', methods=['POST'])
@jwt_required()
def api_agendamentos_criar():
    """Cria novo agendamento."""
    user_id = int(get_jwt_identity())
    data = request.get_json()

    if not data or not data.get("hora"):
        return jsonify({"erro": "Horario obrigatorio"}), 400

    schedule_id = beka_scheduler.adicionar_agendamento(user_id, data)
    if schedule_id:
        return jsonify({"mensagem": "Agendamento criado", "id": schedule_id})
    return jsonify({"erro": "Erro ao criar agendamento"}), 500


@app.route('/api/agendamentos/<int:schedule_id>', methods=['PUT'])
@jwt_required()
def api_agendamentos_atualizar(schedule_id):
    """Atualiza agendamento existente."""
    user_id = int(get_jwt_identity())
    # Verificar propriedade
    sched = Schedule.query.filter_by(id=schedule_id, user_id=user_id).first()
    if not sched:
        return jsonify({"erro": "Agendamento nao encontrado"}), 404

    data = request.get_json()
    if beka_scheduler.atualizar_agendamento(schedule_id, data):
        return jsonify({"mensagem": "Agendamento atualizado"})
    return jsonify({"erro": "Erro ao atualizar"}), 500


@app.route('/api/agendamentos/<int:schedule_id>', methods=['DELETE'])
@jwt_required()
def api_agendamentos_remover(schedule_id):
    """Remove agendamento."""
    user_id = int(get_jwt_identity())
    sched = Schedule.query.filter_by(id=schedule_id, user_id=user_id).first()
    if not sched:
        return jsonify({"erro": "Agendamento nao encontrado"}), 404

    if beka_scheduler.remover_agendamento(schedule_id):
        return jsonify({"mensagem": "Agendamento removido"})
    return jsonify({"erro": "Erro ao remover"}), 500


@app.route('/api/agendamentos/<int:schedule_id>/pausar', methods=['POST'])
@jwt_required()
def api_agendamentos_pausar(schedule_id):
    """Pausa agendamento."""
    user_id = int(get_jwt_identity())
    sched = Schedule.query.filter_by(id=schedule_id, user_id=user_id).first()
    if not sched:
        return jsonify({"erro": "Agendamento nao encontrado"}), 404

    if beka_scheduler.pausar_agendamento(schedule_id):
        return jsonify({"mensagem": "Agendamento pausado"})
    return jsonify({"erro": "Erro ao pausar"}), 500


@app.route('/api/agendamentos/<int:schedule_id>/retomar', methods=['POST'])
@jwt_required()
def api_agendamentos_retomar(schedule_id):
    """Retoma agendamento pausado."""
    user_id = int(get_jwt_identity())
    sched = Schedule.query.filter_by(id=schedule_id, user_id=user_id).first()
    if not sched:
        return jsonify({"erro": "Agendamento nao encontrado"}), 404

    if beka_scheduler.retomar_agendamento(schedule_id):
        return jsonify({"mensagem": "Agendamento retomado"})
    return jsonify({"erro": "Erro ao retomar"}), 500


@app.route('/api/agendamentos/historico', methods=['GET'])
@jwt_required()
def api_agendamentos_historico():
    """Historico de execucoes."""
    user_id = int(get_jwt_identity())
    limite = request.args.get('limite', 20, type=int)
    return jsonify(beka_scheduler.get_historico(user_id, limite))


def _extrair_lojas_do_contato(contato, user_id):
    """Extrai lojas alvo de um contato, expandindo grupos em lojas reais."""
    lojas = set()
    try:
        for loja in json.loads(getattr(contato, 'lojas_json', None) or '[]'):
            if str(loja).strip():
                lojas.add(str(loja).strip())
    except Exception:
        pass
    # Expandir grupos em lojas
    try:
        grupos_raw = json.loads(getattr(contato, 'grupos_json', None) or '[]')
        grupos = [str(g).strip().lower() for g in grupos_raw if str(g).strip()]
        if grupos:
            estado_tmp = _get_estado(user_id) or {}
            agrup = estado_tmp.get("agrupamentos", []) or []
            for g in agrup:
                nome_g = str((g or {}).get("nome", "") or "").strip().lower()
                if nome_g and nome_g in grupos:
                    for ln in (g or {}).get("nomes_lojas", []) or []:
                        ln = str(ln or "").strip()
                        if ln:
                            lojas.add(ln)
    except Exception:
        pass
    return lojas


@app.route('/api/agendamentos/executar-agora', methods=['POST'])
@jwt_required()
def api_agendamentos_executar_agora():
    """Executa pipeline manualmente (sem agendamento). Aceita modo_pipeline='direto'."""
    user_id = int(get_jwt_identity())
    data = request.get_json(force=True, silent=True) or {}
    modo = data.get("modo_pipeline", "completo")

    if modo == "direto":
        lock_ok, lock_atual = _iniciar_acao_massa(user_id, "direto")
        if not lock_ok:
            return jsonify({
                "erro": (
                    f"Processo em massa em andamento ({(lock_atual or {}).get('acao', 'processando')}). "
                    "Aguarde finalizar antes de iniciar o envio direto."
                )
            }), 409

        def _run_direto():
            with app.app_context():
                print(f"[ExecDireto] Iniciando pipeline direto para user {user_id}...")
                log_exec = ExecutionLog(
                    user_id=user_id, tipo="manual",
                    inicio=_agora_brasil(), status="executando"
                )
                db.session.add(log_exec)
                db.session.commit()
                try:
                    user = db.session.get(User, user_id)
                    if not user:
                        raise RuntimeError("Usuario nao encontrado")

                    # Extrair lojas dos contatos configurados (WhatsApp + Email), expandindo grupos
                    lojas_contatos = set()
                    for c in WhatsAppContact.query.filter_by(user_id=user_id, ativo=True).all():
                        lojas_contatos.update(_extrair_lojas_do_contato(c, user_id))
                    for c in EmailContact.query.filter_by(user_id=user_id, ativo=True).all():
                        lojas_contatos.update(_extrair_lojas_do_contato(c, user_id))
                    lojas_alvo_list = sorted(lojas_contatos) if lojas_contatos else None
                    if lojas_alvo_list:
                        print(f"[ExecDireto] Filtrando por lojas dos contatos: {lojas_alvo_list}")
                    else:
                        print(f"[ExecDireto] Nenhum contato com lojas configuradas, baixando todas")
                    result = _executar_imprimir_direto(user_id, lojas_alvo=lojas_alvo_list)
                    print(f"[ExecDireto] Resultado: {result}")
                    if result.get("ok"):
                        log_exec.etiquetas_processadas = result.get("total_etiquetas", 0)
                        # Salvar sucesso IMEDIATAMENTE antes do envio
                        log_exec.status = "sucesso"
                        log_exec.fim = _agora_brasil()
                        db.session.commit()
                        print(f"[ExecDireto] Processamento OK. Enfileirando envios...")
                        # Enviar WhatsApp
                        try:
                            enq = _enfileirar_envio_whatsapp_resultado(
                                user_id=user_id, origem="manual", respeitar_toggle_auto=False
                            )
                            log_exec.whatsapp_enviados = enq.get("total_entregas", 0) if enq.get("ok") else 0
                            db.session.commit()
                        except Exception as e_wa:
                            print(f"[ExecDireto] Erro WhatsApp: {e_wa}")
                        # Enviar Email
                        try:
                            _enviar_email_resultado_agendado(user_id=user_id)
                        except Exception as e_em:
                            print(f"[ExecDireto] Erro Email: {e_em}")
                        # Resumo geral: enviar JPEG parcial + acumular para consolidado
                        try:
                            estado_tmp = _get_estado(user_id)
                            lojas_res = ((estado_tmp or {}).get("ultimo_resultado") or {}).get("lojas", [])
                            _acumular_resumo_diario(user_id, lojas_res)
                            _enviar_resumo_geral_whatsapp(user_id, user.get_pasta_saida())
                        except Exception as e_rg:
                            print(f"[ExecDireto] Erro resumo geral: {e_rg}")
                    else:
                        log_exec.status = "erro"
                        log_exec.detalhes = json.dumps({"erro": result.get("erro", "")}, ensure_ascii=False)
                        log_exec.fim = _agora_brasil()
                        db.session.commit()
                except Exception as e:
                    import traceback
                    print(f"[ExecDireto] EXCECAO: {e}")
                    traceback.print_exc()
                    log_exec.status = "erro"
                    log_exec.detalhes = json.dumps({"erro": str(e)}, ensure_ascii=False)
                    log_exec.fim = _agora_brasil()
                    db.session.commit()
                finally:
                    _finalizar_acao_massa(user_id, "direto")
                print(f"[ExecDireto] Finalizado. Status: {log_exec.status}")

        def _run_direto_com_timeout():
            worker = threading.Thread(target=_run_direto, daemon=True)
            worker.start()
            worker.join(timeout=900)  # 15 min max
            if worker.is_alive():
                print(f"[ExecDireto] TIMEOUT: pipeline direto excedeu 15 min")
                with app.app_context():
                    stuck = ExecutionLog.query.filter_by(
                        user_id=user_id, status="executando"
                    ).order_by(ExecutionLog.id.desc()).first()
                    if stuck:
                        stuck.status = "erro"
                        stuck.fim = _agora_brasil()
                        stuck.detalhes = json.dumps(
                            {"erro": "Timeout: execucao excedeu 15 minutos"},
                            ensure_ascii=False
                        )
                        db.session.commit()

        thread = threading.Thread(target=_run_direto_com_timeout, daemon=True)
        try:
            thread.start()
        except Exception:
            _finalizar_acao_massa(user_id, "direto")
            raise
        return jsonify({"mensagem": "Envio direto iniciado em background"})

    beka_scheduler.executar_agora(user_id)
    return jsonify({"mensagem": "Pipeline iniciado em background"})


@app.route('/api/contato/<string:tipo>/<int:contato_id>/executar-agora', methods=['POST'])
@jwt_required()
def api_contato_executar_agora(tipo, contato_id):
    """Executa pipeline para um contato individual."""
    user_id = int(get_jwt_identity())
    if tipo not in ('whatsapp', 'email'):
        return jsonify({"erro": "Tipo invalido"}), 400
    if tipo == 'whatsapp':
        contato = WhatsAppContact.query.filter_by(id=contato_id, user_id=user_id).first()
    else:
        contato = EmailContact.query.filter_by(id=contato_id, user_id=user_id).first()
    if not contato:
        return jsonify({"erro": "Contato nao encontrado"}), 404

    lock_ok, lock_atual = _iniciar_acao_massa(user_id, "direto")
    if not lock_ok:
        return jsonify({
            "erro": (
                f"Processo em massa em andamento ({(lock_atual or {}).get('acao', 'processando')}). "
                "Aguarde finalizar antes de iniciar um envio individual."
            )
        }), 409

    def _run_individual():
        with app.app_context():
            lojas_alvo = sorted(_extrair_lojas_do_contato(contato, user_id))

            log_exec = ExecutionLog(
                user_id=user_id, tipo="manual_individual",
                inicio=_agora_brasil(), status="executando",
                detalhes=json.dumps({"contato_id": contato_id, "tipo": tipo}, ensure_ascii=False)
            )
            db.session.add(log_exec)
            db.session.commit()

            try:
                result = _executar_imprimir_direto(user_id, lojas_alvo=lojas_alvo or None)
                if result.get("ok"):
                    log_exec.etiquetas_processadas = result.get("total_etiquetas", 0)
                    # Salvar sucesso IMEDIATAMENTE antes do envio —
                    # assim se o servidor morrer, sabemos que processou OK
                    log_exec.status = "sucesso"
                    log_exec.fim = _agora_brasil()
                    db.session.commit()
                    print(f"[ExecIndividual] Processamento OK ({result.get('total_etiquetas',0)} etiq). Enfileirando envio...")
                    try:
                        if tipo == 'whatsapp':
                            enq = _enfileirar_whatsapp_todos_arquivos_para_contato(user_id, contato_id)
                            if not enq.get("ok"):
                                print(f"[ExecIndividual] Envio direto falhou ({enq.get('erro')}), tentando matching...")
                                enq = _enfileirar_envio_whatsapp_para_contato(user_id, contato_id)
                        else:
                            _enviar_email_para_contato_direto(user_id, contato_id)
                            enq = {"ok": True, "total_entregas": 1}
                        log_exec.whatsapp_enviados = enq.get("total_entregas", 0) if enq.get("ok") else 0
                        db.session.commit()
                    except Exception as e_send:
                        print(f"[ExecIndividual] Erro envio: {e_send}")
                else:
                    log_exec.status = "erro"
                    log_exec.detalhes = json.dumps({"erro": result.get("erro", "")}, ensure_ascii=False)
                    log_exec.fim = _agora_brasil()
                    db.session.commit()
            except Exception as e:
                import traceback
                traceback.print_exc()
                log_exec.status = "erro"
                log_exec.detalhes = json.dumps({"erro": str(e)}, ensure_ascii=False)
                log_exec.fim = _agora_brasil()
                db.session.commit()
            finally:
                _finalizar_acao_massa(user_id, "direto")

    def _run_individual_com_timeout():
        """Wrapper com timeout de 15 minutos para evitar travamento infinito."""
        worker = threading.Thread(target=_run_individual, daemon=True)
        worker.start()
        worker.join(timeout=900)  # 15 minutos max
        if worker.is_alive():
            print(f"[ExecIndividual] TIMEOUT: execucao contato #{contato_id} excedeu 15 min")
            with app.app_context():
                stuck = ExecutionLog.query.filter_by(
                    user_id=user_id, status="executando"
                ).order_by(ExecutionLog.id.desc()).first()
                if stuck:
                    stuck.status = "erro"
                    stuck.fim = _agora_brasil()
                    stuck.detalhes = json.dumps(
                        {"erro": "Timeout: execucao excedeu 15 minutos"},
                        ensure_ascii=False
                    )
                    db.session.commit()

    thread = threading.Thread(target=_run_individual_com_timeout, daemon=True)
    try:
        thread.start()
    except Exception:
        _finalizar_acao_massa(user_id, "direto")
        raise
    return jsonify({"mensagem": f"Execucao iniciada para contato #{contato_id}"})


# Inicia workers globais (fila WhatsApp + supervisor Baileys)
_iniciar_background_workers()


# ----------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------

def _iniciar_servidor_background(port):
    """Inicia o servidor WSGI em thread separada para o pywebview."""
    try:
        from waitress import serve
        print(f"[Waitress] Servidor WSGI iniciando na porta {port}...")
        serve(app, host='0.0.0.0', port=port, threads=8,
              channel_timeout=120, recv_bytes=65536)
    except ImportError:
        print("[AVISO] waitress nao instalado, usando Flask built-in")
        app.run(debug=False, port=port, host='0.0.0.0')


if __name__ == '__main__':
    print("=" * 60)
    print("DASHBOARD - Beka MultiPlace")
    print("=" * 60)

    port = int(os.environ.get('PORT', 5000))
    mode = os.environ.get('BEKA_MODE', 'desktop')  # 'desktop' ou 'server'

    if mode == 'server':
        # Modo servidor puro (para Railway, VPS, ou acesso via navegador)
        print(f"\n  Modo SERVIDOR: http://localhost:{port}\n")
        print("=" * 60)
        _iniciar_servidor_background(port)
    else:
        # Modo desktop: abre janela nativa com pywebview
        print(f"\n  Modo DESKTOP: abrindo janela nativa...\n")
        print("=" * 60)

        try:
            import webview

            # Iniciar servidor em thread background
            server_thread = threading.Thread(
                target=_iniciar_servidor_background,
                args=(port,),
                daemon=True
            )
            server_thread.start()

            # Aguardar servidor ficar pronto
            import time
            import urllib.request
            for i in range(30):
                try:
                    urllib.request.urlopen(f'http://localhost:{port}/', timeout=1)
                    break
                except Exception:
                    time.sleep(0.5)

            # Abrir janela nativa
            window = webview.create_window(
                'Beka MultiPlace',
                f'http://localhost:{port}',
                width=1280,
                height=820,
                min_size=(900, 600),
                resizable=True,
                text_select=True,
            )
            webview.start(
                debug=False,
                private_mode=False,
                storage_path=_WEBVIEW_STORAGE_DIR,
            )

        except ImportError:
            print("[AVISO] pywebview nao instalado, abrindo no navegador...")
            import webbrowser
            threading.Timer(1.5, lambda: webbrowser.open(f'http://localhost:{port}')).start()
            _iniciar_servidor_background(port)
