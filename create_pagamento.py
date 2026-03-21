import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# === STYLES ===
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
title_font = Font(name='Arial', bold=True, size=14, color='2F5496')
label_font = Font(name='Arial', bold=True, size=10)
data_font = Font(name='Arial', size=10)
blue_font = Font(name='Arial', size=10, color='0000FF')  # inputs
black_font = Font(name='Arial', size=10, color='000000')  # formulas
currency_fmt = 'R$ #,##0.00'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
yellow_fill = PatternFill('solid', fgColor='FFFF00')
light_gray = PatternFill('solid', fgColor='F2F2F2')
light_blue = PatternFill('solid', fgColor='DCE6F1')

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_data_cell(cell, is_formula=False, is_input=False):
    cell.font = black_font if is_formula else (blue_font if is_input else data_font)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# =============================================
# TAB 1: FOLHA DE PAGAMENTO
# =============================================
ws1 = wb.active
ws1.title = 'Folha de Pagamento'

ws1.merge_cells('A1:J1')
ws1['A1'] = 'FOLHA DE PAGAMENTO - BEKA MKT'
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('A2:B2')
ws1['A2'] = 'QUINZENA:'
ws1['A2'].font = label_font
ws1['A2'].alignment = Alignment(horizontal='right')
ws1['C2'] = '1ª QUINZ. MAR/2026'
ws1['C2'].font = blue_font
ws1['C2'].fill = yellow_fill
ws1['C2'].border = thin_border

headers1 = ['FUNCIONÁRIO', 'SALÁRIO MENSAL', 'QUINZENA BASE', 'FALTAS (DIAS)',
            'DESC. FALTAS', 'VALE 1', 'VALE 2', 'H. EXTRAS', 'VALOR H.EXTRA', 'A RECEBER']
for i, h in enumerate(headers1, 1):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, 10)

employees = [
    ('EDUARDA', 1800),
    ('GABRIEL', 2150),
    ('DANIEL', 2000),
    ('RONEIDE', 3230),
    ('MARY', 2120),
]

for idx, (name, salary) in enumerate(employees):
    row = 5 + idx
    stripe = light_gray if idx % 2 == 0 else None

    # A - Nome
    c = ws1.cell(row=row, column=1, value=name)
    style_data_cell(c)
    c.alignment = Alignment(horizontal='left', vertical='center')
    if stripe: c.fill = stripe

    # B - Salário mensal (input)
    c = ws1.cell(row=row, column=2, value=salary)
    style_data_cell(c, is_input=True)
    c.number_format = currency_fmt
    if stripe: c.fill = stripe

    # C - Quinzena base = salário / 2
    c = ws1.cell(row=row, column=3)
    c.value = f'=B{row}/2'
    style_data_cell(c, is_formula=True)
    c.number_format = currency_fmt
    if stripe: c.fill = stripe

    # D - Faltas (input, dias - aceita 0.5)
    c = ws1.cell(row=row, column=4, value=0)
    style_data_cell(c, is_input=True)
    c.number_format = '0.0'
    if stripe: c.fill = stripe

    # E - Desconto faltas = faltas * (salário/30) * 2
    c = ws1.cell(row=row, column=5)
    c.value = f'=D{row}*(B{row}/30)*2'
    style_data_cell(c, is_formula=True)
    c.number_format = currency_fmt
    if stripe: c.fill = stripe

    # F - Vale 1 (input)
    c = ws1.cell(row=row, column=6, value=0)
    style_data_cell(c, is_input=True)
    c.number_format = currency_fmt
    if stripe: c.fill = stripe

    # G - Vale 2 (input)
    c = ws1.cell(row=row, column=7, value=0)
    style_data_cell(c, is_input=True)
    c.number_format = currency_fmt
    if stripe: c.fill = stripe

    # H - Horas extras (input)
    c = ws1.cell(row=row, column=8, value=0)
    style_data_cell(c, is_input=True)
    c.number_format = '0.0'
    if stripe: c.fill = stripe

    # I - Valor hora extra = (salário/30)/8 * 1.5 * horas
    c = ws1.cell(row=row, column=9)
    c.value = f'=IF(H{row}>0, H{row}*(B{row}/30/8)*1.5, 0)'
    style_data_cell(c, is_formula=True)
    c.number_format = currency_fmt
    if stripe: c.fill = stripe

    # J - A receber = quinzena - desc.faltas - vale1 - vale2 + valor h.extra
    c = ws1.cell(row=row, column=10)
    c.value = f'=C{row}-E{row}-F{row}-G{row}+I{row}'
    style_data_cell(c, is_formula=True)
    c.number_format = currency_fmt
    if stripe: c.fill = stripe

# Totals row
total_row = 5 + len(employees)
ws1.cell(row=total_row, column=1, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
ws1.cell(row=total_row, column=1).border = thin_border
for col in [2, 3, 5, 6, 7, 9, 10]:
    col_letter = get_column_letter(col)
    c = ws1.cell(row=total_row, column=col)
    c.value = f'=SUM({col_letter}5:{col_letter}{total_row-1})'
    c.font = Font(name='Arial', bold=True, size=10)
    c.number_format = currency_fmt
    c.border = thin_border
    c.fill = light_blue

# Column widths
widths1 = [16, 16, 16, 14, 16, 14, 14, 12, 16, 16]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Legend
legend_row = total_row + 2
ws1.cell(row=legend_row, column=1, value='LEGENDA:').font = label_font
ws1.cell(row=legend_row+1, column=1, value='Texto azul = campos editáveis (inputs)').font = blue_font
ws1.cell(row=legend_row+2, column=1, value='Texto preto = fórmulas (não editar)').font = black_font
ws1.cell(row=legend_row+3, column=1, value='Falta: 1 dia = (salário/30) x 2 de desconto').font = data_font
ws1.cell(row=legend_row+4, column=1, value='Meio dia (0.5) = (salário/30) x 1 de desconto').font = data_font

# =============================================
# TAB 2: VALES E PARCELAS
# =============================================
ws2 = wb.create_sheet('Vales e Parcelas')

ws2.merge_cells('A1:I1')
ws2['A1'] = 'CONTROLE DE VALES E PARCELAS'
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center')

headers2 = ['FUNCIONÁRIO', 'DESCRIÇÃO', 'VALOR TOTAL', 'VALOR PARCELA',
            'Nº PARCELAS', 'PAGAS', 'RESTANTES', 'PRÓX. DESCONTO', 'STATUS']
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, 9)

vales = [
    ('GABRIEL', 'Empréstimo pessoal', 4000, 400, 10, 'Dia 15'),
    ('DANIEL', 'Moto', 2000, 500, 4, 'Último dia do mês'),
    ('DANIEL', 'Pintura', 650, 650, 1, 'Último dia do mês'),
    ('DANIEL', 'Chinelo Slin', 25, 25, 1, 'Último dia do mês'),
    ('RONEIDE', 'Carro', 9000, 0, 0, 'Dezembro/2026'),
]

for idx, (func, desc, total, parcela, num_parc, prox) in enumerate(vales):
    row = 4 + idx
    stripe = light_gray if idx % 2 == 0 else None

    c = ws2.cell(row=row, column=1, value=func)
    style_data_cell(c)
    c.alignment = Alignment(horizontal='left', vertical='center')
    if stripe: c.fill = stripe

    c = ws2.cell(row=row, column=2, value=desc)
    style_data_cell(c)
    c.alignment = Alignment(horizontal='left', vertical='center')
    if stripe: c.fill = stripe

    c = ws2.cell(row=row, column=3, value=total)
    style_data_cell(c, is_input=True)
    c.number_format = currency_fmt
    if stripe: c.fill = stripe

    c = ws2.cell(row=row, column=4, value=parcela)
    style_data_cell(c, is_input=True)
    c.number_format = currency_fmt
    if stripe: c.fill = stripe

    c = ws2.cell(row=row, column=5, value=num_parc)
    style_data_cell(c, is_input=True)
    if stripe: c.fill = stripe

    c = ws2.cell(row=row, column=6, value=0)
    style_data_cell(c, is_input=True)
    if stripe: c.fill = stripe

    # G - Restantes = parcelas - pagas
    c = ws2.cell(row=row, column=7)
    c.value = f'=E{row}-F{row}'
    style_data_cell(c, is_formula=True)
    if stripe: c.fill = stripe

    c = ws2.cell(row=row, column=8, value=prox)
    style_data_cell(c, is_input=True)
    if stripe: c.fill = stripe

    # I - Status
    c = ws2.cell(row=row, column=9)
    c.value = f'=IF(G{row}<=0,"QUITADO","PENDENTE")'
    style_data_cell(c, is_formula=True)
    if stripe: c.fill = stripe

widths2 = [16, 22, 16, 16, 14, 10, 12, 20, 12]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Note about Roneide
note_row = 4 + len(vales) + 1
ws2.cell(row=note_row, column=1, value='OBSERVAÇÕES:').font = label_font
ws2.cell(row=note_row+1, column=1, value='RONEIDE: R$9.000 carro - descontar tudo em dezembro/2026').font = data_font

# =============================================
# TAB 3: HISTÓRICO DE FALTAS
# =============================================
ws3 = wb.create_sheet('Histórico Faltas')

ws3.merge_cells('A1:G1')
ws3['A1'] = 'HISTÓRICO DE FALTAS'
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center')

headers3 = ['DATA', 'FUNCIONÁRIO', 'DIAS', 'SALÁRIO', 'DESCONTO', 'MOTIVO', 'QUINZENA']
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, 7)

# Example row to show formula
row = 4
ws3.cell(row=row, column=1, value='dd/mm/aaaa').font = blue_font
ws3.cell(row=row, column=1).border = thin_border
ws3.cell(row=row, column=1).number_format = 'DD/MM/YYYY'

ws3.cell(row=row, column=2, value='NOME').font = blue_font
ws3.cell(row=row, column=2).border = thin_border

c = ws3.cell(row=row, column=3, value=1)
style_data_cell(c, is_input=True)
c.number_format = '0.0'

c = ws3.cell(row=row, column=4, value=2000)
style_data_cell(c, is_input=True)
c.number_format = currency_fmt

c = ws3.cell(row=row, column=5)
c.value = f'=C{row}*(D{row}/30)*2'
style_data_cell(c, is_formula=True)
c.number_format = currency_fmt

ws3.cell(row=row, column=6, value='').font = blue_font
ws3.cell(row=row, column=6).border = thin_border

ws3.cell(row=row, column=7, value='1ª ou 2ª').font = blue_font
ws3.cell(row=row, column=7).border = thin_border

widths3 = [14, 16, 10, 16, 16, 24, 14]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Print setup
for ws in [ws1, ws2, ws3]:
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = 'landscape'

output = r'C:\Users\Micro\Desktop\4_ADMINISTRACAO\funcionários\PAGAMENTO_v2.xlsx'
wb.save(output)
print(f'Saved to {output}')
