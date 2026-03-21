#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Processa etiquetas de envio igual ao sistema Beka MKT:
  1. Lê todos os PDFs da pasta (pag1=etiqueta, pag2=declaração)
  2. Agrupa por loja (REMETENTE)
  3. Ordena: qtd=1 primeiro (por SKU > Cor > Tamanho), qtd>1 no final
  4. Recorta bordas brancas, adiciona rodapé compacto com produtos
  5. Gera 1 PDF por loja (150x230mm) + resumo XLSX + imagem JPEG
  6. Salva em: pasta/etiquetas_prontas/<loja>/
"""

import fitz  # PyMuPDF
import os
import sys
import glob
import re
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Dimensões iguais ao Beka MKT (150x230mm)
LARGURA_PT = 425.197   # 150mm
ALTURA_PT = 651.969    # 230mm
MARGEM_ESQ = 8
MARGEM_DIR = 8
MARGEM_TOPO = 5
MARGEM_INF = 5


# ─── Extração de dados ───────────────────────────────────────────────

def _eh_endereco_ou_cep(line):
    """Verifica se a linha parece endereço, CEP ou info irrelevante."""
    l = line.strip().upper()
    if not l:
        return True
    if re.match(r'^\d{5}[-\s]?\d{3}$', l):
        return True
    keywords = ['CEP:', 'ENVIO PREVISTO', 'RUA ', 'AV ', 'AVENIDA ', 'TRAVESSA ',
                'ALAMEDA ', 'RODOVIA ', 'ESTRADA ', 'BAIRRO:', 'SOC', 'NOVA SERRANA',
                'MINAS GERAIS', 'DANFE', 'SÉRIE', 'SERIE', 'EMISSÃO', 'EMISSAO',
                'NF:', '1 - SAÍDA', '1 - SAIDA', 'DESTINAT', 'PEDIDO:']
    for k in keywords:
        if k in l:
            return True
    # Endereço com número (Rua X, 123)
    if re.search(r',\s*\d+', l) and len(l) > 15:
        return True
    return False


def extrair_loja_remetente(text):
    """Extrai nome da loja a partir do texto da etiqueta (pag 1)."""
    lines = text.split('\n')
    # Procurar tracking code BR...
    tracking_idx = -1
    for i, l in enumerate(lines):
        l = l.strip()
        if re.match(r'^BR\d{10,}', l):
            tracking_idx = i
            break

    if tracking_idx >= 0:
        for j in range(tracking_idx + 1, min(tracking_idx + 10, len(lines))):
            line = lines[j].strip()
            if not line or len(line) < 2:
                continue
            if _eh_endereco_ou_cep(line):
                continue
            if re.match(r'^[A-Z]{2}\d', line):  # Código de rota tipo MG2
                continue
            return line
    return None


def extrair_produtos_declaracao(page):
    """Extrai dados da tabela IDENTIFICAÇÃO DOS BENS da página de declaração."""
    tables = page.find_tables()
    if not tables or not tables.tables:
        return []

    for t in tables.tables:
        rows = t.extract()
        if not rows:
            continue
        first_text = str(rows[0][0] or '').upper()
        if 'IDENTIFICA' in first_text or 'BENS' in first_text:
            produtos = []
            for row in rows[2:]:
                n = str(row[0] or '').strip()
                if not n or 'Peso' in n:
                    continue
                if 'Totais' in n or 'Total' in n:
                    qtd_total = str(row[4] or '').strip()
                    produtos.append({
                        'tipo': 'totais',
                        'qtd': qtd_total,
                    })
                    continue
                codigo = str(row[1] or '').strip()
                descricao = str(row[2] or '').strip().replace('\n', ' ')
                variacao = str(row[3] or '').strip().replace('\n', ' ')
                try:
                    qtd = int(float(str(row[4] or '1').strip()))
                except:
                    qtd = 1
                produtos.append({
                    'tipo': 'produto',
                    'n': n,
                    'codigo': codigo,
                    'descricao': descricao,
                    'variacao': variacao,
                    'qtd': qtd,
                })
            return produtos
    return []


def total_qtd_etiqueta(etq):
    """Retorna quantidade total de itens de uma etiqueta."""
    return sum(p.get('qtd', 1) for p in etq.get('produtos', []) if p.get('tipo') == 'produto')


# ─── Ordenação igual ao Beka MKT ─────────────────────────────────────

def _separar_cor_numero(variacao):
    partes = re.split(r'[,/]', variacao or '', maxsplit=1)
    cor = partes[0].strip() if partes else ''
    num_str = partes[1].strip() if len(partes) > 1 else ''
    m = re.search(r'(\d+)', num_str)
    num_val = int(m.group(1)) if m else 99999
    return cor, num_val, num_str


def ordenar_etiquetas(etiquetas):
    """Ordena: qtd=1 primeiro, qtd>1 ao final. Dentro de cada bloco: SKU > Cor > Tamanho."""
    def chave(etq):
        produtos = [p for p in etq.get('produtos', []) if p.get('tipo') == 'produto']
        if produtos:
            sku = produtos[0].get('codigo', '')
            var = produtos[0].get('variacao', '')
        else:
            sku = ''
            var = ''
        cor, num_val, num_str = _separar_cor_numero(var)
        tq = total_qtd_etiqueta(etq)
        return (sku.casefold(), tq, cor.casefold(), num_val, num_str.casefold())

    simples = [e for e in etiquetas if total_qtd_etiqueta(e) <= 1]
    multiplos = [e for e in etiquetas if total_qtd_etiqueta(e) > 1]
    simples.sort(key=chave)
    multiplos.sort(key=chave)
    return simples + multiplos


# ─── Leitura dos PDFs ────────────────────────────────────────────────

def ler_pdfs(input_dir):
    """Lê todos os PDFs e retorna lista de etiquetas com metadados."""
    pdfs = sorted(glob.glob(os.path.join(input_dir, "*.pdf")))
    pdfs = [p for p in pdfs if '_etiqueta' not in os.path.basename(p)]

    etiquetas = []
    for pdf_path in pdfs:
        try:
            doc = fitz.open(pdf_path)
        except Exception as e:
            print(f"  ERRO ao abrir {os.path.basename(pdf_path)}: {e}")
            continue

        if len(doc) < 2:
            print(f"  AVISO: {os.path.basename(pdf_path)} sem declaração (< 2 pags)")
            doc.close()
            continue

        p1 = doc[0]
        p2 = doc[1]

        # Extrair loja
        loja = extrair_loja_remetente(p1.get_text())
        if not loja:
            loja = 'DESCONHECIDA'

        # Extrair produtos da declaração
        produtos = extrair_produtos_declaracao(p2)

        # Encontrar bounds reais da etiqueta (sem bordas brancas)
        drawings = p1.get_drawings()
        if drawings:
            all_rects = [d["rect"] for d in drawings]
            clip = fitz.Rect(
                min(r.x0 for r in all_rects),
                min(r.y0 for r in all_rects),
                max(r.x1 for r in all_rects),
                max(r.y1 for r in all_rects)
            )
        else:
            blocks = p1.get_text("blocks")
            if blocks:
                clip = fitz.Rect(
                    min(b[0] for b in blocks),
                    min(b[1] for b in blocks),
                    max(b[2] for b in blocks),
                    max(b[3] for b in blocks)
                )
            else:
                clip = p1.rect

        etiquetas.append({
            'pdf_path': pdf_path,
            'loja': loja,
            'produtos': produtos,
            'clip': clip,
            'page_rect': p1.rect,
        })
        doc.close()

    return etiquetas


# ─── Desenhar rodapé compacto ────────────────────────────────────────

def desenhar_rodape(page, produtos, y_start, page_width):
    """Desenha tabela compacta de produtos (fonte maior, negrito)."""
    if not produtos:
        return y_start

    margin = 8
    x_start = margin
    table_width = page_width - 2 * margin
    row_height = 14
    font_size = 8
    font_header = "hebo"   # Helvetica Bold
    font_data = "hebo"     # Helvetica Bold
    y = y_start + 3

    # Linha separadora
    page.draw_line(fitz.Point(x_start, y), fitz.Point(x_start + table_width, y),
                   color=(0, 0, 0), width=0.8)
    y += 2

    # Colunas: Nº | CÓDIGO | DESCRIÇÃO | VARIAÇÃO | QTD
    col_w = [
        table_width * 0.05,   # Nº
        table_width * 0.14,   # CÓDIGO
        table_width * 0.45,   # DESCRIÇÃO
        table_width * 0.26,   # VARIAÇÃO
        table_width * 0.10,   # QTD
    ]

    # Header
    headers = ['Nº', 'CÓDIGO (SKU)', 'DESCRIÇÃO DO PRODUTO', 'VARIAÇÃO', 'QTD']
    x = x_start
    for i, h in enumerate(headers):
        rect = fitz.Rect(x, y, x + col_w[i], y + row_height)
        page.insert_textbox(rect, h, fontsize=font_size, fontname=font_header,
                           color=(0, 0, 0),
                           align=fitz.TEXT_ALIGN_LEFT if i < 4 else fitz.TEXT_ALIGN_CENTER)
        x += col_w[i]
    y += row_height
    page.draw_line(fitz.Point(x_start, y), fitz.Point(x_start + table_width, y),
                   color=(0, 0, 0), width=0.5)

    # Dados
    for item in produtos:
        if item['tipo'] == 'totais':
            page.draw_line(fitz.Point(x_start, y), fitz.Point(x_start + table_width, y),
                           color=(0, 0, 0), width=0.5)
            total_w = col_w[0] + col_w[1] + col_w[2] + col_w[3]
            rect = fitz.Rect(x_start, y, x_start + total_w, y + row_height)
            page.insert_textbox(rect, "Totais", fontsize=font_size, fontname=font_header,
                               color=(0, 0, 0), align=fitz.TEXT_ALIGN_RIGHT)
            xt = x_start + total_w
            rect = fitz.Rect(xt, y, xt + col_w[4], y + row_height)
            page.insert_textbox(rect, str(item['qtd']), fontsize=font_size, fontname=font_header,
                               color=(0, 0, 0), align=fitz.TEXT_ALIGN_CENTER)
            y += row_height
            continue

        x = x_start
        vals = [str(item.get('n', '')), item.get('codigo', ''), item.get('descricao', ''),
                item.get('variacao', ''), str(item.get('qtd', ''))]
        for i, v in enumerate(vals):
            rect = fitz.Rect(x, y, x + col_w[i], y + row_height)
            align = fitz.TEXT_ALIGN_CENTER if i in (0, 4) else fitz.TEXT_ALIGN_LEFT
            page.insert_textbox(rect, v, fontsize=font_size, fontname=font_data,
                               color=(0, 0, 0), align=align)
            x += col_w[i]
        y += row_height

    return y


# ─── Gerar PDF por loja ──────────────────────────────────────────────

def gerar_pdf_loja(etiquetas, caminho_saida):
    """Gera PDF final com todas as etiquetas de uma loja (150x230mm)."""
    etiquetas_ord = ordenar_etiquetas(etiquetas)
    area_util = LARGURA_PT - MARGEM_ESQ - MARGEM_DIR
    doc_saida = fitz.open()

    for idx, etq in enumerate(etiquetas_ord):
        pdf_path = etq['pdf_path']
        clip = etq['clip']
        produtos = etq.get('produtos', [])

        doc_entrada = fitz.open(pdf_path)

        # Calcular rodapé (row_height=14 para fonte maior)
        n_prod = sum(1 for p in produtos if p.get('tipo') == 'produto')
        n_totais = sum(1 for p in produtos if p.get('tipo') == 'totais')
        rodape_h = (3 + 2 + 14 + n_prod * 14 + n_totais * 14 + 4) if produtos else 0

        # Escalar etiqueta para caber
        escala = area_util / clip.width
        alt_etiqueta = clip.height * escala

        # Se etiqueta + rodapé ultrapassa a página, reduzir
        espaco_disponivel = ALTURA_PT - MARGEM_TOPO - MARGEM_INF - rodape_h
        if alt_etiqueta > espaco_disponivel:
            escala = espaco_disponivel / clip.height
            alt_etiqueta = clip.height * escala

        larg_final = clip.width * escala
        x_offset = MARGEM_ESQ + (area_util - larg_final) / 2

        nova_pag = doc_saida.new_page(width=LARGURA_PT, height=ALTURA_PT)
        dest_rect = fitz.Rect(x_offset, MARGEM_TOPO, x_offset + larg_final, MARGEM_TOPO + alt_etiqueta)

        # Renderizar etiqueta como imagem de alta resolução (300 DPI)
        # para manter código de barras nítido após impressão
        dpi_render = 300
        zoom = dpi_render / 72.0
        mat = fitz.Matrix(zoom, zoom)
        pix = doc_entrada[0].get_pixmap(matrix=mat, clip=clip)
        nova_pag.insert_image(dest_rect, pixmap=pix)

        # Rodapé
        if produtos:
            desenhar_rodape(nova_pag, produtos, MARGEM_TOPO + alt_etiqueta, LARGURA_PT)

        # Número de ordem (como Beka)
        try:
            nova_pag.insert_text(
                (MARGEM_ESQ + 2, ALTURA_PT - MARGEM_INF - 8),
                f"p.{idx + 1}",
                fontsize=9, fontname="hebo", color=(0.4, 0.4, 0.4)
            )
        except:
            pass

        doc_entrada.close()

    doc_saida.save(caminho_saida)
    doc_saida.close()
    return len(etiquetas_ord)


# ─── Gerar resumo XLSX + JPEG ────────────────────────────────────────

def gerar_resumo_xlsx(etiquetas, caminho_saida, nome_loja):
    """Gera resumo XLSX com SKU + Variação + Quantidade."""
    if not HAS_OPENPYXL:
        print("  AVISO: openpyxl não instalado, resumo XLSX não gerado.")
        return

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Resumo {nome_loja}"

    # Agregar SKU + Variação
    sku_var_qtd = defaultdict(int)
    for etq in etiquetas:
        for prod in etq.get('produtos', []):
            if prod.get('tipo') != 'produto':
                continue
            codigo = prod.get('codigo', '') or 'SEM_SKU'
            variacao = prod.get('variacao', '') or ''
            qtd = prod.get('qtd', 1)
            if isinstance(qtd, str):
                try:
                    qtd = int(float(qtd))
                except:
                    qtd = 1
            sku_var_qtd[(codigo, variacao)] += qtd

    # Headers
    ws['A1'] = 'Cod. SKU'
    ws['B1'] = 'Variacao'
    ws['C1'] = 'Soma Quant.'
    for cell in [ws['A1'], ws['B1'], ws['C1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left')

    # Dados ordenados
    row = 2
    for (sku, var) in sorted(sku_var_qtd.keys(), key=lambda x: (x[0].casefold(), *_separar_cor_numero(x[1]))):
        ws.cell(row=row, column=1, value=sku).border = border
        ws.cell(row=row, column=2, value=var).border = border
        ws.cell(row=row, column=3, value=sku_var_qtd[(sku, var)]).border = border
        row += 1

    # Total
    ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2, value='').border = border
    ws.cell(row=row, column=3, value=sum(sku_var_qtd.values())).font = Font(bold=True)
    ws.cell(row=row, column=3).border = border

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15

    wb.save(caminho_saida)
    wb.close()

    # Gerar imagem JPEG do resumo
    gerar_imagem_resumo(caminho_saida)

    return len(sku_var_qtd), sum(sku_var_qtd.values())


def gerar_imagem_resumo(caminho_xlsx):
    """Gera imagem JPEG da planilha de resumo."""
    try:
        wb = openpyxl.load_workbook(caminho_xlsx)
        ws = wb.active
    except:
        return

    rows_data = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        rows_data.append([str(c or '') for c in row])
    wb.close()

    if not rows_data:
        return

    n_cols = len(rows_data[0])
    n_rows = len(rows_data)

    # Criar imagem com PyMuPDF
    col_widths = [180, 280, 100] if n_cols >= 3 else [180, 100]
    total_w = sum(col_widths) + 20
    row_h = 22
    total_h = (n_rows + 1) * row_h + 20

    doc = fitz.open()
    page = doc.new_page(width=total_w, height=total_h)

    y = 10
    for r_idx, row in enumerate(rows_data):
        x = 10
        is_header = (r_idx == 0)
        is_total = (r_idx == n_rows - 1)

        if is_header:
            page.draw_rect(fitz.Rect(10, y, total_w - 10, y + row_h),
                          fill=(0.85, 0.88, 0.95), color=(0, 0, 0), width=0.5)
        elif is_total:
            page.draw_rect(fitz.Rect(10, y, total_w - 10, y + row_h),
                          fill=(0.95, 0.95, 0.95), color=(0, 0, 0), width=0.5)
        else:
            page.draw_rect(fitz.Rect(10, y, total_w - 10, y + row_h),
                          color=(0.7, 0.7, 0.7), width=0.3)

        for c_idx, val in enumerate(row):
            if c_idx >= len(col_widths):
                break
            rect = fitz.Rect(x + 2, y + 2, x + col_widths[c_idx] - 2, y + row_h - 2)
            fs = 10 if is_header or is_total else 9
            page.insert_textbox(rect, val, fontsize=fs, fontname="helv", color=(0, 0, 0))
            x += col_widths[c_idx]
        y += row_h

    # Salvar como JPEG
    pix = page.get_pixmap(dpi=150)
    caminho_img = os.path.splitext(caminho_xlsx)[0] + '.jpeg'
    pix.save(caminho_img)
    doc.close()
    return caminho_img


# ─── Resumo geral ────────────────────────────────────────────────────

def gerar_resumo_geral(lojas_info, caminho_saida):
    """Gera resumo geral com totais por loja."""
    if not HAS_OPENPYXL:
        return

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo Geral"

    ws['A1'] = 'Loja'
    ws['B1'] = 'Etiquetas'
    ws['C1'] = 'SKUs'
    ws['D1'] = 'Unidades'
    for cell in [ws['A1'], ws['B1'], ws['C1'], ws['D1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    row = 2
    total_etq = 0
    total_skus = 0
    total_units = 0
    for loja, info in sorted(lojas_info.items()):
        ws.cell(row=row, column=1, value=loja).border = border
        ws.cell(row=row, column=2, value=info['etiquetas']).border = border
        ws.cell(row=row, column=3, value=info['skus']).border = border
        ws.cell(row=row, column=4, value=info['unidades']).border = border
        total_etq += info['etiquetas']
        total_skus += info['skus']
        total_units += info['unidades']
        row += 1

    ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=2, value=total_etq).font = Font(bold=True)
    ws.cell(row=row, column=2).border = border
    ws.cell(row=row, column=3, value=total_skus).font = Font(bold=True)
    ws.cell(row=row, column=3).border = border
    ws.cell(row=row, column=4, value=total_units).font = Font(bold=True)
    ws.cell(row=row, column=4).border = border

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12

    wb.save(caminho_saida)
    wb.close()
    gerar_imagem_resumo(caminho_saida)


# ─── Main ────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) > 1:
        input_dir = sys.argv[1]
    else:
        input_dir = os.path.dirname(os.path.abspath(__file__))

    output_base = os.path.join(input_dir, "etiquetas_prontas")
    os.makedirs(output_base, exist_ok=True)

    print("Lendo PDFs...")
    etiquetas = ler_pdfs(input_dir)
    if not etiquetas:
        print("Nenhum PDF válido encontrado.")
        return

    # Agrupar por loja
    lojas = defaultdict(list)
    for etq in etiquetas:
        lojas[etq['loja']].append(etq)

    print(f"\nTotal: {len(etiquetas)} etiquetas em {len(lojas)} lojas\n")

    lojas_info = {}

    for loja, etqs in sorted(lojas.items()):
        print(f"{'='*50}")
        print(f"  Loja: {loja} ({len(etqs)} etiquetas)")

        # Criar pasta da loja
        loja_safe = re.sub(r'[<>:"/\\|?*]', '_', loja)
        loja_dir = os.path.join(output_base, loja_safe)
        os.makedirs(loja_dir, exist_ok=True)

        # Gerar PDF
        pdf_path = os.path.join(loja_dir, f"etiquetas_{loja_safe}.pdf")
        n = gerar_pdf_loja(etqs, pdf_path)
        print(f"  PDF: {n} páginas -> {os.path.basename(pdf_path)}")

        # Gerar resumo XLSX + JPEG
        xlsx_path = os.path.join(loja_dir, f"resumo_{loja_safe}.xlsx")
        result = gerar_resumo_xlsx(etqs, xlsx_path, loja)
        if result:
            n_skus, n_units = result
            print(f"  Resumo: {n_skus} SKUs, {n_units} unidades")
            lojas_info[loja] = {'etiquetas': len(etqs), 'skus': n_skus, 'unidades': n_units}
        else:
            lojas_info[loja] = {'etiquetas': len(etqs), 'skus': 0, 'unidades': 0}

    # Resumo geral
    print(f"\n{'='*50}")
    geral_path = os.path.join(output_base, "resumo_geral.xlsx")
    gerar_resumo_geral(lojas_info, geral_path)
    print(f"Resumo geral salvo em: resumo_geral.xlsx")

    print(f"\nTudo pronto em: {output_base}")


if __name__ == "__main__":
    main()
