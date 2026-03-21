# -*- coding: utf-8 -*-
"""
Processador de Etiquetas Shopee
- Processa TODOS os PDFs e XMLs da pasta de entrada
- Identifica a loja (CNPJ/emitente) de cada etiqueta via XML
- Separa etiquetas por loja em pastas nomeadas pelo emitente
- Gera PDF final com 1 etiqueta por pagina (150mm x ~230mm)
- Adiciona secao DANFE: codigo de barras + tabela de produtos
- Organiza por SKU, multi-produto no final, numeracao sequencial
- Gera resumo XLSX por SKU para cada loja
- Salva tudo em C:\\Users\\Micro\\Desktop\\Etiquetas Prontas\\<nome_loja>
"""

import fitz  # PyMuPDF
import re
import os
import io
import glob
import zipfile
import collections
import unicodedata
from datetime import datetime
from collections import defaultdict, OrderedDict

# python-barcode para gerar Code128
import barcode
from barcode.writer import SVGWriter

# openpyxl para gerar XLSX
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class ProcessadorEtiquetasShopee:
    # Dimensoes da pagina de saida em pontos (1mm = 2.835pt)
    LARGURA_PT = 425.197   # 150mm
    ALTURA_PT = 651.969    # 230mm

    # Margens de seguranca para evitar cortes na impressao
    MARGEM_ESQUERDA = 8
    MARGEM_DIREITA = 8
    MARGEM_TOPO = 5
    MARGEM_INFERIOR = 5

    def __init__(self):
        self.dados_xml = {}      # nf -> dados completos do XML
        self.cnpj_nome = {}      # cnpj -> nome do emitente (nome limpo)
        self.cnpj_loja = {}      # cnpj -> nome da loja Shopee (extraido do REMETENTE)
        self.fonte_produto = 7   # tamanho da fonte para tabela de produtos (configuravel)
        self.exibicao_produto = 'sku'  # 'sku', 'titulo' ou 'ambos'
        self.dados_xlsx_global = {}    # order_sn -> {produtos, total_itens, total_qtd}
        self.dados_xlsx_tracking = {}  # tracking_number -> order_sn
        self.dados_lista_global = {}   # chave pedido/tracking -> {produtos, total_itens, total_qtd}
        self.dados_lista_seq_por_pdf = {}  # caminho_pdf -> [dados_pedido em ordem]
        self._easyocr_reader = None    # OCR lazy-load para fallback sem XLSX/XML

    # ----------------------------------------------------------------
    # LEITURA DOS XMLs
    # ----------------------------------------------------------------
    def carregar_todos_xmls(self, pasta):
        """Carrega XMLs de TODOS os ZIPs da pasta."""
        zips = [f for f in os.listdir(pasta) if f.lower().endswith('.zip')]
        total = 0
        for z in zips:
            caminho = os.path.join(pasta, z)
            n = self._carregar_zip(caminho)
            total += n
        print(f"  Total: {total} XMLs carregados de {len(zips)} arquivo(s) ZIP")
        return total

    def _carregar_zip(self, caminho_zip):
        """Carrega XMLs de um ZIP."""
        print(f"  Carregando: {os.path.basename(caminho_zip)}")
        contador = 0
        with zipfile.ZipFile(caminho_zip, 'r') as zf:
            for nome in zf.namelist():
                if not nome.lower().endswith('.xml'):
                    continue
                try:
                    conteudo = zf.read(nome)
                    dados = self._parse_xml(conteudo)
                    if dados and dados.get('nf'):
                        nf = dados['nf']
                        # Se ja existe, manter o mais completo
                        if nf not in self.dados_xml:
                            self.dados_xml[nf] = dados
                            contador += 1
                except Exception:
                    pass
        print(f"    {contador} XMLs novos")
        return contador

    def _limpar_nome_emitente(self, nome_raw):
        """Limpa o nome do emitente para usar como nome de pasta."""
        # Remove numeros de CNPJ do inicio tipo "34.847.700 "
        nome = re.sub(r'^\d[\d.]+\s+', '', nome_raw)
        # Remove CPF tipo "11543563619"
        nome = re.sub(r'\s+\d{11}$', '', nome)
        # Limpa LTDA, MEI, etc
        nome = re.sub(r'\s+(LTDA|ME|MEI|EPP|EIRELI)\s*$', '', nome, flags=re.IGNORECASE)
        # Capitaliza
        nome = nome.strip().title()
        # Remove caracteres invalidos para nome de pasta
        nome = re.sub(r'[<>:"/\\|?*]', '', nome)
        return nome.strip() or 'Loja_Desconhecida'

    def _parse_xml(self, conteudo_xml):
        """Extrai dados relevantes do XML da NFe."""
        try:
            xml = conteudo_xml.decode('utf-8', errors='ignore')
            dados = {}

            def get(tag):
                m = re.search(f'<{tag}>([^<]+)</{tag}>', xml)
                return m.group(1).strip() if m else ''

            dados['nf'] = get('nNF')
            dados['serie'] = get('serie') or '1'

            dhEmi = get('dhEmi')
            if dhEmi:
                dt_part = dhEmi[:19]
                try:
                    dt = datetime.strptime(dt_part, '%Y-%m-%dT%H:%M:%S')
                    dados['data_emissao'] = dt.strftime('%d-%m-%Y %H:%M:%S')
                except ValueError:
                    dados['data_emissao'] = dhEmi[:10]
            else:
                dados['data_emissao'] = ''

            # Chave de acesso
            m = re.search(r'Id="NFe(\d+)"', xml)
            dados['chave'] = m.group(1) if m else ''

            # CNPJ e nome do emitente
            cnpj_m = re.search(r'<emit>.*?<CNPJ>([^<]+)</CNPJ>', xml, re.DOTALL)
            nome_m = re.search(r'<emit>.*?<xNome>([^<]+)</xNome>', xml, re.DOTALL)
            dados['cnpj_emitente'] = cnpj_m.group(1) if cnpj_m else ''
            nome_raw = nome_m.group(1) if nome_m else ''
            dados['nome_emitente'] = nome_raw

            # Registra mapeamento CNPJ -> nome limpo
            if dados['cnpj_emitente'] and dados['cnpj_emitente'] not in self.cnpj_nome:
                self.cnpj_nome[dados['cnpj_emitente']] = self._limpar_nome_emitente(nome_raw)

            # Produtos
            produtos = []
            for m in re.finditer(r'<det[^>]*>.*?<prod>(.*?)</prod>', xml, re.DOTALL):
                bloco = m.group(1)
                codigo = re.search(r'<cProd>([^<]+)</cProd>', bloco)
                descricao = re.search(r'<xProd>([^<]+)</xProd>', bloco)
                qtd = re.search(r'<qCom>([^<]+)</qCom>', bloco)

                produtos.append({
                    'codigo': codigo.group(1) if codigo else '',
                    'descricao': descricao.group(1) if descricao else '',
                    'qtd': qtd.group(1) if qtd else '1',
                })

            dados['produtos'] = produtos
            dados['total_itens'] = len(produtos)
            dados['total_qtd'] = sum(int(float(p['qtd'])) for p in produtos)

            return dados if dados.get('nf') else None
        except Exception:
            return None

    # ----------------------------------------------------------------
    # RECORTE DAS ETIQUETAS DO PDF DA SHOPEE
    # ----------------------------------------------------------------
    def _eh_pdf_shein(self, caminho_pdf):
        """Detecta se um PDF e do tipo Shein (paginas alternadas: etiqueta + DANFE ou Declaracao).
        Retorna True se o PDF tem estrutura Shein.
        """
        try:
            doc = fitz.open(caminho_pdf)
            n_pags = len(doc)
            if n_pags < 2:
                doc.close()
                return False

            # Verificar primeiras 2 paginas: pag0=etiqueta Shein, pag1=DANFE ou Declaracao
            texto_p0 = doc[0].get_text()
            texto_p1 = doc[1].get_text()
            doc.close()

            p0_shein = self._eh_etiqueta_shein(texto_p0)
            p1_danfe = 'DANFE' in texto_p1.upper() and 'CHAVE' in texto_p1.upper()
            # Detectar Declaracao de Conteudo (formato alternativo Shein)
            texto_p1_upper = texto_p1.upper()
            p1_declaracao = (
                ('DECLARAÇÃO DE CONTEÚDO' in texto_p1 or 'DECLARACAO DE CONTEUDO' in texto_p1_upper)
                and ('IDENTIFICAÇÃO DOS BENS' in texto_p1 or 'IDENTIFICACAO DOS BENS' in texto_p1_upper)
            )

            return p0_shein and (p1_danfe or p1_declaracao)
        except Exception:
            return False

    def carregar_todos_pdfs(self, pasta):
        """Carrega etiquetas sem recorte (compatibilidade).

        Regra atual do sistema: nao recortar etiqueta, apenas organizar ordem.
        """
        return self.carregar_todos_pdfs_sem_recorte(pasta)

    def carregar_todos_pdfs_sem_recorte(self, pasta):
        """Carrega etiquetas da pasta sem recortar pagina (1 etiqueta por pagina).

        Usado no fluxo novo da aba Automacao/UpSeller, onde o PDF ja vem
        pronto em formato 10x15 por pagina.
        """
        especiais_lower = [p.lower() for p in self.PDFS_ESPECIAIS]
        # Carregar lista de separacao (quando presente) para preencher rodape
        # mesmo com XLSX vindo sem product_info.
        try:
            self.carregar_todas_listas_separacao(pasta)
        except Exception as e:
            print(f"  Aviso: falha ao carregar lista de separacao: {e}")

        pdfs = [f for f in os.listdir(pasta)
                if f.lower().endswith('.pdf')
                and not f.startswith('etiquetas_prontas')
                and not f.lower().startswith('lanim')
                and f.lower() not in especiais_lower]

        pdfs_shein_detectados = []
        pdfs_normais = []
        for pdf_name in pdfs:
            caminho = os.path.join(pasta, pdf_name)
            if self._eh_pdf_shein(caminho):
                pdfs_shein_detectados.append(caminho)
                print(f"  Detectado como Shein: {pdf_name}")
            else:
                pdfs_normais.append(pdf_name)

        todas_etiquetas = []
        etiquetas_cpf_auto = []
        for pdf_name in pdfs_normais:
            caminho = os.path.join(pasta, pdf_name)
            etqs = self.carregar_pdf_pagina_inteira(caminho, 'retirada')
            for etq in etqs:
                if etq.get('tipo_especial') == 'cpf':
                    etiquetas_cpf_auto.append(etq)
                else:
                    todas_etiquetas.append(etq)

        if etiquetas_cpf_auto:
            print(f"  CPF detectadas automaticamente: {len(etiquetas_cpf_auto)} etiquetas")
        if pdfs_shein_detectados:
            print(f"  Shein detectados automaticamente: {len(pdfs_shein_detectados)} PDF(s)")
        print(f"  Total (sem recorte): {len(todas_etiquetas)} etiquetas normais de {len(pdfs_normais)} PDF(s)")
        return todas_etiquetas, etiquetas_cpf_auto, pdfs_shein_detectados

    @staticmethod
    def _remover_acentos(texto):
        txt = str(texto or '')
        return ''.join(ch for ch in unicodedata.normalize('NFD', txt) if unicodedata.category(ch) != 'Mn')

    @staticmethod
    def _normalizar_chave_pedido(chave):
        return re.sub(r'[^A-Z0-9]', '', str(chave or '').upper())

    def _extrair_chaves_pedido_texto(self, texto):
        """Extrai possiveis chaves de pedido/tracking de um texto qualquer."""
        txt = str(texto or '').upper()
        txt = txt.replace('[', ' ').replace(']', ' ')
        candidatos = []
        padroes = [
            r'\bBR[0-9A-Z]{10,24}\b',         # tracking BR...
            r'\b\d{6}[A-Z0-9]{6,12}\b',       # order_sn Shopee
            r'\bUP[0-9A-Z]{6,20}\b',          # codigo UP...
            r'\bMEL[0-9A-Z]{6,20}\b',         # codigo MEL...
            r'\b\d{11,20}\b',                 # ids numericos longos
        ]
        for pad in padroes:
            for m in re.finditer(pad, txt):
                candidatos.append(self._normalizar_chave_pedido(m.group(0)))

        out = []
        vistos = set()
        for c in candidatos:
            if c and c not in vistos:
                vistos.add(c)
                out.append(c)
        return out

    @staticmethod
    def _agrupar_linhas_words(words, tol=2.2):
        """Agrupa palavras em linhas por coordenada Y."""
        itens = []
        for w in words:
            try:
                x0, y0, x1, y1, t = w[:5]
                texto = str(t or '').strip()
                if not texto:
                    continue
                yc = (float(y0) + float(y1)) / 2.0
                itens.append({'x0': float(x0), 'x1': float(x1), 'y': yc, 't': texto})
            except Exception:
                continue

        itens.sort(key=lambda k: (k['y'], k['x0']))
        linhas = []
        for it in itens:
            if not linhas or abs(it['y'] - linhas[-1]['y']) > tol:
                linhas.append({'y': it['y'], 'w': [it]})
            else:
                linhas[-1]['w'].append(it)

        for ln in linhas:
            ln['w'].sort(key=lambda k: k['x0'])
            ln['txt'] = ' '.join(x['t'] for x in ln['w'])

        return linhas

    @staticmethod
    def _construir_dados_produtos(produtos, fonte='lista_separacao'):
        lista = []
        for p in (produtos or []):
            if not isinstance(p, dict):
                continue
            codigo = str(p.get('codigo', '') or '').strip()
            descricao = str(p.get('descricao', '') or '').strip()
            variacao = str(p.get('variacao', '') or '').strip()
            qtd_raw = p.get('qtd', '1')
            try:
                qtd_int = max(1, int(float(qtd_raw)))
            except Exception:
                qtd_int = 1
            if codigo or descricao or variacao:
                lista.append({
                    'codigo': codigo,
                    'descricao': descricao,
                    'variacao': variacao,
                    'qtd': str(qtd_int),
                })

        return {
            'produtos': lista,
            'total_itens': len(lista),
            'total_qtd': sum(int(float(x.get('qtd', '1') or '1')) for x in lista),
            'fonte_dados': fonte,
        }

    def _mesclar_dados_produtos(self, existente, novo):
        """Mescla produtos sem duplicar linha (codigo+descricao+variacao)."""
        if not existente:
            return {
                'produtos': list((novo or {}).get('produtos', [])),
                'total_itens': int((novo or {}).get('total_itens', 0) or 0),
                'total_qtd': int((novo or {}).get('total_qtd', 0) or 0),
                'fonte_dados': (novo or {}).get('fonte_dados', 'lista_separacao'),
            }

        produtos = list((existente or {}).get('produtos', []))
        chaves = set()
        for p in produtos:
            chaves.add((p.get('codigo', ''), p.get('descricao', ''), p.get('variacao', '')))

        for p in (novo or {}).get('produtos', []):
            k = (p.get('codigo', ''), p.get('descricao', ''), p.get('variacao', ''))
            if k not in chaves:
                produtos.append(p)
                chaves.add(k)

        return self._construir_dados_produtos(produtos, fonte=(novo or {}).get('fonte_dados', 'lista_separacao'))

    def _extrair_itens_lista_separacao_pagina(self, pagina):
        """Extrai linhas da tabela da Lista de Separacao.

        Retorna lista de dict:
          {'chaves': [...], 'produto': {codigo, descricao, variacao, qtd}}
        """
        words = pagina.get_text("words")
        if not words:
            return []

        linhas = self._agrupar_linhas_words(words)
        if not linhas:
            return []

        larg = float(pagina.rect.width)
        x_titulo = None
        x_sku = None
        x_qtd = None
        y_header = None

        for ln in linhas:
            txt_norm = self._remover_acentos(ln['txt']).upper()
            if ('TITULO' in txt_norm or 'VARIACAO' in txt_norm) and 'SKU' in txt_norm:
                y_header = ln['y']
                for w in ln['w']:
                    wt = self._remover_acentos(w['t']).upper()
                    if x_titulo is None and ('TITULO' in wt or 'VARIACAO' in wt):
                        x_titulo = w['x0']
                    if x_sku is None and wt == 'SKU':
                        x_sku = w['x0']
                    if x_qtd is None and wt.startswith('QTD'):
                        x_qtd = w['x0']
                break

        if x_titulo is None:
            x_titulo = larg * 0.30
        if x_sku is None:
            x_sku = larg * 0.62
        if x_qtd is None:
            x_qtd = larg * 0.90

        itens = []
        chaves_atuais = []

        for ln in linhas:
            if y_header is not None and ln['y'] <= y_header + 2.0:
                continue

            txt_linha = re.sub(r'\s+', ' ', ln['txt']).strip()
            if not txt_linha:
                continue

            up = self._remover_acentos(txt_linha).upper()
            if (
                'NOTAS DO COMPRADOR' in up or
                'OBSERVACOES' in up or
                'CUSTOMER NOTES' in up or
                'INTERNAL NOTES' in up
            ):
                continue

            col_order = []
            col_title = []
            col_sku = []
            col_qtd = []
            for w in ln['w']:
                cx = (w['x0'] + w['x1']) / 2.0
                if cx < x_titulo:
                    col_order.append(w['t'])
                elif cx < x_sku:
                    col_title.append(w['t'])
                elif cx < x_qtd:
                    col_sku.append(w['t'])
                else:
                    col_qtd.append(w['t'])

            txt_order = ' '.join(col_order).strip()
            txt_title = ' '.join(col_title).strip()
            txt_sku = ' '.join(col_sku).strip()
            txt_qtd = ' '.join(col_qtd).strip()

            chaves_linha = self._extrair_chaves_pedido_texto(txt_order or txt_linha)
            if chaves_linha:
                chaves_atuais = chaves_linha

            if not chaves_atuais:
                continue

            m_qtd = re.search(r'(\d{1,4})\s*$', txt_qtd) or re.search(r'\bQTD\.?\s*[:\-]?\s*(\d{1,4})\b', up)
            if not m_qtd:
                m_qtd = re.search(r'\b(\d{1,4})\b', txt_qtd)
            qtd = m_qtd.group(1) if m_qtd else ''

            sku = ''
            for tok in re.split(r'\s+', txt_sku):
                t = tok.strip()
                if not t:
                    continue
                t_up = self._normalizar_chave_pedido(t)
                if t_up in ('SKU', 'ARMAZEM', 'ESTANTE'):
                    continue
                if re.match(r'^[A-Z0-9][A-Z0-9._/-]{1,}$', t_up):
                    sku = t_up
                    break

            variacao = re.sub(r'\s+', ' ', txt_title).strip()
            if not variacao and not sku:
                continue
            if not qtd:
                qtd = '1'

            item_prod = {
                'codigo': sku,
                'descricao': '',
                'variacao': variacao,
                'qtd': qtd,
            }
            itens.append({
                'chaves': list(chaves_atuais),
                'produto': item_prod,
            })

        return itens

    def carregar_todas_listas_separacao(self, pasta):
        """Escaneia PDFs e popula mapa de produtos via Lista de Separacao."""
        self.dados_lista_global = {}
        self.dados_lista_seq_por_pdf = {}

        pdfs = [f for f in os.listdir(pasta)
                if f.lower().endswith('.pdf')
                and not f.startswith('etiquetas_prontas')]
        total_pedidos = 0
        total_produtos = 0

        for pdf_nome in sorted(pdfs):
            caminho_pdf = os.path.join(pasta, pdf_nome)
            try:
                doc = fitz.open(caminho_pdf)
            except Exception:
                continue

            seq = []
            try:
                por_chave = OrderedDict()
                for i in range(len(doc)):
                    pag = doc[i]
                    texto = pag.get_text()
                    if not self._eh_pagina_lista_separacao(texto):
                        continue
                    itens_pag = self._extrair_itens_lista_separacao_pagina(pag)
                    for it in itens_pag:
                        chaves = list(it.get('chaves', []) or [])
                        prod = dict(it.get('produto', {}) or {})
                        if not chaves:
                            continue
                        prim = None
                        chaves_set = set(chaves)
                        for chave_existente, entry_existente in por_chave.items():
                            if chaves_set.intersection(set(entry_existente.get('chaves', []))):
                                prim = chave_existente
                                break

                        if prim is None:
                            prim = chaves[0]
                            por_chave[prim] = {
                                'chaves': list(chaves),
                                'produtos': [],
                            }

                        for k in chaves:
                            if k not in por_chave[prim]['chaves']:
                                por_chave[prim]['chaves'].append(k)
                        por_chave[prim]['produtos'].append(prod)

                if por_chave:
                    for entry in por_chave.values():
                        dados = self._construir_dados_produtos(entry.get('produtos', []), fonte='lista_separacao')
                        if not dados.get('produtos'):
                            continue
                        seq.append(dados)
                        total_pedidos += 1
                        total_produtos += len(dados.get('produtos', []))
                        for chave in entry.get('chaves', []):
                            k = self._normalizar_chave_pedido(chave)
                            if not k:
                                continue
                            existente = self.dados_lista_global.get(k)
                            self.dados_lista_global[k] = self._mesclar_dados_produtos(existente, dados)
            finally:
                doc.close()

            if seq:
                self.dados_lista_seq_por_pdf[os.path.realpath(caminho_pdf)] = seq
                print(f"    Lista separacao: {pdf_nome} -> {len(seq)} pedido(s)")

        if total_pedidos > 0:
            print(f"  Lista separacao: {total_pedidos} pedido(s), {total_produtos} produto(s)")

    def _carregar_pdf(self, caminho_pdf):
        """Carrega e recorta etiquetas de um PDF."""
        print(f"  Carregando: {os.path.basename(caminho_pdf)}")
        doc = fitz.open(caminho_pdf)
        etiquetas = []

        for num_pag in range(len(doc)):
            pagina = doc[num_pag]
            etqs = self._recortar_pagina(pagina, caminho_pdf)
            etiquetas.extend(etqs)

        doc.close()
        print(f"    {len(etiquetas)} etiquetas")
        return etiquetas

    def _extrair_nf_quadrante(self, pagina, clip):
        """Extrai o numero da NF do texto dentro de um quadrante."""
        texto = pagina.get_text(clip=clip)

        m = re.search(r'Emiss.o:\n(\d+)\n', texto)
        if m:
            return m.group(1)

        # Padrao alternativo: "NF: 12345" (comum em etiquetas de retirada do comprador)
        m = re.search(r'NF:\s*(\d+)', texto)
        if m:
            return m.group(1)

        m = re.search(r'(\d{4,6})\n\d\n\d{2}-\d{2}-\d{4}', texto)
        if m:
            return m.group(1)

        return None

    def _extrair_chave_nfe(self, texto):
        """Extrai a chave de acesso da NFe (44 digitos) do texto da etiqueta."""
        m = re.search(r'(\d{44})', texto)
        return m.group(1) if m else ''

    def _extrair_nome_loja_remetente(self, texto):
        """Extrai o nome da loja do campo REMETENTE da etiqueta Shopee.

        Formato tipico da etiqueta Shopee:
            [dados destinatario]
            [tracking BR...]
            NOME_DA_LOJA          <-- queremos este
            [endereco remetente]
            [CEP remetente]

        Estrategia principal: encontrar o tracking code (BR + digitos + letra)
        e pegar a PRIMEIRA linha seguinte que parece um nome de loja.
        """
        def _eh_endereco_ou_cep(linha):
            """Retorna True se a linha parece ser endereco, CEP ou dado irrelevante."""
            l = linha.strip()
            if not l or len(l) < 3:
                return True
            # CEPs: 12345-678 ou 12345678 ou apenas numeros
            if re.match(r'^[\d\s.,-/]+$', l):
                return True
            if re.match(r'^\d{5}-?\d{3}', l):
                return True
            # Enderecos
            if re.match(r'^(Rua|Avenida|Travessa|Alameda|Estrada|Rodovia|Praca|Largo|R\.|Av\.|Rod\.|Est\.)\s', l, re.IGNORECASE):
                return True
            # Bairro / Complemento com numeros no meio: "Bloco A, 123" etc
            if re.match(r'^(Bloco|Lote|Quadra|Qd|Lt|Sl|Sala|Apto|Apt|Conj|Casa|Galpao|N[°o]?\s)', l, re.IGNORECASE):
                return True
            # Linha que e so "Cidade, Estado" ou "Cidade - UF" ou "UF"
            if re.match(r'^[A-Z]{2}$', l):
                return True
            # Formato "Cidade, Estado" ou "Cidade - UF" (ex: "Italva, Rio de Janeiro")
            if re.match(r'^[A-Za-z\s]+,\s*[A-Za-z\s]+$', l) and len(l.split(',')) == 2:
                parte2 = l.split(',')[1].strip()
                # Se a segunda parte parece um estado brasileiro
                estados = ['acre', 'alagoas', 'amapa', 'amazonas', 'bahia', 'ceara',
                    'distrito federal', 'espirito santo', 'goias', 'maranhao', 'mato grosso',
                    'mato grosso do sul', 'minas gerais', 'para', 'paraiba', 'parana',
                    'pernambuco', 'piaui', 'rio de janeiro', 'rio grande do norte',
                    'rio grande do sul', 'rondonia', 'roraima', 'santa catarina',
                    'sao paulo', 'sergipe', 'tocantins']
                if parte2.lower() in estados or len(parte2) == 2:
                    return True
            # Texto com CEP embutido
            if re.search(r'\d{5}-?\d{3}', l):
                return True
            # Palavras-chave de etiqueta
            if re.match(r'^(envio previsto|peso|volume|frete|destinat|remet|cep)', l, re.IGNORECASE):
                return True
            return False

        # Estrategia 1: Buscar apos tracking code tipo BR261920610412I
        m = re.search(r'BR\d{10,}[A-Z]\s*\n((?:[^\n]+\n){1,5})', texto)
        if m:
            linhas = [l.strip() for l in m.group(1).split('\n') if l.strip()]
            for linha in linhas:
                if not _eh_endereco_ou_cep(linha):
                    return linha

        # Estrategia 2: Buscar apos tracking generico (XX + digitos)
        m2 = re.search(r'[A-Z]{2}\d{9,}[A-Z]?\s*\n((?:[^\n]+\n){1,5})', texto)
        if m2:
            linhas = [l.strip() for l in m2.group(1).split('\n') if l.strip()]
            for linha in linhas:
                if not _eh_endereco_ou_cep(linha):
                    return linha

        return None

    def _inferir_loja_por_nome_arquivo(self, caminho_pdf):
        """Infere nome da loja a partir do nome do arquivo de etiqueta do UpSeller.

        Ex.: etiquetas_BEKA_20260227_171152.pdf -> BEKA
        """
        try:
            base = os.path.splitext(os.path.basename(caminho_pdf))[0].strip()
            m = re.match(r'^etiquetas_(.+?)_\d{8}_\d{6}$', base, flags=re.IGNORECASE)
            if not m:
                return None
            loja_raw = (m.group(1) or '').strip()
            if not loja_raw or loja_raw.lower() in ('todas', 'all', 'geral'):
                return None
            loja = re.sub(r'\s+', ' ', loja_raw.replace('_', ' ')).strip()
            loja = re.sub(r'[<>:"/\\|?*]', '', loja).strip().rstrip('.')
            return loja or None
        except Exception:
            return None

    def _registrar_loja_sintetica(self, nome_loja, prefixo='LOJA'):
        """Registra loja sintética no mapa CNPJ->nome e retorna a chave."""
        nome_loja = (nome_loja or '').strip()
        if not nome_loja:
            return ''
        cnpj_sintetico = f"{prefixo}_{re.sub(r'[^A-Za-z0-9]', '_', nome_loja)}"
        if cnpj_sintetico not in self.cnpj_loja:
            self.cnpj_loja[cnpj_sintetico] = nome_loja
            self.cnpj_nome[cnpj_sintetico] = nome_loja
        return cnpj_sintetico

    def get_nome_loja(self, cnpj):
        """Retorna nome da loja: primeiro tenta cnpj_loja (Shopee), depois cnpj_nome (XML)."""
        nome = self.cnpj_loja.get(cnpj) or self.cnpj_nome.get(cnpj, 'Loja_Desconhecida')
        # Sanitizar para nome de pasta Windows (remover caracteres ilegais)
        nome = re.sub(r'[<>:"/\\|?*]', '', nome).strip().rstrip('.')
        return nome or 'Loja_Desconhecida'

    def remover_duplicatas(self, etiquetas):
        """Remove etiquetas duplicadas usando multiplas chaves de identificacao.

        Prioridade de deduplicacao:
        1. tracking (BR...) - unico por pedido, funciona cross-loja
        2. order_sn (numero do pedido Shopee) - unico por pedido
        3. chave NFe (43 digitos) - unica por nota fiscal
        4. NF (numero da nota fiscal) - fallback (NFs sinteticas nao deduplicam)

        Retorna (etiquetas_unicas, lista_de_duplicadas_removidas).
        """
        vistos_tracking = set()
        vistos_order_sn = set()
        vistos_chave = set()
        vistos_nf = set()
        unicas = []
        duplicadas = []

        for etq in etiquetas:
            tracking = str(etq.get('tracking', '') or '').strip()
            order_sn = str(etq.get('order_sn', '') or '').strip()
            chave_nfe = str((etq.get('dados_xml') or {}).get('chave', '') or '').strip()
            nf = str(etq.get('nf', '') or '').strip()

            is_dup = False

            # 1. Dedup por tracking (mais confiavel, cross-loja)
            if tracking and len(tracking) >= 12:
                if tracking in vistos_tracking:
                    is_dup = True
                else:
                    vistos_tracking.add(tracking)

            # 2. Dedup por order_sn
            if not is_dup and order_sn and len(order_sn) >= 10:
                if order_sn in vistos_order_sn:
                    is_dup = True
                else:
                    vistos_order_sn.add(order_sn)

            # 3. Dedup por chave NFe
            if not is_dup and chave_nfe and len(chave_nfe) >= 40:
                if chave_nfe in vistos_chave:
                    is_dup = True
                else:
                    vistos_chave.add(chave_nfe)

            # 4. Dedup por NF (exceto sinteticas SEM_NF_)
            if not is_dup and nf and not nf.startswith('SEM_NF_'):
                if nf in vistos_nf:
                    is_dup = True
                else:
                    vistos_nf.add(nf)

            if is_dup:
                duplicadas.append(etq)
            else:
                unicas.append(etq)

        if duplicadas:
            print(f"  [Dedup] {len(duplicadas)} duplicata(s) removida(s) "
                  f"(tracking={len(vistos_tracking)}, order_sn={len(vistos_order_sn)}, "
                  f"chave={len(vistos_chave)}, nf={len(vistos_nf)})")

        return unicas, duplicadas

    def _contar_etiquetas_regiao(self, pagina, clip):
        """Verifica se uma regiao contem uma etiqueta Shopee.
        Detecta por marcadores: 'Pedido:' ou 'REMETENTE' ou NF numerica.
        """
        texto = pagina.get_text(clip=clip).strip()
        if len(texto) < 10:
            return False
        # Marcadores de etiqueta Shopee
        tem_pedido = 'Pedido:' in texto or 'Pedido\n' in texto
        tem_remetente = 'REMETENTE' in texto
        tem_danfe = 'DANFE' in texto
        tem_nf = self._extrair_nf_quadrante(pagina, clip) is not None
        # Basta ter 1 marcador forte ou NF
        return tem_nf or tem_pedido or (tem_remetente and tem_danfe)

    def _detectar_layout_pagina(self, pagina):
        """Detecta quantas etiquetas ha na pagina analisando o conteudo.
        Retorna lista de clips (regioes) para recortar.
        Layouts possiveis:
          - 4 etiquetas (grid 2x2) - padrao Shopee para pedidos simples e multi-produto
          - 2 etiquetas (2 linhas, largura total) - formato alternativo
          - 1 etiqueta (pagina inteira) - fallback
        Deteccao baseada em marcadores de etiqueta (Pedido:, REMETENTE, DANFE), nao apenas NFs.
        Paginas pequenas (<= 420pt largura) sao sempre 1 etiqueta (tamanho de 1 quadrante A4).
        """
        rect = pagina.rect
        larg = rect.width
        alt = rect.height

        # Pagina pequena = 1 etiqueta (ex: 297x419, tamanho de 1 quadrante A4)
        # A4 = 595x842, metade = ~297x421. Se largura <= 420, e uma etiqueta individual.
        if larg <= 420:
            return [fitz.Rect(0, 0, larg, alt)]

        meio_x = larg / 2
        meio_y = alt / 2

        # Testar grid 2x2 primeiro (padrao Shopee)
        quadrantes_2x2 = [
            fitz.Rect(0, 0, meio_x, meio_y),
            fitz.Rect(meio_x, 0, larg, meio_y),
            fitz.Rect(0, meio_y, meio_x, alt),
            fitz.Rect(meio_x, meio_y, larg, alt),
        ]

        etiquetas_2x2 = sum(1 for clip in quadrantes_2x2 if self._contar_etiquetas_regiao(pagina, clip))

        # Se encontrou 2+ etiquetas no grid 2x2, usar este layout
        if etiquetas_2x2 >= 2:
            return quadrantes_2x2

        # Testar layout 2 etiquetas (metade superior + metade inferior, largura total)
        quadrantes_2x1 = [
            fitz.Rect(0, 0, larg, meio_y),
            fitz.Rect(0, meio_y, larg, alt),
        ]

        etiquetas_2x1 = sum(1 for clip in quadrantes_2x1 if self._contar_etiquetas_regiao(pagina, clip))

        if etiquetas_2x1 >= 2:
            return quadrantes_2x1

        # Se grid 2x2 achou pelo menos 1, usar 2x2 (pode ter quadrantes vazios)
        if etiquetas_2x2 >= 1:
            return quadrantes_2x2

        # Se 2x1 achou pelo menos 1, usar 2x1
        if etiquetas_2x1 >= 1:
            return quadrantes_2x1

        # Fallback: pagina inteira
        return [fitz.Rect(0, 0, larg, alt)]

    def _eh_etiqueta_shein(self, texto):
        """Verifica se o texto pertence a uma etiqueta Shein.
        Etiquetas Shein tem padroes distintos: PUDO-PGK, Ref.No:GSH, Ref.No:GC,
        codigos GC seguidos de numeros longos, etc.
        """
        # Etiqueta de envio Shein
        if 'PUDO-PGK' in texto or 'PUDO' in texto:
            return True
        if 'Ref.No:GSH' in texto or 'Ref.No:GC' in texto:
            return True
        # Codigo GC longo (tracking Shein)
        if re.search(r'GC\d{15,}', texto):
            return True
        return False

    def _eh_declaracao_conteudo(self, texto):
        """Detecta se o quadrante é uma Declaração de Conteúdo.
        DESABILITADO: retorna sempre False para evitar conflitos.
        """
        return False  # DESABILITADO - use XLSX para dados de produtos

    def _extrair_produtos_declaracao(self, texto):
        """Extrai produtos da Declaração de Conteúdo.
        Retorna lista de dicionários com codigo, descricao, variacao, qtd.
        """
        produtos = []
        linhas = texto.split('\n')
        
        # Procurar tabela de produtos (N° CÓDIGO DESCRIÇÃO VARIAÇÃO)
        em_tabela = False
        for i, linha in enumerate(linhas):
            linha_upper = linha.upper()
            
            # Detectar início da tabela
            if 'C' in linha_upper and 'DIGO' in linha_upper and 'DESCRI' in linha_upper:
                em_tabela = True
                continue
            
            # Parar na seção de totais
            if em_tabela and ('TOTAIS' in linha_upper or 'PESO TOTAL' in linha_upper):
                break
            
            # Extrair produtos (formato: número + código + descrição)
            if em_tabela and linha.strip():
                # Remover linhas de cabeçalho/observações
                if any(x in linha_upper for x in ['OBSERV', 'ASSINATURA', 'CRIME', 'CONSTITUI']):
                    break
                
                # Tentar extrair: número, código, descrição
                partes = linha.split(None, 2)  # Split máximo 3 partes
                if len(partes) >= 3:
                    try:
                        num = int(partes[0])  # Primeiro campo é número
                        codigo = partes[1]
                        descricao = partes[2] if len(partes) > 2 else ''
                        
                        # Limpar descrição
                        descricao = descricao.strip()
                        
                        produtos.append({
                            'codigo': codigo,
                            'descricao': descricao,
                            'variacao': '',
                            'qtd': '1'
                        })
                    except (ValueError, IndexError):
                        continue
        
        # Tentar extrair quantidade total (Peso Total ou quantidade)
        for linha in linhas:
            # Buscar "2" sozinho que geralmente indica quantidade
            m = re.search(r'^(\d+)\s*$', linha.strip())
            if m and produtos:
                qtd_total = m.group(1)
                # Distribuir quantidade pelos produtos
                if len(produtos) == 1:
                    produtos[0]['qtd'] = qtd_total
        
        return produtos

    def _detectar_area_tabela_declaracao(self, pag):
        """Encontra o retangulo da tabela 'IDENTIFICACAO DOS BENS' na pagina de declaracao.
        Retorna fitz.Rect cobrindo desde o cabecalho da tabela ate a linha Total.
        """
        try:
            pag_rect = pag.rect
            # Buscar "IDENTIFICAÇÃO DOS BENS" ou variantes
            rects_id = pag.search_for("IDENTIFICAÇÃO DOS BENS")
            if not rects_id:
                rects_id = pag.search_for("IDENTIFICACAO DOS BENS")
            if not rects_id:
                # Fallback: buscar cabecalho "SKU"
                rects_id = pag.search_for("SKU")

            # Buscar "Total" (ultima ocorrencia na pagina)
            rects_total = pag.search_for("Total")

            if not rects_id:
                # Nao encontrou cabecalho — retornar terco inferior da pagina como fallback
                return fitz.Rect(
                    pag_rect.x0 + 5, pag_rect.height * 0.28,
                    pag_rect.x1 - 5, pag_rect.height * 0.52
                )

            y_topo = rects_id[0].y0 - 3  # Pequena margem acima

            if rects_total:
                # Usar ultimo "Total" encontrado (pode haver varios)
                y_base = rects_total[-1].y1 + 4  # Margem abaixo
            else:
                # Fallback: pegar area abaixo do cabecalho
                y_base = min(y_topo + 120, pag_rect.height - 20)

            return fitz.Rect(
                pag_rect.x0 + 5,   # margem esquerda
                y_topo,
                pag_rect.x1 - 5,   # margem direita
                y_base
            )
        except Exception as e:
            print(f"  Aviso: falha ao detectar area tabela declaracao: {e}")
            # Fallback seguro
            pag_rect = pag.rect
            return fitz.Rect(
                pag_rect.x0 + 5, pag_rect.height * 0.28,
                pag_rect.x1 - 5, pag_rect.height * 0.52
            )

    def _parse_declaracao_conteudo(self, texto):
        """Extrai dados basicos da Declaracao de Conteudo Shein.
        Retorna dict compativel com dados_danfe para sorting.
        Nota: esses dados sao para ordenacao, nao para display (display usa imagem recortada).

        Formato do texto extraido pelo PyMuPDF (campos em linhas separadas):
          IDENTIFICAÇÃO DOS BENS
          Nº
          SKU
          DESCRIÇÃO
          VARIAÇÃO
          QTD
          1                    <- item number
          Rakka-Pink-          <- SKU (pode ser multi-linha)
          BR35/36
          -                    <- descricao
          Pink-                <- variacao (pode ser multi-linha)
          BR35/36
          1                    <- quantidade
          10-03-2026           <- data
          Total
          1                    <- total
        """
        resultado = {
            'tracking': '',
            'nf': '',
            'chave': '',
            'cnpj_emitente': '',
            'nome_emitente': '',
            'produtos_shein': [],
            'total_itens': 0,
            'total_qtd': 0,
        }

        # Extrair tracking do "Codigo de Rastreamento: GC..."
        m_track = re.search(r'Rastreamento[:\s]*([A-Z0-9]{10,})', texto, re.IGNORECASE)
        if m_track:
            resultado['tracking'] = m_track.group(1).strip()

        # Extrair CNPJ do remetente (apenas da secao REMETENTE, antes de DESTINATARIO)
        m_remetente = re.search(r'REMETENTE(.*?)DESTINAT', texto, re.DOTALL | re.IGNORECASE)
        if m_remetente:
            texto_rem = m_remetente.group(1)
            m_cnpj = re.search(r'CPF/CNPJ[:\s]*(\d[\d./-]+)', texto_rem)
            if m_cnpj:
                cnpj_raw = re.sub(r'[.\-/]', '', m_cnpj.group(1))
                if len(cnpj_raw) >= 11:
                    resultado['cnpj_emitente'] = cnpj_raw
            # Extrair nome remetente (da secao REMETENTE)
            m_nome = re.search(r'NOME[:\s]*([^\n]+)', texto_rem)
            if m_nome:
                resultado['nome_emitente'] = m_nome.group(1).strip()

        # Extrair produtos da tabela IDENTIFICACAO DOS BENS
        # O PyMuPDF extrai cada celula da tabela em uma linha separada.
        # Cabecalhos: Nº, SKU, DESCRIÇÃO, VARIAÇÃO, QTD (cada em sua linha)
        # Dados: linhas de valores intercalados (numero, SKU multi-linha, desc, var multi-linha, qtd)
        linhas = texto.split('\n')

        # 1. Encontrar "QTD" (fim dos cabecalhos) e "Total" (fim dos dados)
        idx_qtd_header = -1
        idx_total = -1
        for i, linha in enumerate(linhas):
            ls = linha.strip().upper()
            if ls == 'QTD':
                idx_qtd_header = i
            if linha.strip() == 'Total' or linha.strip() == 'total':
                idx_total = i

        if idx_qtd_header < 0:
            # Fallback: tentar encontrar "SKU" como marcador
            for i, linha in enumerate(linhas):
                if linha.strip().upper() == 'SKU':
                    # Pular mais 3 linhas (DESCRIÇÃO, VARIAÇÃO, QTD)
                    idx_qtd_header = i + 3
                    break

        if idx_qtd_header < 0 or idx_total < 0:
            return resultado

        # 2. Extrair total_qtd da linha apos "Total"
        if idx_total + 1 < len(linhas):
            try:
                resultado['total_qtd'] = int(linhas[idx_total + 1].strip())
            except (ValueError, IndexError):
                pass

        # 3. Extrair linhas de dados entre cabecalhos e Total
        data_lines = [l.strip() for l in linhas[idx_qtd_header + 1:idx_total] if l.strip()]

        # 4. Parsear items: a estrutura e sequencial por colunas
        #    Para cada item: N (1 linha), SKU (1+ linhas), DESC (1 linha, tipicamente "-"),
        #    VARIACAO (1+ linhas), QTD (1 linha)
        #    Seguido opcionalmente por data (dd-mm-yyyy)
        produtos = []
        i = 0
        while i < len(data_lines):
            # Pular linhas de data (dd-mm-yyyy)
            if re.match(r'^\d{2}-\d{2}-\d{4}$', data_lines[i]):
                i += 1
                continue

            # Item number: linha contendo apenas um numero pequeno (1-99)
            if re.match(r'^\d{1,2}$', data_lines[i]):
                item_num = data_lines[i]
                i += 1

                # Coletar SKU (linhas ate encontrar "-" sozinho ou proximo item number)
                sku_parts = []
                while i < len(data_lines):
                    if data_lines[i] == '-':
                        i += 1  # pular a descricao "-"
                        break
                    if re.match(r'^\d{1,2}$', data_lines[i]) and not sku_parts:
                        break  # proximo item sem descricao
                    sku_parts.append(data_lines[i])
                    i += 1

                # Coletar VARIACAO (linhas ate encontrar QTD = numero sozinho 1-4 digitos)
                var_parts = []
                qtd = '1'
                while i < len(data_lines):
                    # Se e uma data, parar
                    if re.match(r'^\d{2}-\d{2}-\d{4}$', data_lines[i]):
                        break
                    # Se e um numero sozinho de 1-4 digitos (QTD)
                    if re.match(r'^\d{1,4}$', data_lines[i]):
                        qtd = data_lines[i]
                        i += 1
                        break
                    var_parts.append(data_lines[i])
                    i += 1

                sku = ' '.join(sku_parts) if sku_parts else ''
                variacao = ' '.join(var_parts) if var_parts else ''
                # atributos = SKU completo para sorting (compativel com _parsear_atributos_shein)
                atrib = sku if not variacao else f"{sku}/{variacao}"

                produtos.append({
                    'codigo_item': sku,
                    'descricao': '',
                    'atributos': atrib,
                    'qtd': qtd,
                })
            else:
                i += 1  # linha nao reconhecida, pular

        resultado['produtos_shein'] = produtos
        resultado['total_itens'] = len(produtos)
        if not resultado['total_qtd']:
            resultado['total_qtd'] = sum(int(p.get('qtd', 1)) for p in produtos)

        return resultado

    def _detectar_tipo_etiqueta(self, texto, nf_encontrada=None):
        """Detecta o tipo de etiqueta pelo conteudo do texto.
        Retorna: 'retirada', 'shein', 'cnpj', 'cpf' ou 'declaracao'

        Criterios:
        - 'retirada': contem "RETIRADA PELO COMPRADOR"
        - 'shein': tem marcadores Shein (PUDO-PGK, Ref.No:GSH, etc.)
        - 'declaracao': contem "DECLARAÇÃO DE CONTEÚDO" (não é etiqueta, é anexo)
        - 'cnpj': tem DANFE SIMPLIFICADO E tem NF numerica real
        - 'cpf': tem DANFE SIMPLIFICADO mas SEM NF numerica (loja CPF/pessoa fisica)
                  OU nao tem DANFE SIMPLIFICADO (declaracao de conteudo)
        """
        texto_upper = texto.upper()
        
        if 'RETIRADA PELO' in texto_upper and 'COMPRADOR' in texto_upper:
            return 'retirada'
        # Detectar Shein antes de CNPJ/CPF
        if self._eh_etiqueta_shein(texto):
            return 'shein'
        if 'DANFE SIMPLIFICADO' in texto_upper:
            # Tem DANFE SIMPLIFICADO, mas tem NF real?
            if nf_encontrada and nf_encontrada.isdigit():
                return 'cnpj'
            # Sem NF real = loja CPF (pessoa fisica)
            return 'cpf'
        
        # Detectar Declaração de Conteúdo por ÚLTIMO (só se não for etiqueta)
        if self._eh_declaracao_conteudo(texto):
            return 'declaracao'
        
        return 'cpf'

    def _recortar_pagina(self, pagina, caminho_pdf):
        """Recorta etiquetas de uma pagina, detectando automaticamente o layout."""
        quadrantes = self._detectar_layout_pagina(pagina)

        etiquetas = []
        quadrantes_vazios = 0
        
        for idx, clip in enumerate(quadrantes):
            # Verificar se o quadrante tem conteudo (nao esta vazio)
            texto_quad = pagina.get_text(clip=clip).strip()
            if len(texto_quad) < 10:
                quadrantes_vazios += 1
                continue  # Quadrante vazio, pular

            # Extrair NF primeiro (necessario para detectar tipo)
            nf = self._extrair_nf_quadrante(pagina, clip)

            # Detectar tipo de etiqueta (passa nf para distinguir CNPJ vs CPF)
            tipo_etiqueta = self._detectar_tipo_etiqueta(texto_quad, nf_encontrada=nf)
            
            # Declaração de conteúdo desabilitada - ignorar
            if tipo_etiqueta == 'declaracao':
                print(f"    Pag {pagina.number} Q{idx}: AVISO - Declaração de Conteúdo detectada mas ignorada (use XLSX para dados de produtos)")
                continue

            # Gerar etiqueta MESMO sem NF - usar identificador sintetico (incluir nome PDF para unicidade)
            if nf is None:
                pdf_id = os.path.splitext(os.path.basename(caminho_pdf))[0].replace(' ', '_')
                nf = f"SEM_NF_{pdf_id}_p{pagina.number}_q{idx}"
                dados_nf = {}
                print(f"    Pag {pagina.number} Q{idx}: NF nao encontrada (tipo={tipo_etiqueta})")
            else:
                dados_nf = {}

            # Extrair order_sn para usar como NF em etiquetas CPF
            order_sn_txt = self._extrair_pedido_texto(texto_quad)

            # Para etiquetas CPF, usar order_sn como identificador (nao tem NF real)
            if tipo_etiqueta == 'cpf' and order_sn_txt:
                nf = order_sn_txt

            # FONTE PRIMARIA: XLSX (buscar por order_sn ou tracking)
            if self.dados_xlsx_global or self.dados_lista_global:
                dados_xlsx, chave_dados = self._buscar_dados_xlsx(texto_quad)
                if dados_xlsx:
                    origem_dados = dados_xlsx.get('fonte_dados', 'xlsx')
                    dados_nf = {
                        'nf': nf,
                        'serie': '',
                        'data_emissao': '',
                        'chave': self._extrair_chave_nfe(texto_quad),
                        'cnpj_emitente': '',
                        'nome_emitente': '',
                        'produtos': dados_xlsx['produtos'],
                        'total_itens': dados_xlsx['total_itens'],
                        'total_qtd': dados_xlsx['total_qtd'],
                        'fonte_dados': origem_dados,
                    }
                    print(f"    Pag {pagina.number} Q{idx}: Dados {origem_dados} ({chave_dados})")

            # FALLBACK: XML (se XLSX nao encontrou produtos) - apenas para CNPJ
            if tipo_etiqueta == 'cnpj' and not dados_nf.get('produtos') and nf in self.dados_xml:
                dados_nf = self.dados_xml.get(nf, {})

            sku = ''
            num_produtos = 1
            cnpj = dados_nf.get('cnpj_emitente', '')
            if dados_nf.get('produtos'):
                sku = dados_nf['produtos'][0].get('codigo', '')
                num_produtos = len(dados_nf['produtos'])

            # Extrair nome da loja do REMETENTE do texto da etiqueta
            nome_loja = self._extrair_nome_loja_remetente(texto_quad)
            if not cnpj:
                if not nome_loja:
                    # Fallback para fluxo UpSeller: usar nome da loja do arquivo.
                    nome_loja = self._inferir_loja_por_nome_arquivo(caminho_pdf)
                if nome_loja:
                    prefixo = 'CPF' if tipo_etiqueta == 'cpf' else 'LOJA'
                    cnpj = self._registrar_loja_sintetica(nome_loja, prefixo=prefixo)
            elif cnpj not in self.cnpj_loja:
                if nome_loja:
                    self.cnpj_loja[cnpj] = nome_loja

            etiquetas.append({
                'nf': nf,
                'sku': sku,
                'num_produtos': num_produtos,
                'cnpj': cnpj,
                'clip': clip,
                'pagina_idx': pagina.number,
                'caminho_pdf': caminho_pdf,
                'dados_xml': dados_nf,
                'tipo_especial': tipo_etiqueta if tipo_etiqueta != 'cnpj' else None,
            })

        # Avisar sobre quadrantes não processados
        if quadrantes_vazios > 0:
            print(f"    Pag {pagina.number}: AVISO - {quadrantes_vazios} quadrante(s) vazio(s) ou ignorado(s)")

        return etiquetas

    # ----------------------------------------------------------------
    # SEPARACAO POR LOJA
    # ----------------------------------------------------------------
    def separar_por_loja(self, etiquetas):
        """Separa etiquetas por CNPJ do emitente (loja)."""
        lojas = defaultdict(list)
        sem_loja = []

        for etq in etiquetas:
            cnpj = etq.get('cnpj', '')
            if cnpj:
                lojas[cnpj].append(etq)
            else:
                sem_loja.append(etq)

        if sem_loja:
            lojas['SEM_CNPJ'] = sem_loja

        return dict(lojas)

    # ----------------------------------------------------------------
    # ORDENACAO
    # ----------------------------------------------------------------
    def _ordenar_etiquetas(self, etiquetas):
        """Ordena etiquetas na mesma ordem do resumo XLSX:
        - qtd=1 primeiro, qtd>1 ao final
        - dentro de cada bloco: SKU > Quantidade > Cor > Numero
        """
        def _total_qtd(etq):
            dados = etq.get('dados_xml', {}) or {}
            total = dados.get('total_qtd', None)
            try:
                if total is not None:
                    return max(1, int(float(total)))
            except Exception:
                pass
            soma = 0
            for p in (dados.get('produtos', []) or []):
                try:
                    soma += int(float(p.get('qtd', 1) or 1))
                except Exception:
                    soma += 1
            return max(1, soma or 1)

        def _chave_etiqueta(etq):
            """Extrai (sku, cor, numero) do primeiro produto da etiqueta."""
            dados = etq.get('dados_xml', {})
            produtos = dados.get('produtos', [])
            if produtos:
                sku = produtos[0].get('codigo', '')
                variacao = produtos[0].get('variacao', '')
            else:
                sku = etq.get('sku', '')
                variacao = ''
            # Separar variacao em cor e numero
            partes = re.split(r',', variacao, maxsplit=1)
            if len(partes) == 2:
                cor = partes[0].strip()
                num_str = partes[1].strip()
            else:
                cor = variacao.strip()
                num_str = ''
            # Extrair valor numerico para ordenacao correta (35 antes de 36)
            m = re.search(r'(\d+)', num_str)
            num_val = int(m.group(1)) if m else 99999
            return ((sku or '').casefold(), (cor or '').casefold(), num_val, (num_str or '').casefold())

        def _chave_ordenacao(etq):
            sku, cor, num_val, num_str = _chave_etiqueta(etq)
            return (sku, _total_qtd(etq), cor, num_val, num_str)

        simples = []
        multiplos = []
        for e in etiquetas:
            total_qtd = _total_qtd(e)
            if total_qtd > 1:
                multiplos.append(e)
            else:
                simples.append(e)

        # Simples: SKU > Quantidade > Cor > Numero
        simples.sort(key=_chave_ordenacao)

        # Multiplos: continuam no final, ordenados por SKU > Quantidade > Cor > Numero
        multiplos.sort(key=_chave_ordenacao)

        return simples + multiplos, len(simples), len(multiplos)

    @staticmethod
    def _ordenar_produtos(produtos):
        """Ordena produtos na tabela: qtd=1 por SKU > Cor > Numero, qtd>1 no final por qtd crescente.

        A variacao costuma vir como "Cor,Tamanho" (ex: "Preto,35") ou "Cor/Tamanho".
        """
        def _separar_cor_numero(variacao):
            """Separa variacao em (cor, numero) para ordenacao."""
            if not variacao:
                return ('', '')
            # Tentar separar por virgula primeiro, depois barra
            partes = re.split(r'[,/]', variacao, maxsplit=1)
            if len(partes) == 2:
                return (partes[0].strip(), partes[1].strip())
            return (variacao.strip(), '')

        def _numero_sort_key(num_str):
            """Converte numero/tamanho para valor numerico para ordenacao correta."""
            # Extrair primeiro numero encontrado (ex: "BR39/40" -> 39, "35" -> 35)
            m = re.search(r'(\d+)', num_str)
            if m:
                return (0, int(m.group(1)))
            return (1, num_str)  # sem numero vai depois

        unitarios = []
        multiplos = []
        for prod in produtos:
            qtd = int(float(prod.get('qtd', '1')))
            if qtd > 1:
                multiplos.append(prod)
            else:
                unitarios.append(prod)

        # Unitarios: ordenar por SKU > Cor > Numero
        unitarios.sort(key=lambda p: (
            p.get('codigo', ''),
            _separar_cor_numero(p.get('variacao', ''))[0],
            _numero_sort_key(_separar_cor_numero(p.get('variacao', ''))[1]),
        ))

        # Multiplos: SKU > Quantidade > Cor > Numero
        multiplos.sort(key=lambda p: (
            p.get('codigo', ''),
            int(float(p.get('qtd', '1'))),
            _separar_cor_numero(p.get('variacao', ''))[0],
            _numero_sort_key(_separar_cor_numero(p.get('variacao', ''))[1]),
        ))

        return unitarios + multiplos

    @staticmethod
    def _limpar_texto_tabela(texto):
        """Normaliza texto para exibicao em tabela."""
        return re.sub(r'\s+', ' ', str(texto or '')).strip()

    @staticmethod
    def _extrair_qtd_sufixo_item(texto):
        """Extrai quantidade em sufixos como '(*2)', '*2', 'x2' e remove do texto."""
        txt = re.sub(r'\s+', ' ', str(texto or '')).strip()
        if not txt:
            return txt, None

        padroes = (
            r'^(.*?)\s*\(\s*[\*xX×-]\s*(\d{1,3})\s*\)\s*$',
            r'^(.*?)\s*[\*xX×-]\s*(\d{1,3})\s*$',
        )
        for patt in padroes:
            m = re.match(patt, txt)
            if not m:
                continue
            corpo = re.sub(r'\s+', ' ', (m.group(1) or '')).strip()
            try:
                qtd = max(1, int(m.group(2)))
            except Exception:
                qtd = 1
            return corpo, qtd

        return txt, None

    @staticmethod
    def _parece_texto_metadado(texto):
        """Detecta textos tecnicos/ruidosos do rodape original da etiqueta."""
        t = (texto or '').upper()
        if not t:
            return False
        marcadores = (
            'SKU:',
            'TOTAL ITEMS',
            'TOTAL ITENS',
            'DEADLINE',
            '#UP',
            'PEDIDO:'
        )
        return any(m in t for m in marcadores)

    @staticmethod
    def _truncate_por_largura(texto, max_largura, fontname, fontsize):
        """Trunca string por largura real em pontos (fallback por caracteres)."""
        txt = str(texto or '').strip()
        if not txt:
            return ''
        try:
            if fitz.get_text_length(txt, fontname=fontname, fontsize=fontsize) <= max_largura:
                return txt
            base = txt
            while base and fitz.get_text_length(base + '..', fontname=fontname, fontsize=fontsize) > max_largura:
                base = base[:-1]
            return (base + '..') if base else ''
        except Exception:
            max_chars = max(1, int(max_largura / max(1, fontsize * 0.55)))
            return txt if len(txt) <= max_chars else (txt[:max_chars - 2] + '..')

    def _normalizar_linha_tabela(self, produto):
        """Normaliza uma linha de produto para o layout limpo do rodape."""
        codigo = self._limpar_texto_tabela(produto.get('codigo', ''))
        descricao = self._limpar_texto_tabela(produto.get('descricao', ''))
        variacao = self._limpar_texto_tabela(produto.get('variacao', ''))
        qtd_raw = str(produto.get('qtd', '1') or '1').strip()

        if self._parece_texto_metadado(codigo):
            base = variacao or descricao
            base = re.sub(r'^\d+\.\s*', '', base)
            codigo = re.split(r'[-,(;/\s]', base, maxsplit=1)[0].strip() if base else ''

        if not codigo:
            base = variacao or descricao
            base = re.sub(r'^\d+\.\s*', '', base)
            codigo = re.split(r'[-,(;/\s]', base, maxsplit=1)[0].strip() if base else ''

        if codigo and len(codigo) > 16 and '-' in codigo:
            codigo = codigo.split('-', 1)[0].strip()

        detalhe = variacao or descricao or '-'
        detalhe = re.sub(r'^\d+\.\s*', '', detalhe)
        detalhe = re.sub(r'\s*[\*\-xX]\s*\d+\s*$', '', detalhe).strip()
        if self._parece_texto_metadado(detalhe):
            detalhe = descricao or variacao or '-'

        if not codigo:
            codigo = '-'
        if not detalhe:
            detalhe = '-'

        try:
            qtd = str(max(1, int(float(qtd_raw))))
        except Exception:
            qtd = '1'

        return codigo, detalhe, qtd

    @staticmethod
    def _mascarar_rodape_original(pagina, area_etiqueta):
        """Cobre o rodape original da etiqueta para evitar duplicidade de informacao.

        Retorna o retangulo mascarado para permitir posicionar o novo rodape
        encostado na area do codigo de barras.
        """
        try:
            altura_total = float(area_etiqueta.height)
            # Faixa mais ampla para remover completamente o rodape antigo
            # (linha SKU/Deadline e eventuais residuos da tabela anterior).
            faixa_alt = max(66.0, min(120.0, altura_total * 0.18))
            respiro_inf = max(3.0, min(8.0, altura_total * 0.012))
            y1 = max(float(area_etiqueta.y0) + 12.0, float(area_etiqueta.y1) - respiro_inf)
            y0 = max(float(area_etiqueta.y0), y1 - faixa_alt)
            faixa = fitz.Rect(
                float(area_etiqueta.x0) + 1.0,
                y0,
                float(area_etiqueta.x1) - 1.0,
                y1,
            )
            pagina.draw_rect(
                faixa,
                color=(1, 1, 1),
                fill=(1, 1, 1),
                width=0,
                overlay=True,
            )
            return faixa
        except Exception:
            return None

    # ----------------------------------------------------------------
    # GERACAO DO CODIGO DE BARRAS
    # ----------------------------------------------------------------
    def _gerar_barcode_svg(self, chave):
        """Gera um codigo de barras Code128 como SVG em bytes."""
        code128 = barcode.get('code128', chave, writer=SVGWriter())
        buf = io.BytesIO()
        code128.write(buf, options={
            'module_width': 0.2,
            'module_height': 8,
            'write_text': False,
            'quiet_zone': 1,
        })
        return buf.getvalue()

    # ----------------------------------------------------------------
    # GERACAO DO PDF FINAL (POR LOJA)
    # ----------------------------------------------------------------
    def gerar_pdf_loja(self, etiquetas, caminho_saida):
        """Gera o PDF final para uma loja."""
        etiquetas_ord, n_simples, n_multi = self._ordenar_etiquetas(etiquetas)

        # Agrupar etiquetas por PDF de origem
        pdfs_usados = set()
        for etq in etiquetas_ord:
            pdfs_usados.add(etq['caminho_pdf'])

        # Abrir todos os PDFs necessarios
        docs_abertos = {}
        for pdf_path in pdfs_usados:
            docs_abertos[pdf_path] = fitz.open(pdf_path)

        doc_saida = fitz.open()
        area_util_larg = self.LARGURA_PT - self.MARGEM_ESQUERDA - self.MARGEM_DIREITA
        image_crop_cache = {}

        com_xml = 0
        sem_xml = 0

        for idx, etq in enumerate(etiquetas_ord):
            nf = etq['nf']
            clip = etq['clip']
            pag_idx = etq['pagina_idx']
            dados = etq.get('dados_xml', {})
            render_imagem_meta = etq.get('render_imagem_meta') or {}
            numero_ordem = idx + 1
            pdf_path = etq['caminho_pdf']
            doc_entrada = docs_abertos[pdf_path]

            nova_pag = doc_saida.new_page(
                width=self.LARGURA_PT,
                height=self.ALTURA_PT
            )
            pag_principal_idx = len(doc_saida) - 1  # indice da pagina principal

            if render_imagem_meta.get('xref') and render_imagem_meta.get('crop_img'):
                ix0, iy0, ix1, iy1 = render_imagem_meta.get('crop_img')
                quad_larg = max(1.0, float(ix1 - ix0))
                quad_alt = max(1.0, float(iy1 - iy0))
            else:
                quad_larg = clip.width
                quad_alt = clip.height
            escala = area_util_larg / quad_larg
            alt_etiqueta = quad_alt * escala

            # Verificar se tem dados de produto (para contagem).
            num_prods = len(dados.get('produtos', []))
            tem_dados_produto = num_prods > 0

            # Manter rodape original do UpSeller: nao encolher a etiqueta,
            # nao mascarar o rodape e nao redesenhar tabela de produtos.
            # A ordenacao por SKU e >1 itens pro final ja foi feita em _ordenar_etiquetas().

            dest_rect = fitz.Rect(
                self.MARGEM_ESQUERDA,
                self.MARGEM_TOPO,
                self.LARGURA_PT - self.MARGEM_DIREITA,
                self.MARGEM_TOPO + alt_etiqueta
            )

            if render_imagem_meta.get('xref') and render_imagem_meta.get('crop_img'):
                xref = int(render_imagem_meta['xref'])
                ix0, iy0, ix1, iy1 = [int(v) for v in render_imagem_meta['crop_img']]
                key = (pdf_path, xref, ix0, iy0, ix1, iy1)
                pix_crop = image_crop_cache.get(key)
                if pix_crop is None:
                    pix_full = fitz.Pixmap(doc_entrada, xref)
                    if pix_full.n > 3:
                        pix_full = fitz.Pixmap(fitz.csRGB, pix_full)
                    crop_rect = fitz.IRect(ix0, iy0, ix1, iy1)
                    pix_crop = fitz.Pixmap(pix_full.colorspace, crop_rect, pix_full.alpha)
                    pix_crop.copy(pix_full, crop_rect)
                    image_crop_cache[key] = pix_crop
                nova_pag.insert_image(dest_rect, pixmap=pix_crop, keep_proportion=False)
            else:
                nova_pag.show_pdf_page(dest_rect, doc_entrada, pag_idx, clip=clip)

            if tem_dados_produto:
                com_xml += 1
            else:
                sem_xml += 1

            # Re-adquirir referencia da pagina principal (new_page() invalida refs anteriores no PyMuPDF)
            nova_pag = doc_saida[pag_principal_idx]

            # Limpar eventual numeracao antiga no canto inferior esquerdo
            # quando o input ja for um PDF previamente processado.
            try:
                faixa_numero = fitz.Rect(
                    self.MARGEM_ESQUERDA,
                    self.ALTURA_PT - self.MARGEM_INFERIOR - 22,
                    self.MARGEM_ESQUERDA + 34,
                    self.ALTURA_PT - self.MARGEM_INFERIOR + 3,
                )
                nova_pag.draw_rect(
                    faixa_numero,
                    color=(1, 1, 1),
                    fill=(1, 1, 1),
                    width=0,
                    overlay=True,
                )
            except (AttributeError, RuntimeError):
                pass

            # Numero de ordem (subido para nao cortar na impressao)
            try:
                nova_pag.insert_text(
                    (self.MARGEM_ESQUERDA + 2, self.ALTURA_PT - self.MARGEM_INFERIOR - 8),
                    f"p.{numero_ordem}",
                    fontsize=9,
                    fontname="hebo",
                    color=(0.4, 0.4, 0.4)
                )
            except (AttributeError, RuntimeError):
                # Bug PyMuPDF - skipar numero da pagina se der erro
                pass

        # Fechar docs de entrada
        image_crop_cache.clear()
        for doc in docs_abertos.values():
            doc.close()

        total = len(doc_saida)
        doc_saida.save(caminho_saida)
        doc_saida.close()

        return total, n_simples, n_multi, com_xml, sem_xml

    def _desenhar_secao_produtos(self, pagina, dados, y_inicio, prod_inicio=0, alt_pagina=None):
        """Desenha a secao de codigo de barras + tabela de produtos abaixo da etiqueta.
        prod_inicio: indice do primeiro produto a desenhar (para continuacao).
        alt_pagina: altura da pagina (usa self.ALTURA_PT se None).
        Retorna indice do proximo produto nao desenhado (len(produtos) se todos couberam).
        """
        preto = (0, 0, 0)
        cinza_linha = (0.55, 0.55, 0.55)
        fonte = "helv"
        fonte_bold = "hebo"
        margem_esq = self.MARGEM_ESQUERDA
        margem_dir = self.MARGEM_DIREITA
        larg = self.LARGURA_PT

        # Dados de envio/produto 30% maiores para melhorar leitura na separacao.
        fs_base = max(7, int(round(self.fonte_produto * 1.30)))
        fs_header = max(fs_base, int(round(fs_base * 1.05)))
        fs_texto = max(10, int(round(fs_base * 1.35)))
        fs_qtd = max(18, int(round(fs_texto * 1.55)))
        line_h_header = fs_header + 4
        line_h = max(fs_qtd + 2, fs_texto + 4)

        nf = dados.get('nf', '')
        chave = dados.get('chave', '')
        produtos = self._ordenar_produtos(dados.get('produtos', []))
        total_itens = dados.get('total_itens', len(produtos))
        total_qtd = dados.get('total_qtd', sum(int(float(p.get('qtd', 1))) for p in produtos))

        y = y_inicio

        # --- Codigo de barras da chave de acesso (so na primeira pagina) ---
        if chave and prod_inicio == 0:
            try:
                svg_bytes = self._gerar_barcode_svg(chave)
                barcode_rect = fitz.Rect(
                    margem_esq + 5, y,
                    larg - margem_dir - 5, y + 35
                )
                pagina.insert_image(barcode_rect, stream=svg_bytes)
                y += 37
            except Exception:
                y += 5

        # --- Tabela de produtos (layout limpo) ---
        col1_w = 62
        col3_w = 30
        col_codigo = margem_esq + 3
        x_div1 = margem_esq + col1_w
        col_prod = x_div1 + 4
        x_div2 = larg - margem_dir - col3_w
        col_qtd = x_div2 + 7
        x_direita = larg - margem_dir

        w_codigo = max(20, x_div1 - col_codigo - 5)
        w_prod = max(40, x_div2 - col_prod - 6)
        w_qtd = max(10, x_direita - col_qtd - 3)

        y_tabela_topo = y
        pagina.draw_line((margem_esq, y), (x_direita, y), color=preto, width=0.8)

        y += line_h_header
        continuacao_txt = f" (cont. {prod_inicio + 1}-)" if prod_inicio > 0 else ""
        header_col1 = "CODIGO"
        header_prod = f"VAR. (NF: {nf} T-ITENS: {total_itens} T-QUANT: {total_qtd}){continuacao_txt}"
        header_prod = self._truncate_por_largura(header_prod, w_prod, fonte_bold, fs_header)

        pagina.insert_text((col_codigo, y), header_col1, fontsize=fs_header, fontname=fonte_bold, color=preto)
        pagina.insert_text((col_prod, y), header_prod, fontsize=fs_header, fontname=fonte_bold, color=preto)
        pagina.insert_text((col_qtd, y), "Q.", fontsize=fs_header, fontname=fonte_bold, color=preto)

        y += 2
        pagina.draw_line((margem_esq, y), (x_direita, y), color=preto, width=0.5)
        y += line_h

        y_limite = (alt_pagina or self.ALTURA_PT) - self.MARGEM_INFERIOR - 10
        ultimo_desenhado = prod_inicio

        for i_abs in range(prod_inicio, len(produtos)):
            if y + line_h > y_limite:
                break

            codigo, detalhe, qtd = self._normalizar_linha_tabela(produtos[i_abs])
            codigo = self._truncate_por_largura(codigo, w_codigo, fonte_bold, fs_texto)
            detalhe = self._truncate_por_largura(detalhe, w_prod, fonte_bold, fs_texto)
            qtd = self._truncate_por_largura(qtd, w_qtd, fonte_bold, fs_qtd)

            pagina.insert_text((col_codigo, y), codigo or "-", fontsize=fs_texto, fontname=fonte_bold, color=preto)
            pagina.insert_text((col_prod, y), detalhe or "-", fontsize=fs_texto, fontname=fonte_bold, color=preto)
            pagina.insert_text((col_qtd, y), qtd or "1", fontsize=fs_qtd, fontname=fonte_bold, color=preto)

            y += line_h
            ultimo_desenhado = i_abs + 1

            if i_abs < len(produtos) - 1 and y + line_h <= y_limite:
                pagina.draw_line((margem_esq, y - 1), (x_direita, y - 1), color=cinza_linha, width=0.3)

        pagina.draw_line((margem_esq, y), (x_direita, y), color=preto, width=0.8)
        pagina.draw_line((x_div1, y_tabela_topo), (x_div1, y), color=preto, width=0.5)
        pagina.draw_line((x_div2, y_tabela_topo), (x_div2, y), color=preto, width=0.5)

        return ultimo_desenhado

    # ----------------------------------------------------------------
    # ETIQUETAS ESPECIAIS: RETIRADA DO COMPRADOR (BEKA) E CPF
    # ----------------------------------------------------------------

    # Nomes dos PDFs especiais (nao processados no grid 2x2)
    PDFS_ESPECIAIS = ['lanim.pdf', 'shein crua.pdf', 'shein.pdf']
    # CNPJ e nome fixos para etiquetas CPF (sem XML/DANFE)
    LANIM_CNPJ = 'LANIM_CPF'
    LANIM_NOME = 'CPF'

    def carregar_xlsx_pedidos(self, caminho_xlsx):
        """Carrega dados de pedidos do XLSX (lanim2.xlsx) para etiquetas CPF.
        Retorna dict: order_sn -> {produtos: [...], total_itens, total_qtd}
        """
        import openpyxl as xl
        wb = xl.load_workbook(caminho_xlsx)
        ws = wb.active
        dados_pedidos = {}

        # Descobrir indices das colunas pelo cabecalho
        cabecalho = {}
        for idx, cell in enumerate(ws[1]):
            if cell.value:
                cabecalho[str(cell.value).strip().lower()] = idx

        col_order = cabecalho.get('order_sn', None)
        col_info = cabecalho.get('product_info', None)

        if col_order is None or col_info is None:
            print("  AVISO: Colunas order_sn ou product_info nao encontradas no XLSX")
            wb.close()
            return dados_pedidos

        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, values_only=True):
            if row is None or len(row) <= max(col_order, col_info):
                continue
            order_sn = str(row[col_order] or '').strip()
            product_info = str(row[col_info] or '').strip()

            if not order_sn or not product_info:
                continue

            # Parsear product_info: pode ter multiplos blocos [N]
            produtos = []
            # Dividir por blocos [N]
            blocos = re.split(r'\[\d+\]\s*', product_info)
            for bloco in blocos:
                if not bloco.strip():
                    continue

                # Extrair Parent SKU
                m_sku = re.search(r'Parent SKU Reference No\.:\s*([^;]+)', bloco)
                codigo = m_sku.group(1).strip() if m_sku else ''
                # Fallback: SKU Reference No.
                if not codigo:
                    m_sku2 = re.search(r'SKU Reference No\.:\s*([^;]+)', bloco)
                    codigo = m_sku2.group(1).strip() if m_sku2 else ''

                # Extrair quantidade
                m_qtd = re.search(r'Quantity:\s*(\d+)', bloco)
                qtd = m_qtd.group(1) if m_qtd else '1'

                # Extrair nome do produto
                m_nome = re.search(r'Product Name:\s*([^;]+)', bloco)
                descricao = m_nome.group(1).strip() if m_nome else ''

                # Extrair Variation Name (usado para etiquetas CPF)
                m_var = re.search(r'Variation Name:\s*([^;]+)', bloco)
                variacao = m_var.group(1).strip() if m_var else ''

                if codigo or descricao or variacao:
                    produtos.append({
                        'codigo': codigo,
                        'descricao': descricao,
                        'variacao': variacao,
                        'qtd': qtd,
                    })

            # Acumular apenas produtos NOVOS se o mesmo order_sn aparecer em multiplas linhas
            if order_sn in dados_pedidos:
                existentes = dados_pedidos[order_sn]['produtos']
                chaves_existentes = set()
                for p in existentes:
                    chaves_existentes.add((p.get('codigo', ''), p.get('descricao', ''), p.get('variacao', '')))
                for p in produtos:
                    chave_p = (p.get('codigo', ''), p.get('descricao', ''), p.get('variacao', ''))
                    if chave_p not in chaves_existentes:
                        existentes.append(p)
                        chaves_existentes.add(chave_p)
                dados_pedidos[order_sn]['total_itens'] = len(existentes)
                dados_pedidos[order_sn]['total_qtd'] = sum(
                    int(float(p.get('qtd', 1))) for p in existentes
                )
            else:
                total_qtd = sum(int(float(p.get('qtd', 1))) for p in produtos)
                dados_pedidos[order_sn] = {
                    'produtos': produtos,
                    'total_itens': len(produtos),
                    'total_qtd': total_qtd,
                }

        wb.close()
        print(f"  XLSX: {len(dados_pedidos)} pedidos carregados")
        return dados_pedidos

    def _parsear_product_info(self, product_info):
        """Parseia o campo product_info do XLSX da Shopee.
        Retorna lista de produtos: [{codigo, descricao, variacao, qtd}, ...]
        """
        produtos = []
        blocos = re.split(r'\[\d+\]\s*', product_info)
        for bloco in blocos:
            if not bloco.strip():
                continue

            m_sku = re.search(r'Parent SKU Reference No\.:\s*([^;]+)', bloco)
            codigo = m_sku.group(1).strip() if m_sku else ''
            if not codigo:
                m_sku2 = re.search(r'SKU Reference No\.:\s*([^;]+)', bloco)
                codigo = m_sku2.group(1).strip() if m_sku2 else ''

            m_qtd = re.search(r'Quantity:\s*(\d+)', bloco)
            qtd = m_qtd.group(1) if m_qtd else '1'

            m_nome = re.search(r'Product Name:\s*([^;]+)', bloco)
            descricao = m_nome.group(1).strip() if m_nome else ''

            m_var = re.search(r'Variation Name:\s*([^;]+)', bloco)
            variacao = m_var.group(1).strip() if m_var else ''

            if codigo or descricao or variacao:
                produtos.append({
                    'codigo': codigo,
                    'descricao': descricao,
                    'variacao': variacao,
                    'qtd': qtd,
                })

        return produtos

    def _normalizar_coluna_xlsx(self, nome_coluna):
        txt = self._remover_acentos(str(nome_coluna or ''))
        txt = re.sub(r'[\r\n\t]+', ' ', txt)
        txt = re.sub(r'\s+', ' ', txt).strip().lower()
        return txt

    def _escolher_chave_principal_resumo(self, chaves):
        """Escolhe chave principal de pedido e tracking a partir das chaves extraidas."""
        chaves = [c for c in (chaves or []) if c]
        if not chaves:
            return '', ''

        tracking_key = ''
        for c in chaves:
            if c.startswith('BR'):
                tracking_key = c
                break

        # Preferir order_sn Shopee (YYMMDD + alfanumerico)
        for c in chaves:
            if re.match(r'^\d{6}[A-Z0-9]{6,12}$', c):
                return c, tracking_key

        # Depois codigos comuns de pedido
        for c in chaves:
            if c.startswith('UP') or c.startswith('MEL'):
                return c, tracking_key

        # Evitar usar tracking como chave principal se houver outra opcao
        for c in chaves:
            if not c.startswith('BR'):
                return c, tracking_key

        return chaves[0], tracking_key

    def carregar_todos_xlsx(self, pasta):
        """Carrega dados de TODOS os XLSX da pasta para fallback quando XML nao existe.
        Popula self.dados_xlsx_global (order_sn -> dados) e
        self.dados_xlsx_tracking (tracking -> order_sn).
        """
        import openpyxl as xl

        xlsx_files = [f for f in os.listdir(pasta)
                      if f.lower().endswith(('.xlsx', '.xls'))
                      and not f.startswith('_')
                      and not f.startswith('~')
                      and f != 'planilha_custos.xlsx']

        if not xlsx_files:
            return

        for xlsx_nome in sorted(xlsx_files):
            caminho = os.path.join(pasta, xlsx_nome)
            try:
                wb = xl.load_workbook(caminho, read_only=False)
                ws = wb.active

                # Descobrir indices das colunas pelo cabecalho
                cabecalho = {}
                for idx, cell in enumerate(ws[1]):
                    if cell.value:
                        cabecalho[self._normalizar_coluna_xlsx(cell.value)] = idx

                idx_tracking = cabecalho.get('tracking_number', cabecalho.get('tracking number', -1))
                idx_order = cabecalho.get('order_sn', cabecalho.get('order sn', -1))
                idx_product = cabecalho.get('product_info', cabecalho.get('product info', -1))

                def _registrar_produtos(order_key, produtos, tracking_key=''):
                    if not order_key or not produtos:
                        return
                    if order_key not in self.dados_xlsx_global:
                        self.dados_xlsx_global[order_key] = {
                            'produtos': list(produtos),
                            'total_itens': len(produtos),
                            'total_qtd': sum(int(float(p.get('qtd', 1) or 1)) for p in produtos),
                            'fonte_dados': 'xlsx',
                        }
                    else:
                        existentes = self.dados_xlsx_global[order_key]['produtos']
                        chaves_existentes = set()
                        for p in existentes:
                            chaves_existentes.add((p.get('codigo', ''), p.get('descricao', ''), p.get('variacao', '')))
                        for p in produtos:
                            chave_p = (p.get('codigo', ''), p.get('descricao', ''), p.get('variacao', ''))
                            if chave_p not in chaves_existentes:
                                existentes.append(p)
                                chaves_existentes.add(chave_p)
                        self.dados_xlsx_global[order_key]['total_itens'] = len(existentes)
                        self.dados_xlsx_global[order_key]['total_qtd'] = sum(
                            int(float(p.get('qtd', 1) or 1)) for p in existentes
                        )
                        self.dados_xlsx_global[order_key]['fonte_dados'] = 'xlsx'
                    if tracking_key:
                        self.dados_xlsx_tracking[tracking_key] = order_key

                count = 0
                if idx_order != -1 and idx_product != -1:
                    # Formato legado: colunas order_sn/tracking_number/product_info
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row is None:
                            continue
                        order_sn = str(row[idx_order] or '').strip() if len(row) > idx_order else ''
                        tracking = str(row[idx_tracking] or '').strip() if idx_tracking >= 0 and len(row) > idx_tracking else ''
                        product_info = str(row[idx_product] or '').strip() if len(row) > idx_product else ''

                        if not order_sn or not product_info:
                            continue

                        produtos = self._parsear_product_info(product_info)
                        if not produtos:
                            continue

                        order_key = self._normalizar_chave_pedido(order_sn)
                        tracking_key = self._normalizar_chave_pedido(tracking)
                        if not order_key:
                            continue

                        _registrar_produtos(order_key, produtos, tracking_key=tracking_key)
                        count += 1
                else:
                    # Formato "Lista de Resumo" do UpSeller (colunas visuais da tabela)
                    idx_col_pedido = -1
                    idx_col_titulo = -1
                    idx_col_sku = -1
                    idx_col_qtd = -1

                    for nome, idx in cabecalho.items():
                        if idx_col_pedido == -1 and (
                            'pedido' in nome or 'rastreio' in nome or 'order' in nome or 'tracking' in nome
                        ):
                            idx_col_pedido = idx
                        if idx_col_titulo == -1 and (
                            'titulo' in nome or 'title' in nome or 'variacao' in nome or 'variation' in nome
                        ):
                            idx_col_titulo = idx
                        if idx_col_sku == -1 and ('sku' == nome or nome.startswith('sku ' ) or nome.startswith('sku(')):
                            idx_col_sku = idx
                        if idx_col_qtd == -1 and ('qtd' in nome or 'quant' in nome or 'qty' in nome):
                            idx_col_qtd = idx

                    if idx_col_pedido == -1 and ws.max_column >= 2:
                        idx_col_pedido = 1  # coluna 2 no layout padrao da lista
                    if idx_col_titulo == -1 and ws.max_column >= 3:
                        idx_col_titulo = 2
                    if idx_col_sku == -1 and ws.max_column >= 4:
                        idx_col_sku = 3
                    if idx_col_qtd == -1 and ws.max_column >= 6:
                        idx_col_qtd = 5

                    if idx_col_pedido == -1 or (idx_col_titulo == -1 and idx_col_sku == -1):
                        wb.close()
                        continue

                    ultimo_order_key = ''
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row is None:
                            continue

                        pedido_raw = str(row[idx_col_pedido] or '').strip() if len(row) > idx_col_pedido else ''
                        titulo_raw = str(row[idx_col_titulo] or '').strip() if idx_col_titulo >= 0 and len(row) > idx_col_titulo else ''
                        sku_raw = str(row[idx_col_sku] or '').strip() if idx_col_sku >= 0 and len(row) > idx_col_sku else ''
                        qtd_raw = row[idx_col_qtd] if idx_col_qtd >= 0 and len(row) > idx_col_qtd else 1

                        chaves = self._extrair_chaves_pedido_texto(pedido_raw)
                        order_key, tracking_key = self._escolher_chave_principal_resumo(chaves)
                        if order_key:
                            ultimo_order_key = order_key
                        else:
                            order_key = ultimo_order_key

                        if not order_key:
                            continue

                        # Linhas de observacoes/notas nao sao produtos.
                        titulo_norm = self._remover_acentos(titulo_raw).lower()
                        if (
                            'notas do comprador' in titulo_norm or
                            'observacoes' in titulo_norm or
                            'internal notes' in titulo_norm or
                            'customer notes' in titulo_norm
                        ):
                            continue

                        linhas_titulo = [ln.strip() for ln in re.split(r'[\r\n]+', titulo_raw) if str(ln).strip()]
                        descricao = linhas_titulo[0] if linhas_titulo else ''
                        variacao = ' / '.join(linhas_titulo[1:]) if len(linhas_titulo) > 1 else ''

                        # Sem info de produto, ignora a linha.
                        if not sku_raw and not descricao and not variacao:
                            continue

                        try:
                            qtd_val = int(float(str(qtd_raw).replace(',', '.')))
                        except Exception:
                            qtd_val = 1
                        if qtd_val <= 0:
                            qtd_val = 1

                        produto = {
                            'codigo': sku_raw,
                            'descricao': descricao,
                            'variacao': variacao,
                            'qtd': str(qtd_val),
                        }
                        _registrar_produtos(order_key, [produto], tracking_key=tracking_key)

                        # Mapeia outras chaves extraidas (UP/MEL/etc) para o mesmo pedido.
                        for chave_extra in chaves:
                            if chave_extra and chave_extra != order_key:
                                self.dados_xlsx_tracking[chave_extra] = order_key

                        count += 1

                wb.close()
                if count > 0:
                    print(f"    {xlsx_nome}: {count} pedidos")

            except Exception as e:
                print(f"    XLSX erro: {xlsx_nome} - {e}")

        print(f"  Total XLSX: {len(self.dados_xlsx_global)} pedidos, {len(self.dados_xlsx_tracking)} trackings")

    def _extrair_pedido_texto(self, texto):
        """Extrai o numero do pedido (order_sn) do texto da etiqueta."""
        # Padrao Shopee: algo como 2602061BMTVXW0 (alfanumerico ~15 chars)
        m = re.search(r'Pedido[:\s]*\n?([A-Z0-9]{12,20})', texto, re.IGNORECASE)
        if m:
            return m.group(1).strip()
        # Padrao order_sn Shopee: YYMMDD + alfanumerico (ex: 260210A88XUUY8)
        # Aparece em linha propria, 12-16 chars, comeca com 6 digitos (data)
        # Nao confundir com chave NFe (44 digitos puros) nem tracking (BR...)
        m = re.search(r'\n(\d{6}[A-Z0-9]{6,10})\n', texto)
        if m:
            return m.group(1).strip()
        return None

    def _extrair_tracking_quadrante(self, texto):
        """Extrai o tracking number (BR...) do texto de um quadrante."""
        m = re.search(r'(BR\w{10,20})', texto)
        return m.group(1) if m else None

    def _buscar_dados_xlsx(self, texto_quadrante):
        """Busca dados do XLSX usando order_sn ou tracking extraidos do texto.
        Retorna (dados_pedido, order_sn) ou (None, None).
        """
        # Consolidar chaves candidatas presentes na etiqueta.
        chaves = []
        order_sn = self._extrair_pedido_texto(texto_quadrante)
        if order_sn:
            k = self._normalizar_chave_pedido(order_sn)
            if k:
                chaves.append(k)

        tracking = self._extrair_tracking_quadrante(texto_quadrante)
        if tracking:
            k = self._normalizar_chave_pedido(tracking)
            if k and k not in chaves:
                chaves.append(k)

        for k in self._extrair_chaves_pedido_texto(texto_quadrante):
            if k not in chaves:
                chaves.append(k)

        # 1) PRIORIDADE: Lista de separacao (modelo mais confiavel para layout novo)
        dados_lista = None
        chave_lista = None

        for chave in chaves:
            if chave in self.dados_lista_global:
                d = dict(self.dados_lista_global[chave])
                d.setdefault('fonte_dados', 'lista_separacao')
                if dados_lista is None:
                    dados_lista = d
                    chave_lista = chave
                else:
                    dados_lista = self._mesclar_dados_produtos(dados_lista, d)

        if dados_lista:
            dados_lista['fonte_dados'] = 'lista_separacao'
            return dados_lista, (chave_lista or (chaves[0] if chaves else ''))

        # 2) Fallback: XLSX
        acumulado = None
        chave_referencia = None
        def _acumular(dados, chave, fonte):
            nonlocal acumulado, chave_referencia
            if not dados:
                return
            d = dict(dados)
            d.setdefault('fonte_dados', fonte)
            if acumulado is None:
                acumulado = d
                chave_referencia = chave
            else:
                acumulado = self._mesclar_dados_produtos(acumulado, d)

        for chave in chaves:
            if chave in self.dados_xlsx_global:
                _acumular(self.dados_xlsx_global[chave], chave, 'xlsx')
            if chave in self.dados_xlsx_tracking:
                order_k = self.dados_xlsx_tracking[chave]
                if order_k in self.dados_xlsx_global:
                    _acumular(self.dados_xlsx_global[order_k], order_k, 'xlsx')
            # Match parcial de tracking
            for tracking_xlsx, osn in self.dados_xlsx_tracking.items():
                if tracking_xlsx.startswith(chave) or chave.startswith(tracking_xlsx):
                    if osn in self.dados_xlsx_global:
                        _acumular(self.dados_xlsx_global[osn], osn, 'xlsx')

        if acumulado:
            acumulado['fonte_dados'] = 'xlsx'
            return acumulado, (chave_referencia or (chaves[0] if chaves else ''))

        return None, None

    def _eh_pagina_lista_separacao(self, texto):
        """Detecta pagina de 'Lista de Separacao' exportada junto das etiquetas."""
        if not texto:
            return False
        up = (texto or "").upper()
        up = up.replace("Ç", "C").replace("Ã", "A").replace("Á", "A").replace("Â", "A")
        up = up.replace("É", "E").replace("Ê", "E").replace("Í", "I").replace("Ó", "O")
        up = up.replace("Ô", "O").replace("Õ", "O").replace("Ú", "U")

        if "LISTA DE SEPARACAO" not in up:
            return False
        if "SKU" not in up:
            return False
        if "TITULO" not in up and "VARIACAO" not in up:
            return False
        if "PEDIDO" not in up:
            return False
        return True

    def _extrair_dados_rodape_por_texto(self, texto_etiqueta):
        """Fallback leve: extrai produtos do texto nativo do rodape antigo.

        Funciona quando o PDF ainda expõe texto selecionavel (sem precisar OCR).
        """
        if not texto_etiqueta:
            return None

        linhas = []
        for ln in str(texto_etiqueta).splitlines():
            norm = re.sub(r'\s+', ' ', str(ln or '')).strip()
            if norm:
                linhas.append(norm)

        if not linhas:
            return None

        sku_line = ''
        for ln in linhas:
            if 'SKU:' in ln.upper():
                sku_line = ln
                break

        start_idx = 0
        if sku_line:
            for i_ln, ln in enumerate(linhas):
                if ln == sku_line:
                    start_idx = i_ln + 1
                    break

        # Capturar linhas de item no formato "1. PRODUTO ... (*1)".
        produtos = []
        idx = start_idx
        while idx < len(linhas):
            ln = linhas[idx]
            # Exigir "1. " para evitar confundir datas/tempos/indices tecnicos.
            m_item = re.match(r'^\s*(\d{1,2})\.\s+(.+?)\s*$', ln)
            if not m_item:
                idx += 1
                continue

            corpo = (m_item.group(2) or '').strip()
            corpo, qtd = self._extrair_qtd_sufixo_item(corpo)

            # Quantidade pode vir na proxima linha: "*1" ou "(*2)"
            if qtd is None and idx + 1 < len(linhas):
                _, qtd_next = self._extrair_qtd_sufixo_item(linhas[idx + 1])
                if qtd_next is not None:
                    qtd = qtd_next
                    idx += 1

            if qtd is None:
                qtd = 1

            if not corpo or self._parece_texto_metadado(corpo):
                idx += 1
                continue
            # Ignorar linhas numericas/tecnicas que nao sao produto.
            if re.match(r'^\d{1,2}/\d{1,2}$', corpo):
                idx += 1
                continue
            if re.match(r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}', corpo):
                idx += 1
                continue

            # Ex: "FH-Caramelo-37/38 (Caramelo,37/38)"
            variacao = ''
            m_par = re.match(r'^(.*?)\s*\((.*?)\)\s*$', corpo)
            if m_par:
                base = m_par.group(1).strip()
                variacao = m_par.group(2).strip()
            else:
                base = corpo

            base = self._corrigir_texto_ocr_produto(base)
            variacao = self._corrigir_texto_ocr_produto(variacao)

            codigo = re.split(r'[-\s]', base, maxsplit=1)[0].strip()
            if not codigo:
                codigo = re.split(r'[-\s]', corpo, maxsplit=1)[0].strip() or 'SEM_SKU'
            codigo = codigo.upper()

            produtos.append({
                'codigo': codigo,
                'descricao': '',
                'variacao': variacao or base,
                'qtd': str(qtd),
            })
            idx += 1

        if not produtos:
            return None

        total_itens = len(produtos)
        m_tot = re.search(r'TOTAL\s*(?:ITEMS|ITENS)\s*[:;]?\s*(\d+)', sku_line.upper()) if sku_line else None
        if m_tot:
            try:
                total_itens = max(total_itens, int(m_tot.group(1)))
            except Exception:
                pass

        total_qtd = 0
        for p in produtos:
            try:
                total_qtd += max(1, int(float(p.get('qtd', 1))))
            except Exception:
                total_qtd += 1

        return {
            'produtos': produtos,
            'total_itens': total_itens,
            'total_qtd': total_qtd,
            'fonte_dados': 'texto_rodape',
        }

    def _get_easyocr_reader(self):
        """Inicializa leitor OCR sob demanda (quando falta XLSX/XML)."""
        if self._easyocr_reader is not None:
            return self._easyocr_reader
        try:
            import easyocr
            self._easyocr_reader = easyocr.Reader(['pt', 'en'], gpu=False, verbose=False)
        except Exception:
            self._easyocr_reader = None
        return self._easyocr_reader

    def _extrair_dados_rodape_por_ocr(self, pagina, clip):
        """Fallback: extrai produto do rodape original da etiqueta via OCR.

        Usado quando nao existe XML/XLSX para montar o novo rodape.
        """
        try:
            import cv2
            import numpy as np
        except Exception:
            return None

        reader = self._get_easyocr_reader()
        if reader is None:
            return None

        try:
            # Renderizar somente a area da etiqueta para OCR.
            pix = pagina.get_pixmap(
                matrix=fitz.Matrix(2.3, 2.3),
                clip=clip,
                alpha=False
            )
            img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, 3)
            h = img.shape[0]

            # Rodape antigo costuma ficar no terco inferior da etiqueta.
            y0 = int(h * 0.68)
            y1 = int(h * 0.92)
            if y1 <= y0 + 10:
                return None
            crop = img[y0:y1, :]
            crop_bgr = cv2.cvtColor(crop, cv2.COLOR_RGB2BGR)

            linhas_raw = reader.readtext(crop_bgr, detail=0, paragraph=False)
            linhas = []
            for t in linhas_raw:
                t_norm = re.sub(r'\s+', ' ', str(t or '')).strip()
                if t_norm:
                    linhas.append(t_norm)

            if not linhas:
                return None

            sku_line = ''
            produto_line = ''
            for ln in linhas:
                up = ln.upper()
                if 'SKU' in up and not sku_line:
                    sku_line = ln
                if re.search(r'^\s*[0-9IVl]+[\.\-)]\s*', ln):
                    if len(ln) > len(produto_line):
                        produto_line = ln

            # fallback: linha com cara de variacao/produto
            if not produto_line:
                for ln in linhas:
                    up = ln.upper()
                    if 'SKU' in up:
                        continue
                    if any(k in up for k in ('DMEVA', 'MINIE', 'PRETO', 'BRANCO', 'AZUL', 'ROSA', 'BEGE')):
                        produto_line = ln
                        break

            if not produto_line:
                return None

            qtd = 1
            up_sku = sku_line.upper() if sku_line else ''
            # tolerante a OCR: Total Itens / Totan Man2 / etc.
            m_qtd = re.search(r'TOT\w*\s*[I1L]T\w*[:;\s]*([0-9]{1,3})', up_sku)
            if m_qtd:
                try:
                    qtd = max(1, int(m_qtd.group(1)))
                except Exception:
                    qtd = 1
            else:
                m_qtd2 = re.search(r'[\*\-xX]\s*([0-9]{1,3})\s*$', produto_line)
                if m_qtd2:
                    try:
                        qtd = max(1, int(m_qtd2.group(1)))
                    except Exception:
                        qtd = 1

            prod_txt = re.sub(r'^\s*[0-9IVl]+[\.\-)]\s*', '', produto_line).strip()
            prod_txt, qtd_inline = self._extrair_qtd_sufixo_item(prod_txt)
            if qtd_inline is not None:
                qtd = max(qtd, qtd_inline)
            if not prod_txt:
                return None

            codigo = prod_txt.split('-', 1)[0].strip()
            if not codigo:
                codigo = prod_txt.split(' ', 1)[0].strip()
            if not codigo:
                codigo = 'SEM_SKU'

            return {
                'produtos': [{
                    'codigo': codigo,
                    'descricao': '',
                    'variacao': self._corrigir_texto_ocr_produto(prod_txt),
                    'qtd': str(qtd),
                }],
                'total_itens': 1,
                'total_qtd': qtd,
                'fonte_dados': 'ocr_rodape',
            }
        except Exception:
            return None

    @staticmethod
    def _levenshtein(a, b):
        """Distancia de Levenshtein pequena para correcao de tokens OCR."""
        if a == b:
            return 0
        if not a:
            return len(b)
        if not b:
            return len(a)
        prev = list(range(len(b) + 1))
        for i, ca in enumerate(a, 1):
            curr = [i]
            for j, cb in enumerate(b, 1):
                curr.append(min(
                    prev[j] + 1,
                    curr[j - 1] + 1,
                    prev[j - 1] + (0 if ca == cb else 1)
                ))
            prev = curr
        return prev[-1]

    def _corrigir_texto_ocr_produto(self, texto):
        """Corrige erros comuns de OCR em descricao/variacao de produto."""
        txt = re.sub(r'\s+', ' ', str(texto or '')).strip()
        if not txt:
            return txt

        # Correcoes diretas mais frequentes observadas nas etiquetas.
        diretas = {
            r'\bplelo\b': 'preto',
            r'\bprelo\b': 'preto',
            r'\bp?reio\b': 'preto',
            r'\bslilch\b': 'stitch',
            r'\bslich\b': 'stitch',
            r'\bstilch\b': 'stitch',
            r'\ba2ul\b': 'azul',
            r'\bazui\b': 'azul',
            r'\bazu1\b': 'azul',
            r'\bminie\b': 'minnie',
            r'\bmmnie\b': 'minnie',
            r'\broial\b': 'royal',
            r'\broval\b': 'royal',
            r'\broyai\b': 'royal',
            r'\broya1\b': 'royal',
            r'\besquerdo\b': 'esquerdo',
        }
        low = txt.lower()
        for patt, rep in diretas.items():
            low = re.sub(patt, rep, low, flags=re.IGNORECASE)

        # Correcao fuzzy por vocabulario (cores / termos de calcado).
        vocab = [
            'preto', 'branco', 'azul', 'rosa', 'royal', 'bege', 'marrom', 'nude', 'caramelo',
            'stitch', 'minnie', 'mickey', 'esquerdo', 'direito',
            'infantil', 'adulto', 'menina', 'menino',
        ]

        def _corrigir_token(tok):
            if len(tok) < 4 or not tok.isalpha():
                return tok
            # Blindagem para cor "royal": evita cair em "rosa" por fuzzy.
            if tok.startswith('roy'):
                if self._levenshtein(tok, 'royal') <= 2:
                    return 'royal'
                return tok
            best = tok
            best_d = 99
            for v in vocab:
                d = self._levenshtein(tok, v)
                if d < best_d:
                    best_d = d
                    best = v
            if best_d <= 2:
                return best
            return tok

        tokens = re.split(r'(\W+)', low)
        tokens = [_corrigir_token(t) for t in tokens]
        corr = ''.join(tokens)

        # Ajustes de formato comuns.
        corr = corr.replace(' ,', ',').replace(' /', '/').replace('/ ', '/')
        corr = re.sub(r'\s+', ' ', corr).strip()

        # Preserva caixa inicial semelhante ao texto original.
        return corr

    def _calcular_clip_conteudo_pagina(self, pagina):
        """Detecta a area util da etiqueta em pagina inteira.

        Nao divide a etiqueta em quadrantes; apenas remove margens vazias
        quando o PDF vem em A4/preview com a etiqueta pequena no canto.
        """
        # 1) Caso especial: etiqueta dentro de imagem grande (preview renderizado)
        clip_img = self._detectar_clip_por_imagem_principal(pagina)
        if clip_img is not None:
            return clip_img

        # 2) Caso especial: PDF de preview (fundo escuro + etiqueta branca)
        clip_preview = self._detectar_clip_preview_escuro(pagina)
        if clip_preview is not None:
            return clip_preview

        # 3) Fallback por objetos da pagina (texto/imagem)
        rects = []

        # Blocos de texto
        try:
            for b in pagina.get_text("blocks"):
                x0, y0, x1, y1 = b[:4]
                if (x1 - x0) > 2 and (y1 - y0) > 2:
                    rects.append((x0, y0, x1, y1))
        except Exception:
            pass

        # Imagens (QR/barcode rasterizados entram aqui)
        try:
            for img in pagina.get_images(full=True):
                xref = img[0]
                for r in pagina.get_image_rects(xref):
                    if r and r.width > 2 and r.height > 2:
                        rects.append((r.x0, r.y0, r.x1, r.y1))
        except Exception:
            pass

        if not rects:
            return pagina.rect

        x0 = min(r[0] for r in rects)
        y0 = min(r[1] for r in rects)
        x1 = max(r[2] for r in rects)
        y1 = max(r[3] for r in rects)

        # Margem de seguranca para nao cortar elementos de borda
        pad_x = max(10, pagina.rect.width * 0.03)
        pad_y = max(12, pagina.rect.height * 0.03)

        clip = fitz.Rect(
            max(0, x0 - pad_x),
            max(0, y0 - pad_y),
            min(pagina.rect.width, x1 + pad_x),
            min(pagina.rect.height, y1 + pad_y),
        )

        # Fallback defensivo
        if clip.width < 30 or clip.height < 30:
            return pagina.rect

        return clip

    def _detectar_clip_por_imagem_principal(self, pagina, retornar_meta=False):
        """Detecta etiqueta dentro de imagem grande incorporada no PDF.

        Quando retornar_meta=True, retorna (clip, meta_imagem).
        meta_imagem contem xref e crop em coordenadas da imagem.
        """
        try:
            imgs = pagina.get_images(full=True)
            if not imgs:
                return (None, None) if retornar_meta else None

            page_rect = pagina.rect
            page_area = page_rect.width * page_rect.height
            melhor_ref = None  # (vis_area, xref, src_w, src_h, rect)

            for img in imgs:
                xref = img[0]
                src_w = int(img[2] or 0)
                src_h = int(img[3] or 0)
                if src_w < 250 or src_h < 250:
                    continue
                try:
                    rects = pagina.get_image_rects(xref)
                except Exception:
                    rects = []
                for r in rects:
                    vis = r & page_rect
                    vis_area = vis.width * vis.height if vis else 0
                    if vis_area < page_area * 0.12:
                        continue
                    if melhor_ref is None or vis_area > melhor_ref[0]:
                        melhor_ref = (vis_area, xref, src_w, src_h, r)

            if melhor_ref is None:
                return (None, None) if retornar_meta else None

            _, xref, src_w, src_h, rect_img = melhor_ref
            doc = pagina.parent
            if doc is None:
                return (None, None) if retornar_meta else None

            pix = fitz.Pixmap(doc, xref)
            if pix.n > 3:
                pix = fitz.Pixmap(fitz.csRGB, pix)
            iw, ih = pix.width, pix.height
            if iw < 20 or ih < 20:
                return (None, None) if retornar_meta else None

            # Downsample simples para reduzir custo em imagens grandes
            step = 2 if (iw * ih) > 300000 else 1
            dw = max(1, iw // step)
            dh = max(1, ih // step)
            total = dw * dh
            data = pix.samples
            thr = 242

            bright = bytearray(total)
            bright_count = 0
            for y in range(dh):
                yy = y * step
                row_off = yy * iw * 3
                off = y * dw
                for x in range(dw):
                    xx = x * step
                    i = row_off + xx * 3
                    r, g, b = data[i:i + 3]
                    if r >= thr and g >= thr and b >= thr:
                        bright[off + x] = 1
                        bright_count += 1

            dark_fraction = 1.0 - (bright_count / max(1, total))
            if dark_fraction < 0.15:
                return (None, None) if retornar_meta else None

            visited = bytearray(total)
            min_bbox_area = int(total * 0.04)
            max_bbox_area = int(total * 0.80)
            melhor_comp = None  # (score, minx, miny, maxx, maxy)

            for idx in range(total):
                if not bright[idx] or visited[idx]:
                    continue

                dq = collections.deque([idx])
                visited[idx] = 1
                area = 0
                minx = maxx = idx % dw
                miny = maxy = idx // dw

                while dq:
                    cur = dq.popleft()
                    area += 1
                    x = cur % dw
                    y = cur // dw
                    if x < minx:
                        minx = x
                    if x > maxx:
                        maxx = x
                    if y < miny:
                        miny = y
                    if y > maxy:
                        maxy = y

                    if x > 0:
                        n = cur - 1
                        if bright[n] and not visited[n]:
                            visited[n] = 1
                            dq.append(n)
                    if x < dw - 1:
                        n = cur + 1
                        if bright[n] and not visited[n]:
                            visited[n] = 1
                            dq.append(n)
                    if y > 0:
                        n = cur - dw
                        if bright[n] and not visited[n]:
                            visited[n] = 1
                            dq.append(n)
                    if y < dh - 1:
                        n = cur + dw
                        if bright[n] and not visited[n]:
                            visited[n] = 1
                            dq.append(n)

                bw = maxx - minx + 1
                bh = maxy - miny + 1
                bbox_area = bw * bh
                if bbox_area < min_bbox_area or bbox_area > max_bbox_area:
                    continue

                bright_fill = area / max(1, bbox_area)
                if bright_fill > 0.88 or bright_fill < 0.40:
                    continue

                width_frac = bw / max(1, dw)
                height_frac = bh / max(1, dh)
                if width_frac < 0.15 or width_frac > 0.85:
                    continue
                if height_frac < 0.20 or height_frac > 0.95:
                    continue

                score = bbox_area * (1.0 - abs(bright_fill - 0.65))
                if melhor_comp is None or score > melhor_comp[0]:
                    melhor_comp = (score, minx, miny, maxx, maxy)

            if melhor_comp is None:
                return (None, None) if retornar_meta else None

            _, minx, miny, maxx, maxy = melhor_comp

            # Converter bbox (imagem) -> bbox na pagina
            ix0 = minx * step
            iy0 = miny * step
            ix1 = min(iw, (maxx + 1) * step)
            iy1 = min(ih, (maxy + 1) * step)

            px0 = rect_img.x0 + (ix0 / max(1, iw)) * rect_img.width
            py0 = rect_img.y0 + (iy0 / max(1, ih)) * rect_img.height
            px1 = rect_img.x0 + (ix1 / max(1, iw)) * rect_img.width
            py1 = rect_img.y0 + (iy1 / max(1, ih)) * rect_img.height

            clip = fitz.Rect(px0, py0, px1, py1) & page_rect

            # margem de seguranca
            pad_x = max(8.0, page_rect.width * 0.02)
            pad_y = max(10.0, page_rect.height * 0.02)
            clip = fitz.Rect(
                max(0.0, clip.x0 - pad_x),
                max(0.0, clip.y0 - pad_y),
                min(page_rect.width, clip.x1 + pad_x),
                min(page_rect.height, clip.y1 + pad_y),
            )

            if clip.width < 30 or clip.height < 30:
                return (None, None) if retornar_meta else None

            if retornar_meta:
                return clip, {
                    "xref": int(xref),
                    "crop_img": (int(ix0), int(iy0), int(ix1), int(iy1)),
                    "img_size": (int(iw), int(ih)),
                }
            return clip
        except Exception:
            return (None, None) if retornar_meta else None

    def _detectar_clip_preview_escuro(self, pagina):
        """Detecta bloco da etiqueta em PDFs de preview com fundo escuro.

        Estrategia:
        - Render reduzido (0.5x) para performance
        - Encontrar componentes conectados de pixels claros
        - Escolher o maior bloco com densidade de branco compativel com etiqueta
          (nao painel quase vazio / nao pagina inteira)
        """
        try:
            scale = 0.5
            inv_scale = 1.0 / scale
            pix = pagina.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
            w, h = pix.width, pix.height
            if w < 20 or h < 20:
                return None

            data = pix.samples
            total = w * h
            thr = 242

            bright = bytearray(total)
            bright_count = 0
            for y in range(h):
                row = data[y * w * 3:(y + 1) * w * 3]
                off = y * w
                for x in range(w):
                    r, g, b = row[x * 3:x * 3 + 3]
                    if r >= thr and g >= thr and b >= thr:
                        bright[off + x] = 1
                        bright_count += 1

            dark_fraction = 1.0 - (bright_count / max(1, total))
            # Se nao houver fundo escuro relevante, nao e caso de preview.
            if dark_fraction < 0.20:
                return None

            visited = bytearray(total)
            min_bbox_area = int(total * 0.02)   # ignora blocos muito pequenos
            max_bbox_area = int(total * 0.70)   # ignora blocos quase pagina inteira
            melhor = None  # (score, minx, miny, maxx, maxy)

            for idx in range(total):
                if not bright[idx] or visited[idx]:
                    continue

                dq = collections.deque([idx])
                visited[idx] = 1
                area = 0
                minx = maxx = idx % w
                miny = maxy = idx // w

                while dq:
                    cur = dq.popleft()
                    area += 1
                    x = cur % w
                    y = cur // w
                    if x < minx:
                        minx = x
                    if x > maxx:
                        maxx = x
                    if y < miny:
                        miny = y
                    if y > maxy:
                        maxy = y

                    if x > 0:
                        n = cur - 1
                        if bright[n] and not visited[n]:
                            visited[n] = 1
                            dq.append(n)
                    if x < w - 1:
                        n = cur + 1
                        if bright[n] and not visited[n]:
                            visited[n] = 1
                            dq.append(n)
                    if y > 0:
                        n = cur - w
                        if bright[n] and not visited[n]:
                            visited[n] = 1
                            dq.append(n)
                    if y < h - 1:
                        n = cur + w
                        if bright[n] and not visited[n]:
                            visited[n] = 1
                            dq.append(n)

                bw = maxx - minx + 1
                bh = maxy - miny + 1
                bbox_area = bw * bh
                if bbox_area < min_bbox_area or bbox_area > max_bbox_area:
                    continue

                width_frac = bw / max(1, w)
                height_frac = bh / max(1, h)
                # Blocos muito largos/altos normalmente sao painel/fundo do preview
                if width_frac > 0.78 or height_frac > 0.95:
                    continue

                bright_fill = area / max(1, bbox_area)
                # Etiqueta costuma ser bloco claro com bastante texto/codigo,
                # portanto o preenchimento branco nao e extremo.
                if bright_fill > 0.84 or bright_fill < 0.30:
                    continue

                # Penalizar componentes colados nas bordas da pagina
                borda = 0
                if minx <= 1:
                    borda += 1
                if miny <= 1:
                    borda += 1
                if maxx >= w - 2:
                    borda += 1
                if maxy >= h - 2:
                    borda += 1

                score = bbox_area * (1.0 - max(0, borda - 1) * 0.25)
                if melhor is None or score > melhor[0]:
                    melhor = (score, minx, miny, maxx, maxy)

            if melhor is None:
                return None

            _, minx, miny, maxx, maxy = melhor

            # Converter para coordenadas da pagina original + margem de seguranca
            pad_x = max(8.0, pagina.rect.width * 0.02)
            pad_y = max(10.0, pagina.rect.height * 0.02)

            x0 = max(0.0, (minx * inv_scale) - pad_x)
            y0 = max(0.0, (miny * inv_scale) - pad_y)
            x1 = min(pagina.rect.width, ((maxx + 1) * inv_scale) + pad_x)
            y1 = min(pagina.rect.height, ((maxy + 1) * inv_scale) + pad_y)

            clip = fitz.Rect(x0, y0, x1, y1)
            if clip.width < 30 or clip.height < 30:
                return None
            return clip
        except Exception:
            return None

    def carregar_pdf_pagina_inteira(self, caminho_pdf, tipo, dados_xlsx=None):
        """Carrega etiquetas de PDF com 1 etiqueta por pagina (pagina inteira).
        tipo: 'retirada' (beka) ou 'cpf' (lanim)
        dados_xlsx: dict order_sn -> dados (apenas para tipo 'cpf')
        Retorna lista de etiquetas no mesmo formato do pipeline existente.
        """
        print(f"  Carregando ({tipo}): {os.path.basename(caminho_pdf)}")
        doc = fitz.open(caminho_pdf)
        etiquetas = []
        caminho_pdf_real = os.path.realpath(caminho_pdf)
        lista_seq = self.dados_lista_seq_por_pdf.get(caminho_pdf_real, [])
        idx_seq_rotulo = 0

        for num_pag in range(len(doc)):
            pagina = doc[num_pag]
            texto = pagina.get_text()
            render_imagem_meta = None

            # Pagina de lista de separacao nao e etiqueta de envio.
            if tipo == 'retirada' and self._eh_pagina_lista_separacao(texto):
                print(f"    Pag {num_pag}: lista de separacao detectada (ignorada no PDF de etiquetas)")
                continue

            if tipo == 'cpf':
                # Auto-crop: detectar bounding box do conteudo real
                # (lanim.pdf e A4 mas conteudo esta no canto superior esquerdo)
                blocks = pagina.get_text("blocks")
                if blocks:
                    x0 = min(b[0] for b in blocks)
                    y0 = min(b[1] for b in blocks)
                    x1 = max(b[2] for b in blocks)
                    y1 = max(b[3] for b in blocks)
                    # Pequena margem de seguranca (2pt)
                    clip = fitz.Rect(
                        max(0, x0 - 2), max(0, y0 - 2),
                        min(pagina.rect.width, x1 + 2), min(pagina.rect.height, y1 + 2)
                    )
                else:
                    clip = pagina.rect
            else:
                # Mantem 1 etiqueta por pagina, com trim automatico de margens vazias
                clip_img, meta_img = self._detectar_clip_por_imagem_principal(
                    pagina, retornar_meta=True
                )
                if clip_img is not None:
                    clip = clip_img
                    render_imagem_meta = meta_img
                else:
                    clip = self._calcular_clip_conteudo_pagina(pagina)

            if tipo == 'retirada':
                # Log de debug quando houver ajuste de margem relevante
                try:
                    pw, ph = pagina.rect.width, pagina.rect.height
                    cw, ch = clip.width, clip.height
                    if cw < (pw * 0.97) or ch < (ph * 0.97):
                        print(
                            f"    Pag {num_pag}: ajuste de margem "
                            f"({int(cw)}x{int(ch)} de {int(pw)}x{int(ph)})"
                        )
                except Exception:
                    pass

            if tipo == 'retirada':
                # Extrair NF do texto
                nf = None
                m = re.search(r'Emiss.o:\n(\d+)\n', texto)
                if m:
                    nf = m.group(1)
                else:
                    m = re.search(r'NF:\s*(\d+)', texto)
                    if m:
                        nf = m.group(1)
                    else:
                        m = re.search(r'(\d{4,6})\n\d\n\d{2}-\d{2}-\d{4}', texto)
                        if m:
                            nf = m.group(1)

                # Classificar tipo real da etiqueta pelo conteudo
                # (mesmo criterio usado no path com recorte de 4-quadrantes).
                # Etiquetas com DANFE SIMPLIFICADO + NF real sao 'cnpj' (regular)
                # e devem ir para gerar_pdf_loja (com ordenacao e mascara de rodape).
                tipo_classificado = self._detectar_tipo_etiqueta(texto, nf_encontrada=nf)

                if nf is None:
                    pdf_id = os.path.splitext(os.path.basename(caminho_pdf))[0].replace(' ', '_')
                    nf = f"SEM_NF_ret_{pdf_id}_p{num_pag}"
                    dados_nf = {}
                    print(f"    Pag {num_pag}: NF nao encontrada, gerando com ID sintetico")
                else:
                    dados_nf = {}

                # FONTE PRIMARIA: XLSX (buscar por order_sn ou tracking)
                if self.dados_xlsx_global or self.dados_lista_global:
                    dados_xlsx_ret, chave_dados = self._buscar_dados_xlsx(texto)
                    if dados_xlsx_ret:
                        origem_dados = dados_xlsx_ret.get('fonte_dados', 'xlsx')
                        dados_nf = {
                            'nf': nf,
                            'serie': '',
                            'data_emissao': '',
                            'chave': self._extrair_chave_nfe(texto),
                            'cnpj_emitente': '',
                            'nome_emitente': '',
                            'produtos': dados_xlsx_ret['produtos'],
                            'total_itens': dados_xlsx_ret['total_itens'],
                            'total_qtd': dados_xlsx_ret['total_qtd'],
                            'fonte_dados': origem_dados,
                        }
                        print(f"    Pag {num_pag}: Retirada usando dados {origem_dados} ({chave_dados})")

                # FALLBACK: lista de separacao sequencial no mesmo PDF.
                if not dados_nf.get('produtos') and idx_seq_rotulo < len(lista_seq):
                    dados_seq = lista_seq[idx_seq_rotulo]
                    if dados_seq and dados_seq.get('produtos'):
                        dados_nf = {
                            'nf': nf,
                            'serie': '',
                            'data_emissao': '',
                            'chave': self._extrair_chave_nfe(texto),
                            'cnpj_emitente': '',
                            'nome_emitente': '',
                            'produtos': dados_seq.get('produtos', []),
                            'total_itens': dados_seq.get('total_itens', 0),
                            'total_qtd': dados_seq.get('total_qtd', 0),
                            'fonte_dados': 'lista_separacao_seq',
                        }
                        print(f"    Pag {num_pag}: Retirada usando lista de separacao sequencial")

                # FALLBACK: texto nativo do rodape antigo (quando disponivel).
                if not dados_nf.get('produtos'):
                    dados_txt = self._extrair_dados_rodape_por_texto(texto)
                    if dados_txt:
                        dados_nf = {
                            'nf': nf,
                            'serie': '',
                            'data_emissao': '',
                            'chave': self._extrair_chave_nfe(texto),
                            'cnpj_emitente': '',
                            'nome_emitente': '',
                            'produtos': dados_txt.get('produtos', []),
                            'total_itens': dados_txt.get('total_itens', 0),
                            'total_qtd': dados_txt.get('total_qtd', 0),
                            'fonte_dados': dados_txt.get('fonte_dados', 'texto_rodape'),
                        }
                        print(f"    Pag {num_pag}: Retirada usando fallback de texto do rodape")

                # FALLBACK final: OCR no proprio rodape da etiqueta quando nao ha XLSX/XML.
                if not dados_nf.get('produtos'):
                    dados_ocr = self._extrair_dados_rodape_por_ocr(pagina, clip)
                    if dados_ocr:
                        dados_nf = {
                            'nf': nf,
                            'serie': '',
                            'data_emissao': '',
                            'chave': self._extrair_chave_nfe(texto),
                            'cnpj_emitente': '',
                            'nome_emitente': '',
                            'produtos': dados_ocr.get('produtos', []),
                            'total_itens': dados_ocr.get('total_itens', 0),
                            'total_qtd': dados_ocr.get('total_qtd', 0),
                            'fonte_dados': dados_ocr.get('fonte_dados', 'ocr_rodape'),
                        }
                        print(f"    Pag {num_pag}: Retirada usando fallback OCR do rodape")

                sku = ''
                num_produtos = 1
                cnpj = dados_nf.get('cnpj_emitente', '')
                if dados_nf.get('produtos'):
                    sku = dados_nf['produtos'][0].get('codigo', '')
                    num_produtos = len(dados_nf['produtos'])

                # Extrair nome da loja do REMETENTE
                if not cnpj:
                    nome_loja = self._extrair_nome_loja_remetente(texto)
                    if not nome_loja:
                        # Fallback para PDF baixado por loja no UpSeller.
                        nome_loja = self._inferir_loja_por_nome_arquivo(caminho_pdf)
                    if nome_loja:
                        cnpj = self._registrar_loja_sintetica(nome_loja, prefixo='LOJA')

                # tipo_especial baseado no conteudo real:
                # - 'cnpj' → None (regular) → gerar_pdf_loja (com ordenacao + mascara rodape)
                # - 'retirada' → 'retirada' → gerar_pdf_cpf
                # - 'cpf' → 'cpf' → gerar_pdf_cpf
                tipo_especial_real = tipo_classificado if tipo_classificado != 'cnpj' else None

                # Extrair tracking (BR...) e order_sn para deduplicacao cross-loja
                _tracking = ''
                _order_sn = ''
                try:
                    _m_track = re.search(r'(BR[0-9A-Z]{10,20}BR|BR[0-9A-Z]{10,20})', texto)
                    if _m_track:
                        _tracking = _m_track.group(1)
                    _m_order = re.search(r'\b(\d{6}[A-Z0-9]{6,12})\b', texto.upper())
                    if _m_order:
                        _order_sn = _m_order.group(1)
                except Exception:
                    pass

                etiquetas.append({
                    'nf': nf,
                    'sku': sku,
                    'num_produtos': num_produtos,
                    'cnpj': cnpj,
                    'clip': clip,
                    'pagina_idx': num_pag,
                    'caminho_pdf': caminho_pdf,
                    'dados_xml': dados_nf,
                    'tipo_especial': tipo_especial_real,
                    'render_imagem_meta': render_imagem_meta,
                    'tracking': _tracking,
                    'order_sn': _order_sn,
                })
                idx_seq_rotulo += 1

            elif tipo == 'cpf':
                # Extrair order_sn do texto
                order_sn = self._extrair_pedido_texto(texto)

                dados_pedido = {}
                if order_sn and dados_xlsx:
                    dados_pedido = dados_xlsx.get(order_sn, {})

                # Extrair nome real da loja do REMETENTE
                nome_loja_cpf = self._extrair_nome_loja_remetente(texto)
                cpf_cnpj = self.LANIM_CNPJ  # fallback
                cpf_nome = self.LANIM_NOME

                if nome_loja_cpf:
                    # Procurar CNPJ real correspondente ao nome da loja
                    cnpj_encontrado = None
                    for cnpj_real, nome_real in self.cnpj_loja.items():
                        if nome_real.lower().strip() == nome_loja_cpf.lower().strip():
                            cnpj_encontrado = cnpj_real
                            break
                    if cnpj_encontrado:
                        cpf_cnpj = cnpj_encontrado
                        cpf_nome = nome_loja_cpf
                    else:
                        # Nome encontrado mas sem CNPJ correspondente
                        cpf_cnpj = f"CPF_{re.sub(r'[^A-Za-z0-9]', '_', nome_loja_cpf)}"
                        cpf_nome = nome_loja_cpf
                        self.cnpj_loja[cpf_cnpj] = nome_loja_cpf
                        self.cnpj_nome[cpf_cnpj] = nome_loja_cpf

                # Registrar mapeamento fallback se necessario
                if cpf_cnpj == self.LANIM_CNPJ and self.LANIM_CNPJ not in self.cnpj_nome:
                    self.cnpj_nome[self.LANIM_CNPJ] = self.LANIM_NOME

                # Criar dados_xml ficticio com dados do XLSX
                produtos = dados_pedido.get('produtos', [])
                sku = produtos[0].get('codigo', '') if produtos else ''
                num_produtos = len(produtos) if produtos else 1

                # Dados simulados (sem chave/NF real)
                dados_ficticio = {
                    'nf': order_sn or f'CPF_pag{num_pag}',
                    'serie': '',
                    'data_emissao': '',
                    'chave': '',  # sem chave = sem barcode
                    'cnpj_emitente': cpf_cnpj,
                    'nome_emitente': cpf_nome,
                    'produtos': produtos,
                    'total_itens': dados_pedido.get('total_itens', len(produtos)),
                    'total_qtd': dados_pedido.get('total_qtd', 0),
                }

                # Extrair tracking para deduplicacao cross-loja
                _cpf_tracking = ''
                try:
                    _m_cpf_track = re.search(r'(BR[0-9A-Z]{10,20}BR|BR[0-9A-Z]{10,20})', texto)
                    if _m_cpf_track:
                        _cpf_tracking = _m_cpf_track.group(1)
                except Exception:
                    pass

                etiquetas.append({
                    'nf': order_sn or f'CPF_pag{num_pag}',
                    'sku': sku,
                    'num_produtos': num_produtos,
                    'cnpj': cpf_cnpj,
                    'clip': clip,
                    'pagina_idx': num_pag,
                    'caminho_pdf': caminho_pdf,
                    'dados_xml': dados_ficticio,
                    'tipo_especial': 'cpf',
                    'tracking': _cpf_tracking,
                    'order_sn': order_sn or '',
                })

        doc.close()
        print(f"    {len(etiquetas)} etiquetas ({tipo})")
        return etiquetas

    def processar_beka(self, pasta_entrada):
        """Processa etiquetas de retirada do comprador (beka.pdf).
        Retorna lista de etiquetas associadas a lojas via XML/CNPJ.
        """
        caminho = os.path.join(pasta_entrada, 'beka.pdf')
        if not os.path.exists(caminho):
            return []

        print(f"\n  Processando etiquetas RETIRADA DO COMPRADOR...")
        etiquetas = self.carregar_pdf_pagina_inteira(caminho, 'retirada')
        return etiquetas

    def processar_cpf(self, pasta_entrada):
        """Processa etiquetas CPF (lanim*.pdf) usando dados de XLSX de declaracao.
        Detecta automaticamente qualquer XLSX com coluna order_sn + product_info.
        Processa todos os PDFs com nome iniciando em 'lanim' (lanim.pdf, lanim 2.pdf, etc.)

        Todas as etiquetas em lanim*.pdf sao CPF.
        Paginas com grid (2x2/2x1) sao recortadas individualmente.
        Paginas com 1 etiqueta sao processadas como pagina inteira.
        Cada etiqueta recebe dados do XLSX pelo order_sn.

        Retorna lista de etiquetas CPF.
        """
        # Detectar todos os PDFs lanim*.pdf
        pdfs_cpf = []
        for f in os.listdir(pasta_entrada):
            if f.lower().startswith('lanim') and f.lower().endswith('.pdf'):
                pdfs_cpf.append(f)

        if not pdfs_cpf:
            return []

        print(f"\n  Processando etiquetas CPF ({len(pdfs_cpf)} PDF(s))...")

        # Buscar XLSX de declaracao: primeiro lanim2.xlsx, depois qualquer XLSX com order_sn
        dados_xlsx = {}
        caminho_xlsx_especifico = os.path.join(pasta_entrada, 'lanim2.xlsx')
        if os.path.exists(caminho_xlsx_especifico):
            dados_xlsx = self.carregar_xlsx_pedidos(caminho_xlsx_especifico)
            print(f"  Usando lanim2.xlsx")
        else:
            # Buscar qualquer XLSX que nao seja planilha de custos ou config
            xlsx_encontrados = []
            for f in os.listdir(pasta_entrada):
                if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('_') and f != 'planilha_custos.xlsx':
                    xlsx_encontrados.append(f)

            for xlsx_nome in xlsx_encontrados:
                caminho_xlsx = os.path.join(pasta_entrada, xlsx_nome)
                try:
                    dados_temp = self.carregar_xlsx_pedidos(caminho_xlsx)
                    if dados_temp:
                        dados_xlsx.update(dados_temp)
                        print(f"  Declaracao encontrada: {xlsx_nome} ({len(dados_temp)} pedidos)")
                except Exception as e:
                    print(f"  Erro ao ler {xlsx_nome}: {e}")

            if not dados_xlsx and not xlsx_encontrados:
                print(f"  AVISO: Nenhum XLSX de declaracao encontrado, etiquetas CPF sem dados de produto")

        etiquetas = []
        for pdf_cpf in sorted(pdfs_cpf):
            caminho_pdf = os.path.join(pasta_entrada, pdf_cpf)
            etqs = self._carregar_pdf_cpf_smart(caminho_pdf, dados_xlsx)
            etiquetas.extend(etqs)
            if len(pdfs_cpf) > 1:
                print(f"  {pdf_cpf}: {len(etqs)} etiquetas CPF")
        return etiquetas

    def _carregar_pdf_cpf_smart(self, caminho_pdf, dados_xlsx=None):
        """Carrega etiquetas CPF de um PDF, detectando layout de cada pagina.
        Paginas com grid (2x2/2x1) sao recortadas por quadrante.
        Paginas com 1 etiqueta sao processadas inteiras (com auto-crop).
        Todas sao marcadas como tipo_especial='cpf' e recebem dados do XLSX.
        """
        print(f"  Carregando (cpf): {os.path.basename(caminho_pdf)}")
        doc = fitz.open(caminho_pdf)
        etiquetas = []

        for num_pag in range(len(doc)):
            pagina = doc[num_pag]

            # Detectar layout da pagina
            # Apenas paginas grandes (A4 ~595x842) podem ter grid 2x2/2x1
            # Paginas pequenas (~297x419) sao sempre 1 etiqueta por pagina
            rect = pagina.rect
            eh_pagina_grande = rect.width > 400 and rect.height > 600

            if eh_pagina_grande:
                quadrantes = self._detectar_layout_pagina(pagina)
            else:
                quadrantes = [pagina.rect]  # pagina inteira = 1 etiqueta

            for idx, clip in enumerate(quadrantes):
                texto_quad = pagina.get_text(clip=clip).strip()
                if len(texto_quad) < 10:
                    continue  # Quadrante vazio

                # Extrair order_sn do texto deste quadrante
                order_sn = self._extrair_pedido_texto(texto_quad)

                dados_pedido = {}
                if order_sn and dados_xlsx:
                    dados_pedido = dados_xlsx.get(order_sn, {})

                # Extrair nome da loja do REMETENTE
                nome_loja_cpf = self._extrair_nome_loja_remetente(texto_quad)
                cpf_cnpj = self.LANIM_CNPJ
                cpf_nome = self.LANIM_NOME

                if nome_loja_cpf:
                    cnpj_encontrado = None
                    for cnpj_real, nome_real in self.cnpj_loja.items():
                        if nome_real.lower().strip() == nome_loja_cpf.lower().strip():
                            cnpj_encontrado = cnpj_real
                            break
                    if cnpj_encontrado:
                        cpf_cnpj = cnpj_encontrado
                        cpf_nome = nome_loja_cpf
                    else:
                        cpf_cnpj = f"CPF_{re.sub(r'[^A-Za-z0-9]', '_', nome_loja_cpf)}"
                        cpf_nome = nome_loja_cpf
                        self.cnpj_loja[cpf_cnpj] = nome_loja_cpf
                        self.cnpj_nome[cpf_cnpj] = nome_loja_cpf

                if cpf_cnpj == self.LANIM_CNPJ and self.LANIM_CNPJ not in self.cnpj_nome:
                    self.cnpj_nome[self.LANIM_CNPJ] = self.LANIM_NOME

                produtos = dados_pedido.get('produtos', [])
                sku = produtos[0].get('codigo', '') if produtos else ''
                num_produtos = len(produtos) if produtos else 1

                nf_id = order_sn or f'CPF_pag{num_pag}_q{idx}'

                dados_ficticio = {
                    'nf': nf_id,
                    'serie': '',
                    'data_emissao': '',
                    'chave': '',
                    'cnpj_emitente': cpf_cnpj,
                    'nome_emitente': cpf_nome,
                    'produtos': produtos,
                    'total_itens': dados_pedido.get('total_itens', len(produtos)),
                    'total_qtd': dados_pedido.get('total_qtd', 0),
                }

                etiquetas.append({
                    'nf': nf_id,
                    'sku': sku,
                    'num_produtos': num_produtos,
                    'cnpj': cpf_cnpj,
                    'clip': clip,
                    'pagina_idx': num_pag,
                    'caminho_pdf': caminho_pdf,
                    'dados_xml': dados_ficticio,
                    'tipo_especial': 'cpf',
                })

        doc.close()
        print(f"    {len(etiquetas)} etiquetas (cpf)")
        return etiquetas

    # ----------------------------------------------------------------
    # GERACAO DO PDF CPF - formato 150x225mm
    # ----------------------------------------------------------------
    ALTURA_CPF_PT = 637.795    # 225mm em pontos

    def _desenhar_secao_produtos_cpf(self, pagina, dados, y_inicio, larg_pagina, alt_pagina=None):
        """Desenha tabela de produtos para etiquetas CPF (usa Variation Name)."""
        preto = (0, 0, 0)
        fonte = "helv"
        fonte_bold = "hebo"
        margem_esq = self.MARGEM_ESQUERDA
        margem_dir = self.MARGEM_DIREITA
        fs = self.fonte_produto
        fs_destaque = int(round(fs * 1.5))  # 50% maior para SKU
        fs_qtd = int(round(fs_destaque * 1.5))  # quantidade 50% maior que destaque
        line_h = fs_qtd + 2

        # Limite inferior da pagina para nao cortar rodape
        y_limite = (alt_pagina or self.ALTURA_CPF_PT) - self.MARGEM_INFERIOR - 5

        order_sn = dados.get('nf', '')
        produtos = self._ordenar_produtos(dados.get('produtos', []))
        total_itens = dados.get('total_itens', len(produtos))
        total_qtd = dados.get('total_qtd', sum(int(float(p.get('qtd', 1))) for p in produtos))

        y = y_inicio

        # Cabecalho do pedido
        header_pedido = f"Pedido: {order_sn}    Total Itens: {total_itens}    Total Quantidade: {total_qtd}"
        pagina.insert_text(
            (margem_esq + 2, y), header_pedido,
            fontsize=fs, fontname=fonte_bold, color=preto
        )
        y += 2

        # Modo de exibicao
        modo = getattr(self, 'exibicao_produto', 'sku')

        # Colunas: depende do modo
        col_sku = margem_esq + 2
        col_var = margem_esq + 50
        col_qtd = larg_pagina - margem_dir - 35

        # Linha superior
        pagina.draw_line(
            (margem_esq, y), (larg_pagina - margem_dir, y),
            color=preto, width=0.8
        )
        y += line_h

        # Cabecalho tabela - depende do modo
        if modo == 'titulo':
            header1, header2 = "PRODUTO", "VARIAÇÃO"
        elif modo == 'ambos':
            header1, header2 = "SKU", "PRODUTO"
        else:
            header1, header2 = "SKU", "VARIAÇÃO"

        pagina.insert_text(
            (col_sku, y), header1,
            fontsize=fs, fontname=fonte_bold, color=preto
        )
        pagina.insert_text(
            (col_var, y), header2,
            fontsize=fs, fontname=fonte_bold, color=preto
        )
        pagina.insert_text(
            (col_qtd, y), "Quant",
            fontsize=fs, fontname=fonte_bold, color=preto
        )
        y += 2
        pagina.draw_line(
            (margem_esq, y), (larg_pagina - margem_dir, y),
            color=preto, width=0.5
        )
        y += line_h

        y_top_tabela = y_inicio + 2

        # Linhas de produtos
        for i_prod, prod in enumerate(produtos[:10]):
            if y + line_h > y_limite:
                break

            codigo = prod.get('codigo', '')
            descricao = prod.get('descricao', '')
            variacao = prod.get('variacao', '')
            qtd = str(int(float(prod.get('qtd', '1'))))

            if modo == 'titulo':
                # Col1: descricao/titulo, Col2: variacao
                texto1 = descricao or codigo
                max_t1 = 10
                if len(texto1) > max_t1:
                    texto1 = texto1[:max_t1 - 2] + '..'
                texto2 = variacao
                max_t2 = 45
                if len(texto2) > max_t2:
                    texto2 = texto2[:max_t2 - 2] + '..'
            elif modo == 'ambos':
                # Col1: SKU, Col2: descricao/titulo
                texto1 = codigo
                max_t1 = 10
                if len(texto1) > max_t1:
                    texto1 = texto1[:max_t1 - 2] + '..'
                texto2 = descricao or variacao
                max_t2 = 45
                if len(texto2) > max_t2:
                    texto2 = texto2[:max_t2 - 2] + '..'
            else:
                # Modo SKU (padrao): Col1: SKU, Col2: variacao
                texto1 = codigo
                max_t1 = 10
                if len(texto1) > max_t1:
                    texto1 = texto1[:max_t1 - 2] + '..'
                texto2 = variacao
                max_t2 = 45
                if len(texto2) > max_t2:
                    texto2 = texto2[:max_t2 - 2] + '..'

            pagina.insert_text(
                (col_sku, y), texto1 if texto1 else "-",
                fontsize=fs_destaque, fontname=fonte_bold, color=preto
            )
            pagina.insert_text(
                (col_var, y), texto2.upper() if texto2 else "-",
                fontsize=fs, fontname=fonte, color=preto
            )
            pagina.insert_text(
                (col_qtd, y), qtd,
                fontsize=fs_qtd, fontname=fonte_bold, color=preto
            )
            y += line_h

            # Linha divisoria entre produtos (exceto apos o ultimo)
            if i_prod < len(produtos) - 1 and y + line_h <= y_limite:
                pagina.draw_line(
                    (margem_esq, y - 1), (larg_pagina - margem_dir, y - 1),
                    color=(0.6, 0.6, 0.6), width=0.3
                )

        # Garantir que linhas nao ultrapassem o limite
        y_final = min(y, y_limite)

        # Linha inferior
        pagina.draw_line(
            (margem_esq, y_final), (larg_pagina - margem_dir, y_final),
            color=preto, width=0.8
        )

        # Linhas verticais
        pagina.draw_line(
            (col_var - 5, y_top_tabela), (col_var - 5, y_final),
            color=preto, width=0.5
        )
        pagina.draw_line(
            (col_qtd - 5, y_top_tabela), (col_qtd - 5, y_final),
            color=preto, width=0.5
        )

    def gerar_pdf_cpf(self, etiquetas_cpf, caminho_saida):
        """Gera PDF para etiquetas CPF/Retirada no formato 150x225mm.
        Renderiza a etiqueta original + tabela de produtos com Variation Name.
        """
        larg = self.LARGURA_PT       # 150mm
        alt = self.ALTURA_CPF_PT     # 225mm
        area_util_larg = larg - self.MARGEM_ESQUERDA - self.MARGEM_DIREITA

        # Ordenar etiquetas (mesma logica de gerar_pdf_loja)
        etiquetas_ord, _, _ = self._ordenar_etiquetas(etiquetas_cpf)

        # Abrir PDFs de origem
        pdfs_usados = set(e['caminho_pdf'] for e in etiquetas_ord)
        docs_abertos = {p: fitz.open(p) for p in pdfs_usados}

        doc_saida = fitz.open()

        for idx, etq in enumerate(etiquetas_ord):
            clip = etq['clip']
            pag_idx = etq['pagina_idx']
            dados = etq.get('dados_xml', {})
            pdf_path = etq['caminho_pdf']
            doc_entrada = docs_abertos[pdf_path]

            nova_pag = doc_saida.new_page(width=larg, height=alt)

            # Escalar a pagina original para caber na largura util
            quad_larg = clip.width
            quad_alt = clip.height
            escala = area_util_larg / quad_larg
            alt_etiqueta = quad_alt * escala

            # Limitar altura da etiqueta para reservar espaco para tabela de produtos
            min_espaco_tabela = 120  # pt, suficiente para cabecalho + alguns produtos
            alt_max_etiqueta = alt - self.MARGEM_TOPO - self.MARGEM_INFERIOR - min_espaco_tabela
            if alt_etiqueta > alt_max_etiqueta:
                alt_etiqueta = alt_max_etiqueta

            dest_rect = fitz.Rect(
                self.MARGEM_ESQUERDA,
                self.MARGEM_TOPO,
                larg - self.MARGEM_DIREITA,
                self.MARGEM_TOPO + alt_etiqueta
            )

            nova_pag.show_pdf_page(dest_rect, doc_entrada, pag_idx, clip=clip)

            # Sem rodape adicional — manter etiqueta como vem do UpSeller

            # Numero de ordem (subido para nao cortar na impressao)
            nova_pag.insert_text(
                (larg - self.MARGEM_DIREITA - 15, alt - self.MARGEM_INFERIOR - 14),
                f"p.{idx + 1}",
                fontsize=9, fontname="hebo", color=(0.4, 0.4, 0.4)
            )

        for doc in docs_abertos.values():
            doc.close()

        total = len(doc_saida)
        doc_saida.save(caminho_saida)
        doc_saida.close()
        return total

    # ----------------------------------------------------------------
    # PROCESSAMENTO SHEIN
    # ----------------------------------------------------------------

    def _parse_shein_danfe(self, texto):
        """Extrai dados de uma pagina DANFE do shein crua.pdf."""
        dados = {}

        # NF
        m = re.search(r'N[uú]mero:\s*\n?(\d+)', texto, re.IGNORECASE)
        dados['nf'] = m.group(1) if m else ''

        # Chave de acesso (44 digitos)
        m = re.search(r'(\d{44})', texto)
        dados['chave'] = m.group(1) if m else ''

        # CNPJ emitente
        m = re.search(r'CNPJ[^:]*:\s*(\d+)', texto)
        dados['cnpj_emitente'] = m.group(1) if m else ''

        # Nome emitente
        m = re.search(r'NOME/RAZ.O SOCIAL[^:]*:\s*(.+)', texto)
        nome_raw = m.group(1).strip() if m else ''
        dados['nome_emitente'] = nome_raw

        # Registrar CNPJ -> nome
        if dados['cnpj_emitente'] and dados['cnpj_emitente'] not in self.cnpj_nome:
            self.cnpj_nome[dados['cnpj_emitente']] = self._limpar_nome_emitente(nome_raw)

        # Produtos: ITEM | CONTEUDO | ATRIBUTOS | QUANT
        produtos = []
        # Pegar tudo apos "ITEM" header
        m_secao = re.search(r'ITEM\s+CONTE.*?QUANT\.\s*\n(.*)', texto, re.DOTALL)
        if m_secao:
            secao = m_secao.group(1).strip()
            linhas = [l.strip() for l in secao.split('\n') if l.strip()]

            # Dividir linhas em blocos por item code (I + alfanumerico, ex: I38cnk94dfzb)
            blocos = []
            bloco_atual = []
            for l in linhas:
                if re.match(r'^[Il][A-Za-z0-9]{6,}$', l) and bloco_atual:
                    blocos.append(bloco_atual)
                    bloco_atual = [l]
                else:
                    bloco_atual.append(l)
            if bloco_atual:
                blocos.append(bloco_atual)

            for bloco in blocos:
                if not bloco:
                    continue
                item_code = bloco[0]
                # Quantidade e a ultima linha que e so um numero
                qtd_str = '1'
                idx_qtd = len(bloco)
                for j in range(len(bloco) - 1, 0, -1):
                    if re.match(r'^\d+$', bloco[j]):
                        qtd_str = bloco[j]
                        idx_qtd = j
                        break

                # Juntar linhas entre item_code e qtd
                meio = ''.join(bloco[1:idx_qtd])

                # Separar descricao de atributos
                # Atributos comecam onde aparece algo com / (tipo Rakka/Roxo...)
                m_atrib = re.search(r'([A-Z][a-z]*/.+)$', meio)
                if m_atrib:
                    atrib = m_atrib.group(1)
                    desc = meio[:m_atrib.start()].strip()
                else:
                    atrib = ''
                    desc = meio

                produtos.append({
                    'codigo_item': item_code,
                    'descricao': desc,
                    'atributos': atrib,
                    'qtd': qtd_str,
                })

        dados['produtos_shein'] = produtos
        dados['total_itens'] = len(produtos)
        dados['total_qtd'] = sum(int(float(p.get('qtd', 1))) for p in produtos)

        return dados

    def _gerar_codigo_shein(self, atributos):
        """Converte ATRIBUTOS Shein em codigo limpo.
        Ex: 'Rakka/Roxo(紫色)-BR41/42' -> 'RakkaRoxoBR4142'
        Ex: 'Rakka/Preto/Dourado-L7(黑/金-L7)-BR41/42' -> 'RakkaPretoDouradoL7BR4142'
        """
        if not atributos:
            return ''
        # Remover texto entre parenteses (incluindo chines)
        limpo = re.sub(r'\([^)]*\)', '', atributos)
        # Remover / e espacos
        limpo = limpo.replace('/', '')
        # Remover espacos
        limpo = limpo.replace(' ', '')
        # Manter apenas letras, numeros e hifen
        limpo = re.sub(r'[^A-Za-z0-9\-]', '', limpo)
        # Remover hifens extras
        limpo = re.sub(r'-+', '', limpo)
        return limpo

    def _parsear_atributos_shein(self, atributos):
        """Parseia atributos Shein em modelo, cor e tamanho.
        Ex: 'Rakka/Dourado(金色)-BR39/40' -> ('Rakka', 'Dourado', 'BR39/40')
        Ex: 'Rakka/Rosa/Rosa(粉色/粉色)-BR39/40' -> ('Rakka', 'Rosa/Rosa', 'BR39/40')
        """
        if not atributos:
            return '', '', ''
        # Remover texto chines entre parenteses
        limpo = re.sub(r'\([^)]*\)', '', atributos).strip()
        # Separar tamanho: -BR39/40
        m = re.match(r'^(.+?)(?:-BR(\d+/?\d*))$', limpo)
        if m:
            modelo_cor = m.group(1)
            tamanho = 'BR' + m.group(2)
        else:
            modelo_cor = limpo
            tamanho = ''
        # Separar modelo e cor pelo primeiro /
        partes = modelo_cor.split('/', 1)
        if len(partes) >= 2:
            modelo = partes[0].strip()
            cor = partes[1].strip()
        else:
            modelo = modelo_cor.strip()
            cor = ''
        return modelo, cor, tamanho

    def processar_shein(self, pasta_entrada, pdfs_extras=None):
        """Processa etiquetas Shein de 'shein crua.pdf', 'shein.pdf' ou PDFs auto-detectados.
        O PDF tem paginas alternadas: etiqueta + DANFE ou etiqueta + Declaracao de Conteudo.
        pdfs_extras: lista de caminhos de PDFs Shein auto-detectados.
        Retorna lista de dicts com dados pareados.
        """
        caminhos_shein = list(pdfs_extras) if pdfs_extras else []

        # Buscar arquivo shein nomeado (case-insensitive para funcionar no Linux/Railway)
        nomes_shein = ['shein crua.pdf', 'shein.pdf']
        for f in os.listdir(pasta_entrada):
            if f.lower() in nomes_shein:
                caminho_fixo = os.path.join(pasta_entrada, f)
                if caminho_fixo not in caminhos_shein:
                    caminhos_shein.append(caminho_fixo)

        if not caminhos_shein:
            return []

        print(f"\n  Processando etiquetas SHEIN ({len(caminhos_shein)} PDF(s))...")
        etiquetas = []
        n_danfe = 0
        n_declaracao = 0
        n_skip = 0

        for caminho in caminhos_shein:
            print(f"    Shein: {os.path.basename(caminho)}")
            doc = fitz.open(caminho)
            n_pags = len(doc)

            # Processar em pares: pag_par (etiqueta) + pag_impar (DANFE ou Declaracao)
            for i in range(0, n_pags - 1, 2):
                pag_etiqueta = doc[i]
                pag_par = doc[i + 1]
                texto_par = pag_par.get_text()
                texto_par_upper = texto_par.upper()

                # Determinar tipo do par
                eh_danfe = 'DANFE' in texto_par_upper and 'CHAVE' in texto_par_upper
                eh_declaracao = (
                    ('DECLARAÇÃO DE CONTEÚDO' in texto_par or 'DECLARACAO DE CONTEUDO' in texto_par_upper)
                    and ('IDENTIFICAÇÃO DOS BENS' in texto_par or 'IDENTIFICACAO DOS BENS' in texto_par_upper)
                )

                if eh_danfe:
                    # --- Pipeline DANFE existente (inalterado) ---
                    dados_danfe = self._parse_shein_danfe(texto_par)
                    nf = dados_danfe.get('nf', '')
                    cnpj = dados_danfe.get('cnpj_emitente', '')

                    if not nf:
                        print(f"      Par {i}/{i+1}: NF nao encontrada no DANFE, pulando")
                        n_skip += 1
                        continue

                    # Buscar dados completos do XML se disponivel
                    dados_xml = self.dados_xml.get(nf, {})
                    if dados_xml:
                        cnpj = dados_xml.get('cnpj_emitente', cnpj)

                    etiquetas.append({
                        'nf': nf,
                        'cnpj': cnpj,
                        'pag_etiqueta_idx': i,
                        'pag_danfe_idx': i + 1,
                        'caminho_pdf': caminho,
                        'clip_etiqueta': pag_etiqueta.rect,
                        'clip_danfe': pag_par.rect,
                        'dados_danfe': dados_danfe,
                        'dados_xml': dados_xml,
                        'tipo_especial': 'shein',
                        'tipo_par': 'danfe',
                    })
                    n_danfe += 1

                elif eh_declaracao:
                    # --- Pipeline Declaracao de Conteudo (novo) ---
                    dados_decl = self._parse_declaracao_conteudo(texto_par)
                    tracking = dados_decl.get('tracking', '')
                    cnpj = dados_decl.get('cnpj_emitente', '')

                    # Detectar area da tabela "IDENTIFICACAO DOS BENS" para recorte
                    clip_tabela = self._detectar_area_tabela_declaracao(pag_par)

                    # Usar tracking como identificador (declaracoes nao tem NF)
                    nf_id = tracking if tracking else f'DECL_{i}_{os.path.basename(caminho)}'

                    etiquetas.append({
                        'nf': nf_id,
                        'cnpj': cnpj,
                        'pag_etiqueta_idx': i,
                        'pag_danfe_idx': i + 1,
                        'caminho_pdf': caminho,
                        'clip_etiqueta': pag_etiqueta.rect,
                        'clip_danfe': pag_par.rect,
                        'clip_tabela_declaracao': clip_tabela,
                        'dados_danfe': dados_decl,
                        'dados_xml': {},
                        'tipo_especial': 'shein',
                        'tipo_par': 'declaracao',
                        'tracking': tracking,
                    })
                    n_declaracao += 1

                else:
                    print(f"      Par {i}/{i+1}: pagina {i+1} nao reconhecida (nem DANFE nem Declaracao), pulando")
                    n_skip += 1

            doc.close()

        total = n_danfe + n_declaracao
        desc_parts = []
        if n_danfe:
            desc_parts.append(f"{n_danfe} DANFE")
        if n_declaracao:
            desc_parts.append(f"{n_declaracao} Declaracao")
        if n_skip:
            desc_parts.append(f"{n_skip} pulados")
        print(f"    {total} pares Shein ({', '.join(desc_parts)})")
        return etiquetas

    def _ordenar_etiquetas_shein(self, etiquetas_shein):
        """Ordena etiquetas Shein: qtd=1 primeiro, qtd>1 ao final.
        Dentro de cada bloco: Modelo > Cor > Tamanho (numero).
        Mesma logica de _ordenar_etiquetas mas adaptada para dados_danfe.
        """
        def _total_qtd(etq):
            return etq.get('dados_danfe', {}).get('total_qtd', 1) or 1

        def _chave(etq):
            prods = etq.get('dados_danfe', {}).get('produtos_shein', [])
            if prods:
                atrib = prods[0].get('atributos', '')
                modelo, cor, tamanho = self._parsear_atributos_shein(atrib)
            else:
                modelo, cor, tamanho = '', '', ''
            m = re.search(r'(\d+)', tamanho)
            num_val = int(m.group(1)) if m else 99999
            return (modelo.casefold(), _total_qtd(etq), cor.casefold(), num_val)

        simples = [e for e in etiquetas_shein if _total_qtd(e) <= 1]
        multiplos = [e for e in etiquetas_shein if _total_qtd(e) > 1]
        simples.sort(key=_chave)
        multiplos.sort(key=_chave)
        return simples + multiplos

    def gerar_pdf_shein(self, etiquetas_shein, caminho_saida):
        """Gera PDF final Shein: etiqueta + barcode vertical + tabela de produtos.
        Suporta dois formatos:
        - DANFE: barcode lateral + tabela de texto gerada
        - Declaracao: etiqueta full-width + tabela recortada da declaracao
        Formato: 150x225mm por pagina.
        """
        # Ordenar: qtd=1 primeiro, por Modelo > Cor > Tamanho
        etiquetas_shein = self._ordenar_etiquetas_shein(etiquetas_shein)

        larg = self.LARGURA_PT       # 150mm = 425.2pt
        alt = self.ALTURA_CPF_PT     # 225mm = 637.8pt
        margem_esq = self.MARGEM_ESQUERDA
        margem_dir = self.MARGEM_DIREITA
        margem_topo = self.MARGEM_TOPO
        margem_inf = self.MARGEM_INFERIOR
        preto = (0, 0, 0)
        fonte = "helv"
        fonte_bold = "hebo"

        # Largura da faixa lateral do barcode vertical (so para DANFE)
        faixa_lateral = 28

        # Abrir PDF de origem
        pdfs_usados = set(e['caminho_pdf'] for e in etiquetas_shein)
        docs_abertos = {p: fitz.open(p) for p in pdfs_usados}

        doc_saida = fitz.open()

        for idx, etq in enumerate(etiquetas_shein):
            pdf_path = etq['caminho_pdf']
            doc_entrada = docs_abertos[pdf_path]
            pag_etiq_idx = etq['pag_etiqueta_idx']
            clip_etiq = etq['clip_etiqueta']
            dados_danfe = etq['dados_danfe']
            dados_xml = etq.get('dados_xml', {})
            tipo_par = etq.get('tipo_par', 'danfe')

            nova_pag = doc_saida.new_page(width=larg, height=alt)
            etiq_larg = clip_etiq.width
            etiq_alt = clip_etiq.height

            if tipo_par == 'declaracao':
                # ============================================================
                # DECLARACAO DE CONTEUDO: etiqueta full-width + tabela recortada
                # ============================================================
                area_util_larg = larg - margem_esq - margem_dir

                # Detectar espaco necessario para tabela
                clip_tabela = etq.get('clip_tabela_declaracao')
                espaco_tabela = 80  # default
                if clip_tabela:
                    tab_h = clip_tabela.height
                    tab_w = clip_tabela.width
                    escala_tab_prev = area_util_larg / tab_w if tab_w > 0 else 1
                    espaco_tabela = max(espaco_tabela, tab_h * escala_tab_prev + 8)

                escala = area_util_larg / etiq_larg
                alt_etiqueta = etiq_alt * escala

                # Limitar altura para caber tabela abaixo
                max_alt_etiq = alt - margem_topo - margem_inf - espaco_tabela
                if alt_etiqueta > max_alt_etiq:
                    escala = max_alt_etiq / etiq_alt
                    alt_etiqueta = max_alt_etiq
                    area_etiq_larg_real = etiq_larg * escala
                else:
                    area_etiq_larg_real = area_util_larg

                dest_rect = fitz.Rect(
                    margem_esq, margem_topo,
                    margem_esq + area_etiq_larg_real, margem_topo + alt_etiqueta
                )
                nova_pag.show_pdf_page(dest_rect, doc_entrada, pag_etiq_idx, clip=clip_etiq)

                # --- Colar tabela "IDENTIFICACAO DOS BENS" recortada ---
                if clip_tabela:
                    y_tabela = margem_topo + alt_etiqueta + 3
                    tab_w = clip_tabela.width
                    tab_h = clip_tabela.height
                    escala_tab = area_util_larg / tab_w if tab_w > 0 else 1
                    tab_h_escalada = tab_h * escala_tab

                    # Garantir que nao excede a pagina
                    espaco_disponivel = alt - y_tabela - margem_inf - 14  # 14pt para p.N
                    if tab_h_escalada > espaco_disponivel:
                        escala_tab = espaco_disponivel / tab_h if tab_h > 0 else 1
                        tab_h_escalada = espaco_disponivel
                        tab_w_escalada = tab_w * escala_tab
                    else:
                        tab_w_escalada = area_util_larg

                    dest_tab = fitz.Rect(
                        margem_esq, y_tabela,
                        margem_esq + tab_w_escalada, y_tabela + tab_h_escalada
                    )
                    # Usar show_pdf_page com clip para copiar a tabela exata
                    pag_decl_idx = etq['pag_danfe_idx']
                    nova_pag.show_pdf_page(dest_tab, doc_entrada, pag_decl_idx, clip=clip_tabela)

            else:
                # ============================================================
                # DANFE: pipeline existente (inalterado) — barcode lateral + tabela texto
                # ============================================================
                chave = dados_xml.get('chave', '') or dados_danfe.get('chave', '')
                nf = etq['nf']
                produtos_shein = dados_danfe.get('produtos_shein', [])
                total_itens = dados_danfe.get('total_itens', len(produtos_shein))
                total_qtd = dados_danfe.get('total_qtd', 0)

                area_etiq_larg = larg - margem_esq - margem_dir - faixa_lateral
                escala = area_etiq_larg / etiq_larg
                alt_etiqueta = etiq_alt * escala

                # Limitar altura para deixar espaco para tabela de produtos
                max_alt_etiq = alt - margem_topo - margem_inf - 60
                if alt_etiqueta > max_alt_etiq:
                    escala = max_alt_etiq / etiq_alt
                    alt_etiqueta = max_alt_etiq
                    area_etiq_larg_real = etiq_larg * escala
                else:
                    area_etiq_larg_real = area_etiq_larg

                dest_rect = fitz.Rect(
                    margem_esq, margem_topo,
                    margem_esq + area_etiq_larg_real, margem_topo + alt_etiqueta
                )
                nova_pag.show_pdf_page(dest_rect, doc_entrada, pag_etiq_idx, clip=clip_etiq)

                # --- Faixa lateral direita: barcode vertical + texto ---
                x_lateral = larg - margem_dir - faixa_lateral
                if chave:
                    nova_pag.insert_text(
                        (larg - margem_dir - 5, margem_topo + 8),
                        "DANFE", fontsize=7, fontname=fonte_bold, color=preto, rotate=270
                    )
                    nova_pag.insert_text(
                        (larg - margem_dir - 13, margem_topo + 8),
                        "Simplificado", fontsize=5, fontname=fonte, color=preto, rotate=270
                    )
                    chave_fmt = ' '.join([chave[i:i+4] for i in range(0, len(chave), 4)])
                    nova_pag.insert_text(
                        (x_lateral + 5, margem_topo + 8),
                        f"Chave: {chave_fmt}",
                        fontsize=4, fontname=fonte, color=preto, rotate=270
                    )
                    try:
                        svg_bytes = self._gerar_barcode_svg(chave)
                        barcode_rect = fitz.Rect(
                            x_lateral + 8, margem_topo + 10,
                            x_lateral + faixa_lateral - 2, margem_topo + alt_etiqueta - 10
                        )
                        nova_pag.insert_image(barcode_rect, stream=svg_bytes, rotate=90)
                    except Exception:
                        pass

                # --- Tabela de produtos no rodape ---
                fs = self.fonte_produto
                line_h = fs + 2
                y = margem_topo + alt_etiqueta + 3
                col_codigo = margem_esq + 2
                col_prod = margem_esq + 95
                col_qtd = larg - margem_dir - 25

                nova_pag.draw_line(
                    (margem_esq, y), (larg - margem_dir, y), color=preto, width=0.8
                )
                y += line_h

                nova_pag.insert_text(
                    (col_codigo, y), "CÓDIGO", fontsize=fs, fontname=fonte_bold, color=preto
                )
                header_prod = f"PROD. (NF: {nf} T-ITENS: {total_itens} T-QUANT: {total_qtd})"
                nova_pag.insert_text(
                    (col_prod, y), header_prod, fontsize=fs, fontname=fonte_bold, color=preto
                )
                nova_pag.insert_text(
                    (col_qtd, y), "Q.", fontsize=fs, fontname=fonte_bold, color=preto
                )
                y += 2
                nova_pag.draw_line(
                    (margem_esq, y), (larg - margem_dir, y), color=preto, width=0.5
                )
                y_top_tabela = margem_topo + alt_etiqueta + 3
                y += line_h

                fs_destaque = int(round(fs * 1.5))
                fs_qtd = int(round(fs_destaque * 1.5))
                for prod in produtos_shein[:5]:
                    atrib = prod.get('atributos', '')
                    modelo, cor, tamanho = self._parsear_atributos_shein(atrib)
                    qtd = str(int(float(prod.get('qtd', '1'))))
                    nova_pag.insert_text(
                        (col_codigo, y), modelo, fontsize=fs_destaque, fontname=fonte_bold, color=preto
                    )
                    cor_tam = f"{cor},{tamanho}" if cor and tamanho else (cor or tamanho or '-')
                    if len(cor_tam) > 30:
                        cor_tam = cor_tam[:28] + '..'
                    nova_pag.insert_text(
                        (col_prod, y), cor_tam, fontsize=fs_destaque, fontname=fonte_bold, color=preto
                    )
                    nova_pag.insert_text(
                        (col_qtd, y), qtd, fontsize=fs_qtd, fontname=fonte_bold, color=preto
                    )
                    y += line_h

                nova_pag.draw_line(
                    (margem_esq, y), (larg - margem_dir, y), color=preto, width=0.8
                )
                nova_pag.draw_line(
                    (col_prod - 5, y_top_tabela), (col_prod - 5, y), color=preto, width=0.5
                )
                nova_pag.draw_line(
                    (col_qtd - 5, y_top_tabela), (col_qtd - 5, y), color=preto, width=0.5
                )

            # Numero de ordem (subido para nao cortar na impressao)
            nova_pag.insert_text(
                (larg - margem_dir - 15, alt - margem_inf - 14),
                f"p.{idx + 1}",
                fontsize=9, fontname="hebo", color=(0.4, 0.4, 0.4)
            )

        for doc in docs_abertos.values():
            doc.close()

        total = len(doc_saida)
        doc_saida.save(caminho_saida)
        doc_saida.close()
        print(f"    Shein PDF: {total} paginas geradas")
        return total

    # ----------------------------------------------------------------
    # GERACAO DO RESUMO XLSX
    # ----------------------------------------------------------------
    def gerar_resumo_xlsx(self, etiquetas, caminho_saida, nome_loja, sku_somente=False):
        """Gera resumo XLSX.

        - Padrao: SKU + Variacao + Quantidade (comportamento legado)
        - sku_somente=True: SKU + Quantidade (fluxo novo da automacao UpSeller)
        """
        # Estilos compartilhados
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Resumo {len(etiquetas)} xmls "

        def _salvar_planilha_e_preview():
            wb.save(caminho_saida)
            try:
                wb.close()
            except Exception:
                pass
            try:
                caminho_img = os.path.splitext(caminho_saida)[0] + '.jpeg'
                self.gerar_imagem_resumo_xlsx(caminho_saida, caminho_img)
            except Exception:
                pass

        if sku_somente:
            sku_qtd = defaultdict(int)
            for etq in etiquetas:
                dados = etq.get('dados_xml', {})
                for prod in dados.get('produtos', []):
                    codigo = (prod.get('codigo', '') or '').strip() or 'SEM_SKU'
                    qtd = int(float(prod.get('qtd', '1')))
                    sku_qtd[codigo] += qtd

            ws['A1'] = 'Cod. SKU'
            ws['B1'] = 'Soma Quant.'
            for cell in [ws['A1'], ws['B1']]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='left')

            row = 2
            for sku in sorted(sku_qtd.keys()):
                ws.cell(row=row, column=1, value=sku).border = border
                ws.cell(row=row, column=2, value=sku_qtd[sku]).border = border
                row += 1

            ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=2, value=sum(sku_qtd.values())).font = Font(bold=True)
            ws.cell(row=row, column=2).border = border

            ws.column_dimensions['A'].width = 28
            ws.column_dimensions['B'].width = 15

            _salvar_planilha_e_preview()
            return len(sku_qtd), sum(sku_qtd.values())

        # Contar quantidade por (SKU, Variacao) - modo legado
        sku_var_qtd = defaultdict(int)
        for etq in etiquetas:
            dados = etq.get('dados_xml', {})
            for prod in dados.get('produtos', []):
                codigo = prod.get('codigo', '')
                variacao = prod.get('variacao', '')
                qtd = int(float(prod.get('qtd', '1')))
                if codigo or variacao:
                    chave = (codigo, variacao)
                    sku_var_qtd[chave] += qtd

        ws['A1'] = 'Cod. SKU'
        ws['B1'] = 'Variacao'
        ws['C1'] = 'Soma Quant.'
        for cell in [ws['A1'], ws['B1'], ws['C1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='left')

        def _sort_var(var):
            partes = re.split(r'[,/]', var or '', maxsplit=1)
            cor = partes[0].strip() if partes else ''
            num_str = partes[1].strip() if len(partes) > 1 else ''
            m = re.search(r'(\d+)', num_str)
            num_val = int(m.group(1)) if m else 99999
            return (cor, num_val, num_str)

        row = 2
        for (sku, var) in sorted(sku_var_qtd.keys(), key=lambda x: (x[0] or '', _sort_var(x[1]))):
            ws.cell(row=row, column=1, value=sku).border = border
            ws.cell(row=row, column=2, value=var).border = border
            ws.cell(row=row, column=3, value=sku_var_qtd[(sku, var)]).border = border
            row += 1

        ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value='').border = border
        ws.cell(row=row, column=3, value=sum(sku_var_qtd.values())).font = Font(bold=True)
        ws.cell(row=row, column=3).border = border

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15

        _salvar_planilha_e_preview()
        return len(sku_var_qtd), sum(sku_var_qtd.values())

    def gerar_imagem_resumo_xlsx(
        self,
        caminho_xlsx,
        caminho_imagem=None,
        max_linhas=2000,
        max_pedidos_por_pagina=100
    ):
        """Gera preview em imagem (JPEG) do resumo XLSX.

        - Divide automaticamente em paginas quando excede 100 pedidos/linhas.
        - Mantem compatibilidade salvando a 1a pagina no caminho base informado.

        Retorna caminho da primeira imagem ou string vazia se falhar.
        """
        try:
            from PIL import Image, ImageDraw, ImageFont
        except Exception:
            return ''

        if not caminho_xlsx or not os.path.exists(caminho_xlsx):
            return ''

        if not caminho_imagem:
            caminho_imagem = os.path.splitext(caminho_xlsx)[0] + '.jpeg'

        try:
            wb = openpyxl.load_workbook(caminho_xlsx, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
        except Exception:
            return ''

        rows = []
        try:
            for row in ws.iter_rows(values_only=True):
                vals = [re.sub(r'\s+', ' ', str(v or '')).strip() for v in row]
                if any(vals):
                    rows.append(vals)
        finally:
            try:
                wb.close()
            except Exception:
                pass

        if not rows:
            return ''

        ncols = max(len(r) for r in rows)
        for r in rows:
            if len(r) < ncols:
                r.extend([''] * (ncols - len(r)))

        # Limite de seguranca contra planilhas gigantes.
        if len(rows) > max_linhas:
            rows = rows[:max_linhas]

        header = rows[0]
        corpo = rows[1:]
        total_row = None
        if corpo and str(corpo[-1][0] or '').strip().upper() == 'TOTAL':
            total_row = corpo.pop()

        # Sempre pelo menos 1 pagina com cabecalho.
        if max_pedidos_por_pagina <= 0:
            max_pedidos_por_pagina = 100
        if not corpo:
            chunks = [[]]
        else:
            chunks = [
                corpo[i:i + max_pedidos_por_pagina]
                for i in range(0, len(corpo), max_pedidos_por_pagina)
            ]

        try:
            font = ImageFont.truetype(r"C:\Windows\Fonts\arial.ttf", 15)
            font_bold = ImageFont.truetype(r"C:\Windows\Fonts\arialbd.ttf", 15)
        except Exception:
            font = ImageFont.load_default()
            font_bold = font

        # Larguras fixadas por amostra global para manter colunas consistentes entre paginas.
        rows_medida = [header] + corpo
        if total_row:
            rows_medida.append(total_row)

        char_caps = [0] * ncols
        for r in rows_medida:
            for i, v in enumerate(r):
                char_caps[i] = max(char_caps[i], min(len(v), 60))

        col_widths = []
        for i, c in enumerate(char_caps):
            base = 70 if i == 0 else 90
            px = max(base, min(560, c * 7 + 20))
            col_widths.append(px)

        row_h = 28
        margin = 2
        img_w = sum(col_widths) + margin * 2 + 1

        cor_borda = (120, 120, 120)
        cor_header = (217, 225, 242)
        cor_total = (226, 239, 218)
        cor_texto = (20, 20, 20)

        def _render_pagina(rows_pagina, caminho_destino):
            img_h = len(rows_pagina) * row_h + margin * 2 + 1
            img = Image.new('RGB', (img_w, img_h), 'white')
            draw = ImageDraw.Draw(img)

            y = margin
            for ridx, r in enumerate(rows_pagina):
                x = margin
                is_header = ridx == 0
                is_total = (str(r[0] or '').strip().upper() == 'TOTAL')
                fill = cor_header if is_header else (cor_total if is_total else (255, 255, 255))
                use_font = font_bold if (is_header or is_total) else font

                for cidx, text in enumerate(r):
                    w = col_widths[cidx]
                    draw.rectangle([(x, y), (x + w, y + row_h)], fill=fill, outline=cor_borda, width=1)

                    txt = str(text or '')
                    max_chars = max(1, int((w - 10) / 7))
                    if len(txt) > max_chars:
                        txt = txt[:max_chars - 2] + '..'

                    tx = x + 4
                    # Alinha quantidade a direita na ultima coluna.
                    if cidx == ncols - 1:
                        try:
                            tw = draw.textlength(txt, font=use_font)
                        except Exception:
                            tw = len(txt) * 7
                        tx = x + w - int(tw) - 6
                    ty = y + 6
                    draw.text((tx, ty), txt, fill=cor_texto, font=use_font)
                    x += w

                y += row_h

            img.save(caminho_destino, format='JPEG', quality=92, optimize=True)

        try:
            os.makedirs(os.path.dirname(caminho_imagem), exist_ok=True)
        except Exception:
            pass

        base, ext = os.path.splitext(caminho_imagem)
        ext = ext or '.jpeg'

        # Limpa paginas antigas para evitar confusao quando reduzir quantidade.
        for antigo in glob.glob(f"{base}_p*{ext}"):
            try:
                os.remove(antigo)
            except Exception:
                pass

        total_paginas = len(chunks)
        primeira_imagem = ''
        for idx, chunk in enumerate(chunks, start=1):
            linhas = [header] + chunk
            if idx == total_paginas and total_row:
                linhas.append(total_row)

            if idx == 1:
                destino = caminho_imagem
                primeira_imagem = destino
            else:
                destino = f"{base}_p{idx:02d}{ext}"

            _render_pagina(linhas, destino)

        # Alias p01 removido — causava envio duplicado via WhatsApp
        # (listar_arquivos_loja pegava tanto o base quanto o _p01)

        return primeira_imagem or caminho_imagem

    def gerar_resumo_xlsx_shein(self, etiquetas_shein, caminho_saida, nome_loja='Shein'):
        """Gera resumo XLSX de etiquetas Shein com Modelo, Cor, Tamanho, Quantidade."""
        # Agrupar por (modelo, cor, tamanho)
        modelo_cor_tam_qtd = defaultdict(int)
        for etq in etiquetas_shein:
            dados_danfe = etq.get('dados_danfe', {})
            for prod in dados_danfe.get('produtos_shein', []):
                atrib = prod.get('atributos', '')
                qtd = int(float(prod.get('qtd', '1')))
                modelo, cor, tamanho = self._parsear_atributos_shein(atrib)
                chave = (modelo or prod.get('descricao', '-'), cor, tamanho)
                modelo_cor_tam_qtd[chave] += qtd

        if not modelo_cor_tam_qtd:
            return 0, 0

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Resumo Shein {len(etiquetas_shein)} etiq"

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws['A1'] = 'Modelo'
        ws['B1'] = 'Cor'
        ws['C1'] = 'Tamanho'
        ws['D1'] = 'Soma Quant.'
        for cell in [ws['A1'], ws['B1'], ws['C1'], ws['D1']]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='left')

        # Ordenar: Modelo > Cor > Tamanho (numerico)
        def _sort_key(chave):
            modelo, cor, tamanho = chave
            m = re.search(r'(\d+)', tamanho)
            num_val = int(m.group(1)) if m else 99999
            return (modelo, cor, num_val, tamanho)

        row = 2
        for (modelo, cor, tamanho) in sorted(modelo_cor_tam_qtd.keys(), key=_sort_key):
            ws.cell(row=row, column=1, value=modelo).border = border
            ws.cell(row=row, column=2, value=cor).border = border
            ws.cell(row=row, column=3, value=tamanho).border = border
            ws.cell(row=row, column=4, value=modelo_cor_tam_qtd[(modelo, cor, tamanho)]).border = border
            row += 1

        # Total
        ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value='').border = border
        ws.cell(row=row, column=3, value='').border = border
        ws.cell(row=row, column=4, value=sum(modelo_cor_tam_qtd.values())).font = Font(bold=True)
        ws.cell(row=row, column=4).border = border

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        wb.save(caminho_saida)
        return len(modelo_cor_tam_qtd), sum(modelo_cor_tam_qtd.values())

    def gerar_resumo_geral_xlsx(self, lojas_info, etiquetas_por_cnpj, caminho_saida):
        """Gera resumo geral XLSX com totais de todas as lojas.
        lojas_info: list of dicts com nome, cnpj, etiquetas, skus, total_qtd
        etiquetas_por_cnpj: dict cnpj -> list de etiquetas
        """
        wb = openpyxl.Workbook()

        # Estilos
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # ---- Sheet 1: Resumo por Loja ----
        ws1 = wb.active
        ws1.title = "Resumo Geral"
        headers = ['Loja', 'Etiquetas', 'SKUs', 'Unidades']
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='left')

        row = 2
        sum_etiq = sum_skus = sum_un = 0
        for loja in sorted(lojas_info, key=lambda x: x['nome']):
            ws1.cell(row=row, column=1, value=loja['nome']).border = border
            ws1.cell(row=row, column=2, value=loja['etiquetas']).border = border
            ws1.cell(row=row, column=3, value=loja['skus']).border = border
            ws1.cell(row=row, column=4, value=loja['total_qtd']).border = border
            sum_etiq += loja['etiquetas']
            sum_skus += loja['skus']
            sum_un += loja['total_qtd']
            row += 1

        # Linha total
        for col, val in enumerate(['TOTAL', sum_etiq, sum_skus, sum_un], 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.font = Font(bold=True, size=11)
            cell.fill = total_fill
            cell.border = border

        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 15

        # ---- Sheet 2: SKUs detalhados por Loja ----
        ws2 = wb.create_sheet("SKUs por Loja")
        headers2 = ['Loja', 'Cod. SKU', 'Quantidade']
        for col, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        row2 = 2
        grand_total = 0
        for cnpj in sorted(etiquetas_por_cnpj.keys(), key=lambda c: self.get_nome_loja(c)):
            nome = self.get_nome_loja(cnpj)
            etiquetas = etiquetas_por_cnpj[cnpj]
            sku_qtd = defaultdict(int)
            for etq in etiquetas:
                dados = etq.get('dados_xml', {})
                for prod in dados.get('produtos', []):
                    codigo = prod.get('codigo', '')
                    qtd = int(float(prod.get('qtd', '1')))
                    if codigo:
                        sku_qtd[codigo] += qtd

            for sku in sorted(sku_qtd.keys()):
                ws2.cell(row=row2, column=1, value=nome).border = border
                ws2.cell(row=row2, column=2, value=sku).border = border
                ws2.cell(row=row2, column=3, value=sku_qtd[sku]).border = border
                grand_total += sku_qtd[sku]
                row2 += 1

        # Total geral
        cell_t1 = ws2.cell(row=row2, column=1, value='TOTAL GERAL')
        cell_t1.font = Font(bold=True)
        cell_t1.fill = total_fill
        cell_t1.border = border
        ws2.cell(row=row2, column=2, value='').border = border
        cell_t3 = ws2.cell(row=row2, column=3, value=grand_total)
        cell_t3.font = Font(bold=True)
        cell_t3.fill = total_fill
        cell_t3.border = border

        ws2.column_dimensions['A'].width = 30
        ws2.column_dimensions['B'].width = 35
        ws2.column_dimensions['C'].width = 15

        wb.save(caminho_saida)
        return len(lojas_info), sum_un


def main():
    """Funcao principal."""

    print("=" * 60)
    print("PROCESSADOR DE ETIQUETAS SHOPEE")
    print("=" * 60)

    pasta_entrada = os.environ.get("PASTA_ENTRADA", os.path.join(os.path.expanduser("~"), "Desktop", "Etiquetas"))
    pasta_saida = os.environ.get("PASTA_SAIDA", os.path.join(os.path.expanduser("~"), "Desktop", "Etiquetas Prontas"))

    if not os.path.exists(pasta_entrada):
        print(f"\nPasta de entrada nao encontrada: {pasta_entrada}")
        return

    # Listar arquivos
    arquivos = os.listdir(pasta_entrada)
    pdfs = [f for f in arquivos if f.lower().endswith('.pdf') and not f.startswith('etiquetas_prontas')]
    zips = [f for f in arquivos if f.lower().endswith('.zip')]

    if not pdfs:
        print(f"\nNenhum PDF encontrado em: {pasta_entrada}")
        return

    print(f"\nArquivos encontrados:")
    for f in pdfs:
        print(f"  [PDF] {f}")
    for f in zips:
        print(f"  [ZIP] {f}")

    proc = ProcessadorEtiquetasShopee()

    # 1. Carregar dados dos XLSX de empacotamento
    print(f"\n{'='*40}")
    print("CARREGANDO XLSX...")
    proc.carregar_todos_xlsx(pasta_entrada)

    # 2. Carregar e recortar TODOS os PDFs
    print(f"\n{'='*40}")
    print("CARREGANDO ETIQUETAS...")
    todas_etiquetas, cpf_auto_detectadas, pdfs_shein_auto = proc.carregar_todos_pdfs(pasta_entrada)

    # 2b. Processar etiquetas especiais (CPF e Shein)
    print(f"\n{'='*40}")
    print("CARREGANDO ETIQUETAS ESPECIAIS...")
    etiquetas_cpf_especial = proc.processar_cpf(pasta_entrada)
    etiquetas_shein = proc.processar_shein(pasta_entrada, pdfs_extras=pdfs_shein_auto)
    # Juntar CPF do lanim*.pdf com CPF auto-detectadas de PDFs genericos
    etiquetas_cpf_especial.extend(cpf_auto_detectadas)
    # CPF e Shein serao gerados com PDFs separados, mas ainda entram no resumo XLSX
    todas_etiquetas.extend(etiquetas_cpf_especial)
    if etiquetas_cpf_especial:
        print(f"  CPF: {len(etiquetas_cpf_especial)} etiquetas ({len(cpf_auto_detectadas)} auto-detectadas)")
    if etiquetas_shein:
        print(f"  Shein: {len(etiquetas_shein)} etiquetas")

    # 2c. Remover duplicatas
    print(f"\n{'='*40}")
    print("VERIFICANDO DUPLICATAS...")
    todas_etiquetas, duplicadas = proc.remover_duplicatas(todas_etiquetas)
    if duplicadas:
        print(f"  AVISO: {len(duplicadas)} etiquetas duplicadas removidas:")
        for d in duplicadas:
            print(f"    NF {d.get('nf','')} (tipo: {d.get('tipo_especial','normal')})")
    else:
        print(f"  Nenhuma duplicata encontrada")

    # 2d. Mostrar nomes das lojas encontradas
    print(f"\n  Lojas identificadas (nomes Shopee):")
    for cnpj in set(e.get('cnpj','') for e in todas_etiquetas if e.get('cnpj')):
        nome = proc.get_nome_loja(cnpj)
        print(f"    [{cnpj}] {nome}")

    # 3. Separar por loja
    print(f"\n{'='*40}")
    print("SEPARANDO POR LOJA...")
    lojas = proc.separar_por_loja(todas_etiquetas)

    # 4. Criar pasta de saida
    if not os.path.exists(pasta_saida):
        os.makedirs(pasta_saida)

    # 5. Gerar PDF e XLSX para cada loja
    print(f"\n{'='*40}")
    print("GERANDO ARQUIVOS POR LOJA...")

    for cnpj, etiquetas_loja in lojas.items():
        nome_loja = proc.get_nome_loja(cnpj)
        n_etiquetas = len(etiquetas_loja)

        print(f"\n  --- {nome_loja} ({n_etiquetas} etiquetas) ---")

        # Criar pasta da loja
        pasta_loja = os.path.join(pasta_saida, nome_loja)
        if not os.path.exists(pasta_loja):
            os.makedirs(pasta_loja)

        # Todas as etiquetas juntas no mesmo PDF (regular, cpf, retirada)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        total_pags = 0
        n_simples = n_multi = com_xml = sem_xml = 0

        if etiquetas_loja:
            caminho_pdf = os.path.join(pasta_loja, f"etiquetas_{nome_loja}_{timestamp}.pdf")
            total_pags, n_simples, n_multi, com_xml, sem_xml = proc.gerar_pdf_loja(
                etiquetas_loja, caminho_pdf
            )
            print(f"    PDF: {total_pags} paginas ({n_simples} simples + {n_multi} multi-produto)")
            if sem_xml > 0:
                print(f"    AVISO: {sem_xml} etiquetas sem XML correspondente")

        # Gerar XLSX (inclui regular + CPF)
        caminho_xlsx = os.path.join(pasta_loja, f"resumo_{nome_loja}_{timestamp}.xlsx")
        n_skus, total_qtd = proc.gerar_resumo_xlsx(etiquetas_loja, caminho_xlsx, nome_loja)
        print(f"    XLSX: {n_skus} SKUs, {total_qtd} unidades vendidas")

    # 5b. Gerar PDF Shein separado (formato 150x225mm com barcode vertical)
    if etiquetas_shein:
        print(f"\n  --- Gerando PDF Shein ---")
        # Agrupar shein por CNPJ
        shein_por_cnpj = defaultdict(list)
        for etq in etiquetas_shein:
            shein_por_cnpj[etq.get('cnpj', '')].append(etq)

        for cnpj, etqs_shein in shein_por_cnpj.items():
            nome_loja_shein = proc.get_nome_loja(cnpj)
            pasta_loja_shein = os.path.join(pasta_saida, nome_loja_shein)
            if not os.path.exists(pasta_loja_shein):
                os.makedirs(pasta_loja_shein)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            caminho_shein = os.path.join(pasta_loja_shein, f"shein_{nome_loja_shein}_{timestamp}.pdf")
            total_shein = proc.gerar_pdf_shein(etqs_shein, caminho_shein)
            print(f"    Shein {nome_loja_shein}: {total_shein} paginas")

    # 6. Resumo final
    print(f"\n{'='*60}")
    print("CONCLUIDO!")
    print(f"  Pasta de saida: {pasta_saida}")
    print(f"  Lojas processadas: {len(lojas)}")
    total_etiquetas = sum(len(e) for e in lojas.values())
    print(f"  Total de etiquetas: {total_etiquetas}")
    for cnpj, etiquetas_loja in lojas.items():
        nome = proc.get_nome_loja(cnpj)
        print(f"    {nome}: {len(etiquetas_loja)} etiquetas")
    print("=" * 60)


if __name__ == "__main__":
    main()
