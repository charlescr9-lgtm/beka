# -*- coding: utf-8 -*-
"""
Funcionarios Routes - Blueprint para gestao de funcionarios e folha de pagamento.
Controle quinzenal de salarios, faltas, vales/parcelas e horas extras.
"""

import io
import os
from datetime import datetime, timezone, timedelta

from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db, Funcionario, FolhaPagamento, ValeParcela

funcionarios_bp = Blueprint('funcionarios', __name__)

_FUSO_BRASILIA = timezone(timedelta(hours=-3))


def _agora_brasil():
    return datetime.now(_FUSO_BRASILIA).replace(tzinfo=None)


def _seed_funcionarios(user_id):
    """Insere funcionarios e vales padrao na primeira utilizacao."""
    empregados = [
        ("EDUARDA", 1800), ("GABRIEL", 2150), ("DANIEL", 2000),
        ("RONEIDE", 3230), ("MARY", 2120),
    ]
    func_map = {}
    for nome, sal in empregados:
        f = Funcionario(user_id=user_id, nome=nome, salario_mensal=sal, ativo=True)
        db.session.add(f)
        db.session.flush()
        func_map[nome] = f.id

    vales = [
        ("GABRIEL", "Emprestimo pessoal", 4000, 400, 10, "1a Quinzena"),
        ("DANIEL", "Moto", 2000, 500, 4, "2a Quinzena"),
        ("DANIEL", "Pintura", 650, 650, 1, "2a Quinzena"),
        ("DANIEL", "Chinelo Slin", 25, 25, 1, "2a Quinzena"),
        ("RONEIDE", "Carro", 9000, 0, 0, "2a Quinzena"),
    ]
    for func_nome, desc, total, parc, n, prox in vales:
        v = ValeParcela(
            user_id=user_id, funcionario_id=func_map[func_nome],
            descricao=desc, valor_total=total, valor_parcela=parc,
            num_parcelas=n, pagas=0, prox_desconto=prox,
        )
        db.session.add(v)

    db.session.commit()


# =============================================
# FUNCIONARIOS CRUD
# =============================================

@funcionarios_bp.route('/api/funcionarios', methods=['GET'])
@jwt_required()
def listar_funcionarios():
    uid = get_jwt_identity()
    funcs = Funcionario.query.filter_by(user_id=uid).all()
    if not funcs:
        _seed_funcionarios(uid)
        funcs = Funcionario.query.filter_by(user_id=uid).all()
    return jsonify([f.to_dict() for f in funcs])


@funcionarios_bp.route('/api/funcionarios', methods=['POST'])
@jwt_required()
def adicionar_funcionario():
    uid = get_jwt_identity()
    data = request.get_json(force=True)
    nome = (data.get("nome") or "").strip()
    salario = float(data.get("salario_mensal", 0))
    if not nome:
        return jsonify({"error": "Nome obrigatorio"}), 400
    f = Funcionario(user_id=uid, nome=nome, salario_mensal=salario, ativo=True)
    db.session.add(f)
    db.session.commit()
    return jsonify(f.to_dict()), 201


@funcionarios_bp.route('/api/funcionarios/<int:fid>', methods=['PUT'])
@jwt_required()
def editar_funcionario(fid):
    uid = get_jwt_identity()
    f = Funcionario.query.filter_by(id=fid, user_id=uid).first()
    if not f:
        return jsonify({"error": "Nao encontrado"}), 404
    data = request.get_json(force=True)
    if "nome" in data:
        f.nome = (data["nome"] or "").strip()
    if "salario_mensal" in data:
        f.salario_mensal = float(data["salario_mensal"])
    if "ativo" in data:
        f.ativo = bool(data["ativo"])
    db.session.commit()
    return jsonify(f.to_dict())


@funcionarios_bp.route('/api/funcionarios/<int:fid>', methods=['DELETE'])
@jwt_required()
def excluir_funcionario(fid):
    uid = get_jwt_identity()
    f = Funcionario.query.filter_by(id=fid, user_id=uid).first()
    if not f:
        return jsonify({"error": "Nao encontrado"}), 404
    # Remover vales vinculados ao funcionario
    ValeParcela.query.filter_by(user_id=uid, funcionario_id=fid).delete()
    db.session.delete(f)
    db.session.commit()
    return jsonify({"ok": True})


# =============================================
# FOLHA DE PAGAMENTO
# =============================================

def _calcular_vales_quinzena(user_id, periodo):
    """Calcula o total de parcelas pendentes por funcionario para o periodo.

    Valores de prox_desconto:
      '1a Quinzena'  -> quinzena 1 (dia 15)
      '2a Quinzena'  -> quinzena 2 (ultimo dia do mes)
      'YYYY-MM'      -> data especifica (2a quinzena daquele mes)

    Retorna dict {funcionario_id: total_desconto}
    """
    import re
    vales = ValeParcela.query.filter_by(user_id=user_id).all()

    partes = periodo.split('-')
    if len(partes) == 3:
        p_ano, p_mes, p_quinz = int(partes[0]), int(partes[1]), int(partes[2])
    else:
        return {}

    totais = {}
    for v in vales:
        # Ignorar quitados ou sem parcela definida
        if (v.num_parcelas or 0) <= 0 or (v.valor_parcela or 0) <= 0:
            continue
        if (v.pagas or 0) >= (v.num_parcelas or 0):
            continue

        prox = (v.prox_desconto or '').strip()
        prox_lower = prox.lower()

        # 1) Data especifica: formato YYYY-MM (ex: '2026-12')
        match_data = re.match(r'^(\d{4})-(\d{2})$', prox)
        if match_data:
            vale_ano, vale_mes = int(match_data.group(1)), int(match_data.group(2))
            aplica = (vale_ano == p_ano and vale_mes == p_mes and p_quinz == 2)
        # 2) Quinzena padrao
        elif '1a quinzena' in prox_lower or '1ª quinzena' in prox_lower or 'dia 15' in prox_lower:
            aplica = (p_quinz == 1)
        elif '2a quinzena' in prox_lower or '2ª quinzena' in prox_lower or 'ultimo' in prox_lower or 'último' in prox_lower:
            aplica = (p_quinz == 2)
        else:
            aplica = True

        if aplica:
            fid = v.funcionario_id
            totais[fid] = totais.get(fid, 0) + (v.valor_parcela or 0)

    return totais


@funcionarios_bp.route('/api/folha', methods=['GET'])
@jwt_required()
def obter_folha():
    uid = get_jwt_identity()
    periodo = request.args.get("periodo", "")
    if not periodo:
        agora = _agora_brasil()
        q = 1 if agora.day <= 15 else 2
        periodo = f"{agora.year}-{agora.month:02d}-{q}"

    funcs = Funcionario.query.filter_by(user_id=uid, ativo=True).order_by(Funcionario.nome).all()
    folhas = FolhaPagamento.query.filter_by(user_id=uid, periodo=periodo).all()
    folha_map = {f.funcionario_id: f for f in folhas}

    # Calcular vales automaticos
    vales_auto = _calcular_vales_quinzena(uid, periodo)

    resultado = []
    for func in funcs:
        vale_auto = round(vales_auto.get(func.id, 0), 2)
        if func.id in folha_map:
            d = folha_map[func.id].to_dict()
            d["vales_auto"] = vale_auto
            resultado.append(d)
        else:
            sal = func.salario_mensal
            quinzena_base = round(sal / 2, 2)
            # Pre-popular vale1 com o desconto automatico dos vales
            a_receber = round(quinzena_base - vale_auto, 2)
            resultado.append({
                "id": None,
                "funcionario_id": func.id,
                "nome": func.nome,
                "salario_mensal": sal,
                "periodo": periodo,
                "faltas": 0, "vale1": vale_auto, "vale2": 0, "horas_extras": 0,
                "quinzena_base": quinzena_base,
                "desc_faltas": 0, "valor_h_extra": 0,
                "a_receber": a_receber,
                "vales_auto": vale_auto,
            })
    return jsonify({"periodo": periodo, "folha": resultado})


@funcionarios_bp.route('/api/folha', methods=['POST'])
@jwt_required()
def salvar_folha():
    uid = get_jwt_identity()
    data = request.get_json(force=True)
    periodo = data.get("periodo", "")
    itens = data.get("itens", [])
    if not periodo or not itens:
        return jsonify({"error": "Periodo e itens obrigatorios"}), 400

    for item in itens:
        fid = item.get("funcionario_id")
        if not fid:
            continue
        folha = FolhaPagamento.query.filter_by(
            user_id=uid, funcionario_id=fid, periodo=periodo
        ).first()
        if folha:
            folha.faltas = float(item.get("faltas", 0))
            folha.vale1 = float(item.get("vale1", 0))
            folha.vale2 = float(item.get("vale2", 0))
            folha.horas_extras = float(item.get("horas_extras", 0))
            folha.updated_at = _agora_brasil()
        else:
            folha = FolhaPagamento(
                user_id=uid, funcionario_id=fid, periodo=periodo,
                faltas=float(item.get("faltas", 0)),
                vale1=float(item.get("vale1", 0)),
                vale2=float(item.get("vale2", 0)),
                horas_extras=float(item.get("horas_extras", 0)),
            )
            db.session.add(folha)
    db.session.commit()
    return jsonify({"ok": True, "periodo": periodo})


@funcionarios_bp.route('/api/folha/export', methods=['GET'])
@jwt_required()
def exportar_folha():
    uid = get_jwt_identity()
    periodo = request.args.get("periodo", "")
    if not periodo:
        return jsonify({"error": "Periodo obrigatorio"}), 400

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl nao instalado"}), 500

    funcs = Funcionario.query.filter_by(user_id=uid, ativo=True).order_by(Funcionario.nome).all()
    folhas = FolhaPagamento.query.filter_by(user_id=uid, periodo=periodo).all()
    folha_map = {f.funcionario_id: f for f in folhas}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Folha de Pagamento"

    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    title_font = Font(name='Arial', bold=True, size=14, color='2F5496')
    data_font = Font(name='Arial', size=10)
    blue_font = Font(name='Arial', size=10, color='0000FF')
    currency_fmt = 'R$ #,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    parts = periodo.split("-")
    q_label = f"{'1a' if parts[2] == '1' else '2a'} Quinz. {parts[1]}/{parts[0]}"

    ws.merge_cells('A1:J1')
    ws['A1'] = 'FOLHA DE PAGAMENTO - BEKA MKT'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A2'] = 'QUINZENA:'
    ws['A2'].font = Font(name='Arial', bold=True, size=10)
    ws['C2'] = q_label
    ws['C2'].font = blue_font

    headers = ['FUNCIONARIO', 'SALARIO MENSAL', 'QUINZENA BASE', 'FALTAS',
               'DESC. FALTAS', 'VALE 1', 'VALE 2', 'H. EXTRAS', 'VALOR H.EXTRA', 'A RECEBER']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border

    row = 5
    for func in funcs:
        sal = func.salario_mensal
        folha = folha_map.get(func.id)
        faltas = folha.faltas if folha else 0
        v1 = folha.vale1 if folha else 0
        v2 = folha.vale2 if folha else 0
        he = folha.horas_extras if folha else 0

        quinzena = sal / 2
        desc = faltas * (sal / 30) * 2
        vh = he * (sal / 30 / 8) * 1.5 if he > 0 else 0
        receber = quinzena - desc - v1 - v2 + vh

        vals = [func.nome, sal, quinzena, faltas, desc, v1, v2, he, vh, receber]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = data_font
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')
            if i >= 2:
                c.number_format = currency_fmt
        row += 1

    widths = [16, 16, 16, 10, 16, 14, 14, 10, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Folha_{periodo.replace('-', '_')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# =============================================
# VALES E PARCELAS
# =============================================

@funcionarios_bp.route('/api/vales', methods=['GET'])
@jwt_required()
def listar_vales():
    uid = get_jwt_identity()
    vales = ValeParcela.query.filter_by(user_id=uid).all()
    return jsonify([v.to_dict() for v in vales])


@funcionarios_bp.route('/api/vales', methods=['POST'])
@jwt_required()
def adicionar_vale():
    uid = get_jwt_identity()
    data = request.get_json(force=True)
    fid = data.get("funcionario_id")
    desc = (data.get("descricao") or "").strip()
    if not fid or not desc:
        return jsonify({"error": "Funcionario e descricao obrigatorios"}), 400
    v = ValeParcela(
        user_id=uid, funcionario_id=fid, descricao=desc,
        valor_total=float(data.get("valor_total", 0)),
        valor_parcela=float(data.get("valor_parcela", 0)),
        num_parcelas=int(data.get("num_parcelas", 0)),
        pagas=int(data.get("pagas", 0)),
        prox_desconto=(data.get("prox_desconto") or ""),
    )
    db.session.add(v)
    db.session.commit()
    return jsonify(v.to_dict()), 201


@funcionarios_bp.route('/api/vales/<int:vid>', methods=['PUT'])
@jwt_required()
def editar_vale(vid):
    uid = get_jwt_identity()
    v = ValeParcela.query.filter_by(id=vid, user_id=uid).first()
    if not v:
        return jsonify({"error": "Nao encontrado"}), 404
    data = request.get_json(force=True)
    for campo in ["descricao", "valor_total", "valor_parcela", "num_parcelas", "pagas", "prox_desconto", "funcionario_id"]:
        if campo in data:
            val = data[campo]
            if campo in ("valor_total", "valor_parcela"):
                val = float(val)
            elif campo in ("num_parcelas", "pagas", "funcionario_id"):
                val = int(val)
            else:
                val = str(val).strip()
            setattr(v, campo, val)
    v.updated_at = _agora_brasil()
    db.session.commit()
    return jsonify(v.to_dict())


@funcionarios_bp.route('/api/vales/<int:vid>', methods=['DELETE'])
@jwt_required()
def remover_vale(vid):
    uid = get_jwt_identity()
    v = ValeParcela.query.filter_by(id=vid, user_id=uid).first()
    if not v:
        return jsonify({"error": "Nao encontrado"}), 404
    db.session.delete(v)
    db.session.commit()
    return jsonify({"ok": True})
